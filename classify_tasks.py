"""
classify_tasks.py
─────────────────────────────────────────────────────────────────────────────
HR As-Is 프로세스 분석 템플릿에서 L5 Task를 읽어
LLM(GPT-4o-mini 기본)으로 AI 수행 가능 여부를 2분류한 뒤
결과를 새 엑셀 파일로 저장합니다.

분류 기준:
  ① AI 수행 가능  — 명확한 규칙·기준이 존재, 시스템이 자동 처리하거나
                    정형 데이터 입력/조회/계산/알림 등으로 구성된 Task
  ② 인간 수행 필요 — 맥락 판단·감성 소통·이해관계자 협의·의사결정·
                    예외 처리 등 인간의 고유 판단이 필요한 Task

사용법:
  python classify_tasks.py
  python classify_tasks.py --file "다른파일.xlsx" --output "결과.xlsx"
  python classify_tasks.py --sheet "시트이름"
  python classify_tasks.py --resume  # 이전 작업 이어서 처리
"""

import os
import sys
import time
import json
import argparse
import re
from pathlib import Path

# ── 의존성 경로 확보 ─────────────────────────────────────────────────────────
_DEPS = ["/tmp/pylibs"]
for _p in _DEPS:
    if _p not in sys.path:
        sys.path.insert(0, _p)

try:
    import openpyxl
    import pandas as pd
except ImportError:
    print("[오류] openpyxl / pandas 가 설치되어 있지 않습니다.")
    print("  pip install openpyxl pandas  를 실행해 주세요.")
    sys.exit(1)

try:
    from openai import OpenAI
except ImportError:
    print("[오류] openai 패키지가 설치되어 있지 않습니다.")
    print("  pip install openai  를 실행해 주세요.")
    sys.exit(1)

from config import (
    OPENAI_API_KEY,
    LLM_MODEL,
    BATCH_SIZE,
    SLEEP_BETWEEN_CALLS,
    OUTPUT_FILENAME,
)

# =============================================================================
# 상수
# =============================================================================

LABEL_AI    = "AI 수행 가능"
LABEL_HUMAN = "인간 수행 필요"
LABELS      = [LABEL_AI, LABEL_HUMAN]

# 엑셀 열 인덱스 (1-based, openpyxl 기준)
COL = {
    "l2_id":   2,   # B
    "l2_name": 3,   # C
    "l3_id":   4,   # D
    "l3_name": 5,   # E
    "l4_id":   6,   # F
    "l4_name": 7,   # G
    "l5_id":   9,   # I
    "l5_name": 10,  # J
    "l5_desc": 11,  # K
    "performer": 12, # L 수행주체
}

DATA_START_ROW = 10  # 실제 데이터 시작 행

# =============================================================================
# 프롬프트 정의
# =============================================================================

SYSTEM_PROMPT = """당신은 HR 업무 프로세스 자동화 전문가입니다.
주어진 HR Task 정보를 바탕으로, 해당 업무가 AI(시스템/자동화)로 수행 가능한지
아니면 인간의 판단이 반드시 필요한지를 분류합니다.

## 분류 기준

### AI 수행 가능 (Rule-based / 자동화 가능)
다음 중 하나 이상 해당하면 "AI 수행 가능":
- 명확한 규칙·조건·기준에 따라 처리되는 작업
- 시스템 데이터 입력, 조회, 계산, 집계
- 자동 발송(메일, 알림), 자동 반영(Interface)
- 체크리스트 확인, 정형화된 서류 검토
- 결재 상신/확정 등 워크플로우 처리
- 정기적·반복적으로 동일한 방식으로 처리되는 작업

### 인간 수행 필요 (Human Judgment Required)
다음 중 하나 이상 해당하면 "인간 수행 필요":
- 맥락·상황 판단이 필요한 의사결정
- 대면 면담, 심리 상담, 구성원 소통·협의
- 이해관계자 조율, 협상, 갈등 해결
- 정성적 검토(사유 적절성, 민감 정보 처리 등)
- 예외 상황 처리 및 창의적 문제 해결
- 보상(안) 수립, 기준 변경 협의, 내부 보고 등 정책 결정

## AI 수행 필요 여건 작성 지침
label이 "AI 수행 가능"인 Task에 대해, AI가 실제로 해당 업무를 수행하기 위해
필요한 전제조건·인프라·데이터 요건을 구체적으로 기술합니다.
- Description, Pain point, Input/Output data, 수행주체 정보를 종합적으로 분석
- 예시: "자회사 의견의 배경·맥락을 파악할 수 있는 히스토리 DB 필요",
        "공동 데이터베이스와 표준 Template 구축 필요",
        "기준정보 시스템 연동 및 실시간 데이터 접근 권한 필요"
- "인간 수행 필요" Task는 빈 문자열("")로 출력

## 출력 형식 (JSON만 출력, 다른 텍스트 없이)
{
  "tasks": [
    {
      "id": "L5_ID",
      "label": "AI 수행 가능" 또는 "인간 수행 필요",
      "reason": "30자 이내 근거",
      "ai_prerequisites": "AI 수행 시 필요한 여건 (AI 수행 가능인 경우만)"
    },
    ...
  ]
}"""


def make_user_prompt(tasks: list[dict]) -> str:
    """tasks 리스트를 받아 분류 요청 프롬프트 생성."""
    lines = ["다음 HR Task들을 분류해 주세요:\n"]
    for t in tasks:
        lines.append(f"[ID: {t['id']}]")
        lines.append(f"  프로세스 계층: {t['l2']} > {t['l3']} > {t['l4']}")
        lines.append(f"  Task명: {t['name']}")
        if t.get("desc"):
            lines.append(f"  설명: {t['desc']}")
        if t.get("performer"):
            lines.append(f"  수행주체/내용: {t['performer'][:120]}")
        lines.append("")
    return "\n".join(lines)


# =============================================================================
# 엑셀 읽기
# =============================================================================

def load_tasks(filepath: str, sheet_name: str | None = None) -> list[dict]:
    """엑셀에서 L5 Task 목록을 추출."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        # 'As-Is' 가 포함된 시트 자동 탐색
        target = None
        for name in wb.sheetnames:
            if "As-Is" in name or "as-is" in name.lower():
                target = name
                break
        ws = wb[target] if target else wb.worksheets[1]

    print(f"[읽기] 시트: '{ws.title}' | 최대 행: {ws.max_row}")

    tasks = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row, values_only=True):
        l5_id = row[COL["l5_id"] - 1]
        if l5_id is None:
            continue

        tasks.append({
            "id":        str(l5_id).strip(),
            "l2":        str(row[COL["l2_name"] - 1] or "").strip(),
            "l3":        str(row[COL["l3_name"] - 1] or "").strip(),
            "l4":        str(row[COL["l4_name"] - 1] or "").strip(),
            "name":      str(row[COL["l5_name"] - 1] or "").strip(),
            "desc":      str(row[COL["l5_desc"] - 1] or "").strip(),
            "performer": str(row[COL["performer"] - 1] or "").strip(),
        })

    print(f"[읽기] 총 {len(tasks)}개 Task 로드 완료")
    return tasks


# =============================================================================
# LLM 분류
# =============================================================================

def classify_batch(client: OpenAI, tasks: list[dict]) -> list[dict]:
    """tasks 배치를 LLM으로 분류. 결과 리스트 반환."""
    prompt = make_user_prompt(tasks)

    response = client.chat.completions.create(
        model=LLM_MODEL,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": prompt},
        ],
        temperature=0.0,
        response_format={"type": "json_object"},
    )

    raw = response.choices[0].message.content
    data = json.loads(raw)
    return data.get("tasks", [])


def parse_fallback(raw: str, task_ids: list[str]) -> list[dict]:
    """JSON 파싱 실패 시 텍스트에서 레이블 추출 시도."""
    results = []
    for tid in task_ids:
        label = LABEL_HUMAN  # 기본값
        if re.search(rf'{re.escape(tid)}.*?AI 수행 가능', raw):
            label = LABEL_AI
        results.append({"id": tid, "label": label, "reason": "파싱 오류 - 기본값 적용"})
    return results


def classify_all(tasks: list[dict], resume_path: str | None = None) -> pd.DataFrame:
    """전체 Task를 배치 단위로 분류하고 DataFrame 반환."""
    client = OpenAI(api_key=OPENAI_API_KEY)

    # 이전 결과 로드 (--resume 옵션)
    done_ids: set[str] = set()
    done_rows: list[dict] = []
    if resume_path and Path(resume_path).exists():
        prev_df = pd.read_csv(resume_path)
        done_ids = set(prev_df["L5_ID"].astype(str))
        done_rows = prev_df.to_dict("records")
        print(f"[재개] 이미 처리된 Task {len(done_ids)}개 스킵")

    remaining = [t for t in tasks if t["id"] not in done_ids]
    total = len(remaining)
    print(f"[분류 시작] 처리할 Task: {total}개 | 모델: {LLM_MODEL} | 배치: {BATCH_SIZE}개\n")

    results = list(done_rows)
    for batch_start in range(0, total, BATCH_SIZE):
        batch = remaining[batch_start: batch_start + BATCH_SIZE]
        batch_num = batch_start // BATCH_SIZE + 1
        total_batches = (total + BATCH_SIZE - 1) // BATCH_SIZE

        print(f"  배치 {batch_num:3d}/{total_batches} | Task {batch_start+1}~{min(batch_start+BATCH_SIZE, total)} 처리 중...", end=" ", flush=True)

        try:
            classified = classify_batch(client, batch)
            # id → result 매핑
            id_to_result = {r["id"]: r for r in classified}

            for t in batch:
                r = id_to_result.get(t["id"], {})
                label = r.get("label", LABEL_HUMAN)
                if label not in LABELS:
                    label = LABEL_HUMAN
                results.append({
                    "L5_ID":          t["id"],
                    "L2":             t["l2"],
                    "L3":             t["l3"],
                    "L4":             t["l4"],
                    "L5_Name":        t["name"],
                    "L5_Description": t["desc"],
                    "수행주체":        t["performer"][:100] if t["performer"] else "",
                    "분류결과":        label,
                    "분류근거":        r.get("reason", ""),
                    "AI수행필요여건":   r.get("ai_prerequisites", "") if label == LABEL_AI else "",
                    "AI수행가능":      1 if label == LABEL_AI else 0,
                })
            print(f"완료")

        except json.JSONDecodeError as e:
            print(f"JSON 파싱 오류 → 기본값 적용")
            for t in batch:
                results.append({
                    "L5_ID": t["id"], "L2": t["l2"], "L3": t["l3"],
                    "L4": t["l4"], "L5_Name": t["name"],
                    "L5_Description": t["desc"],
                    "수행주체": t["performer"][:100],
                    "분류결과": LABEL_HUMAN, "분류근거": f"오류: {e}",
                    "AI수행필요여건": "",
                    "AI수행가능": 0,
                })

        except Exception as e:
            print(f"오류 발생: {e}")
            for t in batch:
                results.append({
                    "L5_ID": t["id"], "L2": t["l2"], "L3": t["l3"],
                    "L4": t["l4"], "L5_Name": t["name"],
                    "L5_Description": t["desc"],
                    "수행주체": t["performer"][:100],
                    "분류결과": LABEL_HUMAN, "분류근거": f"API 오류: {e}",
                    "AI수행필요여건": "",
                    "AI수행가능": 0,
                })

        time.sleep(SLEEP_BETWEEN_CALLS)

    return pd.DataFrame(results)


# =============================================================================
# 결과 저장
# =============================================================================

def save_results(df: pd.DataFrame, output_path: str) -> None:
    """분류 결과를 엑셀로 저장 (요약 시트 포함)."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ── 전체 결과 시트 ──
        df.to_excel(writer, sheet_name="분류결과", index=False)
        ws_main = writer.sheets["분류결과"]

        # 열 너비 자동 조정
        col_widths = {
            "A": 12, "B": 10, "C": 14, "D": 14,
            "E": 28, "F": 45, "G": 35, "H": 16, "I": 35, "J": 50, "K": 8,
        }
        for col_letter, width in col_widths.items():
            ws_main.column_dimensions[col_letter].width = width

        # 헤더 스타일
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill("solid", fgColor="1F4E79")
        ai_fill     = PatternFill("solid", fgColor="E2EFDA")   # 연두
        human_fill  = PatternFill("solid", fgColor="FCE4D6")   # 연주황

        for cell in ws_main[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # 분류결과 셀 색상
        label_col_idx = df.columns.get_loc("분류결과") + 1
        for row_idx in range(2, len(df) + 2):
            cell = ws_main.cell(row=row_idx, column=label_col_idx)
            if cell.value == LABEL_AI:
                cell.fill = ai_fill
            elif cell.value == LABEL_HUMAN:
                cell.fill = human_fill

        # ── 요약 시트 ──
        total       = len(df)
        ai_count    = int(df["AI수행가능"].sum())
        human_count = total - ai_count

        summary_data = {
            "항목": ["전체 Task 수", "AI 수행 가능", "인간 수행 필요",
                    "AI 수행 가능 비율", "인간 수행 필요 비율"],
            "값":   [total, ai_count, human_count,
                    f"{ai_count/total*100:.1f}%",
                    f"{human_count/total*100:.1f}%"],
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="요약", index=False)

        ws_sum = writer.sheets["요약"]
        ws_sum.column_dimensions["A"].width = 22
        ws_sum.column_dimensions["B"].width = 14

        # L3별 요약
        l3_summary = (
            df.groupby("L3")
            .agg(전체=("L5_ID", "count"), AI수행가능=("AI수행가능", "sum"))
            .assign(인간수행필요=lambda x: x["전체"] - x["AI수행가능"])
            .assign(AI비율=lambda x: (x["AI수행가능"] / x["전체"] * 100).round(1))
            .sort_values("AI수행가능", ascending=False)
            .reset_index()
        )
        l3_summary.to_excel(writer, sheet_name="L3별요약", index=False)
        ws_l3 = writer.sheets["L3별요약"]
        ws_l3.column_dimensions["A"].width = 22

    print(f"\n[저장 완료] {output_path}")
    print(f"  • 전체: {total}개")
    print(f"  • AI 수행 가능:   {ai_count}개 ({ai_count/total*100:.1f}%)")
    print(f"  • 인간 수행 필요: {human_count}개 ({human_count/total*100:.1f}%)")


# =============================================================================
# 진행 상황 중간 저장
# =============================================================================

def save_checkpoint(df: pd.DataFrame, checkpoint_path: str = "checkpoint.csv") -> None:
    df.to_csv(checkpoint_path, index=False, encoding="utf-8-sig")


# =============================================================================
# 메인
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="HR Task AI/인간 분류기")
    parser.add_argument(
        "--file",
        default="HR_AsIs_Template.xlsx",
        help="입력 엑셀 파일명 (생략 시 폴더 내 .xlsx 자동 탐색)",
    )
    parser.add_argument("--sheet", default=None, help="시트명 (생략 시 자동 탐색)")
    parser.add_argument("--output", default=OUTPUT_FILENAME, help="출력 파일명")
    parser.add_argument("--resume", action="store_true", help="checkpoint.csv에서 이어서 처리")
    parser.add_argument("--dry-run", action="store_true", help="처음 5개만 테스트 실행")
    args = parser.parse_args()

    # ── API 키 확인 ───────────────────────────────────────────────────────────
    if OPENAI_API_KEY in ("여기에_API_키_입력", "", None):
        print("[오류] config.py 의 OPENAI_API_KEY를 설정하거나")
        print("       환경변수 OPENAI_API_KEY를 export 해주세요.")
        sys.exit(1)

    # ── 파일 경로 확인 ────────────────────────────────────────────────────────
    base_dir = Path(__file__).parent
    input_path = base_dir / args.file
    if not input_path.exists():
        # 폴더에서 xlsx 자동 탐색
        xlsx_files = list(base_dir.glob("*.xlsx"))
        if not xlsx_files:
            print(f"[오류] 엑셀 파일을 찾을 수 없습니다: {input_path}")
            sys.exit(1)
        input_path = xlsx_files[0]
        print(f"[자동 탐색] 파일: {input_path.name}")

    output_path = str(base_dir / args.output)
    checkpoint_path = str(base_dir / "checkpoint.csv")

    # ── 데이터 로드 ───────────────────────────────────────────────────────────
    tasks = load_tasks(str(input_path), args.sheet)

    if args.dry_run:
        tasks = tasks[:5]
        print(f"[Dry-run] 처음 {len(tasks)}개 Task만 처리합니다.\n")

    # ── 분류 실행 ─────────────────────────────────────────────────────────────
    resume_file = checkpoint_path if args.resume else None
    df = classify_all(tasks, resume_path=resume_file)

    # 중간 저장
    save_checkpoint(df, checkpoint_path)

    # ── 결과 저장 ─────────────────────────────────────────────────────────────
    save_results(df, output_path)


if __name__ == "__main__":
    main()
