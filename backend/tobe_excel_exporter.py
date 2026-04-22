"""
tobe_excel_exporter.py — To-Be 설계 결과를 As-Is 템플릿 포맷 Excel로 반출

생성 규칙:
  - 사용자 업로드 As-Is 템플릿 xlsx를 베이스로 복사
  - 데이터 행(row 10+)은 기존 내용 제거 후 To-Be L5 Task로 재작성
  - 각 행 = 하나의 To-Be L5 Task:
      · Junior AI task (AI 파트) — performer = agent 이름, cls_final_label = 원본 분류 유지
      · HR 검토 노드 (AI+Human의 Human 파트) — performer = HR 담당자
      · 보존된 As-Is Human L5 — 그대로
"""
from __future__ import annotations
from pathlib import Path
from typing import Any
from copy import copy as _copy
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side

# 데이터 시작 행 (헤더 4~9, 10행은 템플릿 예시/여백으로 비움, 실제 데이터는 11행부터)
_DATA_START_ROW = 11

# As-Is 템플릿의 열 레이아웃 (1-based)
_COL = {
    "l2_id":    2,   # B
    "l2_name":  3,   # C
    "l3_id":    4,   # D
    "l3_name":  5,   # E
    "l4_id":    6,   # F
    "l4_name":  7,   # G
    "l4_desc":  8,   # H
    "l5_id":    9,   # I
    "l5_name":  10,  # J
    "l5_desc":  11,  # K
    "performer":              12,  # L (자유 텍스트)
    "performer_executive":    13,  # M ●
    "performer_hr":           14,  # N ●
    "performer_manager":      15,  # O ●
    "performer_member":       16,  # P ●
    "cls_1st_label":  34,  # AH
    "cls_1st_knockout": 35,  # AI
    "cls_1st_reason": 36,  # AJ
    "cls_1st_ai_prereq": 37, # AK
    "cls_doosan_label": 38,  # AL
    "cls_doosan_feedback": 39, # AM
    "cls_final_label": 40,  # AN
    "cls_final_feedback": 41,  # AO
}


def _find_data_sheet(wb) -> Any | None:
    """As-Is 데이터가 들어있는 메인 시트를 찾아 반환."""
    for sn in wb.sheetnames:
        # 우선 '3. As-Is 분석' 같은 정규 시트 우선
        if "As-Is" in sn or "as-is" in sn.lower() or "분석" in sn:
            return wb[sn]
    # fallback: 가이드 제외 첫 시트
    for sn in wb.sheetnames:
        if "가이드" in sn or "guide" in sn.lower() or "count" in sn.lower():
            continue
        return wb[sn]
    return None


def _clear_data_rows(ws) -> None:
    """10행(템플릿 예시) + 실제 데이터 행 전부 비움 (스타일 유지)."""
    max_row = ws.max_row
    # 10행부터 끝까지 비움 (실제 데이터는 11행부터 다시 쓰기 시작)
    start_clear = 10
    if max_row < start_clear:
        return
    for r in range(start_clear, max_row + 1):
        for c in range(1, min(ws.max_column + 1, 100)):
            cell = ws.cell(r, c)
            if cell.value is not None and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None


def _get_task_row(task: dict) -> dict[str, str]:
    """To-Be task 한 개(dict)를 엑셀 행 dict로 매핑."""
    tid = str(task.get("task_id") or "").strip()
    tparts = tid.split(".") if tid else []

    l2_id = ".".join(tparts[:1]) if len(tparts) >= 1 else task.get("l2_id", "")
    l3_id = ".".join(tparts[:2]) if len(tparts) >= 2 else task.get("l3_id", "")
    l4_id = ".".join(tparts[:3]) if len(tparts) >= 3 else task.get("l4_id", "")
    l5_id = tid

    return {
        "l2_id": l2_id,
        "l2_name": task.get("l2", ""),
        "l3_id": l3_id,
        "l3_name": task.get("l3", ""),
        "l4_id": l4_id,
        "l4_name": task.get("l4", ""),
        "l5_id": l5_id,
        "l5_name": task.get("task_name", ""),
        "l5_desc": task.get("ai_role", "") or task.get("human_role", ""),
    }


def _write_row(ws, row: int, values: dict[str, str | None]) -> None:
    """dict의 (key, value) → 정의된 열에 value를 씀."""
    for key, val in values.items():
        col = _COL.get(key)
        if col is None:
            continue
        try:
            ws.cell(row, col).value = val if val is not None else ""
        except Exception:
            # merged cell일 수 있음 — 무시
            pass


def _derive_scope(wf_tobe_flow_cache: dict | None, step2_cache: dict) -> tuple[set[str], set[str]]:
    """
    현재 To-Be 설계 스코프(타겟 L4 ID·이름)를 추출.
    우선순위:
      1) wf_tobe_flow_cache.tobe_sheets — Swim Lane 생성 시 고정된 스코프
      2) step2_cache 에이전트의 assigned_tasks.l4 / task_id 접두사
    """
    l4_ids: set[str] = set()
    l4_names: set[str] = set()

    if wf_tobe_flow_cache:
        for s in wf_tobe_flow_cache.get("tobe_sheets", []):
            sid = str(s.get("l4_id") or "").strip()
            sname = str(s.get("l4_name") or "").strip()
            # sheet id가 실제 L4 ID(1.1.3)인 경우 추가
            parts = sid.split(".")
            if len(parts) == 3 and all(p.isdigit() for p in parts):
                l4_ids.add(sid)
            if sname:
                l4_names.add(sname)

    # 보강: step2_cache의 task_id prefix + l4 이름
    for agent in step2_cache.get("agents", []):
        for t in agent.get("assigned_tasks", []):
            tid = str(t.get("task_id") or "")
            parts = tid.split(".")
            if len(parts) >= 3:
                l4_ids.add(".".join(parts[:3]))
            l4_text = str(t.get("l4") or "").strip()
            if l4_text:
                l4_names.add(l4_text)

    return l4_ids, l4_names


def _task_in_scope(tid: str, l4_hint: str, scope_ids: set[str], scope_names: set[str]) -> bool:
    """주어진 task_id / l4 이름이 스코프 안에 있는지."""
    if not scope_ids and not scope_names:
        return True  # 스코프 정보 없으면 모두 포함 (fallback)
    if tid:
        parts = tid.split(".")
        if len(parts) >= 3 and ".".join(parts[:3]) in scope_ids:
            return True
    if l4_hint:
        for tn in scope_names:
            if tn and (tn in l4_hint or l4_hint in tn):
                return True
    return False


def export_tobe_excel(
    template_path: str | Path,
    output_path: str | Path,
    step2_cache: dict,
    classification: dict,
    excel_tasks: list,
    wf_tobe_flow_cache: dict | None = None,
) -> None:
    """
    To-Be Excel 반출 메인 엔트리.

    스코프: wf_tobe_flow_cache에 설정된 L4 만 포함 (사용자가 작업한 시트 범위).
    performer: AI+Human의 Human 파트와 Human L5는 원본 As-Is Task의 performer_* 필드 그대로 복사.

    Args:
        template_path: 사용자가 업로드한 As-Is 템플릿 xlsx 경로
        output_path: 반출할 xlsx 경로
        step2_cache: _wf_step2_cache (agents + assigned_tasks)
        classification: _wf_classification (엑셀 분류 + hybrid_note)
        excel_tasks: _wf_excel_tasks (원본 Task 리스트)
        wf_tobe_flow_cache: 생성된 To-Be Flow dict (스코프 확정용)
    """
    template_path = Path(template_path)
    output_path = Path(output_path)
    if not template_path.exists():
        raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = _find_data_sheet(wb)
    if ws is None:
        raise RuntimeError("As-Is 데이터 시트를 템플릿에서 찾지 못했습니다.")

    _clear_data_rows(ws)

    # 스코프 추출 (현재 작업 중인 L4)
    scope_l4_ids, scope_l4_names = _derive_scope(wf_tobe_flow_cache, step2_cache)

    # 엑셀 task를 id 기준 인덱스 (원본 performer 필드 참조)
    task_by_id: dict[str, Any] = {t.id: t for t in (excel_tasks or [])}

    # ── NEW_xxx 같은 임시 ID 를 실제 L5 순번(예: 2.1.4.16)으로 치환하는 카운터 ──
    # L4 별 현재 최대 L5 seq 추적 (excel_tasks 에서 파싱)
    l4_max_seq: dict[str, int] = {}
    for et in (excel_tasks or []):
        if et.id:
            parts = et.id.split(".")
            if len(parts) >= 4:
                l4 = ".".join(parts[:3])
                try:
                    seq = int(parts[3])
                    l4_max_seq[l4] = max(l4_max_seq.get(l4, 0), seq)
                except ValueError:
                    pass

    def _resolve_l5_id(raw_tid: str, l4_hint: str) -> str:
        """
        raw_tid 가 'NEW_xxx' / 빈 값 / 숫자 포맷 아닐 때 L4 기준 다음 순번 할당.
        유효한 숫자 포맷이면 그대로 반환.
        """
        tid = (raw_tid or "").strip()
        # 이미 올바른 L5 숫자 포맷이면 그대로
        if tid:
            parts = tid.split(".")
            if len(parts) == 4 and all(p.replace(' ', '').isdigit() for p in parts):
                return tid
        # L4 id 추정: scope 에서 추정 or l4_hint 에서 찾기
        l4_id = ""
        if l4_hint:
            # l4_hint 가 이름이면 excel_tasks 에서 매칭되는 l4_id 찾기
            for et in (excel_tasks or []):
                if et.l4 and et.l4.strip() == l4_hint.strip():
                    l4_id = et.l4_id or ""
                    break
        # scope 에서 primary L4 추정
        if not l4_id and scope_l4_ids:
            l4_id = next(iter(scope_l4_ids))
        if not l4_id:
            return tid  # L4 를 알 수 없으면 원본 유지
        l4_max_seq[l4_id] = l4_max_seq.get(l4_id, 0) + 1
        return f"{l4_id}.{l4_max_seq[l4_id]}"

    def _performer_fields_from_task(orig_task: Any | None) -> dict[str, str]:
        """원본 As-Is Task의 performer 필드를 dict로 반환."""
        if orig_task is None:
            return {
                "performer": "", "performer_executive": "", "performer_hr": "",
                "performer_manager": "", "performer_member": "",
            }
        return {
            "performer": orig_task.performer or "",
            "performer_executive": orig_task.performer_executive or "",
            "performer_hr": orig_task.performer_hr or "",
            "performer_manager": orig_task.performer_manager or "",
            "performer_member": orig_task.performer_member or "",
        }

    # 🔑 행을 즉시 쓰지 말고 버퍼에 모은 뒤 L4 별로 재번호한다 (1부터 sequential)
    # 각 버퍼 항목 형식: {"l4_id": ..., "original_seq": int|None, "kind": 'ai'|'human_part'|'human_asis', "pair_key": ..., "values": {...}}
    buffered_rows: list[dict] = []

    # ── 1. Senior/Junior AI task — 스코프 내만 ──
    for agent in step2_cache.get("agents", []):
        atype = agent.get("agent_type", "")
        if atype not in ("Senior AI", "Junior AI"):
            continue
        agent_name = agent.get("agent_name", "") or ""
        for t in agent.get("assigned_tasks", []):
            tid = str(t.get("task_id") or "").strip()
            l4_hint = str(t.get("l4") or "").strip()
            if not _task_in_scope(tid, l4_hint, scope_l4_ids, scope_l4_names):
                continue

            cls = classification.get(tid, {}) if classification else {}
            orig_label = cls.get("label", "")
            orig_task = task_by_id.get(tid)

            # 계층 정보 — 원본 task 우선
            base = {
                "l2_id": (orig_task.l2_id if orig_task else "") or "",
                "l2_name": (orig_task.l2 if orig_task else "") or t.get("l2", ""),
                "l3_id": (orig_task.l3_id if orig_task else "") or "",
                "l3_name": (orig_task.l3 if orig_task else "") or t.get("l3", ""),
                "l4_id": (orig_task.l4_id if orig_task else "") or "",
                "l4_name": (orig_task.l4 if orig_task else "") or t.get("l4", ""),
            }
            row_l4 = base["l4_id"]

            # 원본 task_id의 L5 seq 추출 (정렬 기준 — 없으면 None)
            orig_seq: int | None = None
            tparts = tid.split(".")
            if len(tparts) == 4 and all(p.isdigit() for p in tparts):
                try:
                    orig_seq = int(tparts[3])
                except ValueError:
                    orig_seq = None

            # AI 파트 행
            ai_part_desc = t.get("ai_role", "") or t.get("task_name", "")
            values = {
                **base,
                "l5_id": "",   # ← 나중에 재번호
                "l5_name": f"[AI] {t.get('task_name', '')}",
                "l5_desc": ai_part_desc,
                "performer": f"{atype} — {agent_name}",
                "performer_executive": "",
                "performer_hr": "",
                "performer_manager": "",
                "performer_member": "",
                "cls_1st_label":     cls.get("label", "AI"),
                "cls_1st_knockout":  cls.get("criterion", ""),
                "cls_1st_reason":    cls.get("reason", ""),
                "cls_1st_ai_prereq": cls.get("ai_prerequisites", ""),
                "cls_doosan_label":   cls.get("label", ""),
                "cls_doosan_feedback": cls.get("feedback", ""),
                "cls_final_label":    orig_label or "AI",
                "cls_final_feedback": cls.get("feedback", ""),
            }
            buffered_rows.append({
                "l4_id": row_l4,
                "original_seq": orig_seq,
                "kind": "ai",
                "pair_key": tid or f"ai_{len(buffered_rows)}",
                "values": values,
            })

            # AI+Human 인 경우 Human 파트 별도 행
            is_hybrid = (orig_label == "AI + Human") or bool(t.get("human_role"))
            if is_hybrid:
                human_part = (t.get("human_role") or "").strip()
                if not human_part and cls.get("hybrid_note"):
                    hn = cls["hybrid_note"]
                    if "Human 파트:" in hn:
                        human_part = hn.split("Human 파트:", 1)[1].strip()

                if human_part:
                    performer_vals = _performer_fields_from_task(orig_task)
                    values_h = {
                        **base,
                        "l5_id": "",   # 재번호 (AI 행과 같은 번호 부여)
                        "l5_name": f"[Human] {t.get('task_name', '')} 검토",
                        "l5_desc": human_part,
                        **performer_vals,
                        "cls_1st_label":     cls.get("label", "AI + Human"),
                        "cls_1st_knockout":  cls.get("criterion", ""),
                        "cls_1st_reason":    cls.get("reason", ""),
                        "cls_1st_ai_prereq": cls.get("ai_prerequisites", ""),
                        "cls_doosan_label":   cls.get("label", ""),
                        "cls_doosan_feedback": cls.get("feedback", ""),
                        "cls_final_label":    "Human",
                        "cls_final_feedback": cls.get("feedback", ""),
                    }
                    buffered_rows.append({
                        "l4_id": row_l4,
                        "original_seq": orig_seq,
                        "kind": "human_part",
                        "pair_key": tid or f"ai_{len(buffered_rows)}",   # AI 행과 동일 pair_key → 같은 번호
                        "values": values_h,
                    })

    # ── 2. Human으로 분류된 As-Is L5 — 스코프 내만, AI Agent에 미포함 ──
    agent_task_ids: set[str] = set()
    for agent in step2_cache.get("agents", []):
        for t in agent.get("assigned_tasks", []):
            agent_task_ids.add(str(t.get("task_id") or ""))

    for t in (excel_tasks or []):
        cls = classification.get(t.id, {}) if classification else {}
        label = cls.get("label", "")
        if label != "Human":
            continue
        if t.id in agent_task_ids:
            continue
        # 스코프 필터: task의 L4 정보로 판정
        l4_hint = str(getattr(t, "l4", "") or "").strip()
        if not _task_in_scope(t.id, l4_hint, scope_l4_ids, scope_l4_names):
            continue

        # 원본 L5 seq 추출 (정렬용)
        h_orig_seq: int | None = None
        hparts = t.id.split(".") if t.id else []
        if len(hparts) == 4 and all(p.isdigit() for p in hparts):
            try:
                h_orig_seq = int(hparts[3])
            except ValueError:
                h_orig_seq = None

        values = {
            "l2_id": t.l2_id or "", "l2_name": t.l2,
            "l3_id": t.l3_id or "", "l3_name": t.l3,
            "l4_id": t.l4_id or "", "l4_name": t.l4,
            "l5_id": "",   # 재번호
            "l5_name": f"[Human] {t.name}",
            "l5_desc": t.description or "",
            # Human L5 — 원본 As-Is의 performer 필드 그대로
            "performer": t.performer or "",
            "performer_executive": t.performer_executive or "",
            "performer_hr": t.performer_hr or "",
            "performer_manager": t.performer_manager or "",
            "performer_member": t.performer_member or "",
            "cls_1st_label":     label,
            "cls_1st_knockout":  cls.get("criterion", ""),
            "cls_1st_reason":    cls.get("reason", ""),
            "cls_1st_ai_prereq": cls.get("ai_prerequisites", ""),
            "cls_doosan_label":   cls.get("label", ""),
            "cls_doosan_feedback": cls.get("feedback", ""),
            "cls_final_label":    label,
            "cls_final_feedback": cls.get("feedback", ""),
        }
        buffered_rows.append({
            "l4_id": t.l4_id or "",
            "original_seq": h_orig_seq,
            "kind": "human_asis",
            "pair_key": t.id or f"human_{len(buffered_rows)}",
            "values": values,
        })

    # ── 🔑 L4 별 재번호: 1부터 sequential 할당 ───────────────────────────────
    # 정렬 규칙 (각 L4 내부):
    #   1) original_seq 가 있는 행이 먼저 (As-Is 원본 순서 유지)
    #   2) original_seq 가 없는 신규 task 들이 뒤에
    #   3) 같은 pair_key (AI + Human part 쌍) 는 같은 번호 부여 (두 행 연속 배치)
    from collections import defaultdict as _dd
    by_l4: dict[str, list[dict]] = _dd(list)
    for br in buffered_rows:
        by_l4[br["l4_id"]].append(br)

    # L4 정렬 (l4_id 오름차순)
    sorted_l4s = sorted(by_l4.keys())

    final_rows: list[dict] = []
    for l4_id in sorted_l4s:
        items = by_l4[l4_id]
        # 1차 정렬: (original_seq or inf, pair_key)
        items.sort(key=lambda x: (
            x["original_seq"] if x["original_seq"] is not None else 10**9,
            x["pair_key"],
            0 if x["kind"] == "ai" else (1 if x["kind"] == "human_part" else 2),
        ))

        # pair_key 단위로 그룹핑하여 번호 부여 (같은 pair → 같은 번호)
        seq_counter = 0
        pair_to_seq: dict[str, int] = {}
        for br in items:
            pk = br["pair_key"]
            if pk not in pair_to_seq:
                seq_counter += 1
                pair_to_seq[pk] = seq_counter
            # 최종 L5 ID 할당
            br["values"]["l5_id"] = f"{l4_id}.{pair_to_seq[pk]}" if l4_id else ""
            final_rows.append(br)

    # ── Excel 쓰기 (11행부터) ────────────────────────────────────────────────
    row = _DATA_START_ROW
    for br in final_rows:
        _write_row(ws, row, br["values"])
        row += 1

    wb.save(output_path)
