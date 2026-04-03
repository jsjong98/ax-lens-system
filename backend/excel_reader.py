"""
excel_reader.py — 엑셀 파일에서 L5 Task 목록을 추출합니다.

시트 자동 감지 전략:
  1. 이름에 가이드/guide/작성 등이 포함된 시트는 제외
  2. 나머지 시트 중 열 I에서 L5 ID 패턴(숫자.숫자.숫자)이 가장 많이 발견되는 시트 선택
  3. 데이터 시작 행도 열 I의 첫 번째 유효한 ID 행을 스캔해 자동 결정
"""
from __future__ import annotations
import re
from pathlib import Path

import openpyxl

from models import Task

# ── 기본 열 인덱스 (1-based, 헤더 자동 감지 실패 시 fallback) ──────────────
_COL_DEFAULT = {
    "l2_id": 2, "l2_name": 3, "l3_id": 4, "l3_name": 5,
    "l4_id": 6, "l4_name": 7, "l4_desc": 8,
    "l5_id": 9, "l5_name": 10, "l5_desc": 11, "performer": 12,
    "performer_executive": 13, "performer_hr": 14,
    "performer_manager": 15, "performer_member": 16,
    "pain_time": 17, "pain_accuracy": 18, "pain_repetition": 19,
    "pain_data": 20, "pain_system": 21, "pain_communication": 22, "pain_other": 23,
    "output_system": 24, "output_document": 25, "output_communication": 26,
    "output_decision": 27, "output_other": 28,
    "logic_rule_based": 29, "logic_human_judgment": 30, "logic_mixed": 31,
    "remark": 32, "standard_or_specialized": 33,
    "cls_1st_label": 34, "cls_1st_knockout": 35,
    "cls_1st_reason": 36, "cls_1st_ai_prereq": 37,
    "cls_doosan_label": 38, "cls_doosan_feedback": 39,
    "cls_final_label": 40, "cls_final_feedback": 41,
}

# ── 헤더 텍스트 → 필드 매핑 (부분 일치로 검색) ──────────────────────────────
# 행 8~9에 걸친 다단 헤더의 텍스트를 매칭합니다.
# 각 항목: (필드명, [매칭할 키워드 목록]) — 키워드 중 하나라도 포함되면 매칭
_HEADER_PATTERNS: list[tuple[str, list[str]]] = [
    # 프로세스 계층 — 행 8의 "ID", "Name" 등은 여러 열에 반복되므로
    # L2~L5는 위치 기반으로 남기고, 특수 헤더만 이름으로 매핑
    # L5 ID는 가장 핵심 — "Task (L5)" 하위의 "ID"
    # → L2~L5 계층은 위치가 고정적이므로 fallback 사용

    # ── 분류 결과 (가장 구체적인 키워드 → 먼저 매칭해서 선점) ──
    # 최종 PwC 검토
    ("cls_final_label",    ["최종 분류 결과", "최종 분류"]),
    ("cls_final_feedback", ["PwC Feedback", "PwC 피드백"]),
    # 두산 검토 ("Feedback"은 PwC 뒤에 매칭되므로 "두산"이 필수)
    ("cls_doosan_feedback", ["두산 Feedback", "두산 피드백"]),
    ("cls_doosan_label",    ["변경 필요", "두산 검토"]),
    # 1차 평가
    ("cls_1st_label",      ["1차 분류결과", "1차 분류"]),
    ("cls_1st_knockout",   ["적용기준", "Knock-out", "Knockout"]),
    ("cls_1st_reason",     ["1차 판단", "판단 근거", "판단근거"]),
    ("cls_1st_ai_prereq",  ["AI 수행 필요", "필요여건", "필요 여건"]),

    # ── F-2, F-3 ──
    ("remark",                  ["F-2", "비고"]),
    ("standard_or_specialized", ["F-3", "표준 vs", "특화"]),

    # ── 업무 판단 로직 (F-1) ──
    ("logic_rule_based",     ["Rule-based", "규칙 기반"]),
    ("logic_human_judgment", ["사람 판단"]),
    ("logic_mixed",          ["혼합"]),

    # ── Output (E-2) — "시스템 반영"을 먼저 잡아서 Pain의 "시스템/도구"와 구분 ──
    ("output_system",        ["시스템 반영"]),
    ("output_document",      ["문서/보고서"]),
    ("output_communication", ["커뮤니케이션"]),
    ("output_decision",      ["의사결정"]),

    # ── Pain Point (D-1) — "시스템/도구"는 구체적 키워드 사용 ──
    ("pain_time",           ["시간/속도"]),
    ("pain_accuracy",       ["정확성"]),
    ("pain_repetition",     ["반복/수작업"]),
    ("pain_data",           ["정보/데이터"]),
    ("pain_system",         ["시스템/도구"]),
    ("pain_communication",  ["의사소통/협업"]),

    # ── 수행주체 (A-1) ──
    ("performer_executive", ["임원"]),
    ("performer_manager",   ["현업 팀장"]),
    ("performer_member",    ["현업 구성원"]),
]

# Pain Point / Output의 마지막 "기타" 열은 컨텍스트로 구분
_CONTEXT_OTHERS: list[tuple[str, str, list[str]]] = [
    # (필드명, 이 열 앞에 와야 하는 필드의 키워드, ["기타"])
    ("pain_other",   "pain_",   ["기타"]),
    ("output_other", "output_", ["기타"]),
]


def _detect_columns(ws, scan_rows: int = 15) -> dict[str, int]:
    """
    시트의 헤더 영역(행 1~scan_rows)을 스캔하여 열 이름 → 1-based 열 번호 매핑을 반환.
    다단 헤더(행 7~9 등)를 모두 합쳐서 매칭합니다.
    """
    # 각 열의 헤더 텍스트를 행 전체에서 수집 (다단 헤더 대응)
    col_texts: dict[int, str] = {}  # col_idx(1-based) → 합친 헤더 텍스트
    for row in ws.iter_rows(min_row=1, max_row=scan_rows, values_only=True):
        for col_idx, val in enumerate(row, start=1):
            if val is not None:
                text = str(val).strip().replace("\n", " ")
                if col_idx in col_texts:
                    col_texts[col_idx] += " " + text
                else:
                    col_texts[col_idx] = text

    # 매핑 결과
    detected: dict[str, int] = {}
    used_cols: set[int] = set()

    # 1단계: 특수 패턴 매칭 (구체적인 키워드)
    for field, keywords in _HEADER_PATTERNS:
        if field in detected:
            continue
        for col_idx, text in col_texts.items():
            if col_idx in used_cols:
                continue
            for kw in keywords:
                if kw in text:
                    detected[field] = col_idx
                    used_cols.add(col_idx)
                    break
            if field in detected:
                break

    # 2단계: "기타" 열 — 컨텍스트 기반 (앞 열이 pain_ 계열이면 pain_other)
    for field, prefix, keywords in _CONTEXT_OTHERS:
        if field in detected:
            continue
        # 해당 prefix를 가진 필드들의 최대 열 번호를 찾음
        prefix_cols = [v for k, v in detected.items() if k.startswith(prefix) and k != field]
        if not prefix_cols:
            continue
        max_col = max(prefix_cols)
        # max_col 바로 다음 열들에서 "기타"를 찾음
        for col_idx in range(max_col + 1, max_col + 4):
            if col_idx in used_cols:
                continue
            text = col_texts.get(col_idx, "")
            if any(kw in text for kw in keywords):
                detected[field] = col_idx
                used_cols.add(col_idx)
                break

    # 3단계: L2~L5 계층 — 위치 기반 (L5 ID 열을 찾고 거기서 역산)
    # L5 ID는 _L5_ID_RE 패턴이 있는 열
    l5_id_col = None
    for col_idx, text in col_texts.items():
        if "Task" in text and "L5" in text:
            l5_id_col = col_idx
            break
    if not l5_id_col:
        # ID 열 직접 스캔 (데이터 행에서 L5 ID 패턴이 있는 열)
        for row in ws.iter_rows(min_row=scan_rows + 1, max_row=scan_rows + 10, values_only=True):
            for col_idx, val in enumerate(row, start=1):
                if val and _L5_ID_RE.match(str(val).strip()):
                    l5_id_col = col_idx
                    break
            if l5_id_col:
                break

    if l5_id_col:
        detected["l5_id"] = l5_id_col
        detected["l5_name"] = l5_id_col + 1
        detected["l5_desc"] = l5_id_col + 2

        # L5 앞: 수행주체(performer) 열은 l5_desc 다음
        detected["performer"] = l5_id_col + 3

        # L4: L5 ID 바로 앞 3열
        detected["l4_desc"] = l5_id_col - 1
        detected["l4_name"] = l5_id_col - 2
        detected["l4_id"]   = l5_id_col - 3

        # L3: L4 앞 2열
        detected["l3_name"] = l5_id_col - 4
        detected["l3_id"]   = l5_id_col - 5

        # L2: L3 앞 2열
        detected["l2_name"] = l5_id_col - 6
        detected["l2_id"]   = l5_id_col - 7

    # 4단계: performer_hr — "HR"은 너무 짧아서 다른 열에도 걸림
    # performer_executive(임원) 바로 다음 열이 HR
    if "performer_hr" not in detected and "performer_executive" in detected:
        hr_col = detected["performer_executive"] + 1
        if hr_col not in used_cols:
            detected["performer_hr"] = hr_col
            used_cols.add(hr_col)
    # 또는 performer_manager 바로 앞 열
    if "performer_hr" not in detected and "performer_manager" in detected:
        hr_col = detected["performer_manager"] - 1
        if hr_col not in used_cols:
            detected["performer_hr"] = hr_col
            used_cols.add(hr_col)

    # 5단계: fallback — 감지 못한 필드는 기본값 사용
    result = dict(_COL_DEFAULT)
    result.update(detected)

    # 로그
    auto_count = len(detected)
    fallback_count = len(_COL_DEFAULT) - auto_count
    print(f"[excel] 헤더 자동 감지: {auto_count}개 매핑, {fallback_count}개 기본값 사용")
    for field, col_idx in sorted(result.items(), key=lambda x: x[1]):
        source = "자동" if field in detected else "기본"
        header_text = col_texts.get(col_idx, "")[:30]
        print(f"  {field:30s} → 열 {col_idx:2d}  [{source}] {header_text}")

    return result

# ── 시트 자동 감지 ────────────────────────────────────────────────────────────

# 데이터가 없는 가이드/설명 시트 식별 키워드
_GUIDE_KEYWORDS = {
    "가이드", "guide", "작성", "설명", "manual", "instruction", "readme", "index",
    "backup", "백업", "template", "템플릿",  # 백업/템플릿 시트
    "count", "집계", "통계",                 # 집계용 시트
    "sheet1", "sheet2", "sheet3",           # 기본 빈 시트
    "lv3", "lv4", "lv5",                    # 계층 목록 시트
    "task별", "정보 요청",                   # 기타 보조 시트
}

# L5 ID 패턴: "1.1.1", "1.1.1.1", "2.3.4.5" 등 숫자.숫자.숫자로 시작
_L5_ID_RE = re.compile(r"^\d+\.\d+\.\d+")

# L5 ID가 위치하는 열 (I = 9번째, 0-indexed = 8)
_L5_ID_COL_IDX = 8   # 0-based (openpyxl values_only row 기준)
_L5_ID_COL_1B  = 9   # 1-based  (openpyxl cell 직접 접근 기준)

# 시트 스캔 시 확인할 최대 행 수 (헤더 영역 포함)
_SCAN_ROWS = 50


def _is_guide_sheet(name: str) -> bool:
    lower = name.lower()
    return any(kw in lower for kw in _GUIDE_KEYWORDS)


def _score_sheet(ws) -> int:
    """시트에서 L5 ID 패턴이 몇 개 발견되는지 반환 (스캔은 최대 _SCAN_ROWS행)."""
    count = 0
    for row in ws.iter_rows(
        min_row=1, max_row=_SCAN_ROWS,
        min_col=_L5_ID_COL_1B, max_col=_L5_ID_COL_1B,
        values_only=True,
    ):
        val = str(row[0]).strip() if row[0] is not None else ""
        if _L5_ID_RE.match(val):
            count += 1
    return count


def _find_data_sheet(wb):
    """
    가이드 시트를 제외하고 L5 Task 데이터가 가장 많은 시트를 반환합니다.
    동점이면 시트 목록 순서가 앞선 것을 선택합니다.
    """
    best_ws = None
    best_score = -1

    for name in wb.sheetnames:
        if _is_guide_sheet(name):
            print(f"[sheet] '{name}' — 가이드 시트 제외")
            continue
        ws = wb[name]
        score = _score_sheet(ws)
        print(f"[sheet] '{name}' — L5 ID {score}개 감지")
        if score > best_score:
            best_score = score
            best_ws = ws

    if best_ws is None:
        # 모두 가이드로 판단된 경우 첫 시트로 폴백
        print("[sheet] 적합한 시트 없음 — 첫 번째 시트로 폴백")
        return wb.worksheets[0]

    print(f"[sheet] 선택됨: '{best_ws.title}' (L5 ID {best_score}개)")
    return best_ws


def _find_data_start_row(ws) -> int:
    """
    열 I(L5 ID)에서 첫 번째 유효한 L5 ID가 있는 행 번호를 반환합니다.
    발견하지 못하면 기본값 10을 반환합니다.
    """
    for row in ws.iter_rows(
        min_row=1, max_row=100,
        min_col=_L5_ID_COL_1B, max_col=_L5_ID_COL_1B,
        values_only=True,
    ):
        # iter_rows는 (value,) 튜플을 반환하므로 row[0]
        pass  # iter_rows는 행 번호를 직접 주지 않으므로 enumerate 사용

    start = 10
    for r_idx, row in enumerate(
        ws.iter_rows(
            min_row=1, max_row=100,
            min_col=_L5_ID_COL_1B, max_col=_L5_ID_COL_1B,
            values_only=True,
        ),
        start=1,
    ):
        val = str(row[0]).strip() if row[0] is not None else ""
        if _L5_ID_RE.match(val):
            start = r_idx
            break

    print(f"[sheet] 데이터 시작 행: {start}")
    return start


def _find_excel(base_dir: Path) -> Path:
    """base_dir 또는 상위 디렉토리에서 엑셀 파일을 자동 탐색."""
    # 상위 디렉토리 포함해서 탐색
    for search_dir in [base_dir, base_dir.parent]:
        xlsx_files = list(search_dir.glob("*.xlsx"))
        if xlsx_files:
            # As-Is 파일 우선 선택
            for f in xlsx_files:
                if "As-is" in f.name or "as-is" in f.name.lower():
                    return f
            return xlsx_files[0]
    raise FileNotFoundError(
        f"엑셀 파일을 찾을 수 없습니다. 경로: {base_dir}"
    )


def _cell(row: tuple, col: int) -> str:
    """1-based 열 번호로 행에서 값 추출 후 문자열 변환."""
    idx = col - 1
    if idx < 0 or idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    return str(val).strip()


def list_sheets(excel_path: str | Path) -> list[dict]:
    """
    엑셀 파일의 시트 목록과 각 시트의 L5 Task 개수를 반환합니다.

    Returns:
        [{"name": "시트1", "task_count": 42, "is_guide": False, "recommended": True}, ...]
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"파일 없음: {excel_path}")

    wb = openpyxl.load_workbook(str(excel_path), data_only=True, read_only=True)

    best_score = -1
    best_name = ""
    sheets = []

    for name in wb.sheetnames:
        is_guide = _is_guide_sheet(name)
        ws = wb[name]
        score = _score_sheet(ws) if not is_guide else 0
        if score > best_score:
            best_score = score
            best_name = name
        sheets.append({
            "name": name,
            "task_count": score,
            "is_guide": is_guide,
            "recommended": False,
        })

    # 추천 시트 마킹
    for s in sheets:
        if s["name"] == best_name and best_score > 0:
            s["recommended"] = True

    wb.close()
    return sheets


def load_tasks(
    excel_path: str | Path | None = None,
    sheet_name: str | None = None,
) -> list[Task]:
    """
    엑셀 파일을 읽어 Task 목록을 반환합니다.

    Args:
        excel_path: 엑셀 파일 경로. None이면 현재 파일 기준으로 자동 탐색.
        sheet_name: 시트 이름. None이면 자동 감지.
    """
    if excel_path is None:
        base_dir = Path(__file__).parent
        excel_path = _find_excel(base_dir)

    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"파일 없음: {excel_path}")

    wb = openpyxl.load_workbook(str(excel_path), data_only=True, read_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"[sheet] 사용자 지정 시트: '{sheet_name}'")
    else:
        ws = _find_data_sheet(wb)
    data_start_row = _find_data_start_row(ws)

    # 헤더 자동 감지 (read_only 재오픈 전에 수행)
    COL = _detect_columns(ws)

    # read_only 모드에서는 ws를 재사용할 수 없으므로 다시 열기
    wb.close()
    wb = openpyxl.load_workbook(str(excel_path), data_only=True, read_only=True)
    ws = wb[ws.title]

    tasks: list[Task] = []
    seen_ids: dict[str, int] = {}   # ID 중복 카운터

    for row in ws.iter_rows(
        min_row=data_start_row,
        max_row=ws.max_row,
        values_only=True,
    ):
        l5_id = _cell(row, COL["l5_id"])

        # L5 ID가 없으면 L4/L3/L2 레벨 행으로 인식 (선행 컨텍스트 행)
        if not l5_id:
            l4_id = _cell(row, COL["l4_id"])
            l3_id = _cell(row, COL["l3_id"])
            l2_id = _cell(row, COL["l2_id"])
            if l4_id:
                row_id, row_level, row_name = l4_id, "L4", _cell(row, COL["l4_name"]) or l4_id
            elif l3_id:
                row_id, row_level, row_name = l3_id, "L3", _cell(row, COL["l3_name"]) or l3_id
            elif l2_id:
                row_id, row_level, row_name = l2_id, "L2", _cell(row, COL["l2_name"]) or l2_id
            else:
                continue  # 모든 레벨 ID 없으면 빈 행으로 skip

            if not row_name:
                continue
            if row_id in seen_ids:
                continue  # 상위 레벨 중복은 skip (L5 중복과 달리 첫 번째만 유지)
            seen_ids[row_id] = 1

            tasks.append(
                Task(
                    id=row_id,
                    level=row_level,
                    l2_id=_cell(row, COL["l2_id"]),
                    l2=_cell(row, COL["l2_name"]) or "",
                    l3_id=_cell(row, COL["l3_id"]) or "",
                    l3=_cell(row, COL["l3_name"]) or "",
                    l4_id=l4_id or "",
                    l4=_cell(row, COL["l4_name"]) or "",
                    l4_description=_cell(row, COL["l4_desc"]) or "",
                    name=row_name,
                    description=_cell(row, COL.get("l5_desc", COL["l4_desc"])) or "",
                )
            )
            continue

        # 중복 ID 처리: 두 번째부터 접미사 붙이기 (예: 1.8.1.2_2)
        if l5_id in seen_ids:
            seen_ids[l5_id] += 1
            unique_id = f"{l5_id}_{seen_ids[l5_id]}"
            print(f"[excel] 중복 ID 발견: '{l5_id}' → '{unique_id}'로 변경")
        else:
            seen_ids[l5_id] = 1
            unique_id = l5_id

        tasks.append(
            Task(
                id=unique_id,
                level="L5",
                l2_id=_cell(row, COL["l2_id"]),
                l2=_cell(row, COL["l2_name"]),
                l3_id=_cell(row, COL["l3_id"]),
                l3=_cell(row, COL["l3_name"]),
                l4_id=_cell(row, COL["l4_id"]),
                l4=_cell(row, COL["l4_name"]),
                l4_description=_cell(row, COL["l4_desc"]),
                name=_cell(row, COL["l5_name"]),
                description=_cell(row, COL["l5_desc"]),
                performer=_cell(row, COL["performer"]),
                # A-1. 수행주체 체크박스
                performer_executive=_cell(row, COL["performer_executive"]),
                performer_hr=_cell(row, COL["performer_hr"]),
                performer_manager=_cell(row, COL["performer_manager"]),
                performer_member=_cell(row, COL["performer_member"]),
                # D-1. Pain Point
                pain_time=_cell(row, COL["pain_time"]),
                pain_accuracy=_cell(row, COL["pain_accuracy"]),
                pain_repetition=_cell(row, COL["pain_repetition"]),
                pain_data=_cell(row, COL["pain_data"]),
                pain_system=_cell(row, COL["pain_system"]),
                pain_communication=_cell(row, COL["pain_communication"]),
                pain_other=_cell(row, COL["pain_other"]),
                # E-2. Output
                output_system=_cell(row, COL["output_system"]),
                output_document=_cell(row, COL["output_document"]),
                output_communication=_cell(row, COL["output_communication"]),
                output_decision=_cell(row, COL["output_decision"]),
                output_other=_cell(row, COL["output_other"]),
                # F-1. 업무 판단 로직
                logic_rule_based=_cell(row, COL["logic_rule_based"]),
                logic_human_judgment=_cell(row, COL["logic_human_judgment"]),
                logic_mixed=_cell(row, COL["logic_mixed"]),
                # F-2~F-3
                remark=_cell(row, COL["remark"]),
                standard_or_specialized=_cell(row, COL["standard_or_specialized"]),
                # 분류 결과 (있으면 읽기)
                cls_1st_label=_cell(row, COL["cls_1st_label"]),
                cls_1st_knockout=_cell(row, COL["cls_1st_knockout"]),
                cls_1st_reason=_cell(row, COL["cls_1st_reason"]),
                cls_1st_ai_prereq=_cell(row, COL["cls_1st_ai_prereq"]),
                cls_doosan_label=_cell(row, COL["cls_doosan_label"]),
                cls_doosan_feedback=_cell(row, COL["cls_doosan_feedback"]),
                cls_final_label=_cell(row, COL["cls_final_label"]),
                cls_final_feedback=_cell(row, COL["cls_final_feedback"]),
            )
        )

    wb.close()
    return tasks
