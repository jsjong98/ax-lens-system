"""
main.py — FastAPI 애플리케이션 엔트리포인트

실행:
  uvicorn main:app --reload --port 8000
"""
from __future__ import annotations
import asyncio
import io
import json
import os
import shutil
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

# KST 시간대 (UTC+9)
_KST = timezone(timedelta(hours=9))

def _now_kst() -> str:
    """현재 KST 시각을 ISO 8601 문자열로 반환."""
    return datetime.now(_KST).isoformat()

def _mtime_kst(path) -> str:
    """파일 수정 시각을 KST ISO 8601 문자열로 반환."""
    return datetime.fromtimestamp(path.stat().st_mtime, tz=_KST).isoformat()


def _natural_key(s: str) -> list[int | str]:
    """자연수 정렬 키: '1.1.1.2' < '1.1.1.10' 순서를 보장합니다."""
    import re
    return [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', s)]

# .env 파일 자동 로드 (python-dotenv 없어도 동작하는 fallback 포함)
def _load_dotenv() -> None:
    env_path = Path(__file__).parent / ".env"
    if not env_path.exists():
        return
    try:
        from dotenv import load_dotenv
        load_dotenv(env_path, override=False)
    except ImportError:
        for line in env_path.read_text("utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, _, val = line.partition("=")
                key = key.strip()
                if key and key not in os.environ:
                    os.environ[key] = val.strip()

_load_dotenv()

from fastapi import FastAPI, HTTPException, Query, Request, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from classifier import get_classifier
from excel_reader import load_tasks
from models import (
    ClassificationResult,
    ClassificationResultUpdate,
    ClassifierSettings,
    ClassifyRequest,
    ResultsResponse,
    StatsResponse,
    Task,
    TaskListResponse,
)
from settings_store import (
    clear_results,
    load_results,
    load_settings,
    save_results,
    save_settings,
    set_current_file,
    upsert_result,
)
from usage_store import get_usage, reset_usage
from auth_store import (
    init_default_users,
    authenticate,
    get_session_user,
    change_password,
    logout,
    generate_reset_code,
    verify_reset_code,
    reset_password,
    send_reset_email,
    get_all_sessions,
    get_all_users_info,
    force_logout_user,
    update_session_info,
    get_user_project,
    get_user_projects,
    is_pm,
    request_transfer,
    get_pending_transfers,
    approve_transfer,
    reject_transfer,
    get_all_transfers,
    ADMIN_EMAIL,
    ALL_PROJECTS,
    PROJECT_PMS,
)
import audit_log
from data_store import (
    save_data, load_data, clear_data, get_saved_status,
    set_current_project, get_current_project, list_projects, save_meta,
    delete_project, save_meta_team, list_projects_for_user,
)

# ── 앱 초기화 ────────────────────────────────────────────────────────────────

app = FastAPI(
    title="PwC AX Lens — Process Innovation API",
    description="AI 기반 업무 혁신 설계 플랫폼",
    version="2.0.0",
)

# CORS: ALLOWED_ORIGINS 환경변수 + Railway 도메인 자동 허용
_default_origins = [
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "https://pwc-ax-lens.com",
    "https://www.pwc-ax-lens.com",
    "https://pwc-ax-lens.up.railway.app",
]
_extra_origins = [
    o.strip() for o in os.getenv("ALLOWED_ORIGINS", "").split(",") if o.strip()
]
_all_origins = _default_origins + _extra_origins

app.add_middleware(
    CORSMiddleware,
    allow_origins=_all_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization", "X-Session-Token", "X-User-Id", "X-Team-Id"],
)

# 500 에러에도 CORS 헤더 보장
from fastapi.responses import JSONResponse as _JSONResponse
@app.exception_handler(Exception)
async def _global_exception_handler(request: Request, exc: Exception):
    origin = request.headers.get("origin", "")
    cors_origin = origin if origin in _all_origins else (_all_origins[0] if _all_origins else "*")
    import logging
    logging.getLogger("uvicorn.error").error(f"Unhandled exception: {exc}", exc_info=True)
    return _JSONResponse(
        status_code=500,
        content={"detail": "서버 내부 오류가 발생했습니다."},
        headers={"Access-Control-Allow-Origin": cors_origin,
                 "Access-Control-Allow-Credentials": "true"},
    )

# ── 기본 사용자 초기화 ─────────────────────────────────────────────────────────
init_default_users()

# ── 인증 엔드포인트 ──────────────────────────────────────────────────────────

from pydantic import BaseModel as _BaseModel


def _get_user_context(request: Request) -> dict | None:
    """요청에서 사용자 정보 + 프로젝트 정보를 추출."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    return get_session_user(token)


def _get_client_ip(request: Request) -> str:
    """클라이언트 IP 추출 (프록시 헤더 우선)."""
    forwarded = request.headers.get("x-forwarded-for", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else ""


class _LoginRequest(_BaseModel):
    email: str
    password: str


class _ChangePasswordRequest(_BaseModel):
    old_password: str
    new_password: str


@app.post("/api/auth/login", tags=["Auth"])
async def api_login(body: _LoginRequest, request: Request):
    ip = _get_client_ip(request)
    ua = request.headers.get("user-agent", "")
    token = authenticate(body.email, body.password, ip=ip, user_agent=ua)
    if not token:
        raise HTTPException(401, "이메일 또는 비밀번호가 올바르지 않습니다.")
    user = get_session_user(token)
    return {"ok": True, "token": token, "user": user}


@app.get("/api/auth/me", tags=["Auth"])
async def api_me(request: Request):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    # 세션에 IP/UA가 없으면 갱신 (기존 세션 호환)
    update_session_info(token, ip=_get_client_ip(request), user_agent=request.headers.get("user-agent", ""))
    user = get_session_user(token)
    if not user:
        raise HTTPException(401, "인증이 필요합니다.")
    # PM에게 대기 중인 이동 요청 수 포함
    pending_count = len(get_pending_transfers(user["email"])) if user.get("is_pm") else 0
    return {"ok": True, "user": user, "pending_transfers": pending_count}


# ── 프로젝트 이동 요청/승인 ──────────────────────────────────────────────────

@app.post("/api/auth/transfer-request", tags=["Auth"])
async def api_transfer_request(request: Request):
    """프로젝트 이동 요청."""
    user = _get_user_context(request)
    if not user:
        raise HTTPException(401, "인증이 필요합니다.")
    body = await request.json()
    target = body.get("target_project", "")
    reason = body.get("reason", "")
    if not target:
        raise HTTPException(400, "이동할 프로젝트를 지정하세요.")
    result = request_transfer(user["email"], target, reason)
    if "error" in result:
        raise HTTPException(400, result["error"])
    audit_log.log_event("transfer_request", email=user["email"], ip=_get_client_ip(request),
                        detail=f"{result.get('current_project', '?')} → {target}")
    return {"ok": True, "request": result}


@app.get("/api/auth/pending-transfers", tags=["Auth"])
async def api_pending_transfers(request: Request):
    """대기 중인 이동 요청 목록 (PM/Admin)."""
    user = _get_user_context(request)
    if not user:
        raise HTTPException(401, "인증이 필요합니다.")
    if not user.get("is_pm") and not user.get("is_admin"):
        return {"ok": True, "requests": []}
    return {"ok": True, "requests": get_pending_transfers(user["email"])}


@app.post("/api/auth/approve-transfer", tags=["Auth"])
async def api_approve_transfer(request: Request):
    """이동 요청 승인 (PM/Admin만)."""
    user = _get_user_context(request)
    if not user:
        raise HTTPException(401, "인증이 필요합니다.")
    if not user.get("is_pm") and not user.get("is_admin"):
        raise HTTPException(403, "승인 권한이 없습니다.")
    body = await request.json()
    request_id = body.get("request_id", "")
    result = approve_transfer(request_id, user["email"])
    if "error" in result:
        raise HTTPException(400, result["error"])
    audit_log.log_event("transfer_approved", email=user["email"], ip=_get_client_ip(request),
                        detail=f"{result.get('name', '?')}: {result.get('current_project', '?')} → {result.get('target_project', '?')}")
    return {"ok": True, "request": result}


@app.post("/api/auth/reject-transfer", tags=["Auth"])
async def api_reject_transfer(request: Request):
    """이동 요청 거절 (PM/Admin만)."""
    user = _get_user_context(request)
    if not user:
        raise HTTPException(401, "인증이 필요합니다.")
    if not user.get("is_pm") and not user.get("is_admin"):
        raise HTTPException(403, "승인 권한이 없습니다.")
    body = await request.json()
    request_id = body.get("request_id", "")
    result = reject_transfer(request_id, user["email"])
    if "error" in result:
        raise HTTPException(400, result["error"])
    audit_log.log_event("transfer_rejected", email=user["email"], ip=_get_client_ip(request),
                        detail=f"{result.get('name', '?')}: 거절")
    return {"ok": True, "request": result}


@app.post("/api/auth/change-password", tags=["Auth"])
async def api_change_password(body: _ChangePasswordRequest, request: Request):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = get_session_user(token)
    if not user:
        raise HTTPException(401, "인증이 필요합니다.")
    ok = change_password(user["email"], body.old_password, body.new_password)
    if not ok:
        raise HTTPException(400, "기존 비밀번호가 올바르지 않습니다.")
    return {"ok": True}


@app.post("/api/auth/logout", tags=["Auth"])
async def api_logout(request: Request):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    logout(token)
    return {"ok": True}


class _ResetRequestBody(_BaseModel):
    email: str


class _ResetVerifyBody(_BaseModel):
    email: str
    code: str


class _ResetPasswordBody(_BaseModel):
    email: str
    code: str
    new_password: str


@app.post("/api/auth/reset/request", tags=["Auth"])
async def api_reset_request(body: _ResetRequestBody):
    """비밀번호 재설정 인증번호를 이메일로 발송합니다."""
    code = generate_reset_code(body.email)
    if not code:
        raise HTTPException(404, "등록되지 않은 이메일입니다. 관리자에게 문의해 주세요.")
    sent = await send_reset_email(body.email, code)
    if not sent:
        raise HTTPException(500, "이메일 발송에 실패했습니다. 잠시 후 다시 시도해 주세요.")
    return {"ok": True, "message": "인증번호가 이메일로 발송되었습니다."}


@app.post("/api/auth/reset/verify", tags=["Auth"])
async def api_reset_verify(body: _ResetVerifyBody):
    """인증번호를 검증합니다."""
    valid = verify_reset_code(body.email, body.code)
    if not valid:
        raise HTTPException(400, "인증번호가 올바르지 않거나 만료되었습니다.")
    return {"ok": True}


@app.post("/api/auth/reset/confirm", tags=["Auth"])
async def api_reset_confirm(body: _ResetPasswordBody):
    """인증번호 확인 후 새 비밀번호를 설정합니다."""
    ok = reset_password(body.email, body.code, body.new_password)
    if not ok:
        raise HTTPException(400, "인증번호가 올바르지 않거나 만료되었습니다.")
    return {"ok": True, "message": "비밀번호가 재설정되었습니다."}


# ── 인메모리 Task 캐시 ────────────────────────────────────────────────────────
_tasks_cache: list[Task] = []


# ── 영속 저장 헬퍼 ──────────────────────────────────────────────────────────

def _persist_cache(key: str, cache: dict) -> None:
    """캐시 데이터를 JSON 파일로 영속 저장합니다. (Task 분류 전용)"""
    save_data(key, dict(cache))


def _restore_cache(key: str, cache: dict) -> None:
    """서버 시작 시 JSON 파일에서 캐시를 복원합니다. (Task 분류 전용)"""
    data = load_data(key)
    if data and isinstance(data, dict):
        cache.update(data)
        print(f"[data_store] '{key}' 복원 완료 ({len(data)} keys)")


# ── New Workflow 전용 파일 I/O (_NW_DIR 고정 경로) ───────────────────────────

def _save_nw_state(key: str, data: dict) -> None:
    """New Workflow / 과제 정의서 / 과제 설계서를 _NW_DIR에 직접 저장."""
    try:
        (_NW_DIR / f"{key}.json").write_text(
            json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
        )
    except Exception as e:
        print(f"[NW_STORE] 저장 실패 ({key}): {e}", flush=True)


def _load_nw_state(key: str) -> dict:
    """New Workflow / 과제 정의서 / 과제 설계서를 _NW_DIR에서 로드."""
    p = _NW_DIR / f"{key}.json"
    if not p.exists():
        return {}
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
        return d if isinstance(d, dict) else {}
    except Exception as e:
        print(f"[NW_STORE] 로드 실패 ({key}): {e}", flush=True)
        return {}


# 데이터 저장 상태 확인 API
@app.get("/api/data-status", tags=["Data"])
async def get_data_status():
    """각 단계별 데이터 저장 상태를 반환합니다."""
    return {
        "ok": True,
        "current_project": get_current_project(),
        "tasks_loaded": len(_tasks_cache) > 0,
        "saved": get_saved_status(),
    }


@app.delete("/api/data/reset-all", tags=["Data"])
async def reset_all_data(request: Request):
    """모든 데이터를 초기화합니다 (Volume 포함)."""
    _require_admin(request)
    import shutil
    # 인메모리 캐시 초기화
    _new_workflow_cache.clear()
    _project_definition_cache.clear()
    _project_design_cache.clear()
    _nw_tasks_cache.clear()
    _nw_projects_cache.clear()
    # Volume 데이터 삭제
    data_dir = _PERSIST_ROOT / "data"
    if data_dir.exists():
        shutil.rmtree(data_dir, ignore_errors=True)
        data_dir.mkdir(exist_ok=True)
    results_dir = _PERSIST_ROOT / "results"
    if results_dir.exists():
        shutil.rmtree(results_dir, ignore_errors=True)
        results_dir.mkdir(exist_ok=True)
    uploads_dir = _PERSIST_ROOT / "uploads"
    if uploads_dir.exists():
        shutil.rmtree(uploads_dir, ignore_errors=True)
        uploads_dir.mkdir(exist_ok=True)
    for f in ["current_project.json", "new_workflow_result.json", "project_definition.json", "project_design.json"]:
        p = _PERSIST_ROOT / f
        if p.exists():
            p.unlink(missing_ok=True)
    return {"ok": True, "message": "모든 데이터가 초기화되었습니다."}


@app.get("/api/projects", tags=["Data"])
async def get_project_list(request: Request):
    """저장된 프로젝트 목록 — 사용자 프로젝트로 필터링."""
    user_ctx = _get_user_context(request)
    if user_ctx:
        if user_ctx.get("project") == "미지정":
            projects = []
        else:
            projects = list_projects_for_user(
                None if user_ctx.get("is_admin") or user_ctx.get("project") is None
                else user_ctx.get("projects")
            )
    else:
        projects = list_projects()
    return {"ok": True, "projects": projects}


@app.post("/api/projects/load", tags=["Data"])
async def load_project(request: Request):
    """이전 프로젝트의 결과를 로드합니다."""
    body = await request.json()
    filename = body.get("filename", "")
    if not filename:
        raise HTTPException(400, "파일명이 필요합니다.")

    set_current_project(filename)

    # 저장된 결과 로드
    loaded = {}

    nw_data = load_data("new_workflow_result", filename)
    if nw_data:
        _new_workflow_cache.clear()
        _new_workflow_cache.update(nw_data)
        loaded["new_workflow"] = True

    pd_data = load_data("project_definition", filename)
    if pd_data:
        _project_definition_cache.clear()
        _project_definition_cache.update(pd_data)
        loaded["project_definition"] = True

    pds_data = load_data("project_design", filename)
    if pds_data:
        _project_design_cache.clear()
        _project_design_cache.update(pds_data)
        loaded["project_design"] = True

    # 벤치마킹 결과도 로드
    bm_data = load_data("benchmark_result", filename)
    if bm_data:
        loaded["benchmark_result"] = True

    return {
        "ok": True,
        "filename": filename,
        "loaded": loaded,
        "saved": get_saved_status(filename),
        "benchmark": bm_data if bm_data else None,
    }


@app.get("/api/new-workflow/benchmark-result", tags=["NewWorkflow"])
async def get_benchmark_result():
    """저장된 벤치마킹 결과를 반환합니다."""
    fn = get_current_project()
    if fn:
        data = load_data("benchmark_result", fn)
        if data:
            return {"ok": True, **data}
    raise HTTPException(404, "저장된 벤치마킹 결과가 없습니다.")


@app.delete("/api/projects/{dirname}", tags=["Data"])
async def remove_project(dirname: str):
    """프로젝트를 삭제합니다."""
    if not dirname:
        raise HTTPException(400, "디렉토리명이 필요합니다.")
    ok = delete_project(dirname)
    if not ok:
        raise HTTPException(404, "프로젝트를 찾을 수 없습니다.")
    return {"ok": True, "deleted": dirname}


# ─────────────────────────────────────────────────────────────────────────────
# Task 엔드포인트
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/tasks", response_model=TaskListResponse, tags=["Tasks"])
async def get_tasks(
    search: Optional[str] = Query(None),
    l2: Optional[str] = Query(None),
    l3: Optional[str] = Query(None),
    l4: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=10000),
):
    tasks = _tasks_cache

    if search:
        q = search.lower()
        tasks = [
            t for t in tasks
            if q in t.name.lower()
            or q in t.description.lower()
            or q in t.performer.lower()
        ]
    if l2:
        tasks = [t for t in tasks if t.l2 == l2]
    if l3:
        tasks = [t for t in tasks if t.l3 == l3]
    if l4:
        tasks = [t for t in tasks if t.l4 == l4]

    total = len(tasks)
    start = (page - 1) * page_size
    tasks = tasks[start : start + page_size]

    return TaskListResponse(total=total, tasks=tasks)


@app.get("/api/tasks/filters", tags=["Tasks"])
async def get_filter_options():
    l2_set: dict[str, str] = {}
    l3_set: dict[str, str] = {}
    l4_set: dict[str, str] = {}

    for t in _tasks_cache:
        l2_set[t.l2_id] = t.l2
        l3_set[t.l3_id] = t.l3
        l4_set[t.l4_id] = t.l4

    return {
        "l2": [{"id": k, "name": v} for k, v in sorted(l2_set.items(), key=lambda x: _natural_key(x[0]))],
        "l3": [{"id": k, "name": v} for k, v in sorted(l3_set.items(), key=lambda x: _natural_key(x[0]))],
        "l4": [{"id": k, "name": v} for k, v in sorted(l4_set.items(), key=lambda x: _natural_key(x[0]))],
    }


@app.get("/api/tasks/{task_id}", response_model=Task, tags=["Tasks"])
async def get_task(task_id: str):
    for t in _tasks_cache:
        if t.id == task_id:
            return t
    raise HTTPException(status_code=404, detail=f"Task '{task_id}' 없음")


# ─────────────────────────────────────────────────────────────────────────────
# 분류 엔드포인트 (SSE 스트리밍)
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/classify", tags=["Classify"])
async def classify_tasks(req: ClassifyRequest):
    """
    Task를 분류합니다. provider 필드로 OpenAI / Anthropic 선택 가능.
    Server-Sent Events (SSE) 형식으로 진행 상황을 스트리밍합니다.
    """
    provider = req.provider  # "openai" | "anthropic"
    settings = req.settings or load_settings()

    # 마스킹된 키가 넘어오면 저장된 실제 키로 교체
    if provider == "openai":
        if settings.api_key and settings.api_key.startswith("sk-" + "*"):
            stored = load_settings()
            settings.api_key = stored.api_key
    elif provider == "anthropic":
        if settings.anthropic_api_key and settings.anthropic_api_key.startswith("sk-ant-" + "*"):
            stored = load_settings()
            settings.anthropic_api_key = stored.anthropic_api_key

    if req.task_ids:
        id_set = set(req.task_ids)
        tasks = [t for t in _tasks_cache if t.id in id_set]
    else:
        tasks = list(_tasks_cache)

    if not tasks:
        raise HTTPException(status_code=400, detail="분류할 Task가 없습니다.")

    classifier = get_classifier(settings, provider)
    results_store = load_results(provider)
    total = len(tasks)

    async def event_generator():
        current = 0
        try:
            async for result in classifier.classify_stream(tasks, settings):
                result.provider = provider
                current += 1
                results_store[result.task_id] = result
                payload = {
                    "type": "progress",
                    "task_id": result.task_id,
                    "current": current,
                    "total": total,
                    "result": result.model_dump(),
                }
                yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"

            save_results(results_store, provider)
            yield f"data: {json.dumps({'type': 'done', 'total': current}, ensure_ascii=False)}\n\n"

        except Exception as e:
            print(f"[ERROR] SSE 스트림 오류: {e}")
            yield f"data: {json.dumps({'type': 'error', 'message': '처리 중 오류가 발생했습니다.'}, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ─────────────────────────────────────────────────────────────────────────────
# 결과 엔드포인트
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/results", response_model=ResultsResponse, tags=["Results"])
async def get_results(
    label: Optional[str] = Query(None),
    provider: str = Query("openai", description="openai | anthropic"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=10000),
):
    results_store = load_results(provider)
    results = sorted(results_store.values(), key=lambda r: _natural_key(r.task_id))

    if label:
        results = [r for r in results if r.label == label]

    total = len(results)
    classified = sum(1 for r in results_store.values() if r.label != "미분류")
    unclassified = len(results_store) - classified

    start = (page - 1) * page_size
    paged = results[start : start + page_size]

    return ResultsResponse(
        total=total,
        classified=classified,
        unclassified=unclassified,
        results=paged,
    )


@app.get("/api/results/compare", tags=["Results"])
async def get_comparison_results(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=10000),
):
    """OpenAI와 Anthropic 분류 결과를 나란히 비교합니다."""
    openai_results    = load_results("openai")
    anthropic_results = load_results("anthropic")

    all_ids = sorted(set(openai_results.keys()) | set(anthropic_results.keys()), key=_natural_key)

    comparison = []
    for task_id in all_ids:
        o = openai_results.get(task_id)
        a = anthropic_results.get(task_id)
        comparison.append({
            "task_id": task_id,
            "openai_label":       o.label      if o else None,
            "openai_reason":      o.reason     if o else None,
            "anthropic_label":    a.label      if a else None,
            "anthropic_reason":   a.reason     if a else None,
            "match": (o.label == a.label) if (o and a) else None,
        })

    total    = len(comparison)
    both     = sum(1 for c in comparison if c["openai_label"] and c["anthropic_label"])
    matching = sum(1 for c in comparison if c.get("match") is True)

    start = (page - 1) * page_size
    paged = comparison[start : start + page_size]

    return {
        "total": total,
        "both_classified": both,
        "matching": matching,
        "match_rate": round(matching / both * 100, 1) if both else 0,
        "comparison": paged,
    }


@app.put("/api/results/{task_id}", response_model=ClassificationResult, tags=["Results"])
async def update_result(
    task_id: str,
    update: ClassificationResultUpdate,
    provider: str = Query("openai"),
):
    results_store = load_results(provider)
    existing = results_store.get(task_id)

    if existing is None:
        existing = ClassificationResult(task_id=task_id, provider=provider)

    existing.label = update.label
    if update.reason is not None:
        existing.reason = update.reason
    existing.manually_edited = True

    upsert_result(existing, provider)
    return existing


@app.delete("/api/results", tags=["Results"])
async def delete_all_results(
    request: Request,
    provider: str = Query("openai", description="openai | anthropic | all"),
):
    """분류 결과를 초기화합니다. provider=all 이면 양쪽 모두 초기화."""
    _require_admin(request)
    if provider == "all":
        clear_results("openai")
        clear_results("anthropic")
    else:
        clear_results(provider)
    return {"message": f"'{provider}' 결과가 초기화되었습니다."}


@app.get("/api/results/stats", response_model=StatsResponse, tags=["Results"])
async def get_stats(
    provider: str = Query("openai", description="openai | anthropic"),
):
    results_store = load_results(provider)
    total        = len(_tasks_cache)
    ai_count     = sum(1 for r in results_store.values() if r.label == "AI")
    hybrid_count = sum(1 for r in results_store.values() if r.label == "AI + Human")
    human_count  = sum(1 for r in results_store.values() if r.label == "Human")
    unclassified_count = total - ai_count - hybrid_count - human_count

    l3_map: dict[str, dict] = {}
    for t in _tasks_cache:
        if t.l3 not in l3_map:
            l3_map[t.l3] = {"l3": t.l3, "total": 0, "ai": 0, "hybrid": 0, "human": 0}
        l3_map[t.l3]["total"] += 1
        r = results_store.get(t.id)
        if r:
            if r.label == "AI":
                l3_map[t.l3]["ai"] += 1
            elif r.label == "AI + Human":
                l3_map[t.l3]["hybrid"] += 1
            elif r.label == "Human":
                l3_map[t.l3]["human"] += 1

    return StatsResponse(
        total=total,
        ai_count=ai_count,
        hybrid_count=hybrid_count,
        human_count=human_count,
        unclassified_count=unclassified_count,
        ai_ratio=round(ai_count / total * 100, 1) if total else 0,
        hybrid_ratio=round(hybrid_count / total * 100, 1) if total else 0,
        human_ratio=round(human_count / total * 100, 1) if total else 0,
        by_l3=sorted(l3_map.values(), key=lambda x: x["ai"], reverse=True),
    )


# ─────────────────────────────────────────────────────────────────────────────
# 설정 엔드포인트
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/settings", response_model=ClassifierSettings, tags=["Settings"])
async def get_settings():
    settings = load_settings()
    masked = settings.model_copy()
    if masked.api_key:
        masked.api_key = "sk-" + "*" * 20
    if masked.anthropic_api_key:
        masked.anthropic_api_key = "sk-ant-" + "*" * 20
    return masked


@app.post("/api/settings", response_model=ClassifierSettings, tags=["Settings"])
async def update_settings(request: Request, settings: ClassifierSettings):
    _require_admin(request)
    existing = load_settings()
    # 마스킹된 키(****) → 기존 키 유지, 빈 문자열 → 삭제, 그 외 → 새 키로 교체
    if "*" in settings.api_key:
        settings.api_key = existing.api_key
    if "*" in settings.anthropic_api_key:
        settings.anthropic_api_key = existing.anthropic_api_key
    save_settings(settings)
    masked = settings.model_copy()
    if masked.api_key:
        masked.api_key = "sk-" + "*" * 20
    if masked.anthropic_api_key:
        masked.anthropic_api_key = "sk-ant-" + "*" * 20
    return masked


# ─────────────────────────────────────────────────────────────────────────────
# 엑셀 내보내기
# ─────────────────────────────────────────────────────────────────────────────

def _build_result_sheet(ws, tasks_cache: list[Task], results_store: dict, provider_label: str):
    """결과 시트 생성 공통 함수."""
    headers = ["L5_ID", "L2", "L3", "L4", "L5_Name", "L5_Description", "수행주체",
               "분류결과", "적용기준(Knock-out)", "판단근거", "AI수행필요여건", "수동수정여부"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    ai_fill     = PatternFill("solid", fgColor="FFE0E0")
    hybrid_fill = PatternFill("solid", fgColor="FFF9DB")
    human_fill  = PatternFill("solid", fgColor="D6F5E3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for t in tasks_cache:
        r = results_store.get(t.id)
        label     = r.label if r else "미분류"
        criterion = r.criterion if r else ""
        reason    = r.reason if r else ""
        ai_prereq = r.ai_prerequisites if r else ""
        edited    = "Y" if (r and r.manually_edited) else ""
        row = [t.id, t.l2, t.l3, t.l4, t.name, t.description, t.performer[:100],
               label, criterion, reason, ai_prereq, edited]
        ws.append(row)

        last_row = ws.max_row
        label_cell = ws.cell(row=last_row, column=8)
        if label == "AI":
            label_cell.fill = ai_fill
        elif label == "AI + Human":
            label_cell.fill = hybrid_fill
        elif label == "Human":
            label_cell.fill = human_fill

    col_widths = [12, 10, 14, 14, 28, 45, 35, 16, 28, 40, 50, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w


@app.get("/api/export", tags=["Export"])
async def export_results(
    provider: str = Query("openai", description="openai | anthropic"),
):
    """분류 결과를 엑셀 파일로 다운로드합니다."""
    results_store = load_results(provider)
    provider_label = "OpenAI" if provider == "openai" else "Claude"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"분류결과_{provider_label}"
    _build_result_sheet(ws, _tasks_cache, results_store, provider_label)

    # 요약 시트
    ws2 = wb.create_sheet("요약")
    total    = len(_tasks_cache)
    ai_cnt   = sum(1 for r in results_store.values() if r.label == "AI")
    hyb_cnt  = sum(1 for r in results_store.values() if r.label == "AI + Human")
    hum_cnt  = sum(1 for r in results_store.values() if r.label == "Human")
    rows = [
        ["항목", "값"],
        ["모델", provider_label],
        ["전체 Task", total],
        ["AI", ai_cnt],
        ["AI + Human", hyb_cnt],
        ["Human", hum_cnt],
        ["미분류", total - ai_cnt - hyb_cnt - hum_cnt],
        ["AI 비율", f"{ai_cnt/total*100:.1f}%" if total else "0%"],
        ["AI + Human 비율", f"{hyb_cnt/total*100:.1f}%" if total else "0%"],
    ]
    for row in rows:
        ws2.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    base = _current_excel_path.stem if _current_excel_path else "results"
    filename = f"{base}_a_results.xlsx"
    from urllib.parse import quote as _quote
    encoded_fn = _quote(filename, safe="")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_fn}"},
    )


@app.get("/api/export/compare", tags=["Export"])
async def export_comparison():
    """OpenAI와 Claude 결과를 비교하는 엑셀 다운로드 (3개 시트)."""
    openai_results    = load_results("openai")
    anthropic_results = load_results("anthropic")

    wb = openpyxl.Workbook()

    # 시트 1: OpenAI 결과
    ws1 = wb.active
    ws1.title = "OpenAI (GPT-5.4)"
    _build_result_sheet(ws1, _tasks_cache, openai_results, "OpenAI")

    # 시트 2: Claude 결과
    ws2 = wb.create_sheet("Claude (Sonnet 4.6)")
    _build_result_sheet(ws2, _tasks_cache, anthropic_results, "Claude")

    # 시트 3: 비교
    ws3 = wb.create_sheet("비교")
    comp_headers = ["L5_ID", "L5_Name", "OpenAI 결과", "Claude 결과", "일치 여부",
                    "OpenAI 근거", "Claude 근거"]
    ws3.append(comp_headers)
    header_fill = PatternFill("solid", fgColor="374151")
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    match_fill    = PatternFill("solid", fgColor="D1FAE5")
    mismatch_fill = PatternFill("solid", fgColor="FEE2E2")

    task_map = {t.id: t for t in _tasks_cache}
    all_ids  = sorted(set(openai_results.keys()) | set(anthropic_results.keys()), key=_natural_key)

    for task_id in all_ids:
        o = openai_results.get(task_id)
        a = anthropic_results.get(task_id)
        task = task_map.get(task_id)
        name = task.name if task else task_id
        match = (o.label == a.label) if (o and a) else None

        row = [
            task_id,
            name,
            o.label if o else "-",
            a.label if a else "-",
            "✓ 일치" if match is True else ("✗ 불일치" if match is False else "-"),
            o.reason if o else "",
            a.reason if a else "",
        ]
        ws3.append(row)

        last = ws3.max_row
        fill = match_fill if match is True else (mismatch_fill if match is False else None)
        if fill:
            for col in range(1, 6):
                ws3.cell(row=last, column=col).fill = fill

    for i, w in enumerate([12, 35, 16, 16, 12, 40, 40], 1):
        ws3.column_dimensions[ws3.cell(1, i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote as _quote
    _cmp_fn = f"{(_current_excel_path.stem if _current_excel_path else 'compare')}_a_results_compare.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{_quote(_cmp_fn, safe='')}"},
    )


# ─────────────────────────────────────────────────────────────────────────────
# 엑셀 업로드
# ─────────────────────────────────────────────────────────────────────────────

_PERSIST_ROOT = Path("/app/persist") if Path("/app/persist").exists() else Path(__file__).parent
_UPLOAD_DIR   = _PERSIST_ROOT / "uploads"        # Task 분류 엑셀
_WF_DIR       = _PERSIST_ROOT / "workflow"       # Workflow 설계 세션
_NW_DIR       = _PERSIST_ROOT / "new_workflow"   # New Workflow / 과제 정의서 / 설계서
for _d in (_UPLOAD_DIR, _WF_DIR, _NW_DIR):
    _d.mkdir(exist_ok=True)
_current_excel_path: Path | None = None

# ── 멀티 세션 지원 ────────────────────────────────────────────────────────────
_SESSIONS_DIR = _WF_DIR / "sessions"
_SESSIONS_DIR.mkdir(exist_ok=True)
_current_session_id: str = ""   # 현재 활성 세션 ID (JSON 파일명 기반)
_sessions_manifest: dict = {}   # {session_id: {name, created_at, updated_at, excel_file, json_file, ppt_file}}


def _get_auth_user(request: Request) -> dict | None:
    """Authorization 헤더 토큰으로 로그인된 사용자 정보 반환."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "").strip()
    if not token:
        return None
    return get_session_user(token)

def _get_team_id(request: Request) -> str:
    """로그인된 사용자의 프로젝트(팀)를 team_id로 반환. 미로그인 or 공통='default'."""
    user = _get_auth_user(request)
    if not user:
        return "default"
    proj = user.get("project")  # None = 공통(전체 접근), str = 'SKI'/'두산' 등
    if not proj:
        return "공통"  # 공통 멤버 — 전체 조회 가능
    return proj

def _get_user_id(request: Request) -> str:
    """로그인된 사용자 이름 반환. 미로그인='unknown'."""
    user = _get_auth_user(request)
    if not user:
        return "unknown"
    return user.get("name", "unknown")


def _get_session_dir(sid: str) -> Path:
    """세션 디렉토리 경로 반환 (없으면 생성)."""
    import re
    safe = re.sub(r'[^\w가-힣\-]', '_', sid)[:80] or "default"
    d = _SESSIONS_DIR / safe
    d.mkdir(exist_ok=True)
    return d


def _load_sessions_manifest() -> dict:
    global _sessions_manifest
    p = _WF_DIR / "sessions.json"
    if p.exists():
        try:
            _sessions_manifest = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            _sessions_manifest = {}
    return _sessions_manifest


def _save_sessions_manifest() -> None:
    p = _WF_DIR / "sessions.json"
    p.write_text(json.dumps(_sessions_manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def _save_session_data(sid: str) -> None:
    """현재 메모리 상태를 세션 디렉토리의 session_data.json에 저장."""
    if not sid:
        return


    def _ser(obj):
        """JSON 직렬화 불가 객체 처리."""
        try:
            return json.dumps(obj, ensure_ascii=False)
        except Exception:
            return "{}"

    data = {
        "classification": _wf_classification,
        "benchmark_table": _wf_benchmark_table,
        "step1": _wf_step1_cache,
        "step2": _wf_step2_cache,
        "gap_analysis": _wf_gap_analysis,
        "user_resources": [
            {k: v for k, v in r.items() if k != "image_b64"}  # b64 제외 (파일로 저장)
            for r in _wf_user_resources
        ],
        "saved_at": _now_kst(),
    }
    try:
        d = _get_session_dir(sid)
        (d / "session_data.json").write_text(
            json.dumps(data, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        # manifest 업데이트
        if sid in _sessions_manifest:
            _sessions_manifest[sid]["updated_at"] = _now_kst()
            _save_sessions_manifest()
    except Exception as e:
        print(f"[SESSION] session_data 저장 실패({sid}): {e}", flush=True)


def _load_session_data(sid: str) -> bool:
    """세션 디렉토리에서 state를 메모리에 복구. 성공 시 True 반환."""
    global _wf_classification, _wf_benchmark_table, _wf_step1_cache, _wf_step2_cache, _wf_gap_analysis
    global _workflow_cache, _wf_excel_tasks, _wf_excel_path, _wf_user_resources
    from workflow_parser import parse_workflow_json, get_workflow_summary
    from excel_reader import load_tasks

    d = _get_session_dir(sid)

    # JSON 복구
    jp = d / "workflow.json"
    if jp.exists():
        try:
            raw = json.loads(jp.read_text(encoding="utf-8"))
            parsed = parse_workflow_json(raw)
            summary = get_workflow_summary(parsed)
            _workflow_cache = {"filename": jp.name, "parsed": parsed, "summary": summary, "raw": raw}
        except Exception as e:
            print(f"[SESSION] JSON 복구 실패({sid}): {e}", flush=True)

    # Excel 복구 — 여러 파일이 있으면 가장 최신 파일 사용
    xls = sorted(d.glob("*.xlsx"), key=lambda f: f.stat().st_mtime)
    if xls:
        try:
            tasks = load_tasks(str(xls[-1]))
            _wf_excel_tasks = tasks
            _wf_excel_path = str(xls[-1])
            # 분류 결과는 session_data.json에서 복구하므로 여기서는 비움
        except Exception as e:
            print(f"[SESSION] Excel 복구 실패({sid}): {e}", flush=True)

    # session_data.json 복구 (classification, benchmark, step1, step2, gap)
    dp = d / "session_data.json"
    if dp.exists():
        try:
            sd = json.loads(dp.read_text(encoding="utf-8"))
            _wf_classification = sd.get("classification", {})
            _wf_benchmark_table = sd.get("benchmark_table", {})
            _wf_step1_cache = sd.get("step1", {})
            _wf_step2_cache = sd.get("step2", {})
            _wf_gap_analysis = sd.get("gap_analysis", {})
            _wf_user_resources = sd.get("user_resources", [])
            # 이미지 리소스의 image_b64 복원 (파일에서 재로드)
            import base64 as _b64mod
            for res in _wf_user_resources:
                if res.get("type") == "image" and res.get("image_path"):
                    try:
                        p = Path(res["image_path"])
                        if p.exists():
                            res["image_b64"] = _b64mod.b64encode(p.read_bytes()).decode()
                    except Exception:
                        pass
        except Exception as e:
            print(f"[SESSION] session_data 복구 실패({sid}): {e}", flush=True)

    return jp.exists() or bool(xls)


@app.get("/api/upload/current", tags=["Upload"])
async def get_current_file():
    """현재 세션에서 업로드한 파일 정보만 반환합니다."""
    if _current_excel_path and _current_excel_path.exists():
        return {
            "filename": _current_excel_path.name,
            "size_kb": round(_current_excel_path.stat().st_size / 1024, 1),
            "task_count": len(_tasks_cache),
        }
    # 이 세션에서 업로드한 파일이 없으면 빈 응답
    return {"filename": None, "size_kb": 0, "task_count": 0}


@app.post("/api/upload", tags=["Upload"])
async def upload_excel(request: Request, file: UploadFile = File(...)):
    """엑셀 파일 업로드 — 시트 목록을 반환합니다. 시트 선택은 /api/upload/select-sheet로."""
    global _tasks_cache, _current_excel_path
    from excel_reader import list_sheets

    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail=".xlsx 파일만 업로드 가능합니다.")

    # 파일명에서 경로 순회 방지 (basename만 사용)
    safe_filename = Path(file.filename).name
    save_path = _UPLOAD_DIR / safe_filename
    try:
        contents = await file.read()
        save_path.write_bytes(contents)
    except Exception as e:
        print(f"[ERROR] 파일 저장 실패: {e}")
        raise HTTPException(status_code=500, detail="파일 저장에 실패했습니다.")

    # 시트 목록 조회
    try:
        sheets = list_sheets(save_path)
    except Exception as e:
        save_path.unlink(missing_ok=True)
        print(f"[ERROR] 엑셀 파싱 실패: {e}")
        raise HTTPException(status_code=422, detail="엑셀 파일을 읽을 수 없습니다. 올바른 형식인지 확인하세요.")

    _current_excel_path = save_path

    # 파일별 결과 분리를 위해 현재 파일 설정
    set_current_file(file.filename)
    set_current_project(file.filename)
    save_meta(file.filename, task_count=len(_tasks_cache), source="tasks_page")

    # 추천 시트가 있으면 자동 로드
    recommended = next((s for s in sheets if s["recommended"]), None)
    if recommended:
        try:
            new_tasks = load_tasks(save_path, sheet_name=recommended["name"])
            _tasks_cache = new_tasks
        except Exception:
            pass

    # 사용자 프로젝트 기록
    user_ctx = _get_user_context(request)
    team_project = user_ctx.get("project") if user_ctx else None
    save_meta_team(file.filename, team_project=team_project, source="tasks_page")
    audit_log.log_event("excel_upload", email=user_ctx.get("email", "") if user_ctx else "",
                        ip=_get_client_ip(request),
                        detail=f"{file.filename} ({len(_tasks_cache)} tasks) [{team_project or '공통'}]")

    return {
        "message": "업로드 성공",
        "filename": file.filename,
        "task_count": len(_tasks_cache),
        "sheets": sheets,
    }


@app.get("/api/upload/sheets", tags=["Upload"])
async def get_excel_sheets():
    """현재 업로드된 엑셀 파일의 시트 목록을 반환합니다."""
    from excel_reader import list_sheets

    if not _current_excel_path or not _current_excel_path.exists():
        raise HTTPException(404, "업로드된 엑셀 파일이 없습니다.")

    sheets = list_sheets(_current_excel_path)
    return {"filename": _current_excel_path.name, "sheets": sheets}


@app.post("/api/upload/select-sheet", tags=["Upload"])
async def select_excel_sheet(request: Request):
    """특정 시트를 선택하여 Task를 로드합니다. Body: {"sheet_name": "시트명"}"""
    global _tasks_cache

    body = await request.json()
    sheet_name = body.get("sheet_name", "")

    if not _current_excel_path or not _current_excel_path.exists():
        raise HTTPException(404, "업로드된 엑셀 파일이 없습니다.")

    try:
        new_tasks = load_tasks(_current_excel_path, sheet_name=sheet_name)
    except Exception as e:
        print(f"[ERROR] 시트 파싱 실패: {e}")
        raise HTTPException(status_code=422, detail="시트를 읽을 수 없습니다. 올바른 형식인지 확인하세요.")

    _tasks_cache = new_tasks

    return {
        "message": f"시트 '{sheet_name}' 로드 완료",
        "sheet_name": sheet_name,
        "task_count": len(_tasks_cache),
    }


# ─────────────────────────────────────────────────────────────────────────────
# 헬스체크
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/health", tags=["System"])
async def health():
    env_openai    = os.environ.get("OPENAI_API_KEY", "").strip()
    env_anthropic = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    s = load_settings()
    return {
        "status": "ok",
        "task_count": len(_tasks_cache),
        "api_key_configured": bool(env_openai or s.api_key.strip()),
        "openai_configured": bool(env_openai or s.api_key.strip()),
        "anthropic_configured": bool(env_anthropic or s.anthropic_api_key.strip()),
    }


@app.get("/api/usage", tags=["System"])
async def get_usage_stats():
    """누적 API 토큰 사용량 및 예상 비용을 반환합니다."""
    return get_usage()


@app.delete("/api/usage", tags=["System"])
async def reset_usage_stats(provider: str = Query("all")):
    """사용량을 초기화합니다. provider=all|openai|anthropic"""
    reset_usage(provider)
    return {"ok": True, "reset": provider}


# ─────────────────────────────────────────────────────────────────────────────
# Workflow 엔드포인트
# ─────────────────────────────────────────────────────────────────────────────

_workflow_cache: dict = {}   # 최근 업로드된 워크플로우 저장


@app.post("/api/workflow/upload", tags=["Workflow"])
async def upload_workflow(request: Request, file: UploadFile = File(...)):
    """hr-workflow-ai에서 내보낸 JSON 파일을 업로드하여 파싱합니다."""
    from workflow_parser import parse_workflow_json, get_workflow_summary

    content = await file.read()
    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        raise HTTPException(400, "유효한 JSON 파일이 아닙니다.")

    parsed = parse_workflow_json(data)
    summary = get_workflow_summary(parsed)

    # 세션 ID = JSON 파일명 stem (e.g. "발령관리.json" → "발령관리")

    raw_fname = file.filename or "workflow.json"
    sid = Path(raw_fname).stem or "default"

    global _workflow_cache, _current_session_id
    _workflow_cache = {
        "filename": raw_fname,
        "parsed": parsed,
        "summary": summary,
        "raw": data,
    }
    _current_session_id = sid

    # 세션 디렉토리에 JSON 저장
    sess_dir = _get_session_dir(sid)
    json_path = sess_dir / "workflow.json"
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    # manifest 등록/갱신
    _load_sessions_manifest()
    now = _now_kst()
    user_id = _get_user_id(request)
    team_id = _get_team_id(request)
    if sid not in _sessions_manifest:
        _sessions_manifest[sid] = {"id": sid, "name": sid, "created_at": now, "user_id": user_id, "team_id": team_id}
    _sessions_manifest[sid].update({
        "updated_at": now,
        "json_file": "workflow.json",
        "user_id": user_id,
        "team_id": _sessions_manifest[sid].get("team_id", team_id),
    })
    _sessions_manifest["_current"] = sid
    _save_sessions_manifest()

    return {
        "ok": True,
        "filename": raw_fname,
        "session_id": sid,
        **_enrich_summary_with_cls(summary),
    }


def _enrich_summary_with_cls(summary: dict) -> dict:
    """summary의 l4_details.child_l5s 각 항목에 엑셀 분류 라벨을 붙입니다.
    엑셀 task id == As-Is L5 node task_id 인 경우 직접 매핑."""
    if not _wf_classification:
        return summary

    import copy
    result = copy.deepcopy(summary)
    for sheet in result.get("sheets", []):
        for l4 in sheet.get("l4_details", []):
            for l5 in l4.get("child_l5s", []):
                tid = l5.get("task_id", "")
                cls = _wf_classification.get(tid, {})
                l5["cls_label"] = cls.get("label", "") if cls else ""
                l5["cls_reason"] = cls.get("reason", "") if cls else ""
    return result


@app.get("/api/workflow/summary", tags=["Workflow"])
async def get_workflow():
    """업로드된 워크플로우의 요약 정보를 반환합니다."""
    from workflow_parser import parse_workflow_json, get_workflow_summary

    global _workflow_cache
    if "summary" not in _workflow_cache:
        # 현재 세션 디렉토리에서 workflow.json 로드 시도
        save_path = (
            _get_session_dir(_current_session_id) / "workflow.json"
            if _current_session_id else None
        )
        if save_path and save_path.exists():
            data = json.loads(save_path.read_text(encoding="utf-8"))
            parsed = parse_workflow_json(data)
            summary = get_workflow_summary(parsed)
            _workflow_cache.update({
                "filename": save_path.name,
                "parsed": parsed,
                "summary": summary,
                "raw": data,
            })
        else:
            raise HTTPException(404, "업로드된 워크플로우가 없습니다.")

    return _enrich_summary_with_cls(_workflow_cache["summary"])


@app.get("/api/workflow/sheets", tags=["Workflow"])
async def list_workflow_sheets():
    """워크플로우 시트 목록을 반환합니다."""
    if "parsed" not in _workflow_cache:
        raise HTTPException(404, "업로드된 워크플로우가 없습니다. JSON 파일을 먼저 업로드하세요.")

    parsed = _workflow_cache["parsed"]
    return {
        "sheets": [
            {
                "sheet_id": s.sheet_id,
                "name": s.name,
                "lanes": s.lanes,
                "l4_count": len(s.l4_nodes),
                "l5_count": len(s.l5_nodes),
                "total_steps": len(s.execution_order),
            }
            for s in parsed.sheets
        ]
    }


@app.get("/api/workflow/sheets/{sheet_id}", tags=["Workflow"])
async def get_workflow_sheet_detail(sheet_id: str):
    """특정 시트의 상세 정보 (노드, 엣지, 실행 순서)를 반환합니다."""
    if "parsed" not in _workflow_cache:
        raise HTTPException(404, "업로드된 워크플로우가 없습니다. JSON 파일을 먼저 업로드하세요.")

    parsed = _workflow_cache["parsed"]
    sheet = next((s for s in parsed.sheets if s.sheet_id == sheet_id), None)
    if not sheet:
        raise HTTPException(404, f"시트 '{sheet_id}'를 찾을 수 없습니다.")

    return {
        "sheet_id": sheet.sheet_id,
        "name": sheet.name,
        "lanes": sheet.lanes,
        "nodes": [
            {
                "node_id": n.id,
                "level": n.level,
                "task_id": n.task_id,
                "label": n.label,
                "description": n.description,
                "position": {"x": n.position_x, "y": n.position_y},
                "metadata": n.metadata,
            }
            for n in sorted(sheet.nodes.values(), key=lambda n: (n.y, n.x))
        ],
        "edges": [
            {
                "id": e.id,
                "source": e.source,
                "target": e.target,
                "label": e.label,
            }
            for e in sheet.edges
        ],
        "execution_order": [
            {
                "step": s.step_number,
                "nodes": [
                    {
                        "node_id": nid,
                        "task_id": sheet.nodes[nid].task_id if nid in sheet.nodes else "",
                        "label": sheet.nodes[nid].label if nid in sheet.nodes else "",
                        "level": sheet.nodes[nid].level if nid in sheet.nodes else "",
                    }
                    for nid in s.node_ids
                ],
                "is_parallel": s.is_parallel,
            }
            for s in sheet.execution_order
        ],
    }


@app.get("/api/workflow/execution-order/{sheet_id}", tags=["Workflow"])
async def get_execution_order(sheet_id: str):
    """특정 시트의 실행 순서만 간결하게 반환합니다."""
    if "parsed" not in _workflow_cache:
        raise HTTPException(404, "업로드된 워크플로우가 없습니다. JSON 파일을 먼저 업로드하세요.")

    parsed = _workflow_cache["parsed"]
    sheet = next((s for s in parsed.sheets if s.sheet_id == sheet_id), None)
    if not sheet:
        raise HTTPException(404, f"시트 '{sheet_id}'를 찾을 수 없습니다.")

    steps = []
    for s in sheet.execution_order:
        step_info = {
            "step": s.step_number,
            "type": "병렬" if s.is_parallel else "순차",
            "tasks": [],
        }
        for nid in s.node_ids:
            node = sheet.nodes.get(nid)
            if node:
                step_info["tasks"].append({
                    "task_id": node.task_id,
                    "label": node.label,
                    "level": node.level,
                })
        steps.append(step_info)

    return {
        "sheet_id": sheet_id,
        "sheet_name": sheet.name,
        "total_steps": len(steps),
        "parallel_count": sum(1 for s in steps if s["type"] == "병렬"),
        "sequential_count": sum(1 for s in steps if s["type"] == "순차"),
        "steps": steps,
    }


# ─────────────────────────────────────────────────────────────────────────────
# PPT 업로드 + 태스크 매칭
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/workflow/upload-ppt", tags=["Workflow"])
async def upload_ppt_workflow(request: Request, file: UploadFile = File(...)):
    """PPT 파일을 업로드하여 슬라이드별 노드를 추출하고 태스크와 매칭합니다."""
    from ppt_parser import parse_ppt, match_nodes_to_tasks, ppt_slide_to_react_flow, ppt_to_parsed_workflow


    content = await file.read()
    safe_ppt_name = Path(file.filename or "workflow.pptx").name

    # 세션 디렉토리에 저장
    _load_sessions_manifest()
    sid = _current_session_id or _sessions_manifest.get("_current", "default")
    sess_dir = _get_session_dir(sid)
    ppt_save_path = sess_dir / safe_ppt_name
    ppt_save_path.write_bytes(content)

    # manifest 갱신
    now = _now_kst()
    user_id = _get_user_id(request)
    team_id = _get_team_id(request)
    if sid not in _sessions_manifest:
        _sessions_manifest[sid] = {"id": sid, "name": sid, "created_at": now, "user_id": user_id, "team_id": team_id}
    _sessions_manifest[sid].update({
        "updated_at": now,
        "ppt_file": safe_ppt_name,
        "user_id": user_id,
        "team_id": _sessions_manifest[sid].get("team_id", team_id),
    })
    _save_sessions_manifest()

    try:
        parsed = parse_ppt(content)
    except Exception as e:
        print(f"[ERROR] PPT 파싱 실패: {e}")
        raise HTTPException(400, "PPT 파일을 읽을 수 없습니다.")

    # 태스크 매칭용 데이터 준비
    task_list = [
        {
            "id": t.id, "name": t.name,
            "l4": t.l4, "l4_id": t.l4_id,
            "l3": t.l3, "l3_id": t.l3_id,
        }
        for t in _tasks_cache
    ]

    slides_result = []
    for slide in parsed.slides:
        # 노드-태스크 매칭
        matches = match_nodes_to_tasks(slide.nodes, task_list)

        # React Flow 변환
        react_flow = ppt_slide_to_react_flow(slide)

        # 매칭 결과를 React Flow 노드에 반영
        match_map = {m["node_id"]: m for m in matches}
        for rf_node in react_flow["nodes"]:
            m = match_map.get(rf_node["id"])
            if m and m["matched_task_id"]:
                rf_node["data"]["id"] = m["matched_task_id"]
                rf_node["data"]["matchedTaskName"] = m["matched_task_name"]
                rf_node["data"]["matchConfidence"] = m["match_confidence"]
                rf_node["data"]["matchedLevel"] = m["matched_level"]

        slides_result.append({
            "slide_index": slide.index,
            "title": slide.title,
            "node_count": len(slide.nodes),
            "edge_count": len(slide.edges),
            "matches": matches,
            "react_flow": react_flow,
        })

    # PPT → ParsedWorkflow 변환 (To-Be 생성에서 사용)
    all_matches = [
        match_nodes_to_tasks(slide.nodes, task_list)
        for slide in parsed.slides
    ]
    ppt_workflow = ppt_to_parsed_workflow(parsed, all_matches)

    # 워크플로우 캐시에 PPT 결과 + 변환된 워크플로우 저장
    global _workflow_cache
    _workflow_cache["ppt"] = {
        "filename": file.filename,
        "parsed": parsed,
        "slides": slides_result,
    }
    # JSON 워크플로우와 동일한 키로 저장 → generate-tobe에서 사용 가능
    _workflow_cache["parsed"] = ppt_workflow

    return {
        "ok": True,
        "filename": file.filename,
        "slide_count": parsed.slide_count,
        "slides": slides_result,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Workflow — 엑셀 업로드 (분류 결과 포함) + 2단계 설계
# ─────────────────────────────────────────────────────────────────────────────

_wf_excel_tasks: list = []     # Workflow용 엑셀 Task 캐시
_wf_excel_path: str = ""       # 현재 업로드된 엑셀 파일 경로 (항상 이 경로만 사용)
_wf_classification: dict = {}  # 엑셀에서 추출한 분류 결과
_wf_step1_cache: dict = {}     # Step 1 결과 캐시
_wf_step2_cache: dict = {}     # Step 2 결과 캐시
_wf_tobe_flow_cache: dict = {} # generate-tobe-flow 결과 캐시 (swim lane JSON)
_wf_chat_history: list = []    # Step 1 채팅 이력
_manual_matches: dict = {}     # 수동 매칭: json_task_id → excel_task_id
_wf_user_resources: list = []  # 사용자가 첨부한 URL/이미지 리소스 (누적)

def _build_deprecated_scope() -> tuple[set[str], set[str], set[str]]:
    """Step 1 redesigned_process 에서 change_type='삭제'/'통합' L3/L4/L5 수집.
    Returns: (task_ids, l4_ids, l3_ids) — Swim Lane · AI Service Flow 공통 제외 대상."""
    task_ids: set[str] = set()
    l4_ids: set[str] = set()
    l3_ids: set[str] = set()
    for l3 in (_wf_step1_cache.get("redesigned_process") or []):
        l3_ct = str(l3.get("change_type") or "").strip()
        l3_id = str(l3.get("l3_id") or "").strip()
        if l3_ct in ("삭제", "통합") and l3_id:
            l3_ids.add(l3_id)
        for l4 in (l3.get("l4_list") or []):
            l4_ct = str(l4.get("change_type") or "").strip()
            l4_id = str(l4.get("l4_id") or "").strip()
            if l4_ct in ("삭제", "통합") and l4_id:
                l4_ids.add(l4_id)
            for l5 in (l4.get("l5_list") or []):
                ct = str(l5.get("change_type") or "").strip()
                tid = str(l5.get("task_id") or "").strip()
                if ct in ("삭제", "통합") and tid and not tid.startswith("NEW"):
                    task_ids.add(tid)
    return task_ids, l4_ids, l3_ids


def _make_deprecation_check():
    """폐기/통합 판정 클로저. Returns: is_deprecated(task_id: str) -> bool"""
    tids, l4s, l3s = _build_deprecated_scope()

    def _check(tid: str) -> bool:
        tid = (tid or "").strip()
        if not tid:
            return False
        if tid in tids:
            return True
        parts = tid.split(".")
        if len(parts) >= 3 and ".".join(parts[:3]) in l4s:
            return True
        if len(parts) >= 2 and ".".join(parts[:2]) in l3s:
            return True
        return False
    return _check


def _filter_step2_deprecated(step2: dict) -> dict:
    """Step 2 캐시를 deep-copy 후 assigned_tasks 에서 Step 1 폐기/통합 task 제거.
    AI Service Flow · HTML · PPT · Excel export 직전 공통 필터."""
    if not step2:
        return step2
    import copy
    is_deprecated = _make_deprecation_check()
    cleaned = copy.deepcopy(step2)
    removed = 0
    for agent in cleaned.get("agents", []) or []:
        before = len(agent.get("assigned_tasks", []) or [])
        agent["assigned_tasks"] = [
            t for t in (agent.get("assigned_tasks") or [])
            if not is_deprecated(str(t.get("task_id") or ""))
        ]
        removed += before - len(agent["assigned_tasks"])
    # execution_flow 의 task_ids 에도 동일 필터 적용
    for step in cleaned.get("execution_flow", []) or []:
        step["task_ids"] = [
            tid for tid in (step.get("task_ids") or [])
            if not is_deprecated(str(tid))
        ]
    if removed:
        print(f"[STEP2-FILTER] Step 1 폐기/통합 필터 — assigned_tasks 에서 {removed}건 제거", flush=True)
    return cleaned


def _manual_matches_path() -> Path:
    """현재 세션의 manual_matches.json 경로. 세션 없으면 _WF_DIR 루트 (하위 호환)."""
    if _current_session_id:
        return _get_session_dir(_current_session_id) / "manual_matches.json"
    return _WF_DIR / "manual_matches.json"


def _load_manual_matches():
    global _manual_matches
    path = _manual_matches_path()
    try:
        _manual_matches = json.loads(path.read_text(encoding="utf-8")) if path.exists() else {}
    except Exception:
        _manual_matches = {}


def _save_manual_matches():
    _manual_matches_path().write_text(
        json.dumps(_manual_matches, ensure_ascii=False, indent=2), encoding="utf-8"
    )


@app.post("/api/workflow/upload-excel", tags=["Workflow"])
async def upload_workflow_excel(request: Request, file: UploadFile = File(...)):
    """
    분류 결과가 포함된 엑셀 업로드.
    As-Is 프로세스 + 분류 결과를 읽어 Workflow 설계에 활용합니다.
    """
    from excel_reader import load_tasks, list_sheets

    global _current_session_id

    content = await file.read()
    _WF_DIR.mkdir(exist_ok=True)

    safe_wf_excel_name = Path(file.filename or "workflow_excel.xlsx").name

    # 세션 디렉토리에 저장 (현재 세션 없으면 파일명 stem을 세션 ID로)
    _load_sessions_manifest()
    sid = _current_session_id or _sessions_manifest.get("_current", "")
    if not sid:
        sid = Path(safe_wf_excel_name).stem or "default"
        _current_session_id = sid
    sess_dir = _get_session_dir(sid)

    # 세션 디렉토리에 저장 (같은 이름이면 덮어쓰기, 다른 이름이면 추가 보존)
    save_path = sess_dir / safe_wf_excel_name
    save_path.write_bytes(content)

    # 시트 목록
    sheets = list_sheets(str(save_path))

    # 추천 시트로 Task 로드
    recommended = next((s["name"] for s in sheets if s.get("recommended")), None)
    tasks = load_tasks(str(save_path), sheet_name=recommended)

    global _wf_excel_tasks, _wf_excel_path, _wf_classification, _wf_chat_history
    global _wf_step1_cache, _wf_step2_cache, _wf_benchmark_table, _wf_user_resources
    _wf_excel_tasks = tasks
    _wf_excel_path = str(save_path)   # 현재 파일 경로 고정

    # manifest 갱신
    now = _now_kst()
    user_id = _get_user_id(request)
    team_id = _get_team_id(request)
    if sid not in _sessions_manifest:
        _sessions_manifest[sid] = {"id": sid, "name": sid, "created_at": now, "user_id": user_id, "team_id": team_id}
    _sessions_manifest[sid].update({
        "updated_at": now,
        "excel_file": safe_wf_excel_name,
        "user_id": user_id,
        "team_id": _sessions_manifest[sid].get("team_id", team_id),
    })
    _sessions_manifest["_current"] = sid
    _save_sessions_manifest()
    _wf_classification = {}
    _wf_step1_cache = {}
    _wf_step2_cache = {}
    _wf_chat_history = []
    _wf_benchmark_table = {}
    _wf_user_resources = []

    # 분류 결과 추출 (최종 > 두산 > 1차 순 fallback)
    # '-'(대시)는 "두산 검토: 변경 불필요" 표시이므로 실제 label이 아님 → skip
    def _resolve_label(*candidates: str) -> str:
        for c in candidates:
            v = (c or "").strip()
            if v and v != "-":
                return v
        return ""

    has_classification = False
    for t in tasks:
        label = _resolve_label(t.cls_final_label, t.cls_doosan_label, t.cls_1st_label)
        if label:
            has_classification = True
            reason = t.cls_1st_reason or ""
            knockout = t.cls_1st_knockout or ""
            ai_prereq = t.cls_1st_ai_prereq or ""
            feedback = next((v for x in [t.cls_final_feedback, t.cls_doosan_feedback] if (v := (x or "").strip()) and v != "-"), "")
            _wf_classification[t.id] = {
                "label": label,
                "reason": reason,
                "criterion": knockout,
                "ai_prerequisites": ai_prereq,
                "feedback": feedback,
                "task_name": t.name,
                "hybrid_note": "",
                "input_types": "",
                "output_types": "",
            }

    # 엑셀 파일도 _tasks_cache에 반영 (기존 분류 로직과 호환)
    global _tasks_cache
    _tasks_cache = tasks

    # AI+Human Task 파트 분리 (fire-and-forget — Step 2 진입 시 lazy 보장)
    import asyncio as _asyncio
    _asyncio.create_task(_split_hybrid_tasks_with_llm())

    # 가이드 시트 및 데이터 없는 시트는 선택 목록에서 제외
    data_sheets = [
        s for s in sheets
        if not s.get("is_guide", False) and s.get("task_count", 0) > 0
    ]

    return {
        "ok": True,
        "filename": file.filename,
        "session_id": sid,
        "task_count": len(tasks),
        "has_classification": has_classification,
        "classified_count": len(_wf_classification),
        "sheets": [
            {
                "name": s["name"],
                "recommended": s.get("recommended", False),
                "row_count": s.get("task_count", 0),
                "l5_count": s.get("task_count", 0),
            }
            for s in data_sheets
        ],
    }


@app.post("/api/workflow/select-excel-sheet", tags=["Workflow"])
async def select_workflow_excel_sheet(request: Request):
    """엑셀 시트 선택 (Workflow용)."""
    from excel_reader import load_tasks

    global _wf_excel_tasks, _wf_excel_path, _wf_classification, _tasks_cache

    body = await request.json()
    sheet_name = body.get("sheet_name", "")
    if not sheet_name:
        raise HTTPException(400, "시트 이름이 필요합니다.")

    # 업로드 시 고정된 경로 우선 사용 → 없으면 현재 세션 디렉토리 fallback
    if _wf_excel_path and Path(_wf_excel_path).exists():
        excel_path = _wf_excel_path
    else:
        sess_excels = sorted(
            _get_session_dir(_current_session_id).glob("*.xlsx"),
            key=lambda x: x.stat().st_mtime,
        ) if _current_session_id else []
        if not sess_excels:
            raise HTTPException(404, "업로드된 엑셀 파일이 없습니다.")
        excel_path = str(sess_excels[-1])

    tasks = load_tasks(excel_path, sheet_name=sheet_name)

    _wf_excel_tasks = tasks
    _wf_excel_path = excel_path   # 시트 선택해도 경로 유지
    _tasks_cache = tasks
    _wf_classification = {}

    def _resolve_lbl(*candidates: str) -> str:
        for c in candidates:
            v = (c or "").strip()
            if v and v != "-":
                return v
        return ""

    for t in tasks:
        label = _resolve_lbl(t.cls_final_label, t.cls_doosan_label, t.cls_1st_label)
        if label:
            _wf_classification[t.id] = {
                "label": label,
                "reason": t.cls_1st_reason or "",
                "criterion": t.cls_1st_knockout or "",
                "ai_prerequisites": t.cls_1st_ai_prereq or "",
                "feedback": next((v for x in [t.cls_final_feedback, t.cls_doosan_feedback] if (v := (x or "").strip()) and v != "-"), ""),
                "task_name": t.name,
                "hybrid_note": "",
                "input_types": "",
                "output_types": "",
            }

    # AI+Human Task 파트 분리 (fire-and-forget — Step 2 진입 시 lazy 보장)
    import asyncio as _asyncio
    _asyncio.create_task(_split_hybrid_tasks_with_llm())

    return {
        "ok": True,
        "sheet_name": sheet_name,
        "task_count": len(tasks),
        "classified_count": len(_wf_classification),
    }


@app.get("/api/workflow/excel-tasks", tags=["Workflow"])
async def get_workflow_excel_tasks():
    """엑셀에서 로드한 Task 목록 + 분류 결과를 반환합니다."""
    tasks = []
    for t in _wf_excel_tasks:
        cls = _wf_classification.get(t.id, {})
        tasks.append({
            "id": t.id,
            "l2": t.l2, "l3": t.l3, "l4": t.l4,
            "name": t.name,
            "description": t.description,
            "performer": t.performer,
            "label": cls.get("label", ""),
            "reason": cls.get("reason", ""),
            "criterion": cls.get("criterion", ""),
            "ai_prerequisites": cls.get("ai_prerequisites", ""),
            "feedback": cls.get("feedback", ""),
        })
    return {
        "total": len(tasks),
        "classified": len(_wf_classification),
        "tasks": tasks,
    }


_wf_benchmark_results: list = []   # 벤치마킹 검색 결과 (raw)
_wf_benchmark_table: dict = {}     # 시트별 벤치마킹 결과 테이블 {sheet_id: [rows]}
_wf_search_log: list = []          # 벤치마킹 검색 로그 (thinking process)
_wf_gap_analysis: dict = {}        # Gap 분석 결과 캐시


# ── AI+Human 분리 (hybrid split) ──────────────────────────────────────────────

async def _split_hybrid_tasks_with_llm(target_task_ids: set[str] | None = None) -> int:
    """
    AI+Human으로 분류된 Task에 대해 AI 파트·Human 파트를 LLM으로 자동 분리.
    hybrid_note가 비어있는 Task만 처리하고 결과를 _wf_classification에 저장.

    Args:
        target_task_ids: None이면 전체 분리, set을 주면 해당 task_id만 분리 (Step 2 스코프 한정용)

    Returns:
        분리된 Task 수
    """
    global _wf_classification
    if not _wf_classification or not _wf_excel_tasks:
        return 0

    # 대상: label == "AI + Human" 이면서 hybrid_note 비어있는 것
    task_map = {t.id: t for t in _wf_excel_tasks}
    hybrid_targets: list[dict] = []
    for tid, cls in _wf_classification.items():
        if cls.get("label") != "AI + Human":
            continue
        if cls.get("hybrid_note"):
            continue
        if target_task_ids is not None and tid not in target_task_ids:
            continue   # 스코프 밖 태스크는 건너뜀 (Step 2 에서 필요한 것만 우선 처리)
        task = task_map.get(tid)
        if not task:
            continue
        hybrid_targets.append({
            "task_id": tid,
            "task_name": task.name,
            "description": (task.description or "")[:300],
            "reason": (cls.get("reason") or "")[:200],
            "ai_prereq": (cls.get("ai_prerequisites") or "")[:200],
        })

    if not hybrid_targets:
        return 0

    scope_label = f"스코프 내 {len(hybrid_targets)}개" if target_task_ids else f"전체 {len(hybrid_targets)}개"
    print(f"[HYBRID] {scope_label} AI+Human 태스크 AI/Human 파트 병렬 분리 시작", flush=True)

    BATCH_SIZE = 15
    import asyncio as _asyncio
    import time as _time
    _t0 = _time.time()

    # 배치 구성
    batches = [hybrid_targets[i:i + BATCH_SIZE] for i in range(0, len(hybrid_targets), BATCH_SIZE)]

    async def _process_batch(batch: list[dict]) -> int:
        """한 배치 LLM 호출 + 결과 저장. 반환: 성공 분리 수"""
        task_lines = []
        for idx, t in enumerate(batch, 1):
            task_lines.append(
                f"{idx}. [{t['task_id']}] {t['task_name']}\n"
                f"   설명: {t['description']}\n"
                f"   AI+Human 판정 근거: {t['reason']}\n"
                f"   AI 필요여건: {t['ai_prereq']}"
            )
        task_block = "\n\n".join(task_lines)

        system_prompt = f"""당신은 HR 업무를 AI 파트와 Human 파트로 정확히 분리하는 전문가입니다.

아래 Task들은 이미 'AI + Human' 하이브리드로 분류되어 있습니다. 각 Task에 대해 업무를 둘로 쪼개주세요:
- **AI 파트**: AI가 자동화 가능한 부분 (데이터 수집·분석·추론, 초안 작성, 분류, 예측, 요약 등)
- **Human 파트**: 반드시 사람이 수행해야 하는 부분 (최종 승인, 판단, 조율, 인간적 결정)

## 분리 원칙
- AI 파트는 데이터 처리·추론·생성 중심
- Human 파트는 승인·결정·조율·인간 판단 중심
- 두 파트가 명확히 순차적으로 이어지도록 분리 (예: AI가 초안 → Human이 검토 및 확정)
- 각 파트는 1~2문장, 동사 기반으로 구체적으로 기재

## 분석 대상 Task ({len(batch)}개)
{task_block}

## 출력 형식 (JSON만, 마크다운 코드 블록 없음)
{{
  "tasks": [
    {{
      "task_id": "태스크 ID (입력 그대로)",
      "ai_part": "AI가 수행하는 구체적 업무",
      "human_part": "사람이 반드시 수행하는 구체적 업무"
    }}
  ]
}}"""

        try:
            result = await _call_llm_step1(
                system_prompt,
                [{"role": "user", "content": "각 Task를 AI 파트와 Human 파트로 분리해주세요. JSON만 출력하세요."}],
                max_tokens=4096,   # 15개 간단 JSON 이면 4K 면 충분
            )
            if not result:
                return 0
            count = 0
            items = result.get("tasks") or result.get("results") or []
            for item in items:
                tid = str(item.get("task_id") or "").strip()
                ai_part = str(item.get("ai_part") or "").strip()
                human_part = str(item.get("human_part") or "").strip()
                if tid and tid in _wf_classification and ai_part and human_part:
                    _wf_classification[tid]["hybrid_note"] = (
                        f"AI 파트: {ai_part} / Human 파트: {human_part}"
                    )
                    _wf_classification[tid]["ai_part"] = ai_part
                    _wf_classification[tid]["human_part"] = human_part
                    count += 1
            return count
        except Exception as e:
            print(f"[HYBRID] 배치 실패: {e}", flush=True)
            return 0

    # 동시 실행 제한 — API rate limit 고려해 최대 5개 배치 병렬
    semaphore = _asyncio.Semaphore(5)

    async def _bounded_batch(b: list[dict]) -> int:
        async with semaphore:
            return await _process_batch(b)

    results = await _asyncio.gather(*[_bounded_batch(b) for b in batches], return_exceptions=True)
    split_count = sum(r for r in results if isinstance(r, int))

    elapsed = _time.time() - _t0
    print(f"[HYBRID] 병렬 분리 완료 — {split_count}/{len(hybrid_targets)}개 성공, {len(batches)}배치 {elapsed:.1f}s", flush=True)

    # 세션 저장
    if _current_session_id:
        try:
            _save_session_data(_current_session_id)
        except Exception:
            pass

    return split_count

# ── 벤치마킹 source / URL 검증 필터 ─────────────────────────────────────────

# 익명·컨설팅펌 source 금지 패턴
_FORBIDDEN_SOURCE_PATTERNS = [
    "fortune 500", "글로벌 대기업", "한 제조사", "한 기업", "한 은행", "한 금융사",
    "한 회사", "한 유통", "한 제조업체", "한 it기업",
    "익명", "anonymous", "undisclosed", "leading company", "major company",
    "large enterprise", "nextant",
    # 컨설팅·리서치펌
    "mckinsey", "bcg", "bain", "deloitte", "pwc", "ey", "kpmg", "accenture",
    "gartner", "forrester", "idc", "hbr", "mit sloan",
    # 뉴스 기사 제목이 source로 들어오는 경우 (대괄호·따옴표·공백 많은 긴 문자열)
]

# 한국 저품질 뉴스 도메인만 차단 (글로벌 뉴스는 Sonar Pro citation으로 내용 충분)
_NEWS_DOMAINS = [
    "news.naver.com", "n.news.naver.com", "news.daum.net",
    "chosun.com", "joins.com", "joongang.co.kr", "hani.co.kr",
    "mk.co.kr", "hankyung.com", "etnews.com", "zdnet.co.kr",
    "itworld.co.kr", "dt.co.kr", "bloter.net", "ddaily.co.kr",
    "aitimes.com", "aitimes.kr", "boannews.com", "itbiznews.com",
    "newsis.com", "yonhapnews.co.kr", "yna.co.kr",
]

# URL 경로에서 뉴스 패턴 (공식 기업 블로그/케이스스터디 URL은 허용)
_NEWS_URL_PATTERNS = ["/newsroom/press-release", "/press-release/"]


import re as _re_filter

# 국내 중소·중견 IT기업 패턴 (글로벌 인지도 없는 한국 기업)
_KR_SMALL_COMPANY_PATTERN = _re_filter.compile(
    r"(코퍼레이션|솔루션|시스템즈?|소프트|테크놀로지|테크|인포|데이터|클라우드|"
    r"이노베이션|디지털|네트웍스?|서비스|컨설팅|파트너스?)"
    r"(?!\s*(삼성|현대|SK|LG|포스코|롯데|두산|한화))",  # 대기업 계열사 제외
)

# 글로벌 인지도 있는 한국 대기업 허용 목록
_KR_ALLOWED = {
    "삼성전자", "삼성sds", "현대자동차", "현대차", "sk하이닉스", "sk텔레콤",
    "lg전자", "lg cns", "포스코", "롯데", "두산", "한화", "기아", "kt",
}


# 쿼리 ID 형태의 쓰레기 source 패턴 — [R1-3], r1-1, R2-5 등 검색 round 태그가 source로 들어온 경우
_QUERY_ID_PATTERN = _re_filter.compile(r"^\s*\[?R\s*\d+\s*[-.]\s*\d+\s*\]?\s*$", _re_filter.IGNORECASE)
# 순수 문자 없는 source (숫자·기호만): 회사명 될 수 없음
_NO_ALPHA_PATTERN = _re_filter.compile(r"^[^A-Za-z가-힣]+$")


def _is_valid_benchmark_source(source: str) -> bool:
    """source 필드가 실명 글로벌 대기업인지 검사합니다."""
    src = source.strip()
    s = src.lower()
    if not s or len(s) < 2:
        return False
    # 너무 긴 source는 기사 제목이 들어온 것
    if len(src) > 40:
        return False
    # 금지 패턴
    if any(pattern in s for pattern in _FORBIDDEN_SOURCE_PATTERNS):
        return False
    # 쿼리 ID 형태 차단 (R1-1, [R2-3] 등)
    if _QUERY_ID_PATTERN.match(src):
        return False
    # 알파벳·한글이 하나도 없으면 차단 (숫자·기호만은 회사명 아님)
    if _NO_ALPHA_PATTERN.match(src):
        return False
    # 'R' + 숫자로 시작하는 짧은 문자열 차단 (R1, R2-5 등)
    if _re_filter.match(r"^R\d", src, _re_filter.IGNORECASE) and len(src) <= 8:
        return False
    # 한국어 포함 여부 확인
    has_korean = any("\uAC00" <= c <= "\uD7A3" for c in src)
    if has_korean:
        # 허용된 한국 대기업이면 통과
        if any(allowed in s for allowed in _KR_ALLOWED):
            return True
        # 국내 중소 IT기업 패턴이면 차단
        if _KR_SMALL_COMPANY_PATTERN.search(src):
            return False
    return True


def _is_news_url(url: str) -> bool:
    """URL이 뉴스 기사 URL인지 확인합니다."""
    if not url:
        return False
    url_lower = url.lower()
    if any(domain in url_lower for domain in _NEWS_DOMAINS):
        return True
    if any(pat in url_lower for pat in _NEWS_URL_PATTERNS):
        return True
    return False


def _build_excel_index() -> tuple[dict, dict, dict, dict]:
    """엑셀 Task를 id(L5) / l4_id / l3_id / l2_id 기준으로 인덱싱.
    반환: by_id, by_l4, by_l3, by_l2
    - by_id : 엑셀 task id(L5 ID) → Task  (L5↔L5 직접 매핑)
    - by_l4 : l4_id → [Task, ...]         (L4 그룹핑용)
    """
    by_id: dict[str, object] = {}
    by_l4: dict[str, list] = {}
    by_l3: dict[str, list] = {}
    by_l2: dict[str, list] = {}
    for t in _wf_excel_tasks:
        if t.id:
            by_id[t.id] = t
        if t.l4_id:
            by_l4.setdefault(t.l4_id, []).append(t)
        if t.l3_id:
            by_l3.setdefault(t.l3_id, []).append(t)
        if t.l2_id:
            by_l2.setdefault(t.l2_id, []).append(t)
    return by_id, by_l4, by_l3, by_l2


def _format_task_line(t, show_pain: bool = True) -> str:
    """단일 Task를 한 줄 요약."""
    cls = _wf_classification.get(t.id, {})
    label = cls.get("label", "미분류")
    pain_parts = []
    if t.pain_time: pain_parts.append("시간/속도")
    if t.pain_accuracy: pain_parts.append("정확성")
    if t.pain_repetition: pain_parts.append("반복/수작업")
    if t.pain_data: pain_parts.append("정보/데이터")
    if t.pain_system: pain_parts.append("시스템/도구")
    if t.pain_communication: pain_parts.append("의사소통")
    pain_str = f" | Pain: {', '.join(pain_parts)}" if pain_parts and show_pain else ""
    desc_str = f" | 설명: {t.description[:60]}" if t.description else ""
    return f"      [{t.id}] {t.name} (분류: {label}){pain_str}{desc_str}"


def _build_mapped_asis_context(sheet_id: str = "") -> str:
    """As-Is 워크플로우 노드를 엑셀 Task와 task_id 기준으로 매핑하여 풍부한 컨텍스트 문자열 생성.
    벤치마킹·Step1·Step2에서 공통 사용."""
    if "parsed" not in _workflow_cache:
        # As-Is 없으면 엑셀만으로 L3/L4 기반 컨텍스트 구성
        if not _wf_excel_tasks:
            return ""
        _, by_l4, by_l3, _ = _build_excel_index()
        lines = ["## As-Is 프로세스 (엑셀 기반)\n"]
        cur_l3 = None
        for t in _wf_excel_tasks:
            if t.l3 != cur_l3:
                cur_l3 = t.l3
                lines.append(f"\n### L3: {t.l3} [{t.l3_id}]")
            if t.l4:
                tasks_in_l4 = by_l4.get(t.l4_id, [])
                if tasks_in_l4 and tasks_in_l4[0].id == t.id:  # L4 첫 등장 시 헤더
                    lines.append(f"  - L4 [{t.l4_id}] {t.l4}")
            lines.append(_format_task_line(t))
        return "\n".join(lines)

    parsed = _workflow_cache["parsed"]
    if sheet_id:
        matched_sheets = [s for s in parsed.sheets if s.sheet_id == sheet_id]
        if not matched_sheets:
            print(f"[SCOPE⚠] _build_mapped_asis_context sheet_id='{sheet_id}' 매칭 없음! 전체 시트: {[s.sheet_id for s in parsed.sheets]}", flush=True)
        target_sheets = matched_sheets if matched_sheets else parsed.sheets
    else:
        target_sheets = parsed.sheets
    print(f"[SCOPE] _build_mapped_asis_context sheet_id='{sheet_id}' → target_sheets={[s.sheet_id+'/'+s.name for s in target_sheets]}", flush=True)

    by_id, by_l4, by_l3, by_l2 = _build_excel_index()
    lines = ["## As-Is 워크플로우 + 엑셀 매핑\n"]

    for s in target_sheets:
        lines.append(f"### 시트: {s.name}")
        # L2/L3 노드를 상위 구조로, L4/L5를 세부 내용으로 표시
        l2_nodes = [n for n in s.nodes.values() if n.level == "L2"]
        l3_nodes = [n for n in s.nodes.values() if n.level == "L3"]
        l5_nodes = [n for n in s.nodes.values() if n.level == "L5"]

        # L4 노드 필터링: 이 시트에 L5 자식이 있는 L4만 포함
        # (L5 자식 없는 L4는 다른 L4 연계를 표시하는 connector 노드 → 제외)
        l5_task_id_set = {n.task_id for n in l5_nodes if n.task_id}
        if l5_task_id_set:
            l4_nodes = [
                n for n in s.l4_nodes
                if any(tid.startswith(n.task_id + ".") for tid in l5_task_id_set)
            ]
        else:
            l4_nodes = s.l4_nodes  # L5가 없는 시트는 전체 L4 사용

        # L4를 부모 L3 기준으로 그룹핑 (task_id prefix로 추정)
        l4_by_l3_tid: dict[str, list] = {}
        for n in l4_nodes:
            # task_id "1.2.3" → L3 tid "1.2"
            parts = n.task_id.rsplit(".", 1)
            parent_tid = parts[0] if len(parts) > 1 else n.task_id
            l4_by_l3_tid.setdefault(parent_tid, []).append(n)

        # 엣지 맵 (노드 ID → 나가는 엣지 목록)
        outgoing_map: dict[str, list] = {}
        for e in s.edges:
            outgoing_map.setdefault(e.source, []).append(e)

        # 양방향 엣지 맵 (노드 ID → 협의/상호작용 관계 상대 노드 ID 목록)
        bidir_map: dict[str, list[str]] = {}
        for e in s.edges:
            if e.bidirectional:
                bidir_map.setdefault(e.source, []).append(e.target)
                bidir_map.setdefault(e.target, []).append(e.source)

        decision_nodes_map = {n.id: n for n in s.nodes.values() if n.level == "DECISION"}

        def _render_branches(node_id: str, indent: str = "    ") -> list[str]:
            """노드 → Decision/조건 엣지 분기 구조를 텍스트로 렌더링."""
            result = []
            for e in outgoing_map.get(node_id, []):
                tgt = s.nodes.get(e.target)
                if not tgt:
                    continue
                if tgt.level == "DECISION":
                    dec_label = tgt.label or "분기 조건"
                    result.append(f"{indent}→ [Decision] {dec_label}")
                    for de in outgoing_map.get(tgt.id, []):
                        dtgt = s.nodes.get(de.target)
                        cond = de.label or "(조건 없음)"
                        next_label = dtgt.label if dtgt else de.target
                        next_level = dtgt.level if dtgt else ""
                        result.append(f"{indent}   ├─ ({cond}) → [{next_level}] {next_label}")
                elif e.label:
                    result.append(f"{indent}→ 조건: \"{e.label}\" → [{tgt.level}] {tgt.label}")
            return result

        # 외부 업체/시스템 키워드 (재설계 불가) vs 두산 내부 조직 (재설계 가능)
        _EXTERNAL_KEYWORDS = {"큐벡스", "업체", "외부", "벤더", "vendor"}
        _INTERNAL_ORG_KEYWORDS = {"지주", "자회사", "bg", "계열사", "그룹사"}

        def _classify_role(actor_label: str) -> str:
            """'그 외' swim lane의 수행주체를 외부/내부로 분류한다."""
            lower = actor_label.lower()
            # "그 외:" 접두사 포함 여부 확인
            is_extra = "그 외" in actor_label
            if not is_extra:
                return ""  # 일반 HR/임원 swim lane — 분류 불필요
            # 콜론 뒤 실제 이름 추출 (예: "그 외:큐벡스" → "큐벡스")
            name_part = actor_label.split(":")[-1].strip().lower() if ":" in actor_label else lower
            if any(k in name_part for k in _EXTERNAL_KEYWORDS):
                return "외부 업체/시스템 — 재설계 제외 (HR 업무 범위 밖)"
            if any(k in name_part for k in _INTERNAL_ORG_KEYWORDS):
                return "두산 내부 조직 — 재설계 가능 (HR 협의·협업 범위)"
            # "그 외"이지만 키워드 불명확 → 외부로 보수적 처리
            return "외부/비HR 주체 — 재설계 여부 검토 필요"

        def _node_meta_lines(node: "WorkflowNode", indent: str = "      ") -> list[str]:
            """노드 메타데이터(수행주체·시스템·협의관계)를 텍스트 라인으로 반환."""
            meta_lines = []
            # 수행주체 (role > actors 우선)
            role_val = node.metadata.get("role", "")
            actors_val = node.metadata.get("actors", "")
            actor_label = role_val or (
                ", ".join(actors_val) if isinstance(actors_val, list) else str(actors_val)
            )
            if actor_label:
                # "/ " 로 구분된 복수 수행주체 → 협의 관계 명시
                actor_parts = [a.strip() for a in actor_label.replace("，", ",").split("/") if a.strip()]
                # 각 파트별 외부/내부 분류
                classified_parts = []
                redesign_notes = []
                for part in actor_parts:
                    cls = _classify_role(part)
                    if cls:
                        classified_parts.append(f"{part} [{cls}]")
                        redesign_notes.append(cls)
                    else:
                        classified_parts.append(part)
                if len(classified_parts) > 1:
                    meta_lines.append(f"{indent}수행주체: {' / '.join(classified_parts)} (복수 주체 — 협의 관계)")
                else:
                    meta_lines.append(f"{indent}수행주체: {' / '.join(classified_parts)}")
            # 사용 시스템 / 고유명사 (큐벡스 등)
            sys_val = node.metadata.get("system", "") or node.metadata.get("systems", "")
            sys_label = (
                ", ".join(sys_val) if isinstance(sys_val, list) else str(sys_val)
            ) if sys_val else ""
            if sys_label:
                meta_lines.append(f"{indent}사용시스템: {sys_label}")
            # 양방향 협의/상호작용 관계
            bidir_partners = bidir_map.get(node.id, [])
            if bidir_partners:
                partner_labels = [s.nodes[pid].label for pid in bidir_partners if pid in s.nodes]
                if partner_labels:
                    meta_lines.append(f"{indent}↔ 협의/상호작용: {', '.join(partner_labels)}")
            return meta_lines

        # L3 노드 순회
        if l3_nodes:
            for l3n in sorted(l3_nodes, key=lambda x: x.task_id):
                l3_excel = by_l3.get(l3n.task_id, [])
                l3_pain = set()
                for t in l3_excel:
                    if t.pain_time: l3_pain.add("시간/속도")
                    if t.pain_repetition: l3_pain.add("반복/수작업")
                    if t.pain_accuracy: l3_pain.add("정확성")
                    if t.pain_data: l3_pain.add("정보/데이터")
                pain_tag = f" | Pain: {', '.join(sorted(l3_pain))}" if l3_pain else ""
                lines.append(f"\n  [L3 {l3n.task_id}] {l3n.label}{pain_tag}")

                # 이 L3 아래 L4 노드들
                children = l4_by_l3_tid.get(l3n.task_id, [])
                for l4n in sorted(children, key=lambda x: x.task_id):
                    excel_tasks = by_l4.get(l4n.task_id, [])
                    lines.append(f"    - L4 [{l4n.task_id}] {l4n.label}")
                    lines.extend(_node_meta_lines(l4n, indent="      "))
                    for t in excel_tasks[:5]:
                        lines.append(_format_task_line(t))
                    # L5 자식
                    l5_children = [n for n in l5_nodes if n.task_id.startswith(l4n.task_id + ".")]
                    for l5n in sorted(l5_children, key=lambda x: x.task_id)[:5]:
                        lines.append(f"      └ L5 [{l5n.task_id}] {l5n.label}")
                        lines.extend(_node_meta_lines(l5n, indent="          "))
                    # Decision 분기
                    branch_lines = _render_branches(l4n.id)
                    lines.extend(branch_lines)
        else:
            # L3 노드가 없으면 L4 직접 표시
            for l4n in sorted(l4_nodes, key=lambda x: x.task_id):
                excel_tasks = by_l4.get(l4n.task_id, [])
                lines.append(f"\n  - L4 [{l4n.task_id}] {l4n.label}")
                lines.extend(_node_meta_lines(l4n, indent="      "))
                for t in excel_tasks[:5]:
                    lines.append(_format_task_line(t))
                # Decision 분기
                branch_lines = _render_branches(l4n.id)
                lines.extend(branch_lines)

    return "\n".join(lines)


def _build_task_and_pain_summary(sheet_id: str = "") -> tuple[str, str, str]:
    """엑셀 Task 데이터로 task_summary, pain_summary, process_name을 만든다.
    구조: JSON/PPT 파일 1개 = L3, 파일 안의 시트 1개 = L4
    계층: L2=x, L3=x.y, L4=x.y.z, L5=x.y.z.w (L1 없음)
    - L3 이름: 시트 노드의 task_id 앞 두 자리(x.y) → 엑셀 l3_id 매핑으로 추적
    - sheet_id 지정 시: 해당 L4 시트의 task_id 범위로 필터링
    - sheet_id="" : JSON의 모든 시트(= L3 전체, 모든 L4) 합산
    - JSON 없음: 엑셀 전체 반환"""

    relevant_tasks = _wf_excel_tasks
    json_l3_ids: set[str] = set()  # 시트 내 노드 task_id의 L3 접두사 (x.y)

    if "parsed" in _workflow_cache:
        parsed = _workflow_cache["parsed"]
        all_sheet_ids = [s.sheet_id for s in parsed.sheets]
        # sheet_id 지정 시 해당 시트(=L4)만, 아니면 전체 시트(=L3 전체)
        if sheet_id:
            matched = [s for s in parsed.sheets if s.sheet_id == sheet_id]
            if not matched:
                print(f"[SCOPE⚠] sheet_id='{sheet_id}' 매칭 시트 없음! 전체 시트 목록: {all_sheet_ids}", flush=True)
            target_sheets = matched if matched else parsed.sheets  # fallback: 전체(L3)
        else:
            target_sheets = parsed.sheets  # L3 전체: 모든 L4 시트 합산

        print(f"[SCOPE] _build_task_and_pain_summary sheet_id='{sheet_id}' → target_sheets={[s.sheet_id+'/'+s.name for s in target_sheets]}", flush=True)

        # JSON 노드의 task_id에서 L5→L4→L3 접두사 수집
        # L4 노드는 흐름 연결용 connector이므로 항상 L5 노드 task_id만 사용
        # (L4 노드 task_id를 쓰면 다른 시트 태스크가 오염됨)
        json_tids: set[str] = set()
        json_l4_ids: set[str] = set()
        for s in target_sheets:
            l5_only = [n for n in s.nodes.values() if n.level == "L5"]
            source_nodes = l5_only if l5_only else s.nodes.values()  # L5 없으면 전체 fallback
            for n in source_nodes:
                if n.task_id:
                    json_tids.add(n.task_id)
                    parts = n.task_id.split(".")
                    if len(parts) >= 3:
                        json_l4_ids.add(".".join(parts[:3]))  # x.y.z = L4
                    if len(parts) >= 2:
                        json_l3_ids.add(".".join(parts[:2]))  # x.y = L3

        print(f"[SCOPE] json_tids={sorted(json_tids)[:10]} json_l4_ids={sorted(json_l4_ids)}", flush=True)

        # JSON 파일 전체 = 하나의 L3 → 모든 시트에서 L3 이름·ID 수집 (L3 범위 확정)
        # L4 이름이 여러 L3에 걸쳐 중복 사용될 수 있으므로 L3 범위로 반드시 좁혀야 함
        all_l3_names_in_json: set[str] = set()
        all_l3_ids_in_json: set[str] = set()
        for s_all in parsed.sheets:
            for n in s_all.nodes.values():
                if n.level == "L3" and n.label:
                    all_l3_names_in_json.add(n.label.strip())
                if n.level == "L5" and n.task_id:
                    parts = n.task_id.split(".")
                    if len(parts) >= 2:
                        all_l3_ids_in_json.add(".".join(parts[:2]))

        sheet_names = [s.name.strip() for s in target_sheets if s.name]
        print(f"[SCOPE] L3 범위: names={all_l3_names_in_json} ids={all_l3_ids_in_json}", flush=True)

        # 0순위: L4 이름 + L3 이름 조합 (가장 안정적 — 번호 변경 무관)
        if all_l3_names_in_json:
            filtered = [t for t in _wf_excel_tasks
                        if t.l4 and t.l4.strip() in sheet_names
                        and t.l3 and t.l3.strip() in all_l3_names_in_json]
            print(f"[SCOPE] 0순위(l4명+l3명) matched={len(filtered)}건: sheet_names={sheet_names}", flush=True)
        elif all_l3_ids_in_json:
            # L3 노드 없으면 task_id prefix로 L3 ID 추정
            filtered = [t for t in _wf_excel_tasks
                        if t.l4 and t.l4.strip() in sheet_names
                        and t.l3_id in all_l3_ids_in_json]
            print(f"[SCOPE] 0순위(l4명+l3id) matched={len(filtered)}건", flush=True)
        else:
            # L3 정보 없으면 L4 이름만 (중복 가능성 있음, 경고)
            filtered = [t for t in _wf_excel_tasks if t.l4 and t.l4.strip() in sheet_names]
            print(f"[SCOPE] 0순위(l4명만) matched={len(filtered)}건 — L3 컨텍스트 없음!", flush=True)

        if not filtered:
            # 1순위: task_id 직접 매칭 (이름 매칭 실패 시)
            filtered = [t for t in _wf_excel_tasks if t.id in json_tids]
            print(f"[SCOPE] 1순위(task_id) matched={len(filtered)}건: {[t.id for t in filtered[:5]]}", flush=True)
        if not filtered:
            filtered = [t for t in _wf_excel_tasks if t.l4_id in json_l4_ids]
            print(f"[SCOPE] 2순위(l4_id) matched={len(filtered)}건: {[t.name for t in filtered[:5]]}", flush=True)
        if not filtered:
            filtered = [t for t in _wf_excel_tasks if t.l3_id in json_l3_ids]
            print(f"[SCOPE] 3순위(l3_id) matched={len(filtered)}건 — L3 전체 fallback!", flush=True)
        if filtered:
            relevant_tasks = filtered

    # process_name 결정
    # ── L4 scope(sheet_id 지정): 시트 이름(L4명) 사용
    # ── L3 scope: L3 이름 사용
    if sheet_id and "parsed" in _workflow_cache:
        # 시트 이름 = L4 Activity 명
        matched_sheet = next(
            (s for s in _workflow_cache["parsed"].sheets if s.sheet_id == sheet_id), None
        )
        sheet_l4_name = (matched_sheet.name if matched_sheet else "").strip()
        process_name = sheet_l4_name or "HR 프로세스"
    elif json_l3_ids:
        seen_l3n: set[str] = set()
        l3_names: list[str] = []
        for t in _wf_excel_tasks:
            if t.l3_id in json_l3_ids and t.l3 and t.l3 not in seen_l3n:
                seen_l3n.add(t.l3)
                l3_names.append(t.l3)
        process_name = l3_names[0] if l3_names else "HR 프로세스"
    else:
        l3_names = list(dict.fromkeys(t.l3 for t in relevant_tasks if t.l3))
        process_name = l3_names[0] if l3_names else "HR 프로세스"

    task_lines = []
    for t in relevant_tasks:
        cls = _wf_classification.get(t.id, {})
        label = cls.get("label", "미분류")
        line = f"- [{t.id}] {t.name} (L3: {t.l3}, L4: {t.l4}) — 분류: {label}"
        # AI+Human이면 AI 파트 / Human 파트 분리 정보 주입 (LLM이 Step 2에서 참고)
        if label == "AI + Human":
            hn = cls.get("hybrid_note", "")
            if hn:
                line += f"\n    {hn}"
            reason = cls.get("reason", "")
            if reason:
                line += f"\n    판정근거: {reason[:150]}"
        elif label == "Human":
            reason = cls.get("reason", "")
            if reason:
                line += f" — {reason[:100]}"
        task_lines.append(line)

    pain_lines = []
    for t in relevant_tasks:
        pains = []
        if t.pain_time: pains.append("시간/속도")
        if t.pain_accuracy: pains.append("정확성")
        if t.pain_repetition: pains.append("반복/수작업")
        if t.pain_data: pains.append("정보/데이터")
        if t.pain_system: pains.append("시스템/도구")
        if t.pain_communication: pains.append("의사소통")
        if pains:
            pain_lines.append(f"- [{t.id}] {t.name}: {', '.join(pains)}")

    return (
        process_name,
        "\n".join(task_lines),
        "\n".join(pain_lines) if pain_lines else "Pain Point 정보 없음",
    )


def _step1_system_prompt(process_name: str, task_summary: str, pain_summary: str,
                         benchmark_text: str = "",
                         scope: str = "l3") -> str:
    """scope='l4': 특정 L4 Activity 내 L5 Task AI 적용 설계
       scope='l3': 전체 L3 프로세스 재구조화 + L4/L5 AI 적용 설계"""

    redesign_scope_note = (
        "- **재설계 가능**: HR 담당자, HR 임원, 지주, 자회사, BG, 계열사 등 두산 내부 조직이 수행하는 Task\n"
        "- **재설계 불가 (현행 유지)**: 큐벡스, 업체 등 외부 업체/시스템이 수행하는 Task — "
        "`change_type: \"유지\"`, `ai_application: \"해당 없음\"`으로 처리\n"
        "- As-Is 컨텍스트에서 수행주체 라인에 `[외부 업체/시스템 — 재설계 제외]` 표시된 Task는 반드시 현행 유지"
    )

    if scope == "l4":
        # ── L4 단위: 이 L4의 L5 Task AI 적용 설계만 ──────────────────
        design_direction = (
            "- **L5 Task AI 적용 설계**: 이 L4 Activity의 각 L5 Task별로 AI 적용 여부와 방식 기술\n"
            "  - change_type: \"유지\" / \"통합\" / \"세분화\" / \"추가\" / \"삭제\" 중 하나\n"
            "  - AI 미적용 Task는 ai_application을 \"해당 없음\"으로\n"
            "  - automation_level: \"Full-Auto\" / \"Human-in-Loop\" / \"Human-on-the-Loop\" / \"Human\" 중 하나\n"
            "- **L3 재구조화 불필요**: L4 단위 설계이므로 상위 L3 변경은 제안하지 마세요."
        )
        output_format = f"""{{
  "blueprint_summary": "이 L4 Activity 기본 설계 요약 (3~5문장)",
  "process_name": "{process_name}",
  "benchmark_insights": [
    {{"source": "기업명", "insight": "구체적 사례", "application": "두산 적용 방안"}}
  ],
  "redesigned_process": [
    {{
      "l3_id": "",
      "l3_name": "",
      "change_type": "유지",
      "change_reason": "L4 단위 설계 — L3 변경 없음",
      "l4_list": [
        {{
          "l4_id": "기존 L4 ID (예: 3.1.1)",
          "l4_name": "{process_name}",
          "change_type": "유지|통합|세분화|추가|삭제",
          "change_reason": "변경 이유 (1문장)",
          "l5_list": [
            {{
              "task_id": "기존 task_id 또는 NEW_xxx",
              "task_name": "Task명",
              "change_type": "유지|통합|세분화|추가|삭제",
              "ai_application": "AI 적용 내용 또는 해당 없음",
              "automation_level": "Full-Auto|Human-in-Loop|Human-on-the-Loop|Human",
              "ai_technique": "사용 AI 기법 (예: LLM 요약, RAG, 분류모델, 해당 없음)"
            }}
          ]
        }}
      ]
    }}
  ]
}}"""
    else:
        # ── L3 단위: 전체 프로세스 재구조화 + L4/L5 AI 적용 ─────────
        design_direction = (
            "- **L2~L3 재구조화**: 선도사례를 참고하여 L3 프로세스 그룹의 통합·세분화·추가·삭제 방향 제시\n"
            "- **L4~L5 AI 적용 기본 설계**: 각 L4 Activity / L5 Task별로 AI 적용 여부와 방식 기술\n"
            "  - change_type: \"유지\" / \"통합\" / \"세분화\" / \"추가\" / \"삭제\" 중 하나\n"
            "  - AI 미적용 Task는 ai_application을 \"해당 없음\"으로\n"
            "  - automation_level: \"Full-Auto\" / \"Human-in-Loop\" / \"Human-on-the-Loop\" / \"Human\" 중 하나"
        )
        output_format = f"""{{
  "blueprint_summary": "전체 기본 설계 요약 (3~5문장)",
  "process_name": "{process_name}",
  "benchmark_insights": [
    {{"source": "기업명", "insight": "구체적 사례", "application": "두산 적용 방안"}}
  ],
  "l2_restructure": "L2~L3 프로세스 재구조화 방향 설명 (2~3문장)",
  "redesigned_process": [
    {{
      "l3_id": "기존 L3 ID (예: 3.1)",
      "l3_name": "L3 프로세스명",
      "change_type": "유지|통합|세분화|추가|삭제",
      "change_reason": "변경 이유 (1문장)",
      "l4_list": [
        {{
          "l4_id": "기존 L4 ID (예: 3.1.1)",
          "l4_name": "L4 Activity명",
          "change_type": "유지|통합|세분화|추가|삭제",
          "change_reason": "변경 이유 (1문장)",
          "l5_list": [
            {{
              "task_id": "기존 task_id 또는 NEW_xxx",
              "task_name": "Task명",
              "change_type": "유지|통합|세분화|추가|삭제",
              "ai_application": "AI 적용 내용 또는 해당 없음",
              "automation_level": "Full-Auto|Human-in-Loop|Human-on-the-Loop|Human",
              "ai_technique": "사용 AI 기법 (예: LLM 요약, RAG, 분류모델, 해당 없음)"
            }}
          ]
        }}
      ]
    }}
  ]
}}"""

    return f"""당신은 AI 기반 업무 혁신 설계 전문가입니다.
벤치마킹 선도사례와 Gap 분석 결과를 동시에 활용하여, 기존 As-Is 프로세스를 재구조화하고 각 Task에 AI 적용 기본 설계를 수행합니다.

## ⚠️ 핵심 원칙
- 아래 L5 Task 목록({process_name})이 설계의 기준입니다. 기존 Task는 최대한 유지하되, 필요 시 통합·세분화·추가·삭제를 제안할 수 있습니다.
- `task_id`는 반드시 아래 Task 목록의 실제 ID를 사용하세요. 신규 추가 Task는 "NEW_xxx" 형식으로 표기하세요.
- **Gap 분석 결과를 최우선으로 반영**: A.신규(도입), B.전환(AI 전환), C.폐기/통합(삭제/흡수) 방향을 설계에 직접 반영하세요.

## ⚠️ 재설계 범위 (swim lane 기준)
{redesign_scope_note}

## 프로세스: {process_name}

## As-Is L5 Task 목록 (설계의 기준)
{task_summary}

## Pain Point 현황
{pain_summary}

{benchmark_text}

## 설계 방향
{design_direction}

## 출력 형식 (JSON만 출력, 마크다운 코드 블록 없음)
{output_format}
"""


async def _call_llm_step1(system: str, messages: list, max_tokens: int = 16384) -> dict | None:
    """Step 1 LLM 호출 공통 로직.

    max_tokens: Anthropic 응답 한도. 벤치마킹처럼 긴 JSON 이 나올 땐 상향 전달.
    """
    settings = load_settings()
    anthropic_key = os.getenv("ANTHROPIC_API_KEY", "") or settings.anthropic_api_key
    result_data = None

    from usage_store import add_usage as _add_usage

    if anthropic_key:
        try:
            from anthropic import AsyncAnthropic
            client = AsyncAnthropic(api_key=anthropic_key)
            response = await client.messages.create(
                model=settings.anthropic_model or "claude-sonnet-4-6",
                max_tokens=max_tokens,
                system=system,
                messages=messages,
            )
            if response.usage:
                _add_usage("anthropic",
                           input_tokens=response.usage.input_tokens,
                           output_tokens=response.usage.output_tokens)
            raw = response.content[0].text
            # stop_reason 체크 — max_tokens 에서 잘린 경우 경고 로그
            stop_reason = getattr(response, "stop_reason", None)
            if stop_reason and stop_reason != "end_turn":
                print(f"[workflow-step1] ⚠️ Anthropic stop_reason={stop_reason} "
                      f"(max_tokens={max_tokens}, output_len={len(raw)}). JSON 잘림 가능.",
                      flush=True)
            from new_workflow_generator import _extract_json
            result_data = _extract_json(raw)
            if result_data is None:
                print(f"[workflow-step1] ⚠️ JSON 파싱 실패 — raw 앞 500자: {raw[:500]}", flush=True)
            return result_data
        except Exception as e:
            print(f"[workflow-step1] Anthropic 실패: {e}")

    openai_key = os.getenv("OPENAI_API_KEY", "") or settings.api_key
    if openai_key:
        try:
            from openai import AsyncOpenAI
            client = AsyncOpenAI(api_key=openai_key)
            response = await client.chat.completions.create(
                model=settings.model or "gpt-5.4",
                temperature=0.0,
                messages=[{"role": "system", "content": system}, *messages],
                response_format={"type": "json_object"},
            )
            if response.usage:
                _add_usage("openai",
                           input_tokens=response.usage.prompt_tokens,
                           output_tokens=response.usage.completion_tokens)
            raw = response.choices[0].message.content or "{}"
            return json.loads(raw)
        except Exception as e:
            print(f"[workflow-step1] OpenAI 실패: {e}")

    return None


def _save_step1_result(result_data: dict) -> dict:
    """Step 1 LLM 결과를 캐시에 저장하고 반환."""
    redesigned = result_data.get("redesigned_process", [])

    # 자동화 수준별 L5 Task 집계
    full_auto, hil, hs = 0, 0, 0
    for l3 in redesigned:
        for l4 in l3.get("l4_list", []):
            for l5 in l4.get("l5_list", []):
                lvl = l5.get("automation_level", "")
                if lvl == "Full-Auto": full_auto += 1
                elif lvl == "Human-in-Loop": hil += 1
                elif lvl == "Human-on-the-Loop": hs += 1

    result_dict = {
        "ok": True,
        "blueprint_summary": result_data.get("blueprint_summary", ""),
        "process_name": result_data.get("process_name", ""),
        "l2_restructure": result_data.get("l2_restructure", ""),
        "benchmark_insights": result_data.get("benchmark_insights", []),
        "redesigned_process": redesigned,
        "full_auto_count": full_auto,
        "human_in_loop_count": hil,
        "human_supervised_count": hs,
        "total_tasks": full_auto + hil + hs,
        # Step2 호환용 빈 필드
        "agents": [],
        "execution_flow": [],
        "benchmark_table": {k: v for k, v in _wf_benchmark_table.items()},
    }

    global _wf_step1_cache
    _wf_step1_cache = result_dict
    if _current_session_id:
        _save_session_data(_current_session_id)
    return result_dict


# ── 벤치마킹 전용 ──────────────────────────────────────────

@app.delete("/api/workflow/benchmark-table/row", tags=["Workflow"])
async def delete_benchmark_row(request: Request):
    """벤치마킹 결과 테이블에서 특정 행을 삭제합니다.
    body: {source: "기업명", sheet_id: "시트ID"}
    sheet_id가 없으면 모든 시트에서 source가 일치하는 첫 번째 행을 삭제합니다.
    """
    body = await request.json()
    source = (body.get("source") or "").strip()
    sheet_id = (body.get("sheet_id") or "").strip()

    if not source:
        raise HTTPException(400, "source(기업명)가 필요합니다.")

    deleted = False
    if sheet_id and sheet_id in _wf_benchmark_table:
        before = len(_wf_benchmark_table[sheet_id])
        _wf_benchmark_table[sheet_id] = [
            r for r in _wf_benchmark_table[sheet_id]
            if r.get("source", "").strip() != source
        ]
        deleted = len(_wf_benchmark_table[sheet_id]) < before
    else:
        # 모든 시트에서 source 일치 행 제거
        for sid in list(_wf_benchmark_table.keys()):
            before = len(_wf_benchmark_table[sid])
            _wf_benchmark_table[sid] = [
                r for r in _wf_benchmark_table[sid]
                if r.get("source", "").strip() != source
            ]
            if len(_wf_benchmark_table[sid]) < before:
                deleted = True

    if not deleted:
        raise HTTPException(404, f"'{source}' 항목을 찾지 못했습니다.")

    # 세션 상태 persist
    if _current_session_id:
        session_dir = _get_session_dir(_current_session_id)
        (session_dir / "benchmark_table.json").write_text(
            json.dumps(_wf_benchmark_table, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    all_rows = [r for rows in _wf_benchmark_table.values() for r in rows]
    return {
        "ok": True,
        "deleted_source": source,
        "remaining": len(all_rows),
        "benchmark_table": {k: v for k, v in _wf_benchmark_table.items()},
    }


@app.post("/api/workflow/benchmark-table/cleanup", tags=["Workflow"])
async def cleanup_benchmark_table():
    """벤치마킹 테이블에서 유효하지 않은 source(쿼리 ID, 익명, 컨설팅펌 등) 행을 일괄 제거.
    기존 데이터에 R1-1 같은 쓰레기 row 가 남아있을 때 수동으로 정리용.
    """
    removed: list[dict] = []
    for sid in list(_wf_benchmark_table.keys()):
        kept = []
        for r in _wf_benchmark_table[sid]:
            src = (r.get("source") or "").strip()
            url = (r.get("url") or "").strip()
            if not src or not _is_valid_benchmark_source(src) or not url or _is_news_url(url):
                removed.append({"source": src, "url": url, "sheet": sid})
            else:
                kept.append(r)
        _wf_benchmark_table[sid] = kept

    if _current_session_id:
        session_dir = _get_session_dir(_current_session_id)
        (session_dir / "benchmark_table.json").write_text(
            json.dumps(_wf_benchmark_table, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    all_rows = [r for rows in _wf_benchmark_table.values() for r in rows]
    return {
        "ok": True,
        "removed_count": len(removed),
        "removed_samples": removed[:10],
        "remaining": len(all_rows),
        "benchmark_table": {k: v for k, v in _wf_benchmark_table.items()},
    }



@app.get("/api/workflow/benchmark-table/export", tags=["Workflow"])
async def export_benchmark_table_xlsx():
    """벤치마킹 결과 테이블을 xlsx 파일로 내보냅니다."""
    from fastapi.responses import Response

    all_export_rows = [r for rows in _wf_benchmark_table.values() for r in rows]
    if not all_export_rows:
        raise HTTPException(400, "벤치마킹 결과가 없습니다. 먼저 벤치마킹을 실행하세요.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "벤치마킹 결과"

    headers = [
        "기업", "유형", "산업", "적용 L4", "도입 목표",
        "AI 기술", "핵심 데이터", "도입 방식", "적용 사례",
        "성과", "인프라", "두산 시사점", "출처 URL",
    ]
    field_keys = [
        "source", "company_type", "industry", "target_l4", "goal",
        "ai_technology", "key_data", "adoption_method", "use_case",
        "outcome", "infrastructure", "implication", "url",
    ]

    # 헤더 스타일
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # 데이터 행
    row_fill_even = PatternFill("solid", fgColor="EEF2FF")
    data_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, entry in enumerate(all_export_rows, 2):
        fill = row_fill_even if row_idx % 2 == 0 else None
        for col_idx, key in enumerate(field_keys, 1):
            val = entry.get(key, "") or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = data_align
            if fill:
                cell.fill = fill

    # 열 너비 조정
    col_widths = [18, 10, 12, 14, 20, 20, 18, 12, 35, 25, 20, 35, 40]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    now_str = datetime.now(_KST).strftime("%Y%m%d_%H%M")
    filename = f"벤치마킹_결과_{now_str}.xlsx"
    encoded = filename.encode("utf-8")

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded.decode('latin-1', errors='replace')}; filename=\"benchmark_{now_str}.xlsx\"",
        },
    )


@app.post("/api/workflow/benchmark-step1", tags=["Workflow"])
async def benchmark_workflow_step1(request: Request):
    """
    Step 1-A: 벤치마킹 수행.
    Big Tech / Industry 선도사의 AI 적용 사례를 검색하여 결과 테이블을 생성합니다.
    프롬프트와 순서 무관하게 독립 실행 가능.
    """
    from benchmark_search import search_benchmarks

    if not _wf_excel_tasks and "parsed" not in _workflow_cache:
        raise HTTPException(400, "엑셀 또는 As-Is 워크플로우를 먼저 업로드하세요.")

    body = await request.json()
    extra_companies = body.get("companies", "")  # 추가 검색 기업명
    scope = body.get("scope", "l4")  # "l3" = 전체 시트, "l4" = 현재 시트만
    sheet_id_bm = body.get("sheet_id", "") if scope != "l3" else ""

    process_name, task_summary, pain_summary = _build_task_and_pain_summary(sheet_id_bm)

    # ── 공통 인덱스 + L4/L3/L2 세부 정보 구축 ───────────────────────────────
    _, excel_by_l4, excel_by_l3, excel_by_l2 = _build_excel_index()

    _PAIN_MAP = [
        ("pain_time", "시간/속도"), ("pain_accuracy", "정확성"),
        ("pain_repetition", "반복/수작업"), ("pain_data", "정보/데이터"),
        ("pain_system", "시스템/도구"), ("pain_communication", "의사소통/협업"),
    ]

    def _get_pain_points(tasks: list) -> list[str]:
        found = set()
        for t in tasks:
            for field, label in _PAIN_MAP:
                if getattr(t, field, ""):
                    found.add(label)
        return list(found)

    l4_details: list[dict] = []
    l3_details: list[dict] = []
    l2_details: list[dict] = []
    seen_l4: set[str] = set()
    seen_l3: set[str] = set()
    seen_l2: set[str] = set()

    if "parsed" in _workflow_cache:
        parsed = _workflow_cache["parsed"]
        # ── JSON 파일 = L3, 각 시트 = L4 ──────────────────────────────────────
        # scope="l4" + sheet_id: 선택한 시트가 바로 L4 (시트 이름 = L4 이름)
        # scope="l3"           : 모든 시트가 L4들의 집합 (파일 전체 = L3)
        if scope == "l3":
            target_sheets = parsed.sheets
        elif sheet_id_bm:
            matched_bm = [s for s in parsed.sheets if s.sheet_id == sheet_id_bm]
            if not matched_bm:
                print(f"[SCOPE⚠] benchmark_workflow_step1 sheet_id_bm='{sheet_id_bm}' 매칭 없음! 전체: {[s.sheet_id for s in parsed.sheets]}", flush=True)
                target_sheets = parsed.sheets
            else:
                target_sheets = matched_bm
        else:
            target_sheets = parsed.sheets
        print(f"[SCOPE] benchmark_workflow_step1 scope='{scope}' sheet_id_bm='{sheet_id_bm}' → target_sheets={[s.sheet_id+'/'+s.name for s in target_sheets]}", flush=True)

        # JSON 파일 전체 = 하나의 L3 → 모든 시트에서 L3 이름·ID 수집
        # (동일한 L4 이름이 여러 L3에 걸쳐 존재할 수 있으므로 L3로 반드시 좁혀야 함)
        all_l3_names_bm: set[str] = set()
        all_l3_ids_bm: set[str] = set()
        for s_all in parsed.sheets:
            for n in s_all.nodes.values():
                if n.level == "L3" and n.label:
                    all_l3_names_bm.add(n.label.strip())
                if n.level == "L5" and n.task_id:
                    parts = n.task_id.split(".")
                    if len(parts) >= 2:
                        all_l3_ids_bm.add(".".join(parts[:2]))

        for s in target_sheets:
            # 시트 이름 자체 = L4 이름 (시트 = L4 단위)
            sheet_l4_name = (s.name or s.sheet_id).strip()
            if sheet_l4_name and sheet_l4_name not in seen_l4:
                seen_l4.add(sheet_l4_name)
                sheet_tids: set[str] = set()
                sheet_l4_ids: set[str] = set()
                sheet_l3_ids: set[str] = set()
                # L4 노드는 흐름 연결용 connector이므로 항상 L5만 사용
                l5_only_nodes = [n for n in s.nodes.values() if n.level == "L5"]
                source_nodes = l5_only_nodes if l5_only_nodes else s.nodes.values()
                for node in source_nodes:
                    if node.task_id:
                        sheet_tids.add(node.task_id)
                        parts = node.task_id.split(".")
                        if len(parts) >= 3:
                            sheet_l4_ids.add(".".join(parts[:3]))
                        if len(parts) >= 2:
                            sheet_l3_ids.add(".".join(parts[:2]))

                # 0순위: L4 이름 + L3 범위 조합 (동일 L4명이 여러 L3에 걸쳐 있을 때 필수)
                if all_l3_names_bm:
                    matched = [t for t in _wf_excel_tasks
                               if t.l4 and t.l4.strip() == sheet_l4_name
                               and t.l3 and t.l3.strip() in all_l3_names_bm]
                elif all_l3_ids_bm:
                    matched = [t for t in _wf_excel_tasks
                               if t.l4 and t.l4.strip() == sheet_l4_name
                               and t.l3_id in all_l3_ids_bm]
                else:
                    matched = [t for t in _wf_excel_tasks if t.l4 and t.l4.strip() == sheet_l4_name]
                if not matched:
                    matched = [t for t in _wf_excel_tasks if t.id in sheet_tids]
                if not matched:
                    matched = [t for t in _wf_excel_tasks if t.l4_id in sheet_l4_ids]
                if not matched:
                    matched = [t for t in _wf_excel_tasks if t.l3_id in sheet_l3_ids]
                l4_details.append({
                    "name": sheet_l4_name,
                    "task_id": s.sheet_id,
                    "pain_points": _get_pain_points(matched),
                    "description": "; ".join(t.description for t in matched[:3] if t.description),
                    "task_names": [t.name for t in matched[:5]],
                    # matched 직접 저장 → l5_tasks 빌드 시 정확히 이 L4 범위만 사용
                    "_matched_tasks": matched,
                })

            # L3/L2는 시트 내부 노드에서 추출 (파일 수준 컨텍스트)
            for node in s.nodes.values():
                lv, lbl, tid = node.level, node.label.strip(), (node.task_id or "").strip()
                if not lbl:
                    continue
                if lv == "L3" and lbl not in seen_l3:
                    seen_l3.add(lbl)
                    matched_l3 = excel_by_l3.get(tid, [])
                    l3_details.append({"name": lbl, "task_id": tid,
                        "pain_points": _get_pain_points(matched_l3),
                        "description": "; ".join(t.description for t in matched_l3[:3] if t.description)})
                elif lv == "L2" and lbl not in seen_l2:
                    seen_l2.add(lbl)
                    l2_details.append({"name": lbl, "task_id": tid,
                        "pain_points": _get_pain_points(excel_by_l2.get(tid, []))})

    # JSON 없을 때 엑셀만으로 구성 (L4 = 엑셀의 l4 컬럼)
    if not l4_details:
        for l4_id, tasks in excel_by_l4.items():
            name = tasks[0].l4 if tasks else ""
            if name and name not in seen_l4:
                seen_l4.add(name)
                l4_details.append({"name": name, "task_id": l4_id,
                    "pain_points": _get_pain_points(tasks),
                    "description": "; ".join(t.description for t in tasks[:3] if t.description),
                    "task_names": [t.name for t in tasks[:5]]})
    if not l3_details:
        for l3_id, tasks in excel_by_l3.items():
            name = tasks[0].l3 if tasks else ""
            if name and name not in seen_l3:
                seen_l3.add(name)
                l3_details.append({"name": name, "task_id": l3_id,
                    "pain_points": _get_pain_points(tasks), "description": ""})
    if not l2_details:
        for l2_id, tasks in excel_by_l2.items():
            name = tasks[0].l2 if tasks else ""
            if name and name not in seen_l2:
                seen_l2.add(name)
                l2_details.append({"name": name, "task_id": l2_id, "pain_points": _get_pain_points(tasks)})

    l4_details, l3_details, l2_details = l4_details[:8], l3_details[:6], l2_details[:3]
    l4_names = [d["name"] for d in l4_details]

    # L3/L2는 Excel에서 l5_tasks의 상위 계층을 역추적해 가져옴
    # (다이어그램 노드 기반 수집은 다른 L3 노드가 섞여 오염됨)
    _l3_seen: set[str] = set()
    _l2_seen: set[str] = set()
    _l3_from_excel: list[str] = []
    _l2_from_excel: list[str] = []
    for d in l4_details:
        for t in (d.get("_matched_tasks") or excel_by_l4.get(d.get("task_id", ""), [])):
            if t.l3 and t.l3 not in _l3_seen:
                _l3_seen.add(t.l3)
                _l3_from_excel.append(t.l3)
            if t.l2 and t.l2 not in _l2_seen:
                _l2_seen.add(t.l2)
                _l2_from_excel.append(t.l2)
    # Excel 역추적이 비어 있으면 기존 다이어그램 기반 fallback
    l3_names = _l3_from_excel or [d["name"] for d in l3_details]
    l2_names = _l2_from_excel or [d["name"] for d in l2_details]

    # L5 task — 이 L4(시트) 범위에 속하는 엑셀 task 수집
    l5_tasks = []
    for d in l4_details:
        # _matched_tasks: 이 L4 시트 범위에서 정확히 매칭된 엑셀 태스크
        # (L4 scope면 L5 노드 task_id 기반 매칭, L3 scope면 전체 노드 기반)
        direct_tasks = d.pop("_matched_tasks", None)
        task_source = (
            direct_tasks
            or excel_by_l4.get(d.get("task_id", ""), [])
            or [t for t in _wf_excel_tasks if t.name in (d.get("task_names") or [])]
        )
        for t in task_source:
            if t.name and not any(x["name"] == t.name for x in l5_tasks):
                l5_tasks.append({
                    "name": t.name,
                    "description": t.description or "",
                    "l4": d["name"],
                })

    # 검증 로그 — bm_data에 실제 어떤 계층 정보가 들어가는지 확인
    print(f"[benchmark] 메타데이터 검증")
    print(f"  process_name (L3): {process_name}")
    print(f"  l2_names: {l2_names}")
    print(f"  l3_names: {l3_names}")
    print(f"  l4_names: {l4_names}")
    print(f"  l5_tasks ({len(l5_tasks)}개):")
    for lt in l5_tasks[:10]:
        print(f"    [{lt['l4']}] {lt['name']}: {lt['description'][:60] if lt['description'] else '(description 없음)'}")

    bm_data = {
        "process_name": process_name,
        "agents": [],
        "l2_names": l2_names,
        "l3_names": l3_names,
        "l4_names": l4_names,
        "l4_details": l4_details,
        "l3_details": l3_details,
        "l2_details": l2_details,
        "l5_tasks": l5_tasks[:30],
        "blueprint_summary": (
            f"{process_name} 프로세스 (L3: {', '.join(l3_names[:3])}, L4: {', '.join(l4_names[:3])}) "
            f"AI 적용 벤치마킹"
        ),
    }

    # 추가 기업 검색 쿼리
    if extra_companies:
        focus = l4_names[0] if l4_names else (l3_names[0] if l3_names else process_name)
        bm_data["extra_queries"] = [
            f"{extra_companies} '{process_name}' AI automation case study results 2024 2025",
            f"{extra_companies} '{focus}' AI 자동화 도입 사례 성과",
        ]

    # ── SSE 스트리밍 설정 ────────────────────────────────────────────────────
    import asyncio as _asyncio
    from fastapi.responses import StreamingResponse as _StreamingResponse

    progress_queue: _asyncio.Queue = _asyncio.Queue()

    async def _do_benchmark():
        """검색 + LLM 분석을 background에서 실행하고 progress_queue에 이벤트를 넣습니다."""
        try:
            bm_sr = await search_benchmarks(bm_data, progress_cb=progress_queue.put_nowait)
            raw = bm_sr.get("results", [])
            bm_log = bm_sr.get("search_log", [])

            global _wf_benchmark_results, _wf_benchmark_table, _wf_search_log
            _wf_benchmark_results = raw
            _wf_search_log = bm_log

            progress_queue.put_nowait({"type": "llm_analyze", "total": len(raw)})

            # LLM으로 벤치마킹 결과 분석하여 구조화된 테이블 생성
            from benchmark_search import _translate_to_en as _bm_translate
            l4_en_map = "\n".join(
                f"  - {d['name']} (영어: {_bm_translate(d['name'])})"
                for d in l4_details[:6]
            )

            # 두산 HR 고유 용어 설명 (LLM이 맥락을 이해하도록)
            _DOOSAN_HR_TERMS = """
## 두산 HR 프로그램 용어 설명 (벤치마킹 맥락 이해용)
- PDS (Performance Development Survey): 선임→수석 승진 대상자 역량 평가 프로그램. 다면진단·전문성 인터뷰 등으로 구성
- LDS (Leadership Development Survey): 수석→팀장 승진 대상자 리더십 역량 평가 프로그램
- EDS (Executive Development Survey): 임원 승진 대상자 평가 프로그램
- SDS (Specialist Development Survey): Specialist 트랙 승진 대상자 전문성 평가 프로그램
→ 모두 두산그룹 내부의 승진·역량개발 평가 체계이며, 글로벌 벤치마킹 시 "Performance Review", "Leadership Assessment", "Succession Planning", "Talent Assessment" 등으로 대응하여 검색합니다.
"""
            _doosan_term_ctx = _DOOSAN_HR_TERMS if any(
                term in " ".join(l2_names + l3_names) for term in ["PDS", "LDS", "EDS", "SDS"]
            ) else ""

            bm_analysis_system = f"""당신은 글로벌 AI 업무 혁신 벤치마킹 분석 전문가입니다.
영어·한국어 검색 결과를 엄격하게 분석하여 실제 근거가 있는 기업의 AI 적용 사례만 추출합니다.

## 프로세스: {process_name}
## L2 대분류: {', '.join(l2_names)}
## L3 영역: {', '.join(l3_names)}
{_doosan_term_ctx}
## L4 세부 활동 및 영어 대응어 (최우선 매핑 기준):
{l4_en_map if l4_en_map else chr(10).join(f"  - [{d['task_id']}] {d['name']}" for d in l4_details[:6])}

## ⚠️ 최우선 원칙: 모르면 모른다고 명확히 표기
**절대 사실을 꾸며내거나 추측으로 사례를 만들지 마세요.**
- 검색 결과에 해당 기업명이 명확히 나오지 않으면 → 그 사례는 제외
- 수치 성과가 없으면 → outcome에 "수치 미확인" 명시
- URL이 없으면 → url 빈 문자열 (URL 임의 생성 절대 금지)
- 관련 사례가 없으면 → benchmark_table을 빈 배열로 반환하고 no_cases_note에 이유 명시

## 영어 검색 결과 해석 규칙
검색 결과가 영어여도 아래 L4 활동의 영어 대응 관계를 이용해 관련성을 판단하세요:
{chr(10).join(f'- "{_bm_translate(d["name"])}" → {d["name"]}' for d in l4_details[:8]) if l4_details else "- 위 L4 세부 활동 목록 참조"}

## ⚠️ process_area 필드 규칙 (매우 중요)
`process_area` 필드에는 반드시 위 **"L4 세부 활동"** 목록에 있는 이름 중 하나만 입력하세요.
- 목록에 없는 값 절대 금지 (창작 금지)
- 목록: {", ".join(f'"{n}"' for n in l4_names[:8])}
- 해당하는 L4가 없을 때만 L3명 사용: {", ".join(f'"{n}"' for n in l3_names[:3])}

## 관련성 판단 기준
- 완전 동일하지 않아도, **동일한 HR 업무 기능**이면 포함
- L4 단위 매칭이 안 되면 L3 단위로 매핑
- L3도 안 되면 L2 단위로 매핑

## 포함 기준 (4가지 모두 충족해야 포함)
1. **기업명**: 고유 기업명(실명)이 검색 결과에 명확히 언급됨
2. **내용**: AI 적용 방법 또는 성과가 구체적으로 언급됨
3. **URL**: 검색 결과에 실제 URL이 있어야만 포함 — URL이 없으면 해당 항목 DROP (출처 없는 사례 불인정)
4. **출처 유형**: 아래 허용 유형만 인정
   - ✅ 공식 기업 블로그 / 기술 블로그 (engineering.fb.com, cloud.google.com 등)
   - ✅ 공식 케이스 스터디 / 백서 / 연구 보고서
   - ✅ 학술 논문, LinkedIn 공식 게시물
   - ✅ 글로벌 리서치·컨설팅 기관 보고서에서 **특정 기업 사례를 인용한 경우** (weforum.org, shrm.org, gartner.com, mckinsey.com, hbr.org 등) — source는 인용된 기업명으로 기재
   - ✅ SAP/Workday/ServiceNow/Oracle 공식 고객 사례 페이지
   - ✅ techcrunch, reuters, bloomberg, venturebeat 등 글로벌 IT/비즈니스 미디어 — 구체적 기업명·수치가 포함된 경우 허용
   - ❌ 한국 뉴스 포털 (chosun.com, mk.co.kr, hani.co.kr, 네이버뉴스 등) — 내용이 얕고 출처 불명확
   - ❌ 벤더 마케팅 자료, 일반 AI 통계, 기업명 미확인

## source 필드 핵심 규칙
`source`는 반드시 **글로벌 인지도가 있는 대기업**의 고유 명칭이어야 합니다.

**✅ 허용 기업 기준 — 아래 카테고리에 해당하는 기업만 포함:**
- **Big Tech**: Google, Amazon, Meta, Microsoft, Apple, IBM, Oracle, Salesforce, SAP (도입처로서)
- **글로벌 소비재·유통**: Unilever, P&G, Nestlé, Coca-Cola, PepsiCo, Walmart, Target, IKEA, L'Oréal
- **글로벌 제조·산업**: Siemens, GE, Honeywell, 3M, DuPont, ExxonMobil, Shell, Bosch, Caterpillar
- **글로벌 금융**: JPMorgan, Goldman Sachs, Citi, HSBC, Deutsche Bank, Mastercard, Visa
- **글로벌 물류·항공**: DHL, FedEx, UPS, Delta, American Airlines, Maersk
- **한국 대기업 (글로벌 인지도 있는)**: 삼성전자, 현대자동차, SK하이닉스, LG전자, 포스코, 롯데, 두산
- **기타 Forbes Global 500 수준 대기업** — 일반인도 아는 기업

**❌ 절대 금지 — 해당 행 전체 DROP:**
- McKinsey, BCG, Bain, Deloitte, PwC, EY, KPMG, Accenture, Gartner, Forrester, IDC (보고서 작성자)
- **국내 중소·중견 IT기업** (에스앤아이, 인포뱅크, 더존비즈온, 솔트룩스 등 — 글로벌 인지도 없음)
- **스타트업·벤처기업** — 설립 10년 이내 또는 직원 수 1,000명 미만 추정 기업
- "Fortune 500 기업", "글로벌 대기업", "한 제조사", "한 기업", "(익명)", "undisclosed", "anonymous"
- 솔루션 벤더 자체 (Workday, ServiceNow 등이 **도입처**가 아닌 **제공자**로 기재된 경우)
- 기업명이 특정되지 않거나 글로벌 인지도가 없는 모든 기업
→ **모르는 기업 이름이 나오면 → 그 사례는 제외. 애매하면 DROP.**

## 중요 원칙
- **영어 사례를 한국어로 번역하여 설명** — 영어 원문 그대로 두지 말 것
- 글로벌 대기업 사례만 → 소규모 기업·스타트업 제외
- 수치 성과(%, 시간, 비용) 포함 사례 우선
- 관련 사례가 3건 미만이면 억지로 채우지 말고 있는 것만 반환

## ⚠️ 사례 수량 규칙 (필수)
- **적격 조건을 만족하는 모든 기업 사례를 포함하세요. 상한 없음.**
- "상위 N건만", "대표 5건" 같은 임의 제한 절대 금지
- 같은 기업이라도 도입 사례·L4 활동이 다르면 각각 별개 행으로 분리
- 사례가 부족한 건 OK이지만, 사례가 많을 때 **임의로 추려내지 말 것**

## ⚠️ 각 사례는 간결하게 (JSON 토큰 한도 내에 모든 사례 포함되도록)
- use_case: 1~2문장 (요점만)
- outcome: 수치 + 간단 설명 1문장
- ai_technology·infrastructure: 쉼표 구분 키워드 3~5개
- implication: 1~2문장
- 나머지 필드도 장황하지 않게 — **간결성 > 장황함**
- 목표: 한 사례당 300자 이내 → 15건이어도 JSON 전체가 잘리지 않도록

## 출력 형식 (JSON만, 마크다운 코드 블록 없음)
{{
  "benchmark_table": [
    {{
      "source": "AI를 실제 도입·운영한 기업명 (컨설팅펌·리서치펌 절대 금지)",
      "company_type": "Tech 선도 | 非Tech 실제 구현",
      "industry": "산업군",
      "process_area": "매핑된 L4 활동명 (위 L4 목록 중 해당하는 것, 없으면 L3명)",
      "ai_adoption_goal": "AI 도입 목표 (비용절감·속도개선·정확도향상 등)",
      "ai_technology": "적용 AI 기술 (한국어로)",
      "key_data": "핵심 데이터",
      "adoption_method": "도입 방식 (자체개발 | SaaS | API | 파트너십)",
      "use_case": "구체적 적용 사례 (한국어로 번역, 1~2문장, L4 활동 기준)",
      "outcome": "성과/효과 (수치 포함, 없으면 '수치 미확인 — 정성적 효과: ...'로 기술)",
      "infrastructure": "인프라/시스템 기반",
      "implication": "두산 향 시사점 — 해당 L4 활동에 어떻게 적용할 수 있는지",
      "url": "검색 결과 출처 URL 목록에 있는 것만 — URL이 없으면 이 항목 자체를 benchmark_table에서 제외"
    }}
  ],
  "no_cases_note": "관련 사례가 없을 때만 이유 기재 (예: '검색 결과에 해당 프로세스 직접 사례 없음'). 사례가 있으면 빈 문자열.",
  "summary": "벤치마킹 종합 요약 (3~5문장, 영어 사례도 한국어로 설명. 사례가 없으면 없다고 명시)"
}}
"""

            # Sonar Pro 결과는 같은 쿼리의 citation URL들이 각각 별도 row로 들어옴.
            from collections import defaultdict as _defaultdict
            query_groups: dict = _defaultdict(lambda: {"content": "", "urls": []})
            for r in raw[:200]:
                q = r.get("query", r.get("title", ""))
                if r.get("content") and not query_groups[q]["content"]:
                    query_groups[q]["content"] = r.get("content", "")
                if r.get("url"):
                    query_groups[q]["urls"].append(r["url"])

            bm_user = "## 웹 검색 결과\n\n"
            for i, (q, g) in enumerate(query_groups.items(), 1):
                bm_user += f"### [{i}] 검색 쿼리: {q[:100]}\n"
                if g["urls"]:
                    bm_user += "- 출처 URL 목록:\n"
                    for url in g["urls"][:12]:
                        bm_user += f"  - {url}\n"
                bm_user += f"- 내용: {g['content'][:4000]}\n\n"

            bm_user += (
                "\n위 검색 결과에서 벤치마킹 테이블을 작성해주세요.\n"
                "⚠️ **적격 조건을 만족하는 모든 기업 사례를 빠짐없이 포함**하세요 — 상한 없음. "
                "10건이든 20건이든 전부 나열하고, '상위 N건' 같은 임의 제한 절대 금지.\n"
                "각 사례의 url 필드는 위 검색 결과에 실제로 나온 URL만 기재하세요 (임의 생성 금지).\n"
                "관련 사례가 없으면 benchmark_table을 빈 배열로 반환하고 no_cases_note에 이유를 명시하세요."
            )

            # 벤치마킹 LLM 분석 — Anthropic non-streaming 한도(최대 ~16K) 내에서 실행
            # 간결성 지침(프롬프트)으로 15건도 수용 가능
            result_data = await _call_llm_step1(bm_analysis_system, [{"role": "user", "content": bm_user}], max_tokens=16384)

            sheet_key = sheet_id_bm or "__default__"

            if not result_data:
                sheet_rows = [
                    {
                        "source": r.get("title", "")[:30],
                        "industry": "",
                        "process_area": process_name,
                        "ai_technology": "",
                        "use_case": r.get("snippet", "")[:100],
                        "outcome": "",
                        "implication": "",
                        "url": r.get("url", ""),
                    }
                    for r in raw[:10]
                    if r.get("url")
                ]
            else:
                raw_table = result_data.get("benchmark_table", [])
                sheet_rows = [
                    row for row in raw_table
                    if row.get("url")
                    and _is_valid_benchmark_source(row.get("source", ""))
                    and not _is_news_url(row.get("url", ""))
                    and (row.get("use_case") or row.get("outcome"))
                ]
                valid_areas = set(l4_names) | set(l3_names)
                fallback_area = l4_names[0] if l4_names else (l3_names[0] if l3_names else process_name)
                for row in sheet_rows:
                    area = row.get("process_area", "")
                    if area not in valid_areas:
                        matched_l4 = next(
                            (ln for ln in l4_names if ln in area or area in ln), None
                        )
                        row["process_area"] = matched_l4 if matched_l4 else fallback_area

            _wf_benchmark_table[sheet_key] = sheet_rows

            global _wf_chat_history
            summary_text = result_data.get("summary", "") if result_data else f"벤치마킹 검색 완료: {len(raw)}개 결과"
            _wf_chat_history.append({"role": "assistant", "content": f"[벤치마킹 완료 — {process_name}] {summary_text}"})

            if _wf_step1_cache:
                all_rows = [r for rows in _wf_benchmark_table.values() for r in rows]
                _wf_step1_cache["benchmark_table"] = all_rows

            if _current_session_id:
                _save_session_data(_current_session_id)

            progress_queue.put_nowait({
                "type": "final",
                "ok": True,
                "result_count": len(raw),
                "sheet_id": sheet_key,
                "benchmark_table": sheet_rows,
                "all_benchmark_table": {k: v for k, v in _wf_benchmark_table.items()},
                "summary": summary_text,
                "search_log": bm_log,
            })

        except Exception as exc:
            import traceback
            print(f"[benchmark SSE] 오류: {exc}\n{traceback.format_exc()}")
            progress_queue.put_nowait({"type": "error", "message": str(exc)})

    async def _event_stream():
        import json as _json
        task = _asyncio.create_task(_do_benchmark())
        try:
            while True:
                try:
                    event = await _asyncio.wait_for(progress_queue.get(), timeout=300)
                except _asyncio.TimeoutError:
                    yield "data: {\"type\": \"error\", \"message\": \"타임아웃\"}\n\n"
                    break
                yield f"data: {_json.dumps(event, ensure_ascii=False)}\n\n"
                if event.get("type") in ("final", "error"):
                    break
        finally:
            task.cancel()

    return _StreamingResponse(
        _event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── Gap 분석 ─────────────────────────────────────────────────

@app.post("/api/workflow/gap-analysis", tags=["Workflow"])
async def generate_gap_analysis(request: Request):
    """
    벤치마킹 결과(선도사 To-Be) vs 두산 현재(As-Is)를 비교하여 Gap 분석 수행.
    Perplexity 추가 검색으로 Gap 분석 고도화.
    벤치마킹이 먼저 수행되어 있어야 합니다.

    [Gap 유형]
    A. 신규: 벤치마킹에는 존재하나 As-Is에 없는 L4/L5. AI 기반 신규 프로세스.
    B. 전환: 양쪽 다 존재하나 수행 방식이 완전히 다름. AI로 대체/변환.
    C. 폐기/통합: As-Is에만 있고 벤치마킹에는 없음. 폐기되었거나 타 프로세스에 흡수.
    """
    from benchmark_search import search_benchmarks

    global _wf_gap_analysis

    # sheet_id 수신: 벤치마킹한 L4 시트 단위로 Gap 분석 스코프 제한
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    gap_sheet_id: str = body.get("sheet_id", "") or ""

    # 해당 시트 벤치마킹 결과만 사용 (없으면 전체 fallback)
    if gap_sheet_id and gap_sheet_id in _wf_benchmark_table:
        all_bm_rows = _wf_benchmark_table[gap_sheet_id]
    else:
        all_bm_rows = [r for rows in _wf_benchmark_table.values() for r in rows]

    if not all_bm_rows:
        raise HTTPException(400, "벤치마킹을 먼저 수행해주세요.")

    process_name, task_summary, pain_summary = _build_task_and_pain_summary(gap_sheet_id)
    asis_context = _build_mapped_asis_context(gap_sheet_id)

    # 벤치마킹 요약 (선도사 To-Be)
    benchmark_summary_text = f"## 벤치마킹 결과 (전체 {len(all_bm_rows)}건)\n"
    for bm in all_bm_rows[:20]:
        benchmark_summary_text += (
            f"- **{bm.get('source', '')}** ({bm.get('industry', '')}) | "
            f"L4: {bm.get('process_area', '')} | "
            f"기술: {bm.get('ai_technology', '')} | "
            f"사례: {bm.get('use_case', '')} | "
            f"성과: {bm.get('outcome', '')}\n"
        )

    # ── Perplexity 심층 검색: 기존 벤치마킹 결과 기반 고도화 ─────────────────
    # 기존 _wf_benchmark_table 사례를 바탕으로 사례별 검증/심화 쿼리 생성
    # 새로운 사례 탐색이 아니라, 이미 찾은 사례를 더 깊이 파고드는 것이 목적
    gap_search_context = ""
    try:
        # L4 이름 추출 — gap_sheet_id 있으면 해당 시트만, 없으면 전체 4개
        _, excel_by_l4, _, _ = _build_excel_index()
        if gap_sheet_id and "parsed" in _workflow_cache:
            _matched_sheet = next(
                (s for s in _workflow_cache["parsed"].sheets if s.sheet_id == gap_sheet_id),
                None,
            )
            l4_names_for_search = [((_matched_sheet.name or _matched_sheet.sheet_id).strip())] if _matched_sheet else []
        else:
            l4_names_for_search = [d["name"] for d in
                                   (lambda p: [{"name": (s.name or s.sheet_id).strip()}
                                               for s in p.sheets] if p else [])
                                   (_workflow_cache.get("parsed"))][:4]
        if not l4_names_for_search:
            l4_names_for_search = [tasks[0].l4 for tasks in list(excel_by_l4.values())[:4] if tasks]

        # ── 기존 벤치마킹 사례 기반 심화 쿼리 생성 ──
        # 상위 사례별: "{기업} {AI기술} {L4} implementation ROI outcome 2024 2025"
        deepening_queries: list[str] = []
        seen_sources: set[str] = set()
        for bm in all_bm_rows[:8]:
            src = bm.get("source", "").strip()
            tech = bm.get("ai_technology", "").strip()
            l4 = bm.get("process_area", "").strip()
            if not src or src in seen_sources:
                continue
            seen_sources.add(src)
            q = f"{src} {tech} {l4} AI implementation outcome ROI 2024 2025".strip()
            deepening_queries.append(q)
            if len(deepening_queries) >= 4:
                break

        # As-Is L4 단위 Gap 탐색 쿼리 (두산 현재 수준 vs 선도사 비교)
        for l4n in l4_names_for_search[:2]:
            deepening_queries.append(
                f"{process_name} {l4n} AI adoption gap challenge barrier enterprise 2024"
            )

        gap_bm_data = {
            "process_name": process_name,
            "agents": [],
            "l4_details": [{"name": n, "task_id": "", "pain_points": [], "description": "", "task_names": []}
                           for n in l4_names_for_search],
            "l4_names": l4_names_for_search,
            "l3_names": [process_name],
            "l2_names": [],
            "l3_details": [], "l2_details": [],
            "l5_tasks": [],
            "blueprint_summary": f"{process_name} 기존 벤치마킹 사례 심층 검증",
            "extra_queries": deepening_queries,
        }
        _gap_sr = await search_benchmarks(gap_bm_data)
        _gap_raw = _gap_sr.get("results", [])
        if _gap_raw:
            from collections import defaultdict as _ddict2
            _gq: dict = _ddict2(lambda: {"content": "", "urls": []})
            for r in _gap_raw:
                _q2 = r.get("query", r.get("title", ""))
                if r.get("content") and not _gq[_q2]["content"]:
                    _gq[_q2]["content"] = r.get("content", "")
                if r.get("url"):
                    _gq[_q2]["urls"].append(r["url"])
            gap_search_context = (
                f"\n## 기존 벤치마킹 사례 심층 검증 결과 (Perplexity)\n"
                f"※ 벤치마킹 결과표의 기업·사례를 대상으로 추가 검색한 내용입니다.\n"
            )
            for _q2, _g2 in list(_gq.items())[:8]:
                gap_search_context += f"\n### 검색: {_q2[:100]}\n"
                if _g2["urls"]:
                    gap_search_context += "출처: " + ", ".join(_g2["urls"][:2]) + "\n"
                gap_search_context += f"{_g2['content'][:500]}\n"
        print(f"[gap-analysis] 심층 검색 완료: {len(deepening_queries)}개 쿼리, {len(_gap_raw)}건 결과")
    except Exception as _e:
        print(f"[gap-analysis] 추가 검색 실패 (무시): {_e}")

    # 분석 스코프 명시 텍스트 (L4 단위면 해당 L4명 강조)
    scope_notice = ""
    if gap_sheet_id and l4_names_for_search:
        scope_notice = f"\n## ⚠️ 분석 범위 한정\n이번 Gap 분석은 **'{l4_names_for_search[0]}'** L4 Activity 단위 분석입니다.\n- gap_items의 l4_activity는 반드시 **'{l4_names_for_search[0]}'** 또는 그 하위에서 파생된 명칭만 사용하세요.\n- 다른 L4(사내 추천, 면접 등 무관한 활동)를 l4_activity로 작성하지 마세요.\n- 벤치마킹 사례가 다른 영역에 해당하면 해당 항목은 Gap_items에서 제외하세요.\n"

    gap_system = f"""당신은 경영 혁신 전문가입니다. 두산 HR 프로세스의 As-Is 현황과 글로벌 선도사 벤치마킹 결과를 비교하여 Gap 분석을 수행합니다.
{scope_notice}
## 벤치마킹 결과 (선도사 To-Be)
{benchmark_summary_text}

## 두산 As-Is 현황
{task_summary}

{pain_summary}

{asis_context[:2500]}
{gap_search_context}

## Gap 유형 분류 기준 (반드시 준수)
A. 신규: 벤치마킹에는 존재하나 As-Is에 없는 L4/L5. AI 기반으로 새롭게 생겨난 프로세스. 두산에 도입 여부 검토
B. 전환: 양쪽 다 존재하나 수행 방식이 완전히 다름. 벤치마킹에서는 AI로 대체/변환하여 수행 중. 기존 프로세스를 전환 검토
C. 폐기/통합: As-Is에만 있고 벤치마킹에는 없음. 선도 사례에서는 AI 도입과 함께 폐기되었거나 타 프로세스에 흡수. 존치 필요성 재검토

## Gap 분석 지침
- 각 L4 Activity 단위로 As-Is vs To-Be Gap 분석
- gap_type: 반드시 "A. 신규" / "B. 전환" / "C. 폐기/통합" 중 하나로만 분류
- **위 벤치마킹 결과표의 기업·사례를 1차 근거로 사용하고, 심층 검증 결과로 해당 사례의 실제 성과·구현 방식을 보강하여 분석**
- 벤치마킹 사례에서 기업명 구체적으로 인용 (심층 검증 결과에서 추가 수치/사실 발견 시 반영)
- quick_wins: 단기(6개월 내) 즉시 시행 가능한 액션 — 동사로 시작, 1문장 이내, 구체적 실행 단위
- strategic_actions: 장기(5년) 전략 과제 — 동사로 시작, 1문장 이내, 방향성과 기대 효과 포함
- executive_summary: 3문장 이내로 압축. ① Gap 건수·유형 분포 (숫자 포함) ② 최우선 과제 ③ 기대 효과 순서로 작성. 산문이 아닌 팩트 중심

## gap_wrap_up 작성 기준 — MBB 컨설턴트 스타일 (MECE, 간결, 시사점 중심)
gap_items 전체를 종합하여 세 차원(프로세스 / 인프라 / 데이터)의 Gap을 작성합니다.
해당 Gap이 실제로 존재하는 차원만 작성하고, 없으면 반드시 null로 출력합니다.
세 차원은 **MECE**해야 합니다 — 프로세스(What/How) / 인프라(With What) / 데이터(With What Data) 로 중복 없이 분리.

각 차원은 다음 5가지 필드로 구성합니다:
- headline: 이 차원 Gap의 핵심을 압축한 한 줄 (명사형 종결, 20자 이내, 예: "AI 기반 자동화 부재로 수작업 의존도 극심")
- as_is: 두산 현재 수준을 1문장으로 (수치·실태 포함, 주어 없이 서술)
- to_be: 글로벌 선도사 수준을 1문장으로 (기업명 반드시 포함, "~사는 ~를 통해 ~% 달성" 형식)
- gaps: 핵심 Gap 포인트 2~3개, 각각 15자 이내 명사구 bullet (MECE, 중복 금지)
- implication: 두산이 이 차원에서 해야 할 것을 1문장으로 (행동 지향, "~을 통해 ~를 달성해야 함" 형식)

## 출력 형식 (JSON만, 마크다운 없음)
{{
  "process_name": "{process_name}",
  "executive_summary": "경영진 요약 — 3문장 이내, 핵심 숫자(Gap 건수·유형 분포)·최우선 과제·기대 효과 순으로 작성",
  "gap_items": [
    {{
      "l4_activity": "L4 활동명 (10자 이내)",
      "as_is": "현재 수준 — 1문장, 30자 이내, 수치 포함",
      "to_be": "선도사 수준 — 기업명 포함, 1문장, 30자 이내",
      "gap_type": "A. 신규",
      "root_cause": "Gap 원인 — 핵심 명사구, 15자 이내",
      "action_plan": "실행 방향 — 동사 시작, 1문장, 30자 이내",
      "priority": 1
    }}
  ],
  "gap_wrap_up": {{
    "process_gap": {{
      "headline": "핵심 한 줄 (20자 이내)",
      "as_is": "현재 수준 1문장",
      "to_be": "선도사 수준 1문장 (기업명 포함)",
      "gaps": ["Gap 포인트 1 (15자 이내)", "Gap 포인트 2 (15자 이내)"],
      "implication": "두산 시사점 1문장"
    }},
    "infra_gap": {{
      "headline": "핵심 한 줄 (20자 이내)",
      "as_is": "현재 수준 1문장",
      "to_be": "선도사 수준 1문장 (기업명 포함)",
      "gaps": ["Gap 포인트 1 (15자 이내)", "Gap 포인트 2 (15자 이내)"],
      "implication": "두산 시사점 1문장"
    }},
    "data_gap": {{
      "headline": "핵심 한 줄 (20자 이내)",
      "as_is": "현재 수준 1문장",
      "to_be": "선도사 수준 1문장 (기업명 포함)",
      "gaps": ["Gap 포인트 1 (15자 이내)", "Gap 포인트 2 (15자 이내)"],
      "implication": "두산 시사점 1문장"
    }}
  }},
  "quick_wins": ["단기(6개월 내) 즉시 시행 가능한 액션 — 동사 시작, 1문장 이내 (3개 이내)"],
  "strategic_actions": ["장기(5년) 전략 과제 — 동사 시작, 1문장 이내 (3개 이내)"]
}}"""

    gap_user = "위 데이터를 기반으로 Gap 분석을 수행해주세요. JSON 형식으로만 응답하세요."

    result_data = await _call_llm_step1(gap_system, [{"role": "user", "content": gap_user}])

    if not result_data:
        raise HTTPException(500, "Gap 분석 생성에 실패했습니다.")

    _wf_gap_analysis = result_data
    _wf_gap_analysis["ok"] = True

    return _wf_gap_analysis


# ── Step 1 프롬프트 기반 기본 설계 ──────────────────────────

@app.post("/api/workflow/generate-step1", tags=["Workflow"])
async def generate_workflow_step1(request: Request):
    """
    Step 1-B: 벤치마킹 + Gap 분석 결과 기반 Workflow 기본 설계 (Top-Down, Lv.2~5).
    수행 순서: 벤치마킹 → Gap 분석 → 기본 설계 (Gap 분석이 선행되어야 함)
    """
    if not _wf_excel_tasks:
        raise HTTPException(400, "엑셀을 먼저 업로드하세요.")

    all_bm_rows = [r for rows in _wf_benchmark_table.values() for r in rows]
    if not all_bm_rows:
        raise HTTPException(400, "벤치마킹을 먼저 수행해주세요.")
    if not _wf_gap_analysis:
        raise HTTPException(400, "Gap 분석을 먼저 수행해주세요. 벤치마킹 → Gap 분석 → 기본 설계 순서로 진행해야 합니다.")

    body = await request.json()
    user_prompt = body.get("prompt", "")
    process_name_override = body.get("process_name", "")
    # 벤치마킹·Gap 분석과 동일한 sheet_id 스코프 유지
    step1_sheet_id: str = body.get("sheet_id", "") or ""

    # 벤치마킹과 동일한 스코프로 컨텍스트 구성
    # sheet_id 있으면 L4 단위, 없으면 L3 전체
    process_name, task_summary, pain_summary = _build_task_and_pain_summary(step1_sheet_id)
    if process_name_override:
        process_name = process_name_override

    # 같은 스코프의 벤치마킹 결과만 사용
    if step1_sheet_id and step1_sheet_id in _wf_benchmark_table:
        scoped_bm_rows = _wf_benchmark_table[step1_sheet_id]
        benchmark_text = f"## 벤치마킹 결과 (시트: {step1_sheet_id}, {len(scoped_bm_rows)}건)\n"
        for bm in scoped_bm_rows[:10]:
            benchmark_text += (
                f"- **{bm.get('source', '')}** ({bm.get('industry', '')}): "
                f"{bm.get('use_case', '')} → {bm.get('outcome', '')}\n"
            )
    else:
        # L3 전체 or fallback: 전체 시트 합산
        benchmark_text = f"## 벤치마킹 결과 (전체 {len(_wf_benchmark_table)}개 시트, 총 {len(all_bm_rows)}건)\n"
        for sheet_key, rows in _wf_benchmark_table.items():
            if rows:
                benchmark_text += f"\n### 시트: {sheet_key}\n"
                for bm in rows[:5]:
                    benchmark_text += (
                        f"- **{bm.get('source', '')}** ({bm.get('industry', '')}): "
                        f"{bm.get('use_case', '')} → {bm.get('outcome', '')}\n"
                    )

    # Gap 분석 결과 — 기본 설계의 핵심 인풋
    gap_text = ""
    if _wf_gap_analysis:
        gap_items = _wf_gap_analysis.get("gap_items", [])
        gap_wrap = _wf_gap_analysis.get("gap_wrap_up", {}) or {}
        gap_text = f"\n## Gap 분석 결과 (벤치마킹 vs As-Is)\n"
        gap_text += f"### 경영진 요약\n{_wf_gap_analysis.get('executive_summary', '')}\n\n"
        if gap_wrap.get("process_gap"):
            gap_text += f"### 프로세스 Gap\n{gap_wrap['process_gap']}\n\n"
        if gap_wrap.get("infra_gap"):
            gap_text += f"### 인프라 Gap\n{gap_wrap['infra_gap']}\n\n"
        if gap_wrap.get("data_gap"):
            gap_text += f"### 데이터 Gap\n{gap_wrap['data_gap']}\n\n"
        gap_text += f"### L4 단위 Gap 항목 ({len(gap_items)}건)\n"
        for g in gap_items[:10]:
            gap_text += (
                f"- [{g.get('gap_type', '')}] **{g.get('l4_activity', '')}**: "
                f"{g.get('gap_description', '')} → {g.get('action_plan', '')}\n"
            )

    # As-Is + 엑셀 매핑 컨텍스트 (벤치마킹·Gap과 동일 스코프)
    asis_context = _build_mapped_asis_context(step1_sheet_id)

    step1_scope = "l4" if step1_sheet_id else "l3"
    system = _step1_system_prompt(process_name, task_summary, pain_summary,
                                  benchmark_text + gap_text + "\n\n" + asis_context,
                                  scope=step1_scope)

    global _wf_chat_history
    actual_prompt = user_prompt or "선도사례를 분석하여 To-Be Workflow 기본 설계를 수행해주세요."
    _wf_chat_history.append({"role": "user", "content": actual_prompt})

    result_data = await _call_llm_step1(system, list(_wf_chat_history))

    if not result_data:
        raise HTTPException(500, "AI 모델 호출에 실패했습니다. API 키를 확인하세요.")

    _wf_chat_history.append({"role": "assistant", "content": json.dumps(result_data, ensure_ascii=False)[:500]})

    result_dict = _save_step1_result(result_data)

    return {"ok": True, **result_dict}


# ── 사용자 리소스 (URL / 이미지 첨부) ─────────────────────────────

async def _fetch_url_content(url: str) -> dict:
    """URL 페이지 내용을 크롤링하여 텍스트 추출."""
    import httpx
    from bs4 import BeautifulSoup
    from urllib.parse import urlparse

    # URL 스킴 검증 — SSRF 방지
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https"):
        raise ValueError(f"허용되지 않는 URL 스킴: {parsed.scheme}")
    _blocked_hosts = {"localhost", "127.0.0.1", "0.0.0.0", "::1", "169.254.169.254"}
    if parsed.hostname and (parsed.hostname in _blocked_hosts or parsed.hostname.startswith("10.") or parsed.hostname.startswith("192.168.")):
        raise ValueError("내부 네트워크 접근이 차단되었습니다.")

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8",
    }
    async with httpx.AsyncClient(timeout=15, follow_redirects=True, verify=True) as client:
        resp = await client.get(url, headers=headers)
        resp.raise_for_status()
        content_type = resp.headers.get("content-type", "")
        if "html" not in content_type and "text" not in content_type:
            raise ValueError(f"지원하지 않는 콘텐츠 타입: {content_type}")

    soup = BeautifulSoup(resp.text, "lxml")
    title = soup.title.get_text(strip=True) if soup.title else url

    # 불필요한 태그 제거
    for tag in soup(["script", "style", "nav", "footer", "header", "aside",
                     "noscript", "iframe", "form"]):
        tag.decompose()

    # 본문 텍스트 추출 (article > main > body 우선순위)
    body = (soup.find("article") or soup.find("main") or
            soup.find(id="content") or soup.find(class_="content") or
            soup.body)
    text = body.get_text(separator="\n", strip=True) if body else soup.get_text(separator="\n", strip=True)
    # 연속 줄바꿈 정리
    import re as _re
    text = _re.sub(r"\n{3,}", "\n\n", text)[:6000]
    return {"title": title, "content": text}


async def _analyze_image_with_vision(image_b64: str, image_media_type: str) -> str:
    """Claude Vision으로 이미지 내용을 분석하여 텍스트 추출."""
    settings = load_settings()
    anthropic_key = os.getenv("ANTHROPIC_API_KEY", "") or settings.anthropic_api_key
    if not anthropic_key:
        raise ValueError("Anthropic API 키가 없습니다.")

    from anthropic import AsyncAnthropic
    client = AsyncAnthropic(api_key=anthropic_key)
    response = await client.messages.create(
        model=settings.anthropic_model or "claude-sonnet-4-6",
        max_tokens=2048,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": image_media_type,
                        "data": image_b64,
                    },
                },
                {
                    "type": "text",
                    "text": (
                        "이 이미지의 내용을 한국어로 상세히 분석해주세요. "
                        "문서라면 핵심 내용·수치·키워드를 정리하고, "
                        "차트나 표라면 데이터와 인사이트를 서술하며, "
                        "화면 캡처라면 어떤 서비스/정보인지 설명해주세요. "
                        "마지막에 3~5줄 요약을 추가해주세요."
                    ),
                },
            ],
        }],
    )
    return response.content[0].text


@app.get("/api/workflow/resources", tags=["Workflow"])
async def get_workflow_resources():
    """누적 사용자 리소스 목록 반환 (이미지 b64 제외)."""
    safe = [{k: v for k, v in r.items() if k != "image_b64"} for r in _wf_user_resources]
    return {"resources": safe, "total": len(safe)}


@app.post("/api/workflow/resources/url", tags=["Workflow"])
async def add_url_resource(request: Request):
    """URL 크롤링 후 리소스에 추가."""
    global _wf_user_resources
    body = await request.json()
    url = (body.get("url") or "").strip()
    if not url:
        raise HTTPException(400, "url 필드가 필요합니다.")

    try:
        fetched = await _fetch_url_content(url)
    except Exception as e:
        raise HTTPException(400, f"URL 접근 실패: {e}")

    resource = {
        "type": "url",
        "source": url,
        "title": fetched["title"],
        "content": fetched["content"],
        "added_at": _now_kst(),
    }
    _wf_user_resources.append(resource)
    _save_session_data(_current_session_id)
    return {"ok": True, "resource": resource, "total": len(_wf_user_resources)}


@app.post("/api/workflow/resources/image", tags=["Workflow"])
async def add_image_resource(request: Request):
    """이미지 Vision 분석 후 리소스에 추가. 이미지 파일은 세션 디렉토리에 저장."""
    global _wf_user_resources
    body = await request.json()
    image_b64: str = body.get("image_b64", "")
    image_type: str = body.get("image_type", "image/png")  # e.g. "image/png", "image/jpeg"
    filename: str = body.get("filename", "screenshot")
    if not image_b64:
        raise HTTPException(400, "image_b64 필드가 필요합니다.")

    # Vision 분석
    try:
        analysis = await _analyze_image_with_vision(image_b64, image_type)
    except Exception as e:
        raise HTTPException(500, f"이미지 분석 실패: {e}")

    # 세션 디렉토리에 이미지 파일 저장
    img_path = ""
    if _current_session_id:
        import base64 as _b64
        res_dir = _get_session_dir(_current_session_id) / "resources"
        res_dir.mkdir(exist_ok=True)
        ts = _now_kst().replace(":", "-").replace(" ", "_")
        ext = image_type.split("/")[-1] if "/" in image_type else "png"
        img_file = res_dir / f"img_{ts}.{ext}"
        img_file.write_bytes(_b64.b64decode(image_b64))
        img_path = str(img_file)

    resource = {
        "type": "image",
        "source": filename,
        "title": filename,
        "content": analysis,
        "image_path": img_path,
        "image_b64": image_b64,   # 메모리에만 보관 (session_data 저장 시 제외)
        "added_at": _now_kst(),
    }
    _wf_user_resources.append(resource)
    _save_session_data(_current_session_id)
    # 반환 시 b64 포함 (프론트 썸네일 표시용)
    return {"ok": True, "resource": resource, "total": len(_wf_user_resources)}


@app.delete("/api/workflow/resources/{idx}", tags=["Workflow"])
async def delete_workflow_resource(idx: int):
    """인덱스로 리소스 삭제."""
    global _wf_user_resources
    if idx < 0 or idx >= len(_wf_user_resources):
        raise HTTPException(404, "리소스를 찾을 수 없습니다.")
    _wf_user_resources.pop(idx)
    _save_session_data(_current_session_id)
    return {"ok": True, "total": len(_wf_user_resources)}


# ── To-Be Workflow Swim Lane 생성 ─────────────────────────────

# swim lane 표준 액터 (상→하 순서)
_TOBE_ACTOR_ORDER = [
    "임원", "현업 팀장", "HR 임원", "HR 담당자",
    "Senior AI", "Junior AI", "현업 구성원", "그 외",
]

_ACTOR_ALIAS = {
    "HR담당자": "HR 담당자", "HR 담당자": "HR 담당자",
    "HR임원": "HR 임원", "HR 임원": "HR 임원",
    "HR": "HR 담당자",   # 기본은 담당자로 매핑 (임원/담당자 명시 없을 때)
    "팀장": "현업 팀장", "현업 팀장": "현업 팀장", "현업팀장": "현업 팀장",
    "임원": "임원", "현업 임원": "임원", "임원 (=현업 임원)": "임원",
    "직원": "현업 구성원", "사원": "현업 구성원",
    "구성원": "현업 구성원", "현업 구성원": "현업 구성원",
    "현업구성원": "현업 구성원",
}


_ACTORS_DICT_MAP = {
    "exec": "임원",
    "teamlead": "현업 팀장",
    "hr": "HR 담당자",
    "member": "현업 구성원",
    "other": "그 외",
}


def _actors_dict_to_role(actors: dict) -> str:
    """actors dict({exec:'●', hr:'●', ...}) → 콤마 구분 role 문자열로 변환."""
    parts: list[str] = []
    for key, val in actors.items():
        if not val:
            continue
        mapped = _ACTORS_DICT_MAP.get(key)
        if mapped and mapped not in parts:
            parts.append(mapped)
    return ", ".join(parts)


def _parse_role_string(raw) -> tuple[list[str], str]:
    """
    As-Is role 문자열을 분해 — 콤마/슬래시 구분 다중 액터 처리.

    예: "임원 (=현업 임원), HR, 현업 팀장, 현업 구성원, 그 외:지주HR"
        → (["임원","HR 담당자","현업 팀장","현업 구성원","그 외"], "지주HR")

    Returns:
        (정규화된 표준 액터 리스트, "그 외:" 뒤의 customRole 또는 빈 문자열)
    """
    if not raw:
        return [], ""
    # dict인 경우 (actors 객체) → role 문자열로 먼저 변환
    if isinstance(raw, dict):
        raw = _actors_dict_to_role(raw)
        if not raw:
            return [], ""
    if isinstance(raw, list):
        raw = ", ".join(str(x) for x in raw if x)
    s = str(raw).strip()
    if not s:
        return [], ""

    actors: list[str] = []
    custom = ""
    parts = [p.strip() for p in s.replace("·", ",").replace("/", ",").split(",") if p.strip()]
    for part in parts:
        if part.startswith("그 외:") or part.startswith("기타:"):
            custom = part.split(":", 1)[1].strip()
            actors.append("그 외")
            continue
        if part in ("그 외", "기타"):
            actors.append("그 외")
            continue
        # alias / 표준명 매칭
        if part in _TOBE_ACTOR_ORDER:
            actors.append(part)
        elif part in _ACTOR_ALIAS:
            actors.append(_ACTOR_ALIAS[part])
        else:
            # "HR (자회사)", "팀장 (현업)" 같은 부가 텍스트 떼고 재시도
            base = part.split("(", 1)[0].strip()
            if base in _TOBE_ACTOR_ORDER:
                actors.append(base)
            elif base in _ACTOR_ALIAS:
                actors.append(_ACTOR_ALIAS[base])
            else:
                # 끝까지 매칭 안 되면 "그 외"로 + 원본 텍스트를 customRole에 보존
                actors.append("그 외")
                if not custom:
                    custom = part

    # 중복 제거 (순서 보존)
    seen: set[str] = set()
    unique = []
    for a in actors:
        if a not in seen:
            seen.add(a)
            unique.append(a)
    return unique, custom


def _normalize_actor(raw) -> str:
    """단일 액터 정규화 (다중이면 첫 번째 표준 액터). 후방 호환."""
    actors, _ = _parse_role_string(raw)
    return actors[0] if actors else "그 외"


async def _llm_route_tobe_flow(
    process_name: str,
    sheet_name: str,
    asis_nodes: list[dict],
    asis_edges: list[dict],
    senior_agents: list[dict],
    junior_agents: list[dict],
) -> dict:
    """
    As-Is process map + AI Service Flow를 LLM에게 입력 → 지능형 배선 설계.

    Returns dict:
      {
        "ai_to_asis_edges": [{"from": junior_task_id, "to": asis_node_id, "label": str}],
        "asis_to_ai_edges": [{"from": asis_node_id, "to": junior_task_id, "label": str}],
        "senior_triggers":  [{"from": senior_id, "to": junior_first_id, "label": str}],
        "human_actions":    [{"junior_task_id": str, "action": str, "performer": str}]
      }

    LLM 호출 실패 시 빈 dict 반환 (deterministic 기본 엣지만 유지).
    """
    if not junior_agents:
        return {}

    # 입력 요약 구성
    asis_summary_lines = []
    for n in asis_nodes:
        if n["level"] not in ("L4", "L5"):
            continue
        role_hint = n.get("actor") or "미지정"
        desc = (n.get("description") or "")[:80]
        line = f"- [{n['id']}] {n['label']} ({n['level']}, 수행: {role_hint})"
        if desc:
            line += f" — {desc}"
        asis_summary_lines.append(line)
    asis_block = "\n".join(asis_summary_lines) if asis_summary_lines else "(As-Is 노드 없음)"

    senior_lines = []
    for ag in senior_agents:
        senior_id = f"senior_{ag['agent_id'] or 0}"
        senior_lines.append(f"- [{senior_id}] {ag['agent_name']} — {ag.get('description', '')[:100]}")
    senior_block = "\n".join(senior_lines) if senior_lines else "(Senior AI 없음)"

    junior_lines = []
    for ji, ag in enumerate(junior_agents):
        prefix = f"junior_{ag['agent_id'] or ji}"
        junior_lines.append(f"### Agent: {ag['agent_name']} (ID: {prefix})")
        for t in ag["tasks"]:
            tid = t.get("task_id", "")
            jid = f"{prefix}_{tid}"
            alvl = t.get("automation_level", "")
            ai_role = (t.get("ai_role") or "")[:80]
            human_role = (t.get("human_role") or "")[:80]
            inputs = ", ".join((t.get("input_data") or [])[:3])
            outputs = ", ".join((t.get("output_data") or [])[:3])
            junior_lines.append(
                f"  - [{jid}] {t.get('task_name', '')} ({alvl})\n"
                f"    AI 역할: {ai_role}\n"
                f"    Human 역할: {human_role or '없음'}\n"
                f"    Input: {inputs or '-'} / Output: {outputs or '-'}"
            )
    junior_block = "\n".join(junior_lines)

    system_prompt = f"""당신은 AI/Human 협업 워크플로우를 설계하는 전문가입니다.
아래 As-Is 프로세스와 AI Service Flow를 분석하여, **Swim Lane 다이어그램의 엣지(연결선)와 Human 검토 액션을 지능형으로 배선**하세요.

## 프로세스: {process_name} / 시트: {sheet_name}

## As-Is 프로세스 맵 (L4/L5 노드)
{asis_block}

## Senior AI 오케스트레이터
{senior_block}

## Junior AI 에이전트와 태스크
{junior_block}

## 설계 원칙 — **Pipe-through 패턴이 기본**

Junior AI 태스크가 As-Is L5 를 자동화/대체하는 경우가 대부분이므로, 엣지는 다음 패턴으로 흘러야 합니다:
```
[As-Is 선행 L5] → [Junior AI] → [As-Is 후행 L5]
```
즉 Junior AI 가 As-Is 파이프라인의 한 지점을 "통과"하는 형태입니다.

### 규칙 (엄수)
1. **As-Is → Junior AI 트리거 엣지**: Junior AI 태스크의 **직전 As-Is L5** (선행 노드) 에서 연결.
   **같은 task_id 를 공유하는 As-Is 노드 자체는 선행자로 삼지 말 것** (그건 Junior AI 가 대체하는 노드라 loop 됨)
2. **Junior AI → As-Is 결과 엣지**: Junior AI 태스크의 **직후 As-Is L5** (후행 노드) 로 연결.
   **같은 task_id 를 공유하는 As-Is 노드로 돌아가지 말 것** (loop 금지)
3. **Senior AI 오케스트레이션** (Senior AI 가 존재할 때만): Senior AI 가 각 Junior AI 에이전트의 첫 태스크를 '기동' 하는 엣지 (에이전트 그룹당 1개).
   위 "Senior AI 오케스트레이터" 섹션이 `(Senior AI 없음)` 이면 **senior_triggers 는 빈 배열 `[]`** 로 반환 (절대 새로 만들어내지 말 것)
4. **Human 검토 액션**: automation_level 이 'Human-in-Loop' / 'Human-on-the-Loop' 이거나 human_role 이 비어있지 않은 Junior 태스크에 대해,
   **어떤 사람이(performer)**, **무슨 행동을 구체적으로 해야 하는지(action)** 명시
   performer ∈ ["임원", "현업 팀장", "HR 임원", "HR 담당자", "현업 구성원"]

### ❌ 금지 사항 (엄격)
- **task_id 가 같은 As-Is ↔ Junior AI 간 직접 엣지 금지** (양방향 모두). 예시:
  - `junior_agent_1_1.1.4.1` 은 `l5-1.1.4.1-xx` 를 대체하므로 두 노드 간 직접 연결 금지
  - 대신 `l5-1.1.4.0-xx (이전 단계) → junior_agent_1_1.1.4.1 → l5-1.1.4.2-xx (다음 단계)` 로 파이프 통과
- 과잉 엣지 금지 — 각 Junior 태스크당 트리거+결과 각 **최대 1~2개**
- As-Is에 적절한 선/후행 매칭이 없으면 생략 (억지로 연결 금지)

## 제약
- As-Is 노드 ID는 반드시 위 목록에서 그대로 사용 (변경·새 ID 생성 금지)
- Junior 태스크 ID도 위에 적힌 jid 그대로 사용

## 출력 형식 (JSON만, 마크다운 코드 블록 없음)
{{
  "ai_to_asis_edges": [
    {{"from": "junior_xxx_yyy", "to": "asis_l5_id", "label": "결과 전달"}}
  ],
  "asis_to_ai_edges": [
    {{"from": "asis_l5_id", "to": "junior_xxx_yyy", "label": "트리거"}}
  ],
  "senior_triggers": [
    {{"from": "senior_x", "to": "junior_xxx_first_task", "label": "기동"}}
  ],
  "human_actions": [
    {{"junior_task_id": "junior_xxx_yyy", "action": "AI 결과 검토 및 예외 케이스 판단·승인", "performer": "HR 담당자"}}
  ]
}}"""

    try:
        result = await _call_llm_step1(system_prompt, [
            {"role": "user", "content": "AI와 As-Is를 매끄럽게 잇는 엣지와 Human 액션을 설계해주세요. JSON만 출력하세요."}
        ])
        if not isinstance(result, dict):
            return {}
        return {
            "ai_to_asis_edges": result.get("ai_to_asis_edges") or [],
            "asis_to_ai_edges": result.get("asis_to_ai_edges") or [],
            "senior_triggers": result.get("senior_triggers") or [],
            "human_actions": result.get("human_actions") or [],
        }
    except Exception as e:
        print(f"[TOBE-LLM] 배선 LLM 호출 실패: {e}", flush=True)
        return {}


async def _build_tobe_sheet_from_asis(asis_sheet, process_name: str) -> dict:
    """
    단일 As-Is 시트 → To-Be Swim Lane dict.
    - As-Is 노드/엣지/role 그대로 보존 (level·position 포함)
    - Step 2 Senior/Junior AI agent를 새 레인에 주입
    - 엣지는 As-Is 그대로 + AI 노드 지원 엣지 (점선) 추가
    - Step 2 스코프(L4)에 해당하는 As-Is 노드만 포함 (스코프 꼬임 방지)
    - LLM에게 AI Service Flow + As-Is를 입력 → 지능형 배선·Human action 설계 (핵심)
    """
    sheet_name = (asis_sheet.name or asis_sheet.sheet_id).strip()
    sheet_id = asis_sheet.sheet_id

    # 0) Step 2 에이전트 먼저 읽기 → 타겟 L4 스코프 도출
    def _task_matches_sheet(t: dict) -> bool:
        tl4 = (t.get("l4", "") or "").strip()
        if not tl4 or not sheet_name:
            return True
        return tl4 in sheet_name or sheet_name in tl4

    # Step 1 에서 '삭제'/'통합' 처리된 task_id 판정기 (공용 helper)
    _is_deprecated_tid = _make_deprecation_check()

    senior_agents: list[dict] = []
    junior_agents: list[dict] = []
    for agent in _wf_step2_cache.get("agents", []):
        atype = agent.get("agent_type", "")
        matched = [
            t for t in agent.get("assigned_tasks", [])
            if _task_matches_sheet(t) and not _is_deprecated_tid(str(t.get("task_id") or ""))
        ]
        # Senior AI는 오케스트레이터라 assigned_tasks가 비어있어도 항상 포함
        # (Junior AI는 태스크 없으면 제외)
        if atype == "Senior AI":
            senior_agents.append({
                "agent_id": agent.get("agent_id", ""),
                "agent_name": agent.get("agent_name", "") or "Senior AI Orchestrator",
                "ai_technique": agent.get("ai_technique", ""),
                "description": agent.get("description", ""),
                "tasks": matched,  # 비어있을 수 있음
            })
            continue
        if atype != "Junior AI":
            continue
        if not matched and not junior_agents:
            # 시트 매칭 실패 시 fallback — 단, 폐기/통합 task 는 제외
            matched = [
                t for t in agent.get("assigned_tasks", [])
                if not _is_deprecated_tid(str(t.get("task_id") or ""))
            ]
        if not matched:
            continue
        junior_agents.append({
            "agent_id": agent.get("agent_id", ""),
            "agent_name": agent.get("agent_name", ""),
            "ai_technique": agent.get("ai_technique", ""),
            "tasks": matched,
        })

    # 타겟 L4 스코프 추출 (task_id 접두사 + l4 이름)
    target_l4_ids: set[str] = set()
    target_l4_names: set[str] = set()
    for ag in senior_agents + junior_agents:
        for t in ag.get("tasks", []):
            tid = str(t.get("task_id") or "")
            if tid:
                parts = tid.split(".")
                if len(parts) >= 3:
                    target_l4_ids.add(".".join(parts[:3]))
            l4_text = str(t.get("l4") or "").strip()
            if l4_text:
                target_l4_names.add(l4_text)

    def _node_in_scope(node) -> bool:
        """노드가 타겟 L4 스코프 안에 있는지 판단."""
        if not target_l4_ids and not target_l4_names:
            return True  # 스코프 정보 없으면 필터 안 함
        # 1) metadata.l4Id 매칭
        n_l4id = str(node.metadata.get("l4Id") or "").strip()
        if n_l4id and n_l4id in target_l4_ids:
            return True
        # 2) task_id 접두사 매칭 (L5 노드의 task_id가 L5 ID이면 [:3]이 L4 ID)
        if node.task_id:
            tparts = node.task_id.split(".")
            if len(tparts) >= 3 and ".".join(tparts[:3]) in target_l4_ids:
                return True
            # L4 노드는 task_id가 "1.1.4" 형태
            if len(tparts) == 3 and node.task_id in target_l4_ids:
                return True
        # 3) l4Name 매칭 (문자열 포함)
        n_l4n = str(node.metadata.get("l4Name") or "").strip()
        if n_l4n:
            for tn in target_l4_names:
                if tn and (tn in n_l4n or n_l4n in tn):
                    return True
        return False

    # 1) As-Is 노드 변환 (L2/L3/MEMO 제외) — 타겟 L4 스코프만
    asis_nodes_out: list[dict] = []
    role_counts: dict[str, int] = {}
    kept_ids: set[str] = set()    # 스코프 필터 통과한 노드 id
    pending_decisions: list = []   # decision/memo는 연결된 노드 기준으로 2-pass 판단

    # ── L5 ID 순번 할당용: 각 L4 의 max seq 추적 (2.1.4.N 포맷) ──
    # 새로 생성하는 Junior AI / Human 검토 노드에 일관된 L5 ID 부여
    l4_max_seq: dict[str, int] = {}
    for n in asis_sheet.nodes.values():
        if n.level == "L5" and n.task_id:
            parts = n.task_id.split(".")
            if len(parts) >= 4:
                l4_id = ".".join(parts[:3])
                try:
                    seq = int(parts[3])
                    l4_max_seq[l4_id] = max(l4_max_seq.get(l4_id, 0), seq)
                except ValueError:
                    pass

    # 타겟 L4 (스코프의 첫 L4 를 primary 로 — 보통 단일 L4 scope)
    _primary_l4 = next(iter(target_l4_ids), None) if target_l4_ids else None

    def _next_l5_id(l4_id: str | None = None) -> str:
        """다음 L5 ID 할당. 예: '2.1.4' → '2.1.4.16' (기존 최대 seq+1)"""
        lid = l4_id or _primary_l4
        if not lid:
            return ""
        l4_max_seq[lid] = l4_max_seq.get(lid, 0) + 1
        return f"{lid}.{l4_max_seq[lid]}"

    # 스코프 필터 통과 노드 id 먼저 수집
    for n in asis_sheet.nodes.values():
        if n.level in ("L2", "L3", "MEMO"):
            continue
        if n.level == "DECISION":
            pending_decisions.append(n)
            continue
        # Step 1 에서 '삭제'/'통합' 처리된 Task 는 Swim Lane 에서 완전 제외
        if _is_deprecated_tid(n.task_id or ""):
            print(f"[TOBE] '{n.label}' (task_id={n.task_id}) SKIP — "
                  f"Step 1 에서 폐기/통합 처리됨", flush=True)
            continue
        if _node_in_scope(n):
            kept_ids.add(n.id)

    # Decision 노드는 kept_ids와 엣지로 연결된 경우에만 유지
    for n in pending_decisions:
        connected_to_kept = any(
            (e.source in kept_ids and e.target == n.id) or
            (e.target in kept_ids and e.source == n.id)
            for e in asis_sheet.edges
        )
        if connected_to_kept:
            kept_ids.add(n.id)

    # position 범위 — kept만 대상
    xs = [n.position_x for n in asis_sheet.nodes.values() if n.id in kept_ids]
    ys = [n.position_y for n in asis_sheet.nodes.values() if n.id in kept_ids]
    min_x = min(xs) if xs else 0.0
    max_x = max(xs) if xs else 1200.0
    max_y = max(ys) if ys else 600.0

    # ── 분류 기반 As-Is 처리 헬퍼 ─────────────────────────────────
    def _extract_human_part(task_id: str) -> tuple[str, str]:
        """분류 정보에서 Human 파트 문구 추출.
        Returns: (human_label_short, human_description_full)
        """
        cls = _wf_classification.get(task_id, {}) if task_id else {}
        hybrid_note = str(cls.get("hybrid_note") or "")
        ai_part = str(cls.get("ai_part") or "")
        human_part_explicit = str(cls.get("human_part") or "")
        reason = str(cls.get("reason") or "")

        # 1) 명시적 human_part 필드 (hybrid splitter가 채움)
        if human_part_explicit:
            return human_part_explicit[:60], human_part_explicit

        # 2) hybrid_note 에서 "Human 파트: ..." 추출
        if hybrid_note and "Human 파트:" in hybrid_note:
            human_part = hybrid_note.split("Human 파트:", 1)[1]
            # AI 파트 부분 이전까지 끊기 (형식이 "AI 파트: X / Human 파트: Y" 또는 그 반대)
            human_part = human_part.split(" / AI 파트:", 1)[0].strip()
            return human_part[:60], human_part

        # 3) reason 에서 "Human" 관련 문장 추출 (휴리스틱)
        if reason:
            import re as _re
            # "Human이 수행", "사람이 확정", "최종 확정", "승인·검토" 등이 들어간 문장 찾기
            sentences = _re.split(r"[.。·]", reason)
            for s in sentences:
                s = s.strip()
                if any(kw in s for kw in ["Human이", "사람이", "최종 확정", "최종 승인", "예외·", "검토·", "판단 및", "승인"]):
                    return s[:60], s
        return "", ""

    # ── 🔑 노드 y 좌표 → swim lane 추론 (hr-workflow-ai 의 lane band 구조) ──
    # 사용자가 노드를 swim lane 영역에 배치한 행위 자체가 수행주체 정보임.
    # data.role 이 비어있어도 position.y / lane_height 로 lane 인덱스 결정 가능.
    sheet_lanes_raw = list(asis_sheet.lanes or [])
    # 표준 actor 로 매핑된 lane 명 리스트 (Q3 inference 용)
    sheet_lane_actors: list[str] = []
    for ln in sheet_lanes_raw:
        _la, _ = _parse_role_string(ln)
        sheet_lane_actors.append(_la[0] if _la else ln)

    # lane band 추정: hr-workflow-ai 기본 swimHeight=2400, 6 lanes → 400/lane
    # 사용자 데이터의 y 범위에서 자동 계산 (band_height = (y_max - y_min) / num_lanes)
    # 단, hr-workflow-ai 기본값 400 이 더 합리적이면 그것 사용
    _y_to_lane_cache: dict | None = None

    def _y_to_lane_actor(node_y: float) -> tuple[str, str]:
        """노드 y 좌표 기반 swim lane actor 추론.
        우선순위:
          1. sheet.lane_heights (사용자가 드래그로 조정한 누적 높이) → 정확한 lane 매핑
          2. sheet.swim_height / num_lanes (균등 분할)
          3. 데이터 y 범위 자동 추정
        Returns: (actor, raw_lane_name)
        """
        nonlocal _y_to_lane_cache
        if not sheet_lanes_raw or not sheet_lane_actors:
            return "", ""

        if _y_to_lane_cache is None:
            num_lanes = len(sheet_lanes_raw)
            lane_heights = list(getattr(asis_sheet, "lane_heights", []) or [])
            # ① laneHeights 가 있고 길이가 lanes 와 같으면 — 정확한 누적 좌표 사용
            if lane_heights and len(lane_heights) == num_lanes:
                # cumulative_y[i] = i 번째 lane 의 시작 y
                cumulative: list[float] = [0.0]
                for h in lane_heights:
                    cumulative.append(cumulative[-1] + h)
                _y_to_lane_cache = {
                    "valid": True, "mode": "cumulative",
                    "cumulative": cumulative, "num_lanes": num_lanes,
                }
            else:
                # ② swim_height / num_lanes 균등 분할 (또는 ③ 자동 추정)
                swim_h = float(getattr(asis_sheet, "swim_height", 0) or 0)
                if swim_h <= 0:
                    all_ys = [
                        nd.position_y for nd in asis_sheet.nodes.values()
                        if nd.level not in ("L2", "L3", "MEMO")
                    ]
                    if all_ys:
                        y_max = max(all_ys)
                        # hr-workflow-ai default swim_height = 2400 가정,
                        # 데이터 y_max 가 더 크면 그것 사용
                        swim_h = max(2400.0, y_max + 100.0)
                    else:
                        _y_to_lane_cache = {"valid": False}
                if (_y_to_lane_cache or {}).get("valid") is None:
                    band = swim_h / num_lanes
                    _y_to_lane_cache = {
                        "valid": True, "mode": "uniform",
                        "band": band, "y_origin": 0.0, "num_lanes": num_lanes,
                    }

        if not _y_to_lane_cache.get("valid"):
            return "", ""

        if _y_to_lane_cache["mode"] == "cumulative":
            cumu = _y_to_lane_cache["cumulative"]
            # node_y 가 어느 누적 구간에 들어가는지 binary search
            idx = 0
            for i in range(len(cumu) - 1):
                if cumu[i] <= node_y < cumu[i + 1]:
                    idx = i
                    break
            else:
                # 마지막 lane 이후
                idx = _y_to_lane_cache["num_lanes"] - 1
        else:
            idx = int((node_y - _y_to_lane_cache["y_origin"]) / _y_to_lane_cache["band"])

        idx = max(0, min(idx, _y_to_lane_cache["num_lanes"] - 1))
        return sheet_lane_actors[idx], sheet_lanes_raw[idx]

    # ── 노드 lane 상속 헬퍼 — 선행/후행 노드의 actor 를 상속 ──
    def _infer_neighbor_actor(node_id: str) -> tuple[str, list[str], str]:
        """선행 노드(우선) → 후행 노드 actor 를 상속. 못 찾으면 ('','',[]) 반환."""
        # 1) 이 노드로 들어오는 엣지의 source 노드 탐색
        for e in asis_sheet.edges:
            if e.target != node_id:
                continue
            src = asis_sheet.nodes.get(e.source)
            if src is None:
                continue
            src_role = src.metadata.get("role", "") or src.metadata.get("actors", "")
            src_actors, src_custom = _parse_role_string(src_role)
            if src_actors and src_actors[0] != "그 외":
                return src_actors[0], src_actors, src_custom
        # 2) fallback: 이 노드에서 나가는 엣지의 target
        for e in asis_sheet.edges:
            if e.source != node_id:
                continue
            tgt = asis_sheet.nodes.get(e.target)
            if tgt is None:
                continue
            tgt_role = tgt.metadata.get("role", "") or tgt.metadata.get("actors", "")
            tgt_actors, tgt_custom = _parse_role_string(tgt_role)
            if tgt_actors and tgt_actors[0] != "그 외":
                return tgt_actors[0], tgt_actors, tgt_custom
        return "", [], ""

    def _infer_decision_actor(dec_node_id: str) -> tuple[str, list[str], str]:
        """Decision 노드 — 선행 우선, 없으면 HR 담당자 fallback."""
        a, al, cr = _infer_neighbor_actor(dec_node_id)
        if a:
            return a, al, cr
        return "HR 담당자", ["HR 담당자"], ""

    for n in sorted(asis_sheet.nodes.values(), key=lambda nd: (nd.position_y, nd.position_x)):
        if n.id not in kept_ids:
            continue

        # ── 🔑 분류 기반 필터링 (AI/AI+Human/Human) ──────────
        cls_label = ""
        if n.level == "L5" and n.task_id:
            cls_entry = _wf_classification.get(n.task_id, {}) if _wf_classification else {}
            cls_label = str(cls_entry.get("label") or "")

        # AI 로 분류된 L5 는 As-Is 에서 완전 제거 (Junior AI 가 대체)
        if cls_label == "AI":
            continue

        # ── Decision 은 role 이 없음 → 선행·후행 actor 상속 ──
        if n.level == "DECISION":
            primary_actor, actors_list, custom_role = _infer_decision_actor(n.id)
        else:
            actor_raw = n.metadata.get("role", "") or n.metadata.get("actors", "")
            actors_list, custom_role = _parse_role_string(actor_raw)
            primary_actor = actors_list[0] if actors_list else ""

            # 🔑 actor 정보가 정말 없는 경우 (raw role 도 없고 매핑도 안 됨)
            #    → 1) 연결된 선/후행 노드에서 inheritance 시도
            #    → 2) 그래도 없으면 분류(AI/Human) 기준 default
            #    → 3) 그것도 없으면 노드 자체 SKIP (수행주체 정보가 정말로 없으므로
            #         Swim Lane 에 표시할 의미가 없음 — 사용자 요구)
            raw_role_text = ""
            if isinstance(actor_raw, str):
                raw_role_text = actor_raw.strip()
            elif isinstance(actor_raw, dict):
                # actors dict 가 모두 빈 값이면 raw_role_text 는 빈 상태
                raw_role_text = _actors_dict_to_role(actor_raw)
            elif isinstance(actor_raw, list) and actor_raw:
                raw_role_text = ", ".join(str(x) for x in actor_raw if x).strip()

            if not primary_actor:
                # role 이 명시적으로 "그 외"·"기타" 인 경우는 actors_list 에 들어옴
                if raw_role_text and ("그 외" in raw_role_text or "기타" in raw_role_text):
                    primary_actor = "그 외"
                else:
                    # 🔑 1) sheet lanes + 노드 y 좌표 → swim lane 추론 (사용자가 영역에 배치한 정보)
                    lane_actor, raw_lane_name = _y_to_lane_actor(n.position_y)
                    if lane_actor and lane_actor in _TOBE_ACTOR_ORDER:
                        primary_actor = lane_actor
                        actors_list = [lane_actor]
                    elif lane_actor and "그 외" in str(raw_lane_name):
                        primary_actor = "그 외"
                        if not custom_role:
                            custom_role = raw_lane_name.replace("그 외", "").strip(":, ")
                    else:
                        # 2) 선/후행 inheritance
                        inh_actor, inh_list, inh_custom = _infer_neighbor_actor(n.id)
                        if inh_actor:
                            primary_actor = inh_actor
                            actors_list = inh_list
                            custom_role = inh_custom
                        elif cls_label in ("AI", "AI + Human", "Human"):
                            # 3) 분류 라벨 있으면 HR 담당자 default
                            primary_actor = "HR 담당자"
                            actors_list = ["HR 담당자"]
                        else:
                            # 4) 진짜로 정보 없음 — SKIP
                            print(f"[TOBE] '{n.label}' 노드 SKIP — 수행주체 정보 없음 "
                                  f"(role/actors/y-lane/inheritance/classification 모두 없음)", flush=True)
                            continue

            # primary_actor 가 결정된 후 — '그 외' 로 끝났는데 custom_role 비어있으면
            # 원본 raw text 또는 '미지정' 표시 (이미 명시적 그 외인 경우)
            if primary_actor == "그 외" and not custom_role:
                if raw_role_text:
                    custom_role = raw_role_text[:30]
        # 모든 actor를 카운트 (lane 등장 여부 판단용)
        for a in actors_list or [primary_actor]:
            role_counts[a] = role_counts.get(a, 0) + 1

        node_type = "decision" if n.level == "DECISION" else "task"

        # 원본 data 객체를 그대로 보존 (LevelNode가 systems/painPoints/inputs/outputs/logic 등을 사용)
        full_data = dict(n.metadata)
        existing_role = full_data.get("role")
        if not isinstance(existing_role, str):
            # role 이 dict/None 인 경우 — actors_list 기반 재구성
            parts = list(actors_list) if actors_list else [primary_actor]
            if custom_role and "그 외" in parts:
                parts = [f"그 외:{custom_role}" if p == "그 외" else p for p in parts]
            elif custom_role and primary_actor == "그 외" and "그 외" not in parts:
                parts.append(f"그 외:{custom_role}")
            full_data["role"] = ", ".join(parts)
        else:
            # role 이 이미 문자열 — '그 외' default + custom_role 새로 부여한 경우 추가
            # ("그 외:..." 가 원본에 없으면 끝에 부착해서 LevelNode 가 sky-blue 뱃지 표시하게 함)
            if (primary_actor == "그 외" and custom_role
                    and "그 외:" not in existing_role
                    and "기타:" not in existing_role):
                if existing_role.strip():
                    full_data["role"] = f"{existing_role}, 그 외:{custom_role}"
                else:
                    full_data["role"] = f"그 외:{custom_role}"

        # ── AI + Human: label 은 원본 task 이름 유지, 상세 Human action 은 description ───
        # (기존엔 Human 파트 60자 문장을 label 로 사용 → 너무 길어서 가독성 저하)
        # → 원본 짧은 이름 유지 ("서류 합격자 선정") + description 에 전체 Human action
        original_label = n.label or ""
        display_label = original_label
        display_description = n.description or ""
        if cls_label == "AI + Human":
            hp_label, hp_desc = _extract_human_part(n.task_id)
            # label 은 깔끔하게 — 원본 label 을 기본으로 사용
            display_label = original_label
            full_data["label"] = original_label
            # description 에 Human 파트 상세 action 문구 저장 (툴팁/hover 에서 확인)
            display_description = hp_desc if hp_desc else n.description or ""
            full_data["description"] = display_description
            # 분류 메타 정보
            full_data["_classification"] = "AI + Human"
            full_data["_original_label"] = original_label
            full_data["_human_action"] = hp_desc or ""   # 필요 시 프론트에서 별도 표시
        elif cls_label == "Human":
            full_data["_classification"] = "Human"

        asis_nodes_out.append({
            "id": n.id,
            "label": display_label[:60],
            "level": n.level,
            "actor": primary_actor,
            "actors_all": actors_list,
            "custom_role": custom_role,
            "type": node_type,
            "ai_support": None,
            "position": {"x": float(n.position_x), "y": float(n.position_y)},
            "origin": "asis",
            "task_id": n.task_id,
            "description": display_description,
            "classification": cls_label,
            "data": full_data,
            "next": [],
        })

    # As-Is 엣지 → next 배열
    id_set = {n["id"] for n in asis_nodes_out}
    next_map: dict[str, list[str]] = {n["id"]: [] for n in asis_nodes_out}
    asis_edges_out: list[dict] = []
    for e in asis_sheet.edges:
        if e.source in id_set and e.target in id_set:
            next_map[e.source].append(e.target)
            asis_edges_out.append({
                "id": e.id,
                "source": e.source,
                "target": e.target,
                "label": e.label or "",
                "origin": "asis",
            })
    for n in asis_nodes_out:
        n["next"] = next_map.get(n["id"], [])

    # 2) Step 2 AI agent 주입 (senior_agents/junior_agents는 위에서 이미 추출됨)
    ai_nodes_out: list[dict] = []
    ai_edges_out: list[dict] = []

    # Senior AI 레인 y: As-Is 최대 y 아래 + 여백
    # Senior AI 는 필요한 경우에만 생성 (사용자 정의 3가지 기준: 이질적 작업자 / 비선형 / 상태 관리)
    senior_y = max_y + 220.0
    junior_y = senior_y + 300.0 if senior_agents else max_y + 220.0   # Senior 없으면 바로 Junior 부터

    senior_node_ids: list[str] = []   # 생성된 Senior AI 노드 id 추적 (오케스트레이션 엣지용)
    if senior_agents:
        for si, ag in enumerate(senior_agents):
            senior_id = f"senior_{ag['agent_id'] or si}"
            senior_node_ids.append(senior_id)
            ai_nodes_out.append({
                "id": senior_id,
                "label": (ag["agent_name"] or "Senior AI")[:40],
                "level": "L5",
                "actor": "Senior AI",
                "type": "task",
                "ai_support": ag.get("ai_technique", ""),
                "position": {"x": min_x, "y": senior_y + si * 200.0},
                "origin": "ai",
                "next": [],
            })

    # task_id → As-Is L5 노드 맵 (x 정렬·중복 엣지 필터링에 사용)
    asis_by_task_id: dict[str, dict] = {n["task_id"]: n for n in asis_nodes_out if n.get("task_id")}

    if junior_agents:
        total_jtasks = sum(len(ag["tasks"]) for ag in junior_agents)
        # 최소 x_step: L5 노드 폭 380 + gap 60 = 440 (겹침 방지)
        MIN_X_STEP = 440.0
        x_step = max(MIN_X_STEP, (max_x - min_x) / max(total_jtasks, 1)) if total_jtasks else MIN_X_STEP
        task_i = 0
        for ji, ag in enumerate(junior_agents):
            agent_prefix = f"junior_{ag['agent_id'] or ji}"
            prev_id: str | None = None
            for ti, t in enumerate(ag["tasks"]):
                tid = t.get("task_id", f"t{ti}")
                jid = f"{agent_prefix}_{tid}"
                # 🔑 Junior AI x 는 항상 pipeline 순서대로 sequential (좌→우 흐름 보장)
                # 이전에 매칭된 As-Is x 를 상속했더니 원본 좌표가 뒤죽박죽이라 흐름이 꼬임
                # → 순수 task_i 기반 sequential 배치
                x = min_x + x_step * (task_i + 0.5)
                task_i += 1

                matched_asis = asis_by_task_id.get(tid)

                # L5 표시용 ID — 기존 As-Is task_id 와 매칭되면 재사용,
                # 새 task (NEW_xxx 등) 이면 L4 의 다음 순번 할당 (2.1.4.16 형식)
                display_l5_id = tid if matched_asis else (
                    tid if (tid and not tid.startswith("NEW") and tid[0].isdigit())
                    else _next_l5_id()
                )

                ai_nodes_out.append({
                    "id": jid,
                    "label": (t.get("task_name", "") or tid)[:40],
                    "level": "L5",
                    "actor": "Junior AI",
                    "type": "task",
                    "ai_support": t.get("ai_role", ""),
                    # 교대 y 제거: 같은 lane 내 겹침은 프론트가 stack으로 처리
                    "position": {"x": x, "y": junior_y},
                    "origin": "ai",
                    "display_id": display_l5_id,   # LevelNode 가 data.id 로 보여줌
                    "automation_level": t.get("automation_level", ""),
                    "human_role": t.get("human_role", ""),
                    "input_data": t.get("input_data", []),
                    "output_data": t.get("output_data", []),
                    "agent_name": ag["agent_name"],
                    "next": [],
                })
                # Junior AI 내부 sequential 엣지
                if prev_id:
                    ai_edges_out.append({
                        "id": f"aie_seq_{prev_id}_{jid}",
                        "source": prev_id,
                        "target": jid,
                        "label": "",
                        "origin": "ai",
                    })
                prev_id = jid

            # Senior AI → 첫 Junior task (오케스트레이션) — Senior AI 가 존재할 때만
            if senior_node_ids and ag["tasks"]:
                first_jid = f"{agent_prefix}_{ag['tasks'][0].get('task_id', 't0')}"
                senior_first = senior_node_ids[0]   # 명시적으로 첫 Senior id 참조
                ai_edges_out.append({
                    "id": f"aie_orc_{senior_first}_{first_jid}",
                    "source": senior_first,
                    "target": first_jid,
                    "label": "기동",
                    "origin": "ai",
                })

    # 3) LLM 지능형 배선 — AI Service Flow + As-Is Process Map 분석
    llm_routing = await _llm_route_tobe_flow(
        process_name=process_name,
        sheet_name=sheet_name,
        asis_nodes=asis_nodes_out,
        asis_edges=asis_edges_out,
        senior_agents=senior_agents,
        junior_agents=junior_agents,
    )

    # LLM 결과 id 검증용 세트
    asis_id_set = {n["id"] for n in asis_nodes_out}
    ai_id_set = {n["id"] for n in ai_nodes_out}

    # LLM 반환 human_actions (junior_task_id → {action, performer})
    llm_actions: dict[str, dict] = {}
    for ha in llm_routing.get("human_actions", []):
        jid = str(ha.get("junior_task_id") or "").strip()
        if jid:
            llm_actions[jid] = {
                "action": str(ha.get("action") or "").strip(),
                "performer": str(ha.get("performer") or "HR 담당자").strip() or "HR 담당자",
            }

    # ── 🔑 As-Is L5 x 정렬: Junior AI 가 대체하는 As-Is 는 같은 x 로 끌어올림 ──
    # (좌→우 pipeline 흐름 유지 + 대체 관계 세로 정렬 시각화)
    junior_by_taskid: dict[str, dict] = {}
    for jn in ai_nodes_out:
        if jn["actor"] != "Junior AI":
            continue
        # jn["id"] 에서 task_id 추출 (junior_agentid_taskid 형식)
        # 또는 직접 매칭된 As-Is task_id 사용
        for tid_key, asis_node in asis_by_task_id.items():
            if jn["id"].endswith(f"_{tid_key}") or jn["id"].endswith(f"_{tid_key}_AI"):
                junior_by_taskid[tid_key] = jn
                break
    # 매칭된 As-Is 의 x 를 Junior AI 의 x 로 업데이트
    for asis_n in asis_nodes_out:
        tid = asis_n.get("task_id")
        if tid and tid in junior_by_taskid:
            jx = junior_by_taskid[tid]["position"]["x"]
            asis_n["position"]["x"] = jx

    # 매칭 안 된 As-Is (Human 분류 등) 는 원본 x 유지 — 하지만 Junior x 범위 내로 들어와야
    # pipeline 안에서 자연스럽게 섞임. 너무 멀리 떨어진 노드는 normalize.
    if junior_nodes_count := len([n for n in ai_nodes_out if n["actor"] == "Junior AI"]):
        junior_x_list = [n["position"]["x"] for n in ai_nodes_out if n["actor"] == "Junior AI"]
        jr_min_x, jr_max_x = min(junior_x_list), max(junior_x_list)
        unmatched_asis = [n for n in asis_nodes_out
                          if n.get("task_id") not in junior_by_taskid]
        if unmatched_asis:
            # 원본 x 순서 유지하되, junior 범위 뒤에 이어붙임
            unmatched_asis.sort(key=lambda n: n["position"]["x"])
            extra_step = 440.0
            for i, n in enumerate(unmatched_asis):
                n["position"]["x"] = jr_max_x + extra_step * (i + 1)

    # 3-1) Human 행동 노드 생성 — Junior AI 중 사람 개입 필요한 것
    human_nodes_out: list[dict] = []
    junior_nodes = [n for n in ai_nodes_out if n["actor"] == "Junior AI"]
    hr_lane_y = junior_y + 260.0

    for jnode in junior_nodes:
        jid = jnode["id"]
        alvl = str(jnode.get("automation_level", ""))
        human_role = str(jnode.get("human_role", "") or "").strip()
        llm_info = llm_actions.get(jid, {})
        llm_action = llm_info.get("action", "")
        llm_performer = llm_info.get("performer", "HR 담당자")

        needs_human = (
            bool(llm_action)
            or bool(human_role)
            or "Loop" in alvl
            or "Supervised" in alvl
        )
        if not needs_human:
            continue

        # performer는 LLM이 판단한 실제 담당자 (HR 담당자 하드코딩 X)
        if llm_performer in _TOBE_ACTOR_ORDER:
            actor_lane = llm_performer
        else:
            actor_lane = "HR 담당자"

        # 짧은 타이틀 = "{AI task 이름 (AI 파트/(AI 파트) 접미사 제거)} 검토"
        # → 한눈에 무슨 검토인지 드러나도록 (단, 이미 검토/확정/승인 등 종료어가 있으면 중복 안 함)
        raw_task_label = str(jnode.get("label") or "")
        clean_task_name = (
            raw_task_label
            .replace(" (AI 파트)", "")
            .replace("(AI 파트)", "")
            .replace(" - AI 파트", "")
            .strip()
        )
        # 종료어 중복 방지: clean_task_name 이 이미 검토/확정/승인/검증 등으로 끝나면
        # "검토" 접미를 다시 붙이지 않음
        _SUFFIX_KEYWORDS = ("검토", "확정", "승인", "검증", "확인", "최종 결정", "판단")
        if any(clean_task_name.endswith(k) for k in _SUFFIX_KEYWORDS):
            review_title = clean_task_name
        elif clean_task_name:
            review_title = f"{clean_task_name} 검토"
        else:
            review_title = "AI 결과 검토"

        # 상세 설명: LLM이 제시한 구체 action > human_role > 기본 문구
        review_description = llm_action or human_role or f"{clean_task_name} 결과 검토 및 확정"

        hr_id = f"human_{jid}"
        # 🔑 Human 검토 노드는 Junior AI 와 **같은 display_id 공유** (동일 task 의 AI↔Human 쌍)
        # 별도 번호가 아니라 같은 번호 (예: 2.1.4.7 AI → 2.1.4.7 Human 검토)
        review_display_id = jnode.get("display_id", "") or ""

        human_nodes_out.append({
            "id": hr_id,
            "label": review_title[:40],
            "level": "L5",
            "actor": actor_lane,
            "type": "task",
            "ai_support": None,
            "position": {"x": jnode["position"]["x"], "y": hr_lane_y},
            "origin": "ai",
            "display_id": review_display_id,
            "description": review_description,
            "data": {
                "role": actor_lane,
                "label": review_title,
                "level": "L5",
                "id": review_display_id,
                "description": review_description,
            },
            "next": [],
        })
        ai_edges_out.append({
            "id": f"aie_hr_{jid}_{hr_id}",
            "source": jid,
            "target": hr_id,
            "label": "검토",
            "origin": "ai",
        })

    # 3-2) LLM이 제공한 AI ↔ As-Is 엣지 병합
    # jid → 대체하는 As-Is task_id (self-loop 필터링용)
    jid_to_replaced_asis: dict[str, str] = {}
    for jn in ai_nodes_out:
        if jn["actor"] != "Junior AI":
            continue
        # junior_agent_X_TASKID 구조에서 TASKID 추출 후 매칭 As-Is id 찾기
        jid = jn["id"]
        # task_id 가 포함된 junior id 패턴: junior_<agent_id>_<task_id>
        # 매칭된 As-Is 찾기 — asis_by_task_id에서 junior id 끝에 붙은 task_id로
        for tid, asis_node in asis_by_task_id.items():
            if jid.endswith(f"_{tid}") or jid.endswith(f"_{tid}_AI"):
                jid_to_replaced_asis[jid] = asis_node["id"]
                break

    def _is_self_loop(frm: str, to: str) -> bool:
        """task_id 가 같은 As-Is ↔ Junior AI 직접 엣지인지 판정."""
        if frm in jid_to_replaced_asis and jid_to_replaced_asis[frm] == to:
            return True
        if to in jid_to_replaced_asis and jid_to_replaced_asis[to] == frm:
            return True
        return False

    def _add_llm_edge(prefix: str, frm: str, to: str, label: str, valid_from: set, valid_to: set) -> None:
        frm = str(frm or "").strip()
        to = str(to or "").strip()
        if not frm or not to or frm not in valid_from or to not in valid_to:
            return
        # 같은 task_id 쌍 self-loop 차단
        if _is_self_loop(frm, to):
            return
        ai_edges_out.append({
            "id": f"{prefix}_{frm}_{to}",
            "source": frm,
            "target": to,
            "label": str(label or "")[:20],
            "origin": "ai",
        })

    # As-Is → AI (트리거)
    for e in llm_routing.get("asis_to_ai_edges", []):
        _add_llm_edge("llm_a2i", e.get("from"), e.get("to"), e.get("label", ""),
                      asis_id_set, ai_id_set)
    # AI → As-Is (결과 전달)
    for e in llm_routing.get("ai_to_asis_edges", []):
        _add_llm_edge("llm_i2a", e.get("from"), e.get("to"), e.get("label", ""),
                      ai_id_set, asis_id_set)
    # Senior AI → Junior AI (오케스트레이션 — LLM 우선, 기본 배선과 중복되면 중복 제거 대비 dedup)
    senior_ids = {n["id"] for n in ai_nodes_out if n["actor"] == "Senior AI"}
    junior_ids = {n["id"] for n in ai_nodes_out if n["actor"] == "Junior AI"}
    existing_pairs = {(e["source"], e["target"]) for e in ai_edges_out}
    for e in llm_routing.get("senior_triggers", []):
        frm = str(e.get("from") or "").strip()
        to = str(e.get("to") or "").strip()
        if frm in senior_ids and to in junior_ids and (frm, to) not in existing_pairs:
            ai_edges_out.append({
                "id": f"llm_orc_{frm}_{to}",
                "source": frm,
                "target": to,
                "label": str(e.get("label") or "기동")[:20],
                "origin": "ai",
            })
            existing_pairs.add((frm, to))

    # 4) lanes — 표준 순서대로, 실제 등장한 것만 (As-Is에서 내용 있는 레인 + AI 레인 + 새 Human 레인)
    for hn in human_nodes_out:
        role_counts[hn["actor"]] = role_counts.get(hn["actor"], 0) + 1
    actors_used = [a for a in _TOBE_ACTOR_ORDER if (role_counts.get(a, 0) > 0
                   or (a == "Senior AI" and senior_agents)
                   or (a == "Junior AI" and junior_agents))]

    all_nodes = asis_nodes_out + ai_nodes_out + human_nodes_out
    all_edges = asis_edges_out + ai_edges_out

    # ── 🔑 L4 별 1부터 sequential 재번호 (display_id) ─────────────────────
    # 사용자 요구: 같은 task 의 AI/Human 쌍은 같은 번호 + 왼쪽 기준 1부터 시작
    # pair_key 결정 규칙:
    #   - As-Is L5: task_id (예: "2.1.4.3")
    #   - Junior AI: 매칭 As-Is task_id 가 있으면 그것, 없으면 자체 task_id
    #     → Junior id 패턴 "junior_<aid>_<tid>" 에서 tid 추출
    #   - Human review: 대응 Junior 의 pair_key (Human 이미 jnode display_id 공유 받음)
    from collections import defaultdict as _defaultdict

    def _node_pair_key(n: dict) -> str:
        """노드의 pair_key (같은 task 그룹) 반환."""
        # As-Is 는 task_id 그대로
        if n.get("origin") == "asis":
            return n.get("task_id") or n["id"]
        # AI 노드 (Junior/Senior) - id 에서 task_id 추출
        nid = n.get("id", "")
        # junior_<agent>_<task_id> 패턴
        for prefix in ("junior_", "senior_", "human_"):
            if nid.startswith(prefix):
                rest = nid[len(prefix):]
                # _<task_id> 끝 부분 — As-Is task_id 와 매칭되는지 시도
                for tid in asis_by_task_id:
                    if rest.endswith(f"_{tid}") or rest == tid or rest.endswith(f"_{tid}_AI"):
                        return tid   # 매칭 As-Is 의 task_id 공유
                # 매칭 없으면 자체 jid 사용
                return rest
        return nid

    # L4_id → {pair_key → list[node]}
    l4_groups: dict[str, dict[str, list[dict]]] = _defaultdict(lambda: _defaultdict(list))
    for n in all_nodes:
        if n.get("level") not in ("L4", "L5"):
            continue
        # L4_id 결정
        n_l4_id = ""
        n_tid = n.get("task_id") or ""
        if n_tid:
            parts = n_tid.split(".")
            if len(parts) >= 3:
                n_l4_id = ".".join(parts[:3])
        if not n_l4_id:
            n_l4_id = (n.get("data") or {}).get("l4Id") or _primary_l4 or "unknown"
        pkey = _node_pair_key(n)
        l4_groups[n_l4_id][pkey].append(n)

    # 각 L4 내에서 pair_key 들을 leftmost x 기준 정렬 후 1부터 번호 부여
    for l4_id, pair_map in l4_groups.items():
        # pair 별 leftmost x 계산 → 정렬 키
        pair_sort = sorted(
            pair_map.items(),
            key=lambda kv: min(n.get("position", {}).get("x", 0) for n in kv[1])
        )
        for seq, (pkey, nodes) in enumerate(pair_sort, start=1):
            new_l5_id = f"{l4_id}.{seq}"
            for n in nodes:
                n["display_id"] = new_l5_id
                # data.id 도 갱신 (LevelNode 가 표시)
                if "data" not in n or not isinstance(n.get("data"), dict):
                    n["data"] = {}
                n["data"]["id"] = new_l5_id

    return {
        "l4_id": sheet_id,
        "l4_name": sheet_name,
        "actors_used": actors_used,
        "lanes": actors_used,
        "nodes": all_nodes,
        "edges": all_edges,
    }


def _merge_sheets_into_one(sheets: list, merged_id: str, merged_name: str):
    """여러 As-Is 시트를 하나의 가상 시트로 머지 (L3 전체 scope 대응)."""
    from workflow_parser import WorkflowSheet
    merged = WorkflowSheet(sheet_id=merged_id, name=merged_name)
    # 시트별 x 오프셋으로 충돌 방지
    x_offset = 0.0
    for s in sheets:
        if not s.nodes:
            continue
        # 노드 id 충돌 방지 — sheet_id prefix 부여
        for n in s.nodes.values():
            nid = f"{s.sheet_id}::{n.id}"
            cloned = type(n)(
                id=nid,
                level=n.level,
                task_id=n.task_id,
                label=n.label,
                description=n.description,
                position_x=n.position_x + x_offset,
                position_y=n.position_y,
                metadata=dict(n.metadata),
            )
            merged.nodes[nid] = cloned
        for e in s.edges:
            merged.edges.append(type(e)(
                id=f"{s.sheet_id}::{e.id}",
                source=f"{s.sheet_id}::{e.source}",
                target=f"{s.sheet_id}::{e.target}",
                label=e.label,
                animated=getattr(e, "animated", False),
                bidirectional=getattr(e, "bidirectional", False),
            ))
        xs = [n.position_x for n in s.nodes.values()]
        if xs:
            x_offset += (max(xs) - min(xs)) + 400.0
    return merged


@app.post("/api/workflow/generate-tobe-flow", tags=["Workflow"])
async def generate_tobe_flow(request: Request):
    """
    To-Be Workflow Swim Lane 생성 (결정론적 파이프라인).

    입력: sheet_id (optional) — L4 단위면 해당 시트, 비어있으면 L3 전체 머지
    동작:
      1. As-Is 노드/엣지/position/actor role 그대로 보존
      2. 실제 내용 있는 lane만 유지 (빈 lane 제거)
      3. Step 2 Senior AI / Junior AI agent를 새 lane(Senior AI / Junior AI)에 노드로 주입
      4. Junior AI 내부 sequential 엣지 + input_data/output_data 매칭으로 As-Is ↔ AI 엣지 연결

    출력 스키마 (hr-workflow-ai 호환):
      {
        process_name, tobe_sheets: [{
          l4_id, l4_name, lanes, actors_used,
          nodes: [{id, label, level, actor, type, position:{x,y}, ai_support, origin, ...}],
          edges: [{id, source, target, label, origin}]
        }]
      }
    """
    if not _wf_step2_cache:
        raise HTTPException(400, "상세 설계(Step 2)를 먼저 수행해주세요.")

    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    target_sheet_id = (body.get("sheet_id") or "").strip()

    if "parsed" not in _workflow_cache:
        raise HTTPException(400, "As-Is JSON을 먼저 업로드해주세요.")

    parsed = _workflow_cache["parsed"]
    all_sheets = parsed.sheets

    process_name = _wf_step2_cache.get("process_name",
                    _wf_step1_cache.get("process_name", "HR 프로세스"))

    tobe_sheets: list[dict] = []

    if target_sheet_id:
        # L4 scope: 단일 시트만
        picked = [s for s in all_sheets if s.sheet_id == target_sheet_id]
        if not picked:
            print(f"[TOBE⚠] sheet_id='{target_sheet_id}' 매칭 없음 — 전체 머지", flush=True)
            merged = _merge_sheets_into_one(all_sheets, "merged", process_name)
            tobe_sheets.append(await _build_tobe_sheet_from_asis(merged, process_name))
        else:
            for s in picked:
                tobe_sheets.append(await _build_tobe_sheet_from_asis(s, process_name))
    else:
        # L3 scope: 전체를 하나의 통합 시트로 머지 → 한 장짜리 Swim Lane
        if len(all_sheets) == 1:
            tobe_sheets.append(await _build_tobe_sheet_from_asis(all_sheets[0], process_name))
        elif all_sheets:
            merged = _merge_sheets_into_one(all_sheets, "merged", process_name)
            tobe_sheets.append(await _build_tobe_sheet_from_asis(merged, process_name))

    if not tobe_sheets:
        raise HTTPException(500, "To-Be Workflow 생성에 실패했습니다 — 대상 시트가 없습니다.")

    result = {"process_name": process_name, "tobe_sheets": tobe_sheets}

    global _wf_tobe_flow_cache
    _wf_tobe_flow_cache = dict(result)

    print(f"[TOBE] 생성 완료 — 시트 {len(tobe_sheets)}개, "
          f"총 노드 {sum(len(s['nodes']) for s in tobe_sheets)}개", flush=True)

    return {"ok": True, **result}


def _tobe_cache_to_hr_json(cache: dict) -> dict:
    """
    새 generate-tobe-flow 파이프라인 결과(tobe_sheets)를 hr-workflow-ai v2.0 호환 JSON으로 직렬화.
    레거시 tobe_sheets_to_hr_json과 달리 position/level/data/role/actors 등 원본 보존.
    """
    from datetime import datetime, timezone

    process_name = cache.get("process_name", "To-Be Workflow")
    raw_sheets = cache.get("tobe_sheets", [])
    sheets_out: list[dict] = []

    for rs in raw_sheets:
        sheet_id = str(rs.get("l4_id") or f"sheet-{len(sheets_out) + 1}")
        sheet_name = str(rs.get("l4_name") or sheet_id)
        lanes = rs.get("lanes") or rs.get("actors_used") or []

        nodes_out: list[dict] = []
        for n in rs.get("nodes", []):
            level = (n.get("level") or "L5").upper()
            node_type = (
                "decision" if level == "DECISION"
                else "memo" if level == "MEMO"
                else level.lower()  # "l2" | "l3" | "l4" | "l5"
            )
            pos = n.get("position") or {}
            x = float(pos.get("x", 0) or 0)
            y = float(pos.get("y", 0) or 0)

            # LevelNode가 사용하는 data 필드 구성 (원본 n.data + 추가 매핑)
            base_data = dict(n.get("data") or {})
            # label/level/id는 LevelNode 루트에서 참조하는 필드
            base_data.setdefault("label", n.get("label", ""))
            base_data["level"] = level
            if n.get("task_id"):
                base_data.setdefault("id", n["task_id"])
            if n.get("description"):
                base_data.setdefault("description", n["description"])
            # AI 노드는 role을 actor로 강제 (Senior/Junior AI 레인 배치)
            if n.get("origin") == "ai" and n.get("actor"):
                base_data["role"] = n["actor"]
                # ai_support는 description에 저장 (memo로 두면 hr-workflow-ai에서 노란 스티커로 뜸)
                if n.get("ai_support"):
                    base_data.setdefault("description", n["ai_support"])
            # role이 여전히 문자열이 아니면 actors_all + custom_role로 재구성
            if not isinstance(base_data.get("role"), str):
                parts = list(n.get("actors_all") or [])
                cr = n.get("custom_role") or ""
                if cr and "그 외" in parts:
                    parts = [f"그 외:{cr}" if p == "그 외" else p for p in parts]
                base_data["role"] = ", ".join(parts) if parts else (n.get("actor") or "")

            nodes_out.append({
                "id": n["id"],
                "type": node_type,
                "position": {"x": x, "y": y},
                "data": base_data,
            })

        edges_out: list[dict] = []
        for e in rs.get("edges", []) or []:
            is_ai = e.get("origin") == "ai"
            edge_obj = {
                "id": e["id"],
                "source": e["source"],
                "target": e["target"],
                "type": "ortho",
                "animated": is_ai,
                "style": {
                    "stroke": "#00827F" if is_ai else "#64748B",
                    "strokeWidth": 2 if is_ai else 1.5,
                },
                "markerEnd": {
                    "type": "ArrowClosed",
                    "width": 18,
                    "height": 18,
                    "color": "#00827F" if is_ai else "#64748B",
                },
            }
            if e.get("label"):
                edge_obj["label"] = e["label"]
            edges_out.append(edge_obj)

        sheets_out.append({
            "id": sheet_id,
            "name": sheet_name,
            "type": "swimlane",
            "lanes": lanes,
            "nodes": nodes_out,
            "edges": edges_out,
        })

    return {
        "version": "2.0",
        "exportedAt": datetime.now(timezone.utc).isoformat(),
        "processName": process_name,
        "sheets": sheets_out,
    }


@app.get("/api/workflow/export-tobe-flow-json", tags=["Workflow"])
async def export_tobe_flow_json():
    """
    generate-tobe-flow 결과(새 파이프라인)를 hr-workflow-ai v2.0 호환 JSON으로 그대로 직렬화.
    As-Is 노드의 position/level/data (actors/systems/painPoints/inputs/outputs/logic) 전부 보존.
    """
    if not _wf_tobe_flow_cache:
        raise HTTPException(400, "To-Be Flow를 먼저 생성해주세요 (generate-tobe-flow).")

    hr_json = _tobe_cache_to_hr_json(_wf_tobe_flow_cache)

    process_name = _wf_tobe_flow_cache.get("process_name", "tobe_workflow")
    filename = f"{process_name}_ToBeFlow.json"

    from urllib.parse import quote
    encoded_fn = quote(filename)
    return StreamingResponse(
        iter([json.dumps(hr_json, ensure_ascii=False, indent=2)]),
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_fn}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@app.get("/api/workflow/export-tobe-excel", tags=["Workflow"])
async def export_tobe_excel_endpoint():
    """
    To-Be 설계 결과를 As-Is 템플릿과 동일한 포맷의 Excel로 반출.
    - 업로드한 As-Is 엑셀을 베이스로 복사 → 데이터 행 초기화 → To-Be task로 재작성
    - 각 행: Junior AI task (AI 파트) / AI+Human의 Human 파트 검토 행 / 보존된 Human L5
    """
    if not _wf_step2_cache:
        raise HTTPException(400, "Step 2 상세 설계를 먼저 수행해주세요.")
    if not _wf_excel_path or not Path(_wf_excel_path).exists():
        raise HTTPException(400, "As-Is 템플릿 엑셀을 먼저 업로드해주세요.")

    from tobe_excel_exporter import export_tobe_excel
    import tempfile

    process_name = _wf_step2_cache.get("process_name", "ToBe_설계")
    # 세션 임시 파일
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        output_path = tmp.name

    # Step 1 에서 '삭제'/'통합' 처리된 task 는 엑셀 export 에서도 제외
    _filtered_step2_for_excel = _filter_step2_deprecated(_wf_step2_cache)
    try:
        export_tobe_excel(
            template_path=_wf_excel_path,
            output_path=output_path,
            step2_cache=_filtered_step2_for_excel,
            classification=_wf_classification,
            excel_tasks=_wf_excel_tasks,
            wf_tobe_flow_cache=_wf_tobe_flow_cache,
        )
    except Exception as e:
        try:
            Path(output_path).unlink(missing_ok=True)
        except Exception:
            pass
        raise HTTPException(500, f"Excel 반출 실패: {e}")

    filename = f"{process_name}_ToBe설계.xlsx"
    from urllib.parse import quote
    encoded_fn = quote(filename)

    def _iter_file():
        with open(output_path, "rb") as f:
            while True:
                chunk = f.read(8192)
                if not chunk:
                    break
                yield chunk
        try:
            Path(output_path).unlink(missing_ok=True)
        except Exception:
            pass

    return StreamingResponse(
        _iter_file(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_fn}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


# ── Step 1 채팅 ─────────────────────────────────────────────

@app.post("/api/workflow/chat-step1", tags=["Workflow"])
async def chat_workflow_step1(request: Request):
    """
    Step 1 채팅: 기본 설계에 대한 추가 질문/수정 요청.
    특정 기업명이 언급되면 해당 기업의 벤치마킹도 추가로 수행합니다.
    """
    from benchmark_search import search_benchmarks

    body = await request.json()
    user_message = body.get("message", "")
    sheet_id = body.get("sheet_id", "")
    if not user_message:
        raise HTTPException(400, "메시지가 필요합니다.")

    # 채팅도 L3 전체 기준 (JSON의 모든 시트 합산)
    process_name, task_summary, pain_summary = _build_task_and_pain_summary("")

    # 메시지에서 기업명 감지 → 추가 쿼리에 포함
    import re as _re
    company_pattern = _re.findall(
        r'(?:Google|Amazon|Meta|Microsoft|Apple|Unilever|삼성|현대|SK|LG|두산|'
        r'JPMorgan|Goldman|DoorDash|Siemens|GE|SAP|Workday|IBM|Oracle|'
        r'[A-Z][a-z]+(?:\s[A-Z][a-z]+)*)',
        user_message,
    )

    extra_bm_text = ""
    new_bm_entries: list[dict] = []
    raw_search_context = ""   # 채팅 LLM에 직접 전달할 실제 검색 결과 텍스트

    # 엑셀 데이터가 로드된 경우 항상 추가 벤치마킹 검색 수행
    # (채팅의 핵심 목적: Excel/JSON/PPT 업무 내용 기반으로 추가 사례 발굴)
    if _wf_excel_tasks:
        companies = list(set(company_pattern)) if company_pattern else []

        # JSON 전체 시트 기준 L4 ID 수집
        _, task_sum_txt, _ = _build_task_and_pain_summary("")
        _, excel_by_l4, excel_by_l3, excel_by_l2 = _build_excel_index()

        scoped_l4_ids: set[str] = set()
        if "parsed" in _workflow_cache:
            parsed_tmp = _workflow_cache["parsed"]
            for s in parsed_tmp.sheets:  # 모든 시트
                for n in s.nodes.values():
                    if n.task_id:
                        parts = n.task_id.split(".")
                        if len(parts) >= 3:
                            scoped_l4_ids.add(".".join(parts[:3]))

        _PAIN_MAP_BM = [("pain_time","시간/속도"),("pain_accuracy","정확성"),
                        ("pain_repetition","반복/수작업"),("pain_data","정보/데이터"),
                        ("pain_system","시스템/도구"),("pain_communication","의사소통/협업")]

        def _get_pains_bm(tasks):
            found = set()
            for t in tasks:
                for f, lbl in _PAIN_MAP_BM:
                    if getattr(t, f, ""): found.add(lbl)
            return list(found)

        l4_details_chat = []
        l4_items = (
            [(lid, excel_by_l4[lid]) for lid in scoped_l4_ids if lid in excel_by_l4]
            if scoped_l4_ids else list(excel_by_l4.items())[:4]
        )
        for l4_id, tasks in l4_items[:4]:
            name = tasks[0].l4 if tasks else ""
            if name:
                l4_details_chat.append({"name": name, "task_id": l4_id,
                    "pain_points": _get_pains_bm(tasks),
                    "description": "; ".join(t.description for t in tasks[:3] if t.description),
                    "task_names": [t.name for t in tasks[:5]]})

        # 범위 내 Task만으로 L3/L2 이름 구성
        scoped_tasks = [t for t in _wf_excel_tasks if not scoped_l4_ids or t.l4_id in scoped_l4_ids]

        # L5 tasks — 채팅 경로도 동일하게 scope-filtered L5 context 전달
        l5_tasks_chat = []
        for d in l4_details_chat:
            for t in excel_by_l4.get(d.get("task_id", ""), []):
                if t.name:
                    l5_tasks_chat.append({
                        "name": t.name,
                        "description": t.description or "",
                        "l4": t.l4 or "",
                    })

        bm_data = {
            "process_name": process_name,
            "agents": [],
            "l4_details": l4_details_chat,
            "l4_names": [d["name"] for d in l4_details_chat],
            "l3_names": list(dict.fromkeys(t.l3 for t in scoped_tasks if t.l3))[:4],
            "l2_names": list(dict.fromkeys(t.l2 for t in scoped_tasks if t.l2))[:2],
            "l3_details": [], "l2_details": [],
            "l5_tasks": l5_tasks_chat[:30],
            "blueprint_summary": f"{' '.join(companies) + ' ' if companies else ''}{process_name} AI 적용",
        }
        if companies:
            bm_data["extra_queries"] = [
                f"{' '.join(companies)} '{process_name}' AI automation case study 2024 2025",
            ]
        _bm_sr2 = await search_benchmarks(bm_data)
        raw = _bm_sr2.get("results", [])

        if raw:
            # 검색 결과를 쿼리별로 그룹핑 (LLM 추출 + 채팅 컨텍스트 공용)
            # raw 전체를 순회해야 모든 라운드(R1/R2/R3)의 쿼리가 포함됨
            # raw[:12] 처럼 자르면 앞 쿼리의 citation들만 잡혀 나머지 쿼리가 누락됨
            from collections import defaultdict as _ddict
            _qg: dict = _ddict(lambda: {"content": "", "urls": [], "round": 0})
            for r in raw:
                _q = r.get("query", r.get("title", ""))
                if r.get("content") and not _qg[_q]["content"]:
                    _qg[_q]["content"] = r.get("content", "")
                if r.get("url"):
                    _qg[_q]["urls"].append(r["url"])
                if r.get("round", 0) > _qg[_q]["round"]:
                    _qg[_q]["round"] = r.get("round", 0)

            # ── 채팅 LLM에 넘길 실제 검색 결과 텍스트 구성 ──
            # 쿼리별 R1→R2→R3 순서로 정렬, 각 content는 800자로 제한 (총 토큰 절감)
            # ⚠️ 과거에 [R1-1] 같은 round-query 태그가 source로 오인되던 문제 → 쿼리 구분선으로만 표시
            sorted_qg = sorted(_qg.items(), key=lambda x: x[1].get("round", 0))
            raw_search_context = f"## ✅ Perplexity 실시간 검색 결과 ({len(sorted_qg)}개 쿼리 실행)\n"
            raw_search_context += "※ 아래는 실제 웹에서 수집된 내용입니다. 이 내용을 기반으로 답변하세요.\n"
            for idx, (_q, _g) in enumerate(sorted_qg, 1):
                raw_search_context += f"\n────── 쿼리 {idx}: {_q[:100]} ──────\n"
                if _g["urls"]:
                    raw_search_context += "출처 URL:\n" + "\n".join(f"  - {u}" for u in _g["urls"][:4]) + "\n"
                else:
                    raw_search_context += "출처 URL: 없음\n"
                raw_search_context += f"내용:\n{_g['content'][:800]}\n"

            # ── LLM으로 테이블 항목 추출 (URL 있는 것만 테이블에 추가) ──
            bm_sys = f"""벤치마킹 분석 전문가입니다. 검색 결과에서 '{process_name}' 프로세스 관련 사례를 추출합니다.
L4 활동: {', '.join(bm_data['l4_names'][:4])}

## 🚨 source 필드 엄격 규칙
- source 는 반드시 **AI 솔루션을 실제로 도입·운영한 구체 기업명** (실명, 글로벌 인지도)
- ❌ 절대 금지 예시:
  · "R1-1", "[R2-3]", "쿼리 1" 등 검색 라운드 태그 (이건 쿼리 구분선일 뿐 회사명 아님)
  · "한 제조사", "Fortune 500 기업", "글로벌 HR 부서", "익명" 등 추상·익명 표현
  · McKinsey/BCG/Gartner 등 컨설팅·리서치펌 (보고서 작성자이지 도입 기업 아님)
  · Workday/SAP/ServiceNow 등 **솔루션 제공자**가 "도입 기업" 자리에 들어가는 경우
  · 검색 결과에 기업명이 명시되지 않으면 → 해당 항목 **DROP** (억지로 만들지 말 것)
- ✅ 허용 예시: Google, Amazon, Microsoft, Unilever, JPMorgan, Siemens, 삼성전자, 현대자동차, 두산 등
  **특정 기업이 자기 회사에 AI 를 도입한 사례**만

## ⚠️ 절대 규칙:
1. url 필드는 위 검색 결과의 "출처 URL" 목록에 있는 것만 사용. URL 임의 생성 금지.
2. url이 없으면 해당 항목을 benchmark_table에 포함하지 말 것.
3. 학습 지식으로 사례를 만들어내지 말 것. 검색 결과 텍스트에 근거한 내용만.
4. source 에 구체적 실명 회사명이 없으면 **그 사례는 그냥 빼세요** — 억지로 채우지 마세요.

관련성 없는 사례는 제외. 출력 형식 (JSON만):
{{"benchmark_table": [{{"source":"AI솔루션 도입 기업명 (실명)","company_type":"Tech 선도 | 非Tech 실제 구현","industry":"","process_area":"","ai_adoption_goal":"","ai_technology":"","key_data":"","adoption_method":"","use_case":"","outcome":"","infrastructure":"","implication":"","url":"[검색결과URL만]"}}]}}"""
            bm_user_msg = "## 검색 결과\n"
            bm_user_msg += "※ 아래 '쿼리 N' 라벨은 검색 쿼리 구분선이지 회사 이름이 아닙니다. source 에 절대 쓰지 마세요.\n"
            for idx, (_q, _g) in enumerate(sorted_qg, 1):
                bm_user_msg += f"\n────── 쿼리 {idx}: {_q[:100]} ──────\n"
                if _g["urls"]:
                    bm_user_msg += "출처 URL:\n" + "\n".join(f"  - {u}" for u in _g["urls"][:4]) + "\n"
                bm_user_msg += f"내용: {_g['content'][:800]}\n"
            bm_result = await _call_llm_step1(bm_sys, [{"role":"user","content":bm_user_msg}], max_tokens=16384)
            if bm_result and bm_result.get("benchmark_table"):
                new_bm_entries = bm_result["benchmark_table"]
                chat_sheet_key = sheet_id or "__chat__"
                existing_sources = {b.get("source", "") for rows in _wf_benchmark_table.values() for b in rows}
                if chat_sheet_key not in _wf_benchmark_table:
                    _wf_benchmark_table[chat_sheet_key] = []
                for entry in new_bm_entries:
                    src = entry.get("source", "")
                    url = entry.get("url", "")
                    if (url
                            and src not in existing_sources
                            and not _is_news_url(url)
                            and _is_valid_benchmark_source(src)
                            and (entry.get("use_case") or entry.get("outcome"))):
                        _wf_benchmark_table[chat_sheet_key].append(entry)
            extra_bm_text = f"\n\n[✅ Perplexity 검색 완료 — 3라운드 {len(sorted_qg)}개 쿼리 실행, {len(new_bm_entries)}건 사례 분석]"

    # 기존 설계 결과 or 벤치마킹 결과가 있으면 포함
    context = ""
    if _wf_step1_cache:
        context = f"현재 설계 결과:\n{json.dumps(_wf_step1_cache, ensure_ascii=False, indent=2)[:3000]}"
    all_bm_for_context = [r for rows in _wf_benchmark_table.values() for r in rows]
    if all_bm_for_context:
        context += "\n\n현재 벤치마킹 테이블 (전체 시트 합산):\n"
        for bm in all_bm_for_context[:8]:
            context += f"- [{bm.get('company_type','?')}] {bm.get('source','')}: {bm.get('use_case','')}\n"

    # 사용자 첨부 리소스 (URL/이미지 분석 결과) 포함
    if _wf_user_resources:
        context += f"\n\n## 사용자 첨부 리서치 자료 ({len(_wf_user_resources)}건)\n"
        for i, res in enumerate(_wf_user_resources):
            rtype = "🔗 URL" if res.get("type") == "url" else "🖼 이미지"
            context += (f"\n[{i+1}] {rtype} — {res.get('title', res.get('source',''))}\n"
                        f"출처: {res.get('source','')}\n"
                        f"{res.get('content','')[:1500]}\n")


    chat_system = f"""당신은 AI 기반 업무 혁신 벤치마킹 전문가입니다.
현재 '{process_name}' 프로세스의 벤치마킹 리서치를 진행 중입니다.

## 🚨 이 채팅의 역할 (매우 중요)
이 채팅은 **벤치마킹 사례 수집·Q&A 전용** 입니다.
- ✅ 허용: 추가 사례 조사 결과 설명, 사용자 질문에 대한 답변, 벤치마킹 테이블에 새 기업 사례를 찾아 추가
- ❌ **절대 금지**: Step 1 기본 설계(redesigned_process, agents 같은 구조) JSON 반환
- ❌ 금지: L3/L4/L5 재설계, automation_level/ai_technique 등 설계 의사결정을 직접 내리지 말 것

사용자가 설계 반영을 원하면 → "추가된 벤치마킹 사례를 반영하려면 **[기본 설계 재생성]** 버튼을 눌러주세요"로 안내하세요.

채팅은 벤치마킹 enrichment 까지만, 그 다음 Gap 분석·기본 설계 생성은 사용자가 명시적으로 버튼을 눌러 단계별로 진행합니다.

## 출력 형식
- **순수 텍스트 답변만** (JSON/마크다운 코드블록 X)
- 벤치마킹 사례를 언급할 때 → source·use_case·URL 형태로 짧게 정리
- 사용자가 질문한 범위를 벗어나 설계 전체를 다시 쓰지 말 것

## 두산 HR 전문 약어 정의 (반드시 준수)
- **BP** = Business Partner (HR BP, 인사 담당 파트너) — 절대로 'British Petroleum'이 아님
- **ER** = Employee Relations (노사관계/직원관계)
- **L2/L3/L4/L5** = 두산 프로세스 계층 레벨 (L2: 대분류, L3: 프로세스, L4: 활동, L5: Task)
- **발령** = 인사발령 (personnel assignment/job transfer) — 석유화학 용어 아님

## ⚠️ 절대 원칙
1. **아래 제공된 검색 결과만 사용** — 당신의 학습 지식으로 사례를 만들지 말 것
2. **검색 결과에 없으면 솔직하게 "이번 검색에서 확인되지 않았습니다"라고 명시**
3. **URL은 검색 결과에 나온 것만 인용** — 임의 생성 절대 금지
4. **"저는 실시간 검색을 못 합니다" 같은 말 금지** — 아래에 이미 검색 결과가 제공됨

{context}

{raw_search_context}
{extra_bm_text}

## 답변 방식
- 위 "Perplexity 실시간 검색 결과"에 있는 내용을 기반으로 답변
- 기존 벤치마킹 테이블 + 이번 검색 결과를 종합하여 구체적으로 설명
- 검색 결과에 URL이 있으면 반드시 출처 명시
- 검색 결과에 없는 내용은 "이번 검색에서 확인되지 않았습니다"라고 명시
- 벤치마킹 사례가 추가됐으면 "벤치마킹 테이블에 N건 추가됐습니다. Gap 분석·기본 설계에 반영하려면 해당 단계 버튼을 눌러주세요" 라고 사용자에게 안내

## 엑셀 기반 L5 Task 목록 (실제 업무 범위 참고)
{task_summary[:1500]}
"""

    global _wf_chat_history
    _wf_chat_history.append({"role": "user", "content": user_message})

    settings = load_settings()
    anthropic_key = os.getenv("ANTHROPIC_API_KEY", "") or settings.anthropic_api_key
    response_text = ""

    from usage_store import add_usage as _add_usage_chat

    if anthropic_key:
        try:
            from anthropic import AsyncAnthropic
            client = AsyncAnthropic(api_key=anthropic_key)
            response = await client.messages.create(
                model=settings.anthropic_model or "claude-sonnet-4-6",
                max_tokens=8192,
                system=chat_system,
                messages=_wf_chat_history,
            )
            if response.usage:
                _add_usage_chat("anthropic",
                                input_tokens=response.usage.input_tokens,
                                output_tokens=response.usage.output_tokens)
            response_text = response.content[0].text
        except Exception as e:
            print(f"[workflow-chat] Anthropic 실패: {e}")

    if not response_text:
        openai_key = os.getenv("OPENAI_API_KEY", "") or settings.api_key
        if openai_key:
            try:
                from openai import AsyncOpenAI
                client = AsyncOpenAI(api_key=openai_key)
                resp = await client.chat.completions.create(
                    model=settings.model or "gpt-5.4",
                    temperature=0.0,
                    messages=[
                        {"role": "system", "content": chat_system},
                        *_wf_chat_history,
                    ],
                )
                if resp.usage:
                    _add_usage_chat("openai",
                                    input_tokens=resp.usage.prompt_tokens,
                                    output_tokens=resp.usage.completion_tokens)
                response_text = resp.choices[0].message.content or ""
            except Exception as e:
                print(f"[workflow-chat] OpenAI 실패: {e}")

    if not response_text:
        raise HTTPException(500, "AI 모델 호출에 실패했습니다.")

    _wf_chat_history.append({"role": "assistant", "content": response_text})

    # 🔒 채팅은 벤치마킹 enrichment + Q&A 전용 — Step 1 기본 설계(_wf_step1_cache) 자동 덮어쓰기 금지
    # 사용자가 '기본 설계 재생성' 버튼을 눌러야만 Step 1 이 업데이트됨 (채팅에서 직접 반영 X)
    # 벤치마킹 테이블에 새 사례가 추가된 경우만 UI 에 알림으로 표시
    return {
        "ok": True,
        "message": response_text,
        "updated": False,               # 채팅은 Step 1 덮어쓰지 않음
        "result": None,                 # 기본 설계 변경 없음
        "benchmark_updated": len(new_bm_entries) > 0,
        "benchmark_table": {k: v for k, v in _wf_benchmark_table.items()},
    }


@app.post("/api/workflow/generate-step2", tags=["Workflow"])
async def generate_workflow_step2(request: Request):
    """
    Step 2: Pain Point + PwC 분석 기반 상세 설계 (Bottom-Up, Lv.4~5).
    Step 1 결과를 바탕으로 상세 설계를 수행합니다.
    """
    if not _wf_step1_cache:
        raise HTTPException(400, "Step 1을 먼저 실행하세요.")

    body = await request.json() if True else {}
    additional_context = body.get("additional_context", "") if body else ""
    step2_sheet_id = body.get("sheet_id", "")  # L4 단위 분석 시 해당 시트 ID

    process_name = _wf_step1_cache.get("process_name", "HR 프로세스")

    # 상세 설계 스코프: sheet_id 있으면 해당 L4만, 없으면 JSON 전체(L3)
    # (스코프를 먼저 계산해서, hybrid 분리를 스코프 내 태스크로만 한정 → 속도 대폭 향상)
    scoped_tasks_step2, _, _ = _build_task_and_pain_summary(step2_sheet_id or "")
    # _build_task_and_pain_summary가 relevant_tasks를 반환하지 않으므로 직접 구성
    scoped_tasks_step2 = _wf_excel_tasks  # 기본: 전체
    if "parsed" in _workflow_cache:
        _p2 = _workflow_cache["parsed"]
        if step2_sheet_id:
            # L4 scope: _build_task_and_pain_summary와 동일한 0순위 로직 사용
            _target_sheets = [s for s in _p2.sheets if s.sheet_id == step2_sheet_id]
            if _target_sheets:
                _sheet_name = _target_sheets[0].name.strip()
                # 전체 시트에서 L3 이름 수집
                _l3_names = {n.label.strip() for s in _p2.sheets for n in s.nodes.values() if n.level == "L3" and n.label}
                _filtered2 = [t for t in _wf_excel_tasks
                               if t.l4 and t.l4.strip() == _sheet_name
                               and (not _l3_names or (t.l3 and t.l3.strip() in _l3_names))]
                if not _filtered2:
                    _filtered2 = [t for t in _wf_excel_tasks if t.l4 and t.l4.strip() == _sheet_name]
                if _filtered2:
                    scoped_tasks_step2 = _filtered2
                    print(f"[STEP2] L4 scope='{_sheet_name}' → {len(scoped_tasks_step2)}개 태스크", flush=True)
        else:
            # L3 scope: 전체 JSON 시트 범위
            _tids2, _l4ids2, _l3ids2 = set(), set(), set()
            for _s2 in _p2.sheets:
                for _n2 in _s2.nodes.values():
                    if _n2.level == "L5" and _n2.task_id:
                        _tids2.add(_n2.task_id)
                        _pts2 = _n2.task_id.split(".")
                        if len(_pts2) >= 3: _l4ids2.add(".".join(_pts2[:3]))
                        if len(_pts2) >= 2: _l3ids2.add(".".join(_pts2[:2]))
            _filtered2 = [t for t in _wf_excel_tasks if t.id in _tids2]
            if not _filtered2: _filtered2 = [t for t in _wf_excel_tasks if t.l4_id in _l4ids2]
            if not _filtered2: _filtered2 = [t for t in _wf_excel_tasks if t.l3_id in _l3ids2]
            if _filtered2: scoped_tasks_step2 = _filtered2
            print(f"[STEP2] L3 scope → {len(scoped_tasks_step2)}개 태스크", flush=True)

    # AI+Human Task lazy 분리 보증 — 스코프 내 태스크만 대상으로 병렬 실행
    # (이전엔 전체 AI+Human 태스크 329개를 순차 분리해서 수 분간 블록되던 문제 해결)
    scoped_task_ids = {t.id for t in scoped_tasks_step2}
    pending_hybrid = [
        tid for tid, cls in _wf_classification.items()
        if cls.get("label") == "AI + Human"
        and not cls.get("hybrid_note")
        and tid in scoped_task_ids
    ]
    if pending_hybrid:
        print(f"[STEP2] 미분리 AI+Human Task {len(pending_hybrid)}개 (스코프 내) → 병렬 분리 실행", flush=True)
        await _split_hybrid_tasks_with_llm(target_task_ids=scoped_task_ids)

    # 상세 Pain Point 분석 (스코프 내 tasks만)
    pain_detail_lines = []
    for t in scoped_tasks_step2:
        pains = []
        if t.pain_time: pains.append(f"시간/속도: {t.pain_time}")
        if t.pain_accuracy: pains.append(f"정확성: {t.pain_accuracy}")
        if t.pain_repetition: pains.append(f"반복/수작업: {t.pain_repetition}")
        if t.pain_data: pains.append(f"정보/데이터: {t.pain_data}")
        if t.pain_system: pains.append(f"시스템/도구: {t.pain_system}")
        if t.pain_communication: pains.append(f"의사소통: {t.pain_communication}")
        if pains:
            cls = _wf_classification.get(t.id, {})
            lbl = cls.get('label', '미분류')
            header_suffix = ""
            if lbl == "AI + Human":
                hn = cls.get("hybrid_note", "")
                if hn:
                    header_suffix = f"\n  ⚙️ {hn}"
            elif lbl == "Human":
                rsn = cls.get("reason", "")
                if rsn:
                    header_suffix = f"\n  🙋 Human 전용 근거: {rsn[:150]}"
            pain_detail_lines.append(
                f"### [{t.id}] {t.name} (분류: {lbl}){header_suffix}\n"
                f"  Pain Point: {'; '.join(pains)}\n"
                f"  비고: {t.remark or '없음'}"
            )

    pain_detail = "\n".join(pain_detail_lines) if pain_detail_lines else "상세 Pain Point 정보 없음"

    # 전체 분류 집계 — Step 2 프롬프트에서 LLM이 분포를 파악하도록
    cls_stats = {"AI": 0, "AI + Human": 0, "Human": 0}
    scoped_ids = {t.id for t in scoped_tasks_step2}
    for tid, cls in _wf_classification.items():
        if tid not in scoped_ids:
            continue
        lbl = cls.get("label", "")
        if lbl in cls_stats:
            cls_stats[lbl] += 1

    # As-Is + 엑셀 매핑 컨텍스트 — L4 scope면 해당 시트만, L3면 전체
    asis_info = _build_mapped_asis_context(step2_sheet_id or "")

    # 두산 HR 고유 용어 설명 (Step 2 LLM용)
    _s2_l2_names = list({t.l2 for t in _wf_excel_tasks if t.l2})
    _s2_l3_names = list({t.l3 for t in _wf_excel_tasks if t.l3})
    _s2_doosan_ctx = ""
    if any(term in " ".join(_s2_l2_names + _s2_l3_names + [process_name])
           for term in ["PDS", "LDS", "EDS", "SDS"]):
        _s2_doosan_ctx = """
## 두산 HR 프로그램 용어
- PDS (Performance Development Survey): 선임→수석 승진 역량 평가 (다면진단·전문성 인터뷰 등)
- LDS (Leadership Development Survey): 수석→팀장 승진 리더십 역량 평가
- EDS (Executive Development Survey): 임원 승진 대상자 평가
- SDS (Specialist Development Survey): Specialist 트랙 승진 전문성 평가
→ 모두 두산그룹 내부 승진·역량개발 평가 체계입니다. AI 설계 시 평가 자동화, 다면 피드백 분석, Assessor 매칭 등을 적극 고려하세요.
"""

    step2_system = f"""당신은 AI 기반 업무 혁신 설계 전문가입니다.
Step 1에서 도출된 기본 설계를 기반으로, 두산에 최적화된 **AI 기반 To-Be Workflow 상세 설계**를 수행합니다.
{_s2_doosan_ctx}
## 핵심 철학
- **Bottom-Up 접근**: As-Is 프로세스의 Pain Point를 Deep-dive 분석
- **Senior AI 기반 End-to-End 오케스트레이션 구조로 전환**
- Step 1(Top-Down)의 벤치마킹 인사이트 + Step 2(Bottom-Up)의 Pain Point 분석을 융합

## ⚠️ 재설계 범위 (swim lane 기준)
- **재설계 가능**: HR 담당자, HR 임원, 지주, 자회사, BG, 계열사 등 두산 내부 조직이 수행하는 Task
- **재설계 불가 (현행 유지)**: 큐벡스, 업체 등 외부 업체/시스템이 수행하는 Task — AI Agent 설계 대상에서 제외하고 "현행 유지" 처리
- As-Is 컨텍스트에서 수행주체 라인에 `[외부 업체/시스템 — 재설계 제외]` 표시된 Task는 Agent에 할당하지 말 것

## 🚨 엑셀 분류(AI / AI+Human / Human) — Knock-out 제약 (절대 변경 금지)
엑셀 업로드 시 이미 판정·검토된 분류 라벨은 Step 2 설계의 **하드 제약**입니다. 임의로 바꿀 수 없습니다.

스코프 내 태스크 분포: AI {cls_stats["AI"]}개 / AI+Human {cls_stats["AI + Human"]}개 / Human {cls_stats["Human"]}개

- **Human으로 분류된 Task** → AI Agent의 `assigned_tasks`에 절대 포함하지 말 것.
  이 Task들은 "현행 유지"로 간주하고 AI 설계 대상에서 제외. (결과 카운트에도 반영하지 말 것)

- **AI + Human으로 분류된 Task** → 반드시 **AI 파트와 Human 파트를 분리**해서 반영:
  - task_summary / pain_detail에 이미 분리 정보(`AI 파트: ... / Human 파트: ...`)가 주입되어 있음. 이를 그대로 사용.
  - `assigned_tasks`에는 **AI 파트만** task_name·ai_role·input/output으로 기재
  - `human_role` 필드에는 **Human 파트를 반드시 구체적으로 명시** (빈 문자열 금지)
  - `automation_level`은 `"Human-in-Loop"` 또는 `"Human-on-the-Loop"` 중 하나 (절대 Full-Auto 금지)

- **AI로 분류된 Task** → Junior AI의 `assigned_tasks`에 자유 배치.
  `automation_level`은 `"Full-Auto"` 또는 `"Human-on-the-Loop"` 중 선택 (Human-in-Loop도 가능하지만 지양).

위 제약을 어기면 설계가 무효 처리됩니다.

## 프로세스: {process_name}

## Step 1 기본 설계 결과
{json.dumps(_wf_step1_cache, ensure_ascii=False, indent=2)[:4000]}

## As-Is 워크플로우 + 엑셀 매핑 (L3→L4→Task 계층, Pain Point 포함)
{asis_info or "As-Is 워크플로우 정보 없음 (JSON/PPT 미업로드)"}

## 상세 Pain Point Deep-dive (엑셀 기준)
{pain_detail}

{f"## 추가 컨텍스트{chr(10)}{additional_context}" if additional_context else ""}

## 상세 설계 요구사항
1. Lv.4~5 단위로 구체적인 AI 적용 방안 설계
2. 각 Agent의 구체적 역할, 사용 AI 기법, Input/Output 명시
3. Human-on-the-Loop 중심으로 설계 (Senior AI가 전체 관리)
4. 기존 As-Is에 없던 혁신적 Task 추가 가능
5. Pain Point를 근본적으로 해결하는 방향

## ⚠️ agent_type 규칙 (반드시 준수)
- `agent_type`은 **"Senior AI"** 또는 **"Junior AI"** 둘 중 하나만 사용 (다른 텍스트 절대 금지)
- **Senior AI**: 프로세스 전체 오케스트레이션, 상태 관리, 예외 라우팅, 복수 Agent 조율 역할 → 존재 시 항상 첫 번째로 배치
- **Junior AI**: 하나의 목적을 완수하는 순차 파이프라인을 처리하는 실무 Agent

## 🔑 Senior AI 생성 여부 판단 (매우 중요)

**Senior AI는 항상 만들지 않습니다.** 아래 3가지 질문을 평가한 뒤 **하나라도 Yes** 이면 Senior AI 를 생성하고,
**세 가지 모두 No** 이면 Senior AI 를 **생성하지 마세요** (agents 배열에서 Senior AI 항목 자체 제외).

평가 초점: **워크플로우의 구조적 복잡성**

### 📌 판단 근거 데이터 (반드시 참고)
아래 섹션들을 **먼저 읽고** Q1~Q3 을 평가하세요:
- **"As-Is 워크플로우 + 엑셀 매핑"** 섹션의 각 노드에 **`수행주체: ...`** 라인 → Q1 판단 핵심 (수행주체가 1개 vs 2+, 동일 팀 vs 서로 다른 조직)
- **`(복수 주체 — 협의 관계)`** 표기가 있는 노드 → Q1=Yes 강력 시그널
- **`[외부 업체/시스템]`** 표기 → Q1 의 (c) 외부 이해관계자에 해당
- **As-Is 엣지 구조**(decision 노드, 분기 개수) → Q2 판단 데이터
- **엑셀의 `logic.mixed` / `logic.rule` / `outputs.decision`** 필드 → Q2/Q3 판단 힌트
- **Pain Point "의사소통/협업"** 내용 → 여러 주체 간 협의 복잡성 드러나면 Q1/Q3 Yes 시그널

### Q1. 여러 이질적인 작업자 참여
"서로 다른 역량·도구·권한을 가진 2개 이상의 작업자(AI Agent 또는 인간)가 참여하며, 이들 간 **작업 위임 & 결과 통합**이 필요한가?"
- Yes: 이질적 작업자 2+ 참여, 위임·통합 필요
- No: 단일 또는 동종 Agent 로 처리 가능

**"이질적 작업자" 조작적 정의** (아래 중 하나 이상 충족):
  (a) **서로 다른 시스템 접근 권한** — 예: HR 담당자(인사 DB) + 현업 팀장(업무시스템)
  (b) **서로 다른 프로세스** 담당 — 예: 채용 담당자 + 보상 담당자 / 교육 담당자 + 평가 담당자
  (c) **조직 외부 이해관계자** — 예: 헤드헌터, 외부 평가기관, 노무사, 아웃소싱 업체

**예외**: 동일 팀 내에서 동일 시스템·동일 권한으로 업무를 분담하는 경우는 **동종 작업자** (이질적 아님).
  예: 채용팀 A 담당자와 B 담당자가 이력서 검토를 분담 → Q1 = No

### Q2. 비선형 작업 의존성
"Task 간에 **분기 / 병렬 / merge 등 단순 순차(A→B→C)를 넘는 의존 구조**가 존재하는가?"
- Yes: 분기·병렬·합류 구조 존재 (예: 조건에 따라 다른 경로, 병렬 실행 후 결과 합류)
- No: 단순 순차 흐름 (A→B→C→D)

### Q3. Cross-Task 상태 관리
"앞선 Task의 결과가 **후속 Task 의 입력·조건·경로를 결정**하여, 전체 워크플로우의 **누적 상태(state)를 추적·전달**해야 하는가?"
- Yes: 누적 상태 추적·전달 필요 (예: 단계별 승인 상태, 누적 결과 집계, 예외 발생 시 재라우팅)
- No: Task 간 독립적이거나 단순 전달만 하면 충분

### 판단 예시
- ✅ Senior AI 필요: Q1=Yes(HR 담당자 + 현업 팀장 + 외부 평가기관), Q2=Yes(분기·병렬), Q3=Yes(승인 상태 추적)
- ❌ Senior AI 불필요: Q1=No(HR 담당자만 관여), Q2=No(선형), Q3=No(단순 전달) → **agents 배열에 Senior AI 없이 Junior AI 만 나열**

### 출력 규약
- Senior AI 생성 시: agents 배열 맨 앞에 Senior AI, 뒤에 Junior AI 들 나열
- Senior AI 미생성 시: agents 배열에 Junior AI 만 나열
- 생성 여부와 판단 근거를 `design_philosophy` 필드에 한 줄로 기록:
  예: "Senior AI 생성: Q1 Yes (HR+현업+외부 3자 협업), Q2 No, Q3 Yes → 생성함"
  예: "Senior AI 미생성: Q1 No (HR 단일), Q2 No, Q3 No → 단일 Junior AI 로 충분"

## ⚠️ Task 그루핑 규칙 (핵심)
**기본 구조: L5 Task 여러 개 → 하나의 Junior AI Agent**
- assigned_tasks에 들어가는 단위는 **L5 Task**입니다
- 하나의 목적을 향해 순서대로 흘러가는 L5 Task들을 묶어 1개의 Junior AI Agent로 구성합니다
- 앞 L5 Task의 산출물이 다음 L5 Task의 입력이 되는 관계 → 반드시 같은 Agent로 묶음

**묶는 기준:**
- 같은 결과물을 만들기 위해 순서대로 실행되는 L5 Task 묶음
- 예) "지원서 수집 → 항목 추출 → 적합도 분류 → 결과 알림" → "서류 스크리닝기" 1개

**금지 사항:**
- L5 Task 1개 = Agent 1개 금지 (반드시 2개 이상 묶을 것)
- 과도한 세분화 금지 — Junior AI Agent는 **최대 5개**

**분리 기준 (새 Agent로 나누는 경우):**
- 목적이 완전히 달라지는 지점 (예: "초안 작성" 완료 후 "검토·승인·피드백 반영" 시작)
- 성격이 이질적인 Task (예: 문서 생성 계열 vs 일정 최적화 계열)

**Agent 이름 규칙:** "~기" 형태로 목적 명확화 (예: "서류 스크리닝기", "과정 설계기", "결과 보고기")

## 출력 형식 (JSON만 출력, 마크다운 코드 블록 없음)
{{
  "blueprint_summary": "상세 설계 요약 (5~7문장, Step 1 대비 개선점 포함)",
  "process_name": "{process_name}",
  "design_philosophy": "상세 설계 철학/방향 설명",
  "full_auto_count": 정수,
  "human_in_loop_count": 정수,
  "human_supervised_count": 정수,
  "agents": [
    {{
      "agent_id": "agent_0",
      "agent_name": "오케스트레이터 AI 이름",
      "agent_type": "Senior AI",
      "ai_technique": "AI 기법 (구체적)",
      "description": "역할 설명 (상세)",
      "automation_level": "Human-on-the-Loop",
      "assigned_tasks": []
    }},
    {{
      "agent_id": "agent_1",
      "agent_name": "서류 스크리닝 AI",
      "agent_type": "Junior AI",
      "ai_technique": "OCR + LLM + 분류모델",
      "description": "지원서 수집부터 적합도 분류까지 순차 파이프라인 처리",
      "automation_level": "Full-Auto",
      "assigned_tasks": [
        {{
          "task_id": "Task ID 1",
          "task_name": "지원서 항목 추출",
          "l4": "서류 전형",
          "l3": "채용관리",
          "ai_role": "OCR로 지원서 텍스트 추출 후 구조화",
          "human_role": "",
          "input_data": ["지원서 PDF"],
          "output_data": ["구조화된 지원자 데이터"],
          "automation_level": "Full-Auto",
          "ai_technique": "OCR, IDP, LLM 텍스트 추출"
        }},
        {{
          "task_id": "Task ID 2",
          "task_name": "적합도 자동 분류",
          "l4": "서류 전형",
          "l3": "채용관리",
          "ai_role": "JD 기반 지원자 적합도 점수 산정 및 등급 분류",
          "human_role": "최종 합불 검토",
          "input_data": ["구조화된 지원자 데이터", "JD 요건"],
          "output_data": ["적합도 등급, 사유"],
          "automation_level": "Human-in-Loop",
          "ai_technique": "ML 분류모델, 임베딩 매칭, 설명가능 AI(SHAP)"
        }},
        {{
          "task_id": "Task ID 3",
          "task_name": "스크리닝 결과 알림",
          "l4": "서류 전형",
          "l3": "채용관리",
          "ai_role": "합불 결과 지원자 자동 발송 및 내부 보고서 생성",
          "human_role": "",
          "input_data": ["적합도 등급"],
          "output_data": ["합불 통보 메일, 스크리닝 보고서"],
          "automation_level": "Full-Auto",
          "ai_technique": "LLM 메일 생성, BI 리포트 자동화"
        }}
      ]
    }}
  ],

## ⚠️ task_name 작성 규칙 (매우 중요)

**형식**: 짧은 명사구로 10~20자 이내. 동사+목적어 or 명사구. 마침표/문장 금지.

**✅ 올바른 예시**:
- `"DBS 점수 추출"` / `"지원자 데이터 정리"` / `"에세이 검토"` / `"발령품의 진행"`
- `"서류 합격자 확정"` / `"시스템 발령 처리"` / `"Local Job 신설"`

**❌ 금지 예시** (이렇게 쓰지 말 것):
- `"지주와 직접 협의하여 서류합격자 검토 결과를 조율하고 최종 합·불을 확정·승인한다"` — 이건 문장, task_name 아님
- `"확정된 조직개편 결과의 시스템 반영은 AI가 처리하되..."` — 문장
- `"~한다"`, `"~수행한다"`, `"~처리한다"` 로 끝나는 완전 문장 금지

**규칙**:
- task_name: 10~20자 짧은 명사구 or 동사+목적어
- 긴 action 문장·방법·결과 설명은 **ai_role / human_role / description 필드**에 넣을 것
- task_name 은 Swim Lane 뱃지·PPT 상자에 표시되므로 짧고 명확해야 함

## ⚠️ task_id 형식 규칙

- 기존 As-Is L5 를 대체하는 경우: **기존 task_id 그대로 사용** (예: `"2.1.4.3"`)
- 신규 추가 Task 의 경우: 해당 L4 번호 뒤에 순번 (예: `"2.1.4.16"`, `"2.1.4.17"`)
  - L4 가 `2.1.4` 이고 기존 최대 L5 가 `2.1.4.15` 이면 신규는 `2.1.4.16` 부터
- `"NEW_xxx"` 같은 임시 ID 지양 — 가능한 한 실제 L5 숫자 형식으로 작성

⚠️ **task별 ai_technique 필드 필수 규칙**

**형식**: 짧은 기술명만 3~5개 쉼표로 구분. 절대 task 설명·문장·괄호 안 콤마 포함 금지.

**허용되는 값** (예시, 이 중에서 선택하거나 유사 기술명으로):
`LLM` / `RAG` / `OCR` / `IDP` / `NLP 분류` / `ML 회귀` / `ML 분류모델` / `XGBoost` / `시계열 예측`
`SHAP` / `임베딩 매칭` / `RPA` / `ETL` / `Rule-based` / `Chatbot` / `BI 대시보드` / `전자결재 API`
`GenAI 문서 생성` / `프로세스 마이닝` / `최적화 알고리즘` / `이상 탐지` / `Template` / `Speech-to-Text`

**❌ 금지 예시** (이런 식으로 쓰지 말 것):
- `"GenAI 다국어 개인화 문서 생성(한국어, 영어 동시 생성)"` — 너무 길고 괄호 안 쉼표가 파싱 오류 유발
- `"결재 승인 완료 신호를 트리거로 GenAI가..."` — 이건 ai_role에 쓸 내용, ai_technique 아님
- `"공지 게시 자동화(HR 포털, 사내 메신저 연동)"` — 설명이지 기술명 아님

**✅ 올바른 예시**:
- `"GenAI, 다국어 생성, 메일 자동화 API"`
- `"RPA, ETL, 규칙 엔진"`
- `"LLM, RAG, 설명가능 AI"`

**규칙**:
- 각 기술명은 **15자 이내** 짧게
- 괄호와 괄호 안 쉼표 사용 금지 (파싱 시 뱃지가 중간에 쪼개짐)
- 같은 Agent 안이라도 task마다 다른 기법 조합 (OCR 파트 / 분류 파트 / 생성 파트 각각 다름)
- 비워두면 Agent의 ai_technique이 fallback — 모든 task가 동일해지므로 반드시 task별로 작성
  "execution_flow": [
    {{
      "step": 1,
      "step_name": "단계명",
      "step_type": "sequential",
      "description": "상세 설명",
      "agent_ids": ["agent_1"],
      "task_ids": ["Task ID"]
    }}
  ]
}}
"""

    settings = load_settings()
    anthropic_key = os.getenv("ANTHROPIC_API_KEY", "") or settings.anthropic_api_key

    result_data = None

    from usage_store import add_usage as _add_usage_step2

    if anthropic_key:
        try:
            from anthropic import AsyncAnthropic
            client = AsyncAnthropic(api_key=anthropic_key)
            response = await client.messages.create(
                model=settings.anthropic_model or "claude-sonnet-4-6",
                max_tokens=16384,   # 이전 8192에서 상향 — agents/tasks 많을 때 JSON 잘림 방지
                system=step2_system,
                messages=[{"role": "user", "content": "Step 1 결과와 Pain Point를 기반으로 상세 설계를 수행해주세요."}],
            )
            if response.usage:
                _add_usage_step2("anthropic",
                                 input_tokens=response.usage.input_tokens,
                                 output_tokens=response.usage.output_tokens)
            # stop_reason 체크 — max_tokens 초과 시 경고
            stop_reason = getattr(response, "stop_reason", None)
            if stop_reason and stop_reason != "end_turn":
                print(f"[workflow-step2] ⚠️ Anthropic stop_reason={stop_reason} "
                      f"(max_tokens=16384, output_tokens={response.usage.output_tokens if response.usage else '?'}). "
                      f"JSON 잘림 가능 — 프롬프트 크기 또는 max_tokens 추가 조정 필요.", flush=True)
            raw = response.content[0].text
            from new_workflow_generator import _extract_json
            result_data = _extract_json(raw)
            if result_data is None:
                # JSON 파싱 실패 — 진단 로그
                print(f"[workflow-step2] ⚠️ JSON 파싱 실패. "
                      f"output_len={len(raw)}, 마지막 300자: ...{raw[-300:]}", flush=True)
        except Exception as e:
            print(f"[workflow-step2] Anthropic 실패: {e}", flush=True)

    if not result_data:
        openai_key = os.getenv("OPENAI_API_KEY", "") or settings.api_key
        if openai_key:
            try:
                from openai import AsyncOpenAI
                client = AsyncOpenAI(api_key=openai_key)
                response = await client.chat.completions.create(
                    model=settings.model or "gpt-5.4",
                    temperature=0.0,
                    messages=[
                        {"role": "system", "content": step2_system},
                        {"role": "user", "content": "Step 1 결과와 Pain Point를 기반으로 상세 설계를 수행해주세요."},
                    ],
                    response_format={"type": "json_object"},
                )
                if response.usage:
                    _add_usage_step2("openai",
                                     input_tokens=response.usage.prompt_tokens,
                                     output_tokens=response.usage.completion_tokens)
                raw = response.choices[0].message.content or "{}"
                result_data = json.loads(raw)
            except Exception as e:
                print(f"[workflow-step2] OpenAI 실패: {e}")

    if not result_data:
        raise HTTPException(500, "AI 모델 호출에 실패했습니다. API 키를 확인하세요.")

    from new_workflow_generator import _parse_freeform_result, result_to_dict
    parsed = _parse_freeform_result(result_data)
    result_dict = result_to_dict(parsed)
    result_dict["design_philosophy"] = result_data.get("design_philosophy", "")

    # ── task별 ai_technique 준수 여부 진단 로그 ──
    # LLM 이 프롬프트 지침을 무시하고 task.ai_technique 을 비우면 frontend/html 이
    # agent.ai_technique 으로 fallback 해 모든 task 가 동일한 기술 스택으로 보임
    _task_tech_stats = {"total": 0, "has_task_tech": 0, "agent_tech_only": 0}
    for _a in result_dict.get("agents", []):
        _agent_tech = (_a.get("ai_technique") or "").strip()
        for _t in _a.get("assigned_tasks", []):
            _task_tech_stats["total"] += 1
            _task_tech = (_t.get("ai_technique") or "").strip()
            if _task_tech and _task_tech != _agent_tech:
                _task_tech_stats["has_task_tech"] += 1
            else:
                _task_tech_stats["agent_tech_only"] += 1
    print(f"[STEP2] task별 ai_technique 준수: "
          f"{_task_tech_stats['has_task_tech']}/{_task_tech_stats['total']} task 고유 | "
          f"{_task_tech_stats['agent_tech_only']}개 task 는 agent fallback "
          f"(agent와 같거나 빈 값)", flush=True)

    # Pain Point 가시성: Step 2 설계에 실제 반영된 Pain Point + 분류 맥락을 결과에 저장
    pain_context_items = []
    for _t in scoped_tasks_step2:
        _pains = []
        if _t.pain_time:          _pains.append({"type": "시간/속도",    "text": _t.pain_time})
        if _t.pain_accuracy:      _pains.append({"type": "정확성",      "text": _t.pain_accuracy})
        if _t.pain_repetition:    _pains.append({"type": "반복/수작업",  "text": _t.pain_repetition})
        if _t.pain_data:          _pains.append({"type": "정보/데이터", "text": _t.pain_data})
        if _t.pain_system:        _pains.append({"type": "시스템/도구", "text": _t.pain_system})
        if _t.pain_communication: _pains.append({"type": "의사소통",    "text": _t.pain_communication})
        _cls = _wf_classification.get(_t.id, {})
        if _pains or _cls.get("label"):
            pain_context_items.append({
                "task_id": _t.id,
                "task_name": _t.name,
                "l4": _t.l4,
                "l3": _t.l3,
                "classification": _cls.get("label", ""),
                "classification_reason": _cls.get("reason", ""),
                "hybrid_note": _cls.get("hybrid_note", ""),
                "ai_prerequisites": _cls.get("ai_prerequisites", ""),
                "pain_points": _pains,
            })
    result_dict["pain_context"] = pain_context_items
    result_dict["classification_stats"] = cls_stats

    global _wf_step2_cache
    _wf_step2_cache = result_dict
    if _current_session_id:
        _save_session_data(_current_session_id)

    # New Workflow 캐시에도 반영 (과제 정의서/설계서 생성에 활용)
    _new_workflow_cache.clear()
    _new_workflow_cache.update(result_dict)
    _save_nw_state("new_workflow", _new_workflow_cache)

    return {"ok": True, **result_dict}


@app.get("/api/workflow/step-results", tags=["Workflow"])
async def get_workflow_step_results():
    """Step 1 / Step 2 결과를 반환합니다. 벤치마킹·Gap 분석 결과도 포함.

    AI Service Flow 는 프론트엔드 WorkflowEditor 가 step2.agents[].assigned_tasks 를
    기반으로 그리므로, Step 1 에서 '삭제'/'통합' 으로 처리된 task 는 여기서 제거하여
    Swim Lane · AI Service Flow 양쪽에서 일관되게 빠지게 한다.
    """
    filtered_step2 = _filter_step2_deprecated(_wf_step2_cache) if _wf_step2_cache else None
    return {
        "ok": True,
        "has_excel": len(_wf_excel_tasks) > 0,
        "has_asis": "parsed" in _workflow_cache,
        "has_step1": bool(_wf_step1_cache),
        "has_step2": bool(_wf_step2_cache),
        "step1": _wf_step1_cache if _wf_step1_cache else None,
        "step2": filtered_step2,
        "chat_history": _wf_chat_history,
        # 벤치마킹·Gap 분석 복원용
        "benchmark_table": _wf_benchmark_table,
        "gap_analysis": _wf_gap_analysis if _wf_gap_analysis else None,
    }


# ─────────────────────────────────────────────────────────────────────────────
# To-Be Workflow 생성 (기존)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/workflow/export-tobe-json", tags=["Workflow"])
async def export_tobe_workflow_json():
    """
    Step 1 기본 설계 결과를 hr-workflow-ai 호환 JSON으로 내보냅니다.
    Step 2 결과가 있으면 Step 2를, 없으면 Step 1을 사용합니다.
    """
    from new_workflow_generator import (
        result_to_hr_workflow_json,
        NewWorkflowResult,
        AIAgent,
        AssignedTask,
        ExecutionStep as NWExecutionStep,
    )
    import io

    cache = _wf_step2_cache if _wf_step2_cache else _wf_step1_cache
    if not cache:
        raise HTTPException(404, "Step 1 또는 Step 2를 먼저 실행하세요.")

    # Step 1 에서 '삭제'/'통합' 처리된 task 는 export 에서도 제외 (Swim Lane 과 동일)
    if _wf_step2_cache:
        cache = _filter_step2_deprecated(cache)

    agents = []
    for a in cache.get("agents", []):
        assigned = [
            AssignedTask(
                task_id=t.get("task_id", ""),
                task_name=t.get("task_name", ""),
                l4=t.get("l4", ""),
                l3=t.get("l3", ""),
                ai_role=t.get("ai_role", ""),
                human_role=t.get("human_role", ""),
                input_data=t.get("input_data", []),
                output_data=t.get("output_data", []),
                automation_level=t.get("automation_level", "Human-on-the-Loop"),
            )
            for t in a.get("assigned_tasks", [])
        ]
        agents.append(AIAgent(
            agent_id=a.get("agent_id", ""),
            agent_name=a.get("agent_name", ""),
            agent_type=a.get("agent_type", ""),
            ai_technique=a.get("ai_technique", ""),
            description=a.get("description", ""),
            automation_level=a.get("automation_level", "Human-on-the-Loop"),
            assigned_tasks=assigned,
        ))

    flow = [
        NWExecutionStep(
            step=s.get("step", i + 1),
            step_name=s.get("step_name", ""),
            step_type=s.get("step_type", "sequential"),
            description=s.get("description", ""),
            agent_ids=s.get("agent_ids", []),
            task_ids=s.get("task_ids", []),
        )
        for i, s in enumerate(cache.get("execution_flow", []))
    ]

    result = NewWorkflowResult(
        blueprint_summary=cache.get("blueprint_summary", ""),
        process_name=cache.get("process_name", "To-Be Workflow"),
        total_tasks=sum(len(a.assigned_tasks) for a in agents),
        full_auto_count=cache.get("full_auto_count", 0),
        human_in_loop_count=cache.get("human_in_loop_count", 0),
        human_supervised_count=cache.get("human_supervised_count", 0),
        agents=agents,
        execution_flow=flow,
    )

    hr_json = result_to_hr_workflow_json(result)
    step_label = "Step2" if _wf_step2_cache else "Step1"
    filename = f"{result.process_name or 'tobe_workflow'}_{step_label}.json"

    from urllib.parse import quote
    encoded_fn = quote(filename)
    return StreamingResponse(
        io.BytesIO(json.dumps(hr_json, ensure_ascii=False, indent=2).encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_fn}"},
    )


@app.post("/api/workflow/slide-l4-mapping", tags=["Workflow"])
async def save_slide_l4_mapping(request: Request):
    """
    슬라이드별 L4 매핑을 저장합니다.
    Body: { "mappings": { "0": "1.1.1", "1": "1.1.2", ... } }
    key = 슬라이드 인덱스, value = L4 ID
    """
    body = await request.json()
    mappings = body.get("mappings", {})

    global _workflow_cache
    _workflow_cache["slide_l4_mapping"] = mappings

    # PPT 결과가 있으면 매핑 반영하여 ParsedWorkflow 재생성
    if "ppt" in _workflow_cache:
        from ppt_parser import ppt_to_parsed_workflow
        ppt_data = _workflow_cache["ppt"]
        parsed_ppt = ppt_data["parsed"]

        # 매칭 정보 재구성
        slides_result = ppt_data["slides"]
        all_matches = [s.get("matches", []) for s in slides_result]

        ppt_workflow = ppt_to_parsed_workflow(
            parsed_ppt, all_matches, slide_l4_mapping=mappings
        )
        _workflow_cache["parsed"] = ppt_workflow

    return {"ok": True, "mappings": mappings}


def _word_jaccard(a: str, b: str) -> float:
    """두 한국어 문자열의 단어 단위 Jaccard 유사도 (0~1)."""
    a_words = set(a.split())
    b_words = set(b.split())
    if not a_words or not b_words:
        return 0.0
    inter = len(a_words & b_words)
    union = len(a_words | b_words)
    return inter / union if union else 0.0


def _fuzzy_match_excel(label: str, candidates: list, already_matched: set,
                       threshold: float = 0.4) -> tuple | None:
    """
    label(JSON L5 노드명)과 가장 유사한 엑셀 Task를 candidates 안에서 찾아 반환.
    - 이미 정확 매칭된 Task(already_matched)는 제외
    - Jaccard 유사도가 threshold 이상인 최고점 Task 반환
    - 없으면 None 반환. 반환값: (Task, score)
    """
    best, best_score = None, 0.0
    for t in candidates:
        if t.id in already_matched:
            continue
        score = _word_jaccard(label, t.name)
        if score > best_score:
            best, best_score = t, score
    if best and best_score >= threshold:
        return best, best_score
    return None


@app.get("/api/workflow/manual-matches", tags=["Workflow"])
async def get_manual_matches():
    """저장된 수동 매칭 목록 반환."""
    return {"ok": True, "matches": _manual_matches}


@app.post("/api/workflow/manual-match", tags=["Workflow"])
async def set_manual_match(request: Request):
    """JSON L5 task_id → 엑셀 task_id 수동 연결 저장."""
    body = await request.json()
    json_task_id = body.get("json_task_id", "").strip()
    excel_task_id = body.get("excel_task_id", "").strip()
    if not json_task_id or not excel_task_id:
        raise HTTPException(400, "json_task_id와 excel_task_id가 필요합니다.")
    _manual_matches[json_task_id] = excel_task_id
    _save_manual_matches()
    return {"ok": True, "json_task_id": json_task_id, "excel_task_id": excel_task_id}


@app.delete("/api/workflow/manual-match/{json_task_id:path}", tags=["Workflow"])
async def delete_manual_match(json_task_id: str):
    """수동 매칭 삭제."""
    if json_task_id in _manual_matches:
        del _manual_matches[json_task_id]
        _save_manual_matches()
    return {"ok": True}


@app.get("/api/workflow/mapping-check", tags=["Workflow"])
async def get_mapping_check():
    """엑셀 Task ↔ As-Is 워크플로우 노드 매핑 현황을 반환합니다."""
    try:
        return _run_mapping_check()
    except Exception as exc:
        print(f"[ERROR] mapping-check 실패: {exc}", flush=True)
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"매핑 분석 오류: {exc}")


def _run_mapping_check() -> dict:
    has_excel = bool(_wf_excel_tasks)
    has_asis = "parsed" in _workflow_cache

    if not has_excel and not has_asis:
        return {"ok": True, "has_excel": False, "has_asis": False, "sheets": [], "excel_only": [], "stats": {}}

    by_id, by_l4, by_l3, _ = _build_excel_index()
    matched_excel_ids: set[str] = set()
    sheets_result = []

    if has_asis:
        parsed = _workflow_cache["parsed"]
        for s in parsed.sheets:
            l3_nodes = [n for n in s.nodes.values() if n.level == "L3"]
            l4_nodes = s.l4_nodes
            l5_nodes = [n for n in s.nodes.values() if n.level == "L5"]

            # L5 → 부모 L4 그룹핑: 엑셀 매칭 시 엑셀의 l4_id 사용, 미매칭은 prefix fallback
            l5_by_l4: dict[str, list] = {}
            for l5n in l5_nodes:
                if not l5n.task_id:
                    continue
                et = by_id.get(l5n.task_id)
                if et and et.l4_id:
                    # 엑셀 계층 정보 사용 (정확한 L4 배정)
                    l5_by_l4.setdefault(et.l4_id, []).append(l5n)
                else:
                    # 엑셀 매칭 없으면 task_id prefix fallback
                    parts = l5n.task_id.rsplit(".", 1)
                    parent = parts[0] if len(parts) > 1 else l5n.task_id
                    l5_by_l4.setdefault(parent, []).append(l5n)

            # L4 유효 맵: 명시 L4 노드 + 엑셀 l4_id 기반 파생
            eff_l4: dict[str, str] = {}
            for l4n in l4_nodes:
                if l4n.task_id:
                    eff_l4[l4n.task_id] = l4n.label
            for l4_tid in l5_by_l4:
                if l4_tid not in eff_l4:
                    xl = by_l4.get(l4_tid, [])
                    eff_l4[l4_tid] = xl[0].l4 if xl else l4_tid

            # L3 유효 맵: 엑셀 l3_id 기반으로 L4 → L3 배정
            l3_node_map: dict[str, str] = {n.task_id: n.label for n in l3_nodes if n.task_id}
            eff_l3: dict[str, list] = {}  # l3_tid → [l4_tid, ...]
            # 엑셀에서 l4_id → l3_id 매핑 구성
            l4_to_l3_id: dict[str, str] = {}
            for t in _wf_excel_tasks:
                if t.l4_id and t.l3_id and t.l4_id not in l4_to_l3_id:
                    l4_to_l3_id[t.l4_id] = t.l3_id
            for l4_tid in sorted(eff_l4):
                l3_tid = l4_to_l3_id.get(l4_tid)
                if not l3_tid:
                    # 엑셀에 없는 L4면 prefix fallback
                    parts = l4_tid.rsplit(".", 1)
                    l3_tid = parts[0] if len(parts) > 1 else l4_tid
                eff_l3.setdefault(l3_tid, []).append(l4_tid)

            # L4 엔트리 생성 (closure 피하고 인자로 전달)
            def _l4_entry(l4_tid, l4_label, _l5_by_l4=l5_by_l4, _by_id=by_id, _by_l4=by_l4,
                          _manual=_manual_matches):
                children = sorted(_l5_by_l4.get(l4_tid, []), key=lambda x: x.task_id)
                # 이 L4 범위 안의 엑셀 Task 후보 (퍼지 매칭 스코프 제한)
                l4_candidates = _by_l4.get(l4_tid, [])
                entries, cls_counts = [], {}
                exact_matched_ids: set[str] = set()

                # 1차 패스: 정확 매칭 + 수동 매칭 미리 수집 (중복 방지용)
                for l5n in children:
                    et = _by_id.get(l5n.task_id)
                    if et:
                        exact_matched_ids.add(et.id)
                    elif l5n.task_id in _manual:
                        exact_matched_ids.add(_manual[l5n.task_id])

                # 2차 패스: 정확 → 수동 → 퍼지 순서로 매칭
                for l5n in children:
                    et = _by_id.get(l5n.task_id)
                    fuzzy_matched = False
                    manual_matched = False
                    fuzzy_score = 0.0

                    if not et:
                        # 수동 매칭
                        manual_excel_id = _manual.get(l5n.task_id)
                        if manual_excel_id:
                            et = _by_id.get(manual_excel_id)
                            if et:
                                manual_matched = True
                        # 퍼지 매칭 (수동도 없을 때)
                        if not et and l5n.label and l4_candidates:
                            result = _fuzzy_match_excel(l5n.label, l4_candidates, exact_matched_ids)
                            if result:
                                et, fuzzy_score = result
                                fuzzy_matched = True
                                exact_matched_ids.add(et.id)

                    if et:
                        matched_excel_ids.add(et.id)
                        lbl = _wf_classification.get(et.id, {}).get("label", "미분류")
                        cls_counts[lbl] = cls_counts.get(lbl, 0) + 1
                        pains = [p for p, v in [
                            ("시간/속도", et.pain_time), ("정확성", et.pain_accuracy),
                            ("반복/수작업", et.pain_repetition), ("정보/데이터", et.pain_data),
                            ("시스템/도구", et.pain_system),
                        ] if v]
                        entries.append({"task_id": l5n.task_id, "label": l5n.label,
                            "matched": True, "fuzzy_matched": fuzzy_matched,
                            "manual_matched": manual_matched,
                            "fuzzy_score": round(fuzzy_score, 2),
                            "excel_id": et.id, "excel_name": et.name,
                            "cls_label": lbl, "pain_points": pains,
                            "description": (et.description or "")[:80]})
                    else:
                        entries.append({"task_id": l5n.task_id, "label": l5n.label,
                            "matched": False, "fuzzy_matched": False,
                            "manual_matched": False, "fuzzy_score": 0.0,
                            "excel_id": "", "excel_name": "",
                            "cls_label": "", "pain_points": [], "description": ""})
                return {"task_id": l4_tid, "label": l4_label, "level": "L4",
                        "l5_nodes": entries, "cls_summary": cls_counts,
                        "matched_l5": sum(1 for e in entries if e["matched"]),
                        "total_l5": len(entries)}

            l3_groups = []
            for l3_tid in sorted(eff_l3):
                l4_list = [_l4_entry(tid, eff_l4[tid]) for tid in sorted(eff_l3[l3_tid])]
                xl3 = by_l3.get(l3_tid, [])
                l3_label = l3_node_map.get(l3_tid) or (xl3[0].l3 if xl3 else l3_tid)
                l3_groups.append({"task_id": l3_tid, "label": l3_label,
                    "l4_nodes": l4_list,
                    "total_l5": sum(x["total_l5"] for x in l4_list),
                    "matched_l5": sum(x["matched_l5"] for x in l4_list)})

            sheets_result.append({"sheet_id": s.sheet_id, "sheet_name": s.name,
                "l3_count": len(l3_nodes), "l4_count": len(eff_l4),
                "l5_count": len(l5_nodes), "l3_groups": l3_groups})

    # 엑셀 전용 (매핑 안 된 Task)
    excel_only = []
    if has_excel:
        for t in _wf_excel_tasks:
            if t.id not in matched_excel_ids:
                excel_only.append({
                    "id": t.id,
                    "name": t.name,
                    "l2": t.l2, "l2_id": t.l2_id,
                    "l3": t.l3, "l3_id": t.l3_id,
                    "l4": t.l4, "l4_id": t.l4_id,
                    "label": _wf_classification.get(t.id, {}).get("label", "미분류"),
                })

    # 통계 — 분모: JSON As-Is L5 노드 수 (엑셀 기준 아님)
    total_l5_asis = sum(s["l5_count"] for s in sheets_result)
    matched_l5_asis = sum(
        sum(n["matched_l5"] for n in g["l4_nodes"])
        for s in sheets_result
        for g in s["l3_groups"]
    )
    # L5 자식이 없는 phantom L4 (다른 sheet의 L4가 헤더로만 참조된 경우)는 통계에서 제외
    total_l4 = sum(
        sum(1 for n in g["l4_nodes"] if n["total_l5"] > 0)
        for s in sheets_result
        for g in s["l3_groups"]
    )
    matched_l4 = sum(
        sum(1 for n in g["l4_nodes"] if n["matched_l5"] > 0)
        for s in sheets_result
        for g in s["l3_groups"]
    )

    # 연결된 L5 노드 기준 분류별 집계
    cls_matched: dict[str, int] = {}
    for sid in matched_excel_ids:
        lbl = _wf_classification.get(sid, {}).get("label", "미분류")
        cls_matched[lbl] = cls_matched.get(lbl, 0) + 1

    # 전체 엑셀 분류별 집계 (참고용)
    cls_total: dict[str, int] = {}
    for t in _wf_excel_tasks:
        lbl = _wf_classification.get(t.id, {}).get("label", "미분류")
        cls_total[lbl] = cls_total.get(lbl, 0) + 1

    # L4 노드별 분류 집계
    l4_cls_stats = []
    for s in sheets_result:
        for g in s["l3_groups"]:
            for n in g["l4_nodes"]:
                if n["matched_l5"] > 0:
                    l4_cls_stats.append({
                        "task_id": n["task_id"],
                        "label": n["label"],
                        "cls_summary": n.get("cls_summary", {}),
                    })

    return {
        "ok": True,
        "has_excel": has_excel,
        "has_asis": has_asis,
        "stats": {
            # JSON L5 기준 (연결률 분모)
            "total_l5_nodes": total_l5_asis,
            "matched_l5_nodes": matched_l5_asis,
            "unmatched_l5_nodes": total_l5_asis - matched_l5_asis,
            "match_rate": round(matched_l5_asis / total_l5_asis * 100, 1) if total_l5_asis > 0 else 0,
            # L4 기준
            "total_l4_nodes": total_l4,
            "matched_l4_nodes": matched_l4,
            "unmatched_l4_nodes": total_l4 - matched_l4,
            # 엑셀 기준 (참고용)
            "total_excel_tasks": len(_wf_excel_tasks),
            "matched_excel_tasks": len(matched_excel_ids),
            "unmatched_excel_tasks": len(_wf_excel_tasks) - len(matched_excel_ids),
            # 분류별 집계
            "cls_matched": cls_matched,
            "cls_total": cls_total,
        },
        "sheets": sheets_result,
        "excel_only": excel_only,
        "l4_cls_stats": l4_cls_stats,
    }


@app.get("/api/workflow/debug-ids", tags=["Workflow"])
async def debug_id_mapping():
    """엑셀 Task ID와 As-Is L5 task_id 샘플을 비교하여 매핑 문제를 진단합니다."""
    excel_samples = [
        {"id": t.id, "l4_id": t.l4_id, "name": t.name[:30]}
        for t in _wf_excel_tasks[:10]
    ]
    asis_l5_samples = []
    persist_root = str(_PERSIST_ROOT)
    wf_dir = str(_WF_DIR)
    if "parsed" in _workflow_cache:
        parsed = _workflow_cache["parsed"]
        for s in parsed.sheets:
            for n in list(s.nodes.values()):
                if n.level == "L5":
                    asis_l5_samples.append({
                        "node_id": n.id,
                        "task_id": n.task_id,
                        "label": n.label[:30],
                        "sheet": s.name,
                    })
                    if len(asis_l5_samples) >= 10:
                        break
            if len(asis_l5_samples) >= 10:
                break

    by_id, _, _, _ = _build_excel_index()
    match_test = []
    for s in asis_l5_samples[:5]:
        match_test.append({
            "l5_task_id": s["task_id"],
            "found_in_excel": s["task_id"] in by_id,
            "excel_id_sample": s["task_id"],
        })

    return {
        "persist_root": persist_root,
        "wf_dir": wf_dir,
        "persist_exists": Path(persist_root).exists(),
        "wf_dir_files": [f.name for f in Path(wf_dir).iterdir() if f.is_file()] if Path(wf_dir).exists() else [],
        "excel_task_count": len(_wf_excel_tasks),
        "asis_loaded": "parsed" in _workflow_cache,
        "excel_samples": excel_samples,
        "asis_l5_samples": asis_l5_samples,
        "match_test": match_test,
    }


@app.post("/api/workflow/generate-tobe", tags=["Workflow"])
async def generate_tobe_workflow(
    sheet_id: str = Query("default", description="As-Is 워크플로우 시트 ID"),
    provider: str = Query("openai", description="분류 결과 제공자 (openai | anthropic)"),
    process_name: str = Query("", description="프로세스 이름 (빈 값이면 시트 이름 사용)"),
):
    """As-Is 워크플로우 + 분류 결과 → To-Be Workflow 초안을 생성합니다 (Claude 기반)."""
    from workflow_parser import parse_workflow_json, get_workflow_summary
    from tobe_generator import generate_tobe, generate_tobe_with_llm

    # 워크플로우 로드
    global _workflow_cache
    if "parsed" not in _workflow_cache:
        save_path = (
            _get_session_dir(_current_session_id) / "workflow.json"
            if _current_session_id else None
        )
        if save_path and save_path.exists():
            data = json.loads(save_path.read_text(encoding="utf-8"))
            parsed = parse_workflow_json(data)
            _workflow_cache["parsed"] = parsed
            _workflow_cache["summary"] = get_workflow_summary(parsed)
            _workflow_cache["raw"] = data
        else:
            raise HTTPException(404, "업로드된 워크플로우가 없습니다. 먼저 워크플로우를 업로드하세요.")

    parsed = _workflow_cache["parsed"]
    sheet = next((s for s in parsed.sheets if s.sheet_id == sheet_id), None)
    if not sheet:
        if parsed.sheets:
            sheet = parsed.sheets[0]
        else:
            raise HTTPException(404, "워크플로우에 시트가 없습니다.")

    # 분류 결과 로드
    results_store = load_results(provider)
    if not results_store:
        raise HTTPException(
            400,
            f"'{provider}' 분류 결과가 없습니다. 먼저 분류를 실행하세요.",
        )

    # 슬라이드-L4 매핑이 있으면 해당 L4의 L5 태스크만 필터링
    slide_l4_mapping = _workflow_cache.get("slide_l4_mapping", {})
    # sheet_id가 "ppt-slide-N" 형식이면 슬라이드 인덱스 추출
    mapped_l4_id = ""
    if sheet_id.startswith("ppt-slide-"):
        slide_idx = sheet_id.replace("ppt-slide-", "")
        mapped_l4_id = slide_l4_mapping.get(slide_idx, "")

    # ClassificationResult → dict 변환
    classification_dict: dict[str, dict] = {}
    for tid, cr in results_store.items():
        task = next((t for t in _tasks_cache if t.id == tid), None)

        # L4 필터: 매핑이 설정되어 있으면 해당 L4의 태스크만 포함
        if mapped_l4_id and task:
            if task.l4_id != mapped_l4_id:
                continue

        classification_dict[tid] = {
            "label": cr.label,
            "reason": cr.reason,
            "hybrid_note": cr.hybrid_note,
            "input_types": cr.input_types,
            "output_types": cr.output_types,
            "task_name": task.name if task else "",
        }

    # To-Be 생성 (Claude LLM 기반)
    tobe = await generate_tobe_with_llm(
        as_is_sheet=sheet,
        classification_results=classification_dict,
        process_name=process_name or sheet.name,
    )

    return {
        "ok": True,
        "summary": tobe.summary,
        "execution_steps": tobe.execution_steps,
        "react_flow": tobe.react_flow,
    }


# ─────────────────────────────────────────────────────────────────────────────
# New Workflow 엔드포인트 (엑셀 → AI 워크플로우 설계 초안)
# ─────────────────────────────────────────────────────────────────────────────

# New Workflow 전용 캐시 (독립 운영) + 자동 복원
_nw_tasks_cache: list[Task] = []
_nw_projects_cache: list[dict] = []  # 과제 엑셀 형식일 때 사용
_nw_excel_path: Path | None = None
_new_workflow_cache: dict = {}
_new_workflow_cache.update(_load_nw_state("new_workflow"))


@app.post("/api/new-workflow/upload", tags=["NewWorkflow"])
async def upload_new_workflow_excel(file: UploadFile = File(...)):
    """
    New Workflow 전용 엑셀 업로드.
    과제 엑셀 형식 (2행 병합 헤더) 자동 감지 → 과제 데이터로 파싱.
    L5 Task 형식이면 기존 방식으로 파싱.
    """
    global _nw_tasks_cache, _nw_excel_path
    from excel_reader import list_sheets
    from project_excel_reader import parse_project_excel, projects_to_freeform_params

    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(400, ".xlsx 파일만 업로드 가능합니다.")

    safe_fn = Path(file.filename).name
    save_path = _UPLOAD_DIR / f"nw_{safe_fn}"
    try:
        contents = await file.read()
        save_path.write_bytes(contents)
    except Exception as e:
        raise HTTPException(500, f"파일 저장 실패: {e}")

    _nw_excel_path = save_path
    set_current_project(file.filename)

    # 1차: 과제 엑셀 형식 시도 (2행 병합 헤더)
    projects = []
    try:
        projects = parse_project_excel(save_path)
    except Exception:
        pass

    if projects:
        # 과제 형식 감지됨 — 프로젝트 데이터를 별도 저장
        _nw_projects_cache.clear()
        _nw_projects_cache.extend(projects)
        _nw_tasks_cache = []  # L5 Task는 없음

        save_meta(file.filename, task_count=0, project_count=len(projects), source="new_workflow", format="project")

        return {
            "message": "과제 엑셀 업로드 성공",
            "filename": file.filename,
            "format": "project",
            "project_count": len(projects),
            "task_count": 0,
            "projects": [{"no": p.get("project_no", ""), "name": p.get("name", "")} for p in projects],
            "sheets": [],
        }

    # 2차: L5 Task 형식 시도
    try:
        sheets = list_sheets(save_path)
    except Exception as e:
        save_path.unlink(missing_ok=True)
        print(f"[ERROR] 엑셀 파싱 실패: {e}")
        raise HTTPException(422, "엑셀 파일을 읽을 수 없습니다.")

    recommended = next((s for s in sheets if s["recommended"]), None)
    if recommended:
        try:
            _nw_tasks_cache = load_tasks(save_path, sheet_name=recommended["name"])
        except Exception:
            pass

    save_meta(file.filename, task_count=len(_nw_tasks_cache), source="new_workflow", format="l5_tasks")

    return {
        "message": "업로드 성공",
        "filename": file.filename,
        "format": "l5_tasks",
        "project_count": 0,
        "task_count": len(_nw_tasks_cache),
        "sheets": sheets,
    }


@app.post("/api/new-workflow/select-sheet", tags=["NewWorkflow"])
async def select_new_workflow_sheet(request: Request):
    """New Workflow 전용 시트 선택."""
    global _nw_tasks_cache

    body = await request.json()
    sheet_name = body.get("sheet_name", "")

    if not _nw_excel_path or not _nw_excel_path.exists():
        raise HTTPException(404, "업로드된 엑셀 파일이 없습니다.")

    try:
        _nw_tasks_cache = load_tasks(_nw_excel_path, sheet_name=sheet_name)
    except Exception as e:
        print(f"[ERROR] 시트 파싱 실패: {e}")
        raise HTTPException(422, "시트를 읽을 수 없습니다.")

    return {
        "message": f"시트 '{sheet_name}' 로드 완료",
        "sheet_name": sheet_name,
        "task_count": len(_nw_tasks_cache),
    }


@app.get("/api/new-workflow/tasks", tags=["NewWorkflow"])
async def get_new_workflow_tasks():
    """New Workflow에 로드된 Task 목록 반환."""
    if not _nw_tasks_cache:
        return {"total": 0, "tasks": []}
    tasks = [
        {
            "id": t.id, "l2": t.l2, "l3": t.l3, "l3_id": t.l3_id,
            "l4": t.l4, "l4_id": t.l4_id, "name": t.name,
            "description": t.description, "performer": t.performer,
        }
        for t in _nw_tasks_cache
    ]
    return {"total": len(tasks), "tasks": tasks}


@app.get("/api/new-workflow/filters", tags=["NewWorkflow"])
async def get_new_workflow_filters():
    """New Workflow Task에서 L3 필터 옵션 반환."""
    l3_set: dict[str, str] = {}
    for t in _nw_tasks_cache:
        if t.l3 and t.l3_id:
            l3_set[t.l3_id] = t.l3
    return {
        "l3_options": [{"id": k, "name": v} for k, v in sorted(l3_set.items(), key=lambda x: _natural_key(x[0]))],
    }


@app.post("/api/new-workflow/generate", tags=["NewWorkflow"])
async def generate_new_workflow(
    process_name: str = Query("", description="프로세스 이름 (비우면 자동 추론)"),
    project_index: int = Query(-1, description="과제 엑셀에서 선택된 과제 인덱스 (0부터)"),
    l3: Optional[str] = Query(None, description="특정 L3 Unit Process로 필터 (비우면 전체)"),
    l4: Optional[str] = Query(None, description="특정 L4 Activity로 필터 (비우면 전체)"),
):
    """
    New Workflow 전용 Task 또는 과제 데이터를 분석하여 AI 워크플로우 설계 초안을 생성합니다.
    과제 엑셀이 업로드된 경우 project_index로 개별 과제를 선택합니다.
    """
    from new_workflow_generator import generate_new_workflow as _gen, result_to_dict
    from new_workflow_generator import generate_workflow_from_freeform, result_to_dict as _rtd

    # 과제 형식이 로드된 경우 → 선택된 과제만 처리
    if _nw_projects_cache:
        from project_excel_reader import projects_to_freeform_params
        if 0 <= project_index < len(_nw_projects_cache):
            selected = [_nw_projects_cache[project_index]]
        else:
            selected = _nw_projects_cache  # fallback: 전체
        params = projects_to_freeform_params(selected)
        if process_name:
            params["process_name"] = process_name

        settings = load_settings()
        anthropic_key = os.getenv("ANTHROPIC_API_KEY", "") or settings.anthropic_api_key
        openai_key = os.getenv("OPENAI_API_KEY", "") or settings.api_key

        result = await generate_workflow_from_freeform(
            **params,
            api_key=anthropic_key,
            model=settings.anthropic_model or "claude-sonnet-4-6",
            openai_api_key=openai_key,
            openai_model=settings.model or "gpt-5.4",
        )
        result_dict = _rtd(result)
        _new_workflow_cache.update(result_dict)
        _save_nw_state("new_workflow", _new_workflow_cache)
        return {"ok": True, **result_dict}

    # L5 Task 형식
    source_tasks = _nw_tasks_cache if _nw_tasks_cache else _tasks_cache
    if not source_tasks:
        raise HTTPException(400, "로드된 Task가 없습니다. 먼저 엑셀 파일을 업로드하세요.")

    # 필터 적용
    tasks = source_tasks
    if l3:
        tasks = [t for t in tasks if t.l3 == l3 or t.l3_id == l3]
    if l4:
        tasks = [t for t in tasks if t.l4 == l4 or t.l4_id == l4]

    if not tasks:
        raise HTTPException(400, "필터 조건에 맞는 Task가 없습니다.")

    # 프로세스명 자동 추론
    if not process_name:
        l3_names = list(dict.fromkeys(t.l3 for t in tasks if t.l3))
        process_name = l3_names[0] if l3_names else "HR 프로세스"

    # 설정에서 API 키 로드
    settings = load_settings()
    anthropic_key = os.getenv("ANTHROPIC_API_KEY", "") or settings.anthropic_api_key
    openai_key = os.getenv("OPENAI_API_KEY", "") or settings.api_key

    result = await _gen(
        tasks=tasks,
        process_name=process_name,
        api_key=anthropic_key,
        model=settings.anthropic_model or "claude-sonnet-4-6",
        openai_api_key=openai_key,
        openai_model=settings.model or "gpt-5.4",
    )

    result_dict = result_to_dict(result)
    _new_workflow_cache.update(result_dict)
    _save_nw_state("new_workflow", _new_workflow_cache)

    # meta에 agent 수 저장
    agents = result_dict.get("agents", [])
    fn = get_current_project()
    if fn:
        save_meta(fn, agent_count=len(agents),
                  agent_names=[a.get("agent_name", "") for a in agents])

    return {"ok": True, **result_dict}


@app.post("/api/new-workflow/generate-freeform", tags=["NewWorkflow"])
async def generate_new_workflow_freeform(request: Request):
    """
    자유형식 입력을 받아 AI가 새로운 L5 Task를 정의하고 Workflow를 설계합니다.
    Body: { process_name, inputs, outputs, systems, pain_points, additional_info }
    """
    from new_workflow_generator import generate_workflow_from_freeform, result_to_dict

    body = await request.json()
    process_name = body.get("process_name", "")
    if not process_name:
        raise HTTPException(400, "프로세스/주제명을 입력해 주세요.")

    settings = load_settings()
    anthropic_key = os.getenv("ANTHROPIC_API_KEY", "") or settings.anthropic_api_key
    openai_key = os.getenv("OPENAI_API_KEY", "") or settings.api_key

    result = await generate_workflow_from_freeform(
        process_name=process_name,
        inputs=body.get("inputs", ""),
        outputs=body.get("outputs", ""),
        systems=body.get("systems", ""),
        pain_points=body.get("pain_points", ""),
        additional_info=body.get("additional_info", ""),
        api_key=anthropic_key,
        model=settings.anthropic_model or "claude-sonnet-4-6",
        openai_api_key=openai_key,
        openai_model=settings.model or "gpt-5.4",
    )

    result_dict = result_to_dict(result)
    _new_workflow_cache.clear()
    _new_workflow_cache.update(result_dict)
    _save_nw_state("new_workflow", _new_workflow_cache)


    return {"ok": True, **result_dict}


@app.post("/api/new-workflow/benchmark", tags=["NewWorkflow"])
async def benchmark_new_workflow():
    """
    현재 Workflow 결과를 기반으로 웹 벤치마킹 검색 후 LLM으로 개선합니다.
    """
    from benchmark_search import search_benchmarks, refine_workflow_with_benchmarks
    from new_workflow_generator import result_to_dict, _parse_freeform_result

    if not _new_workflow_cache:
        raise HTTPException(400, "Workflow 결과가 없습니다. 먼저 1단계를 실행하세요.")

    # 1. 웹 벤치마킹 검색
    _nw_bm_sr = await search_benchmarks(_new_workflow_cache)
    benchmark_results = _nw_bm_sr.get("results", [])

    if not benchmark_results:
        raise HTTPException(500, "벤치마킹 검색 결과를 가져오지 못했습니다. 잠시 후 다시 시도해 주세요.")

    # 2. LLM으로 Workflow 개선
    settings = load_settings()
    anthropic_key = os.getenv("ANTHROPIC_API_KEY", "") or settings.anthropic_api_key
    openai_key = os.getenv("OPENAI_API_KEY", "") or settings.api_key

    refined_data = await refine_workflow_with_benchmarks(
        workflow_cache=_new_workflow_cache,
        benchmark_results=benchmark_results,
        api_key=anthropic_key,
        model=settings.anthropic_model or "claude-sonnet-4-6",
        openai_api_key=openai_key,
        openai_model=settings.model or "gpt-5.4",
    )

    if "error" in refined_data:
        raise HTTPException(500, refined_data["error"])

    # 벤치마킹 인사이트 저장
    benchmark_insights = refined_data.pop("benchmark_insights", [])
    improvement_summary = refined_data.pop("improvement_summary", "")

    # 개선된 Workflow 캐시 업데이트
    # redesigned_process 형식이면 직접 저장, 구형 agents 형식이면 파싱
    if "redesigned_process" in refined_data:
        refined_dict = {
            "ok": True,
            "blueprint_summary": refined_data.get("blueprint_summary", ""),
            "process_name": refined_data.get("process_name", _new_workflow_cache.get("process_name", "")),
            "redesigned_process": refined_data.get("redesigned_process", []),
            # 원본 agents/execution_flow 보존 — 벤치마킹이 redesigned_process를 업데이트해도 AI Service Flow는 유지
            "agents": _new_workflow_cache.get("agents", []),
            "execution_flow": _new_workflow_cache.get("execution_flow", []),
        }
    else:
        refined_result = _parse_freeform_result(refined_data)
        refined_dict = result_to_dict(refined_result)

    _new_workflow_cache.clear()
    _new_workflow_cache.update(refined_dict)
    _save_nw_state("new_workflow", _new_workflow_cache)

    # 벤치마킹 결과 별도 저장 (재실행 없이 불러오기 위해)
    benchmark_data = {
        "benchmark_insights": benchmark_insights,
        "improvement_summary": improvement_summary,
        "search_count": len(benchmark_results),
    }
    _save_nw_state("benchmark_result", benchmark_data)

    return {
        "ok": True,
        **benchmark_data,
        **refined_dict,
    }


@app.get("/api/new-workflow/result", tags=["NewWorkflow"])
async def get_new_workflow_result():
    """마지막으로 생성된 New Workflow 결과를 반환합니다."""
    if not _new_workflow_cache:
        raise HTTPException(404, "생성된 New Workflow 결과가 없습니다. 먼저 생성을 실행하세요.")
    return {"ok": True, **_new_workflow_cache}


@app.put("/api/new-workflow/result", tags=["NewWorkflow"])
async def save_edited_workflow(request: Request):
    """3단계에서 편집된 Workflow를 저장합니다."""
    body = await request.json()
    if not body:
        raise HTTPException(400, "저장할 데이터가 없습니다.")
    _new_workflow_cache.clear()
    _new_workflow_cache.update(body)
    _save_nw_state("new_workflow", _new_workflow_cache)
    return {"ok": True}


@app.delete("/api/new-workflow/result", tags=["NewWorkflow"])
async def clear_new_workflow_result():
    """New Workflow 결과를 초기화합니다."""
    _new_workflow_cache.clear()
    clear_data("new_workflow")
    return {"ok": True}


@app.get("/api/new-workflow/export-html", tags=["NewWorkflow"])
async def export_new_workflow_as_html():
    """AI Service Flow를 PwC 표준 HTML로 내보냅니다."""
    from html_exporter import export_workflow_html

    if not _new_workflow_cache:
        raise HTTPException(404, "Workflow 결과가 없습니다.")

    html = export_workflow_html(_new_workflow_cache)
    process_name = _new_workflow_cache.get("process_name", "workflow")
    filename = f"AI_Service_Flow_{process_name}.html"
    # 한글 파일명 인코딩
    from urllib.parse import quote
    encoded_filename = quote(filename)

    return StreamingResponse(
        io.BytesIO(html.encode("utf-8")),
        media_type="text/html; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@app.get("/api/new-workflow/export-hr-json", tags=["NewWorkflow"])
async def export_new_workflow_as_hr_json():
    """
    New Workflow 결과를 hr-workflow-ai v2.0 호환 JSON으로 변환하여 다운로드합니다.
    hr-workflow-ai에 직접 Import할 수 있는 형식입니다.
    """
    from new_workflow_generator import (
        result_to_hr_workflow_json,
        NewWorkflowResult,
        AIAgent,
        AssignedTask,
        ExecutionStep as NWExecutionStep,
    )
    import io

    if not _new_workflow_cache:
        raise HTTPException(404, "생성된 New Workflow 결과가 없습니다. 먼저 생성을 실행하세요.")

    # 캐시 dict → NewWorkflowResult 복원
    cache = _new_workflow_cache
    agents = []
    for a in cache.get("agents", []):
        assigned = [
            AssignedTask(
                task_id=t["task_id"],
                task_name=t["task_name"],
                l4=t["l4"],
                l3=t["l3"],
                ai_role=t["ai_role"],
                human_role=t["human_role"],
                input_data=t["input_data"],
                output_data=t["output_data"],
                automation_level=t["automation_level"],
            )
            for t in a.get("assigned_tasks", [])
        ]
        agents.append(AIAgent(
            agent_id=a["agent_id"],
            agent_name=a["agent_name"],
            agent_type=a["agent_type"],
            ai_technique=a["ai_technique"],
            description=a["description"],
            automation_level=a["automation_level"],
            assigned_tasks=assigned,
        ))

    flow = [
        NWExecutionStep(
            step=s["step"],
            step_name=s["step_name"],
            step_type=s["step_type"],
            description=s["description"],
            agent_ids=s["agent_ids"],
            task_ids=s["task_ids"],
        )
        for s in cache.get("execution_flow", [])
    ]

    result = NewWorkflowResult(
        blueprint_summary=cache.get("blueprint_summary", ""),
        process_name=cache.get("process_name", "AI 워크플로우"),
        total_tasks=cache.get("total_tasks", 0),
        full_auto_count=cache.get("full_auto_count", 0),
        human_in_loop_count=cache.get("human_in_loop_count", 0),
        human_supervised_count=cache.get("human_supervised_count", 0),
        agents=agents,
        execution_flow=flow,
    )

    hr_json = result_to_hr_workflow_json(result)
    filename = f"{result.process_name or 'new_workflow'}.json"

    from urllib.parse import quote
    encoded_fn = quote(filename)
    return StreamingResponse(
        io.BytesIO(json.dumps(hr_json, ensure_ascii=False, indent=2).encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_fn}"},
    )


# ─────────────────────────────────────────────────────────────────────────────
# Project Management 엔드포인트 (과제 정의서 / 과제 설계서)
# ─────────────────────────────────────────────────────────────────────────────

# 과제 정의서 캐시
_project_definition_cache: dict = {}
_project_definition_cache.update(_load_nw_state("project_definition"))


def _build_classification_from_workflow(workflow_cache: dict) -> dict[str, dict]:
    """New Workflow 결과에서 분류 데이터를 합성합니다."""
    classification: dict[str, dict] = {}
    for agent in workflow_cache.get("agents", []):
        for task in agent.get("assigned_tasks", []):
            level = task.get("automation_level", "")
            if "on-the-Loop" in level or "out-of-the-Loop" in level:
                label = "AI"
            elif "in-the-Loop" in level:
                label = "AI + Human"
            else:
                label = "Human"
            classification[task["task_id"]] = {
                "label": label,
                "reason": task.get("ai_role", ""),
                "hybrid_note": task.get("human_role", ""),
                "input_types": ", ".join(task.get("input_data", [])),
                "output_types": ", ".join(task.get("output_data", [])),
                "ai_prerequisites": "",
            }
    return classification


@app.post("/api/project-management/definition/generate", tags=["ProjectManagement"])
async def generate_project_definition(
    provider: str = Query("openai", description="분류 결과 제공자 (openai | anthropic)"),
    source: str = Query("", description="데이터 소스 (new-workflow이면 NW 캐시 사용)"),
    process_name: str = Query("", description="프로세스 이름 (빈 값이면 자동 추론)"),
    author: str = Query("", description="작성자 (빈 값이면 'PwC')"),
    l3: Optional[str] = Query(None, description="특정 L3로 필터"),
    l4: Optional[str] = Query(None, description="특정 L4로 필터"),
):
    """
    분류 결과 + To-Be Workflow 결과를 기반으로 과제 정의서를 자동 생성합니다.
    source=new-workflow이면 New Workflow 캐시에서 데이터를 사용합니다.
    """
    from project_definition_generator import (
        generate_project_definition_with_llm,
        generate_project_definition_fallback,
        project_definition_to_dict,
    )

    # 데이터 소스 선택
    use_nw = source == "new-workflow"

    # New Workflow source: Workflow 캐시에서 Task 추출 (과제 엑셀은 L5 Task 없음)
    if use_nw:
        if not _new_workflow_cache:
            raise HTTPException(400, "New Workflow 결과가 없습니다. 먼저 New Workflow에서 설계를 실행하세요.")

        classification_dict = _build_classification_from_workflow(_new_workflow_cache)
        task_dicts = []
        for agent in _new_workflow_cache.get("agents", []):
            for t in agent.get("assigned_tasks", []):
                td = {
                    "id": t.get("task_id", ""), "l2": "", "l2_id": "",
                    "l3": t.get("l3", ""), "l3_id": "",
                    "l4": t.get("l4", ""), "l4_id": "",
                    "name": t.get("task_name", ""),
                    "description": t.get("ai_role", ""),
                    "performer": "",
                }
                for attr in ["pain_time", "pain_accuracy", "pain_repetition",
                              "pain_data", "pain_system", "pain_communication", "pain_other"]:
                    td[attr] = ""
                task_dicts.append(td)

        if not task_dicts:
            raise HTTPException(400, "Workflow에 Task가 없습니다.")

        if not process_name:
            process_name = _new_workflow_cache.get("process_name", "HR 프로세스")

        tobe_data = _new_workflow_cache

        settings = load_settings()
        anthropic_key = os.getenv("ANTHROPIC_API_KEY", "") or settings.anthropic_api_key

        try:
            result = await generate_project_definition_with_llm(
                tasks=task_dicts,
                classification_results=classification_dict,
                tobe_data=tobe_data,
                process_name=process_name,
                api_key=anthropic_key,
                model=settings.anthropic_model or "claude-sonnet-4-6",
                author=author,
            )
        except Exception as e:
            print(f"[과제 정의서] LLM 실패, fallback 사용: {e}")
            result = generate_project_definition_fallback(
                tasks=task_dicts,
                classification_results=classification_dict,
                tobe_data=tobe_data,
                process_name=process_name,
                author=author,
            )

        result_dict = project_definition_to_dict(result)
        _project_definition_cache.clear()
        _project_definition_cache.update(result_dict)
        _save_nw_state("project_definition", _project_definition_cache)

        return {"ok": True, **result_dict}

    # 기존 방식: Tasks 페이지에서 업로드한 파일 기반
    src_tasks = _tasks_cache

    if not src_tasks:
        raise HTTPException(400, "로드된 Task가 없습니다. 먼저 엑셀 파일을 업로드하세요.")

    tasks = src_tasks
    if l3:
        tasks = [t for t in tasks if t.l3 == l3 or t.l3_id == l3]
    if l4:
        tasks = [t for t in tasks if t.l4 == l4 or t.l4_id == l4]

    if not tasks:
        raise HTTPException(400, "필터 조건에 맞는 Task가 없습니다.")

    results_store = load_results(provider)
    if not results_store:
        raise HTTPException(400, f"'{provider}' 분류 결과가 없습니다. 먼저 분류를 실행하세요.")

    task_dicts = []
    classification_dict = {}
    for t in tasks:
        td = {
            "id": t.id, "l2": t.l2, "l2_id": t.l2_id,
            "l3": t.l3, "l3_id": t.l3_id,
            "l4": t.l4, "l4_id": t.l4_id,
            "name": t.name, "description": t.description,
            "performer": t.performer,
        }
        # Pain Point 필드 추가
        for attr in ["pain_time", "pain_accuracy", "pain_repetition",
                      "pain_data", "pain_system", "pain_communication", "pain_other"]:
            td[attr] = getattr(t, attr, "")
        task_dicts.append(td)

        if not use_nw:
            cr = results_store.get(t.id)
            if cr:
                classification_dict[t.id] = {
                    "label": cr.label,
                    "reason": cr.reason,
                    "hybrid_note": cr.hybrid_note,
                    "input_types": cr.input_types,
                    "output_types": cr.output_types,
                    "ai_prerequisites": cr.ai_prerequisites,
                }

    # 프로세스명 자동 추론
    if not process_name:
        l3_names = list(dict.fromkeys(t.l3 for t in tasks if t.l3))
        process_name = l3_names[0] if l3_names else "HR 프로세스"

    # To-Be 데이터 (있으면 활용)
    tobe_data = None
    if _new_workflow_cache:
        tobe_data = _new_workflow_cache

    # 설정에서 API 키 로드
    settings = load_settings()
    anthropic_key = os.getenv("ANTHROPIC_API_KEY", "") or settings.anthropic_api_key

    try:
        result = await generate_project_definition_with_llm(
            tasks=task_dicts,
            classification_results=classification_dict,
            tobe_data=tobe_data,
            process_name=process_name,
            api_key=anthropic_key,
            model=settings.anthropic_model or "claude-sonnet-4-6",
            author=author,
        )
    except Exception as e:
        print(f"[과제 정의서] LLM 실패, fallback 사용: {e}")
        result = generate_project_definition_fallback(
            tasks=task_dicts,
            classification_results=classification_dict,
            tobe_data=tobe_data,
            process_name=process_name,
            author=author,
        )

    result_dict = project_definition_to_dict(result)
    _project_definition_cache.clear()
    _project_definition_cache.update(result_dict)
    _save_nw_state("project_definition", _project_definition_cache)

    return {"ok": True, **result_dict}


@app.get("/api/project-management/definition", tags=["ProjectManagement"])
async def get_project_definition():
    """마지막으로 생성된 과제 정의서를 반환합니다."""
    if not _project_definition_cache:
        raise HTTPException(404, "생성된 과제 정의서가 없습니다. 먼저 생성을 실행하세요.")
    return {"ok": True, **_project_definition_cache}


@app.delete("/api/project-management/definition", tags=["ProjectManagement"])
async def clear_project_definition():
    """과제 정의서 결과를 초기화합니다."""
    _project_definition_cache.clear()
    clear_data("project_definition")
    return {"ok": True}


# ── 과제 설계서 ───────────────────────────────────────────────────────────────

_project_design_cache: dict = {}
_project_design_cache.update(_load_nw_state("project_design"))


@app.post("/api/project-management/design/generate", tags=["ProjectManagement"])
async def generate_project_design(
    provider: str = Query("openai", description="분류 결과 제공자 (openai | anthropic)"),
    source: str = Query("", description="데이터 소스 (new-workflow이면 NW 캐시 사용)"),
    process_name: str = Query("", description="프로세스 이름 (빈 값이면 자동 추론)"),
    l3: Optional[str] = Query(None, description="특정 L3로 필터"),
    l4: Optional[str] = Query(None, description="특정 L4로 필터"),
):
    """
    분류 결과 + To-Be Workflow 결과를 기반으로 과제 설계서를 자동 생성합니다.
    source=new-workflow이면 New Workflow 캐시에서 데이터를 사용합니다.
    """
    from project_design_generator import (
        generate_project_design_with_llm,
        generate_project_design_fallback,
        project_design_to_dict,
    )

    use_nw = source == "new-workflow"

    # New Workflow source: Workflow 캐시에서 직접 Task 추출
    if use_nw:
        if not _new_workflow_cache:
            raise HTTPException(400, "New Workflow 결과가 없습니다. 먼저 New Workflow에서 설계를 실행하세요.")

        classification_dict = _build_classification_from_workflow(_new_workflow_cache)
        task_dicts = []
        for agent in _new_workflow_cache.get("agents", []):
            for t in agent.get("assigned_tasks", []):
                task_dicts.append({
                    "id": t.get("task_id", ""), "l2": "", "l3": t.get("l3", ""),
                    "l4": t.get("l4", ""), "l4_id": "", "name": t.get("task_name", ""),
                    "description": t.get("ai_role", ""), "performer": "",
                })

        if not task_dicts:
            raise HTTPException(400, "Workflow에 Task가 없습니다.")

        if not process_name:
            process_name = _new_workflow_cache.get("process_name", "HR 프로세스")

        tobe_data = _new_workflow_cache
        project_title = _project_definition_cache.get("project_title", f"{process_name} AI 자동화")

        settings = load_settings()
        anthropic_key = os.getenv("ANTHROPIC_API_KEY", "") or settings.anthropic_api_key

        try:
            result = await generate_project_design_with_llm(
                tasks=task_dicts,
                classification_results=classification_dict,
                tobe_data=tobe_data,
                process_name=process_name,
                project_title=project_title,
                api_key=anthropic_key,
                model=settings.anthropic_model or "claude-sonnet-4-6",
            )
        except Exception as e:
            print(f"[과제 설계서] LLM 실패, fallback 사용: {e}")
            result = generate_project_design_fallback(
                tasks=task_dicts,
                classification_results=classification_dict,
                tobe_data=tobe_data,
                process_name=process_name,
                project_title=project_title,
            )

        result_dict = project_design_to_dict(result)
        _project_design_cache.clear()
        _project_design_cache.update(result_dict)
        _save_nw_state("project_design", _project_design_cache)

        return {"ok": True, **result_dict}

    # 기존 방식: Tasks 페이지 기반
    src_tasks = _tasks_cache
    if not src_tasks:
        raise HTTPException(400, "로드된 Task가 없습니다. 먼저 엑셀 파일을 업로드하세요.")

    tasks = src_tasks
    if l3:
        tasks = [t for t in tasks if t.l3 == l3 or t.l3_id == l3]
    if l4:
        tasks = [t for t in tasks if t.l4 == l4 or t.l4_id == l4]

    if not tasks:
        raise HTTPException(400, "필터 조건에 맞는 Task가 없습니다.")

    results_store = load_results(provider)
    if not results_store:
        raise HTTPException(400, f"'{provider}' 분류 결과가 없습니다. 먼저 분류를 실행하세요.")

    task_dicts = []
    classification_dict = {}
    for t in tasks:
        td = {
            "id": t.id, "l2": t.l2, "l3": t.l3, "l4": t.l4,
            "l4_id": t.l4_id, "name": t.name,
            "description": t.description, "performer": t.performer,
        }
        task_dicts.append(td)
        cr = results_store.get(t.id)
        if cr:
            classification_dict[t.id] = {
                "label": cr.label, "reason": cr.reason,
                "hybrid_note": cr.hybrid_note,
                "input_types": cr.input_types, "output_types": cr.output_types,
            }

    if not process_name:
        l3_names = list(dict.fromkeys(t.l3 for t in tasks if t.l3))
        process_name = l3_names[0] if l3_names else "HR 프로세스"

    tobe_data = _new_workflow_cache if _new_workflow_cache else None
    project_title = _project_definition_cache.get("project_title", f"{process_name} AI 자동화")

    settings = load_settings()
    anthropic_key = os.getenv("ANTHROPIC_API_KEY", "") or settings.anthropic_api_key

    try:
        result = await generate_project_design_with_llm(
            tasks=task_dicts,
            classification_results=classification_dict,
            tobe_data=tobe_data,
            process_name=process_name,
            project_title=project_title,
            api_key=anthropic_key,
            model=settings.anthropic_model or "claude-sonnet-4-6",
        )
    except Exception as e:
        print(f"[과제 설계서] LLM 실패, fallback 사용: {e}")
        result = generate_project_design_fallback(
            tasks=task_dicts,
            classification_results=classification_dict,
            tobe_data=tobe_data,
            process_name=process_name,
            project_title=project_title,
        )

    result_dict = project_design_to_dict(result)
    _project_design_cache.clear()
    _project_design_cache.update(result_dict)
    _save_nw_state("project_design", _project_design_cache)

    return {"ok": True, **result_dict}


@app.get("/api/project-management/design", tags=["ProjectManagement"])
async def get_project_design():
    """마지막으로 생성된 과제 설계서를 반환합니다."""
    if not _project_design_cache:
        raise HTTPException(404, "생성된 과제 설계서가 없습니다. 먼저 생성을 실행하세요.")
    return {"ok": True, **_project_design_cache}


@app.delete("/api/project-management/design", tags=["ProjectManagement"])
async def clear_project_design():
    """과제 설계서 결과를 초기화합니다."""
    _project_design_cache.clear()
    clear_data("project_design")
    return {"ok": True}


@app.get("/api/project-management/export-ppt", tags=["ProjectManagement"])
async def export_project_ppt():
    """과제 정의서 + 설계서를 PPT 템플릿에 채워서 다운로드합니다."""
    from ppt_exporter import export_ppt

    definition = _project_definition_cache if _project_definition_cache else None
    design = _project_design_cache if _project_design_cache else None

    if not definition and not design:
        raise HTTPException(404, "과제 정의서 또는 설계서가 없습니다. 먼저 생성을 실행하세요.")

    ppt_bytes = export_ppt(definition=definition, design=design, workflow=_new_workflow_cache or None)

    title = definition.get("project_title", "과제정의서") if definition else "과제정의서"
    filename = f"{title}.pptx"
    from urllib.parse import quote
    encoded_fn = quote(filename)

    return StreamingResponse(
        ppt_bytes,
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_fn}"},
    )


# ─────────────────────────────────────────────────────────────────────────────
# Workflow 세션 관리 API
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/workflow/sessions/overview", tags=["Workflow"])
async def get_workflow_sessions_overview(request: Request):
    """전체 세션을 user_id 기준으로 그룹핑하여 PM 대시보드용으로 반환합니다."""
    _load_sessions_manifest()
    req_team_id = _get_team_id(request)
    auth_user = _get_auth_user(request)
    is_admin = auth_user.get("is_admin", False) if auth_user else False
    is_common = (req_team_id == "공통")

    all_sessions = [
        v for k, v in _sessions_manifest.items()
        if k != "_current" and isinstance(v, dict)
        and (is_admin or is_common or v.get("team_id", "default") == req_team_id)
    ]

    # user_id별 그룹핑
    user_map: dict[str, list] = {}
    for sess in all_sessions:
        uid = sess.get("user_id", "unknown") or "unknown"
        if uid not in user_map:
            user_map[uid] = []

        # session_data.json에서 진행 상태 읽기
        import re as _re
        safe_sid = _re.sub(r'[^\w가-힣\-]', '_', sess["id"])[:80] or "default"
        sess_data_path = _SESSIONS_DIR / safe_sid / "session_data.json"
        has_step1 = False
        has_step2 = False
        has_benchmark = False
        has_gap = False
        try:
            if sess_data_path.exists():
                sd = json.loads(sess_data_path.read_text(encoding="utf-8"))
                has_step1 = bool(sd.get("step1"))
                has_step2 = bool(sd.get("step2"))
                has_benchmark = bool(sd.get("benchmark_table"))
                has_gap = bool(sd.get("gap_analysis"))
        except Exception:
            pass

        user_map[uid].append({
            "id": sess["id"],
            "name": sess.get("name", sess["id"]),
            "created_at": sess.get("created_at", ""),
            "updated_at": sess.get("updated_at"),
            "has_step1": has_step1,
            "has_step2": has_step2,
            "has_benchmark": has_benchmark,
            "has_gap": has_gap,
        })

    users = [
        {"user_id": uid, "sessions": sorted(sess_list, key=lambda s: s.get("updated_at") or s.get("created_at") or "", reverse=True)}
        for uid, sess_list in user_map.items()
    ]
    users.sort(key=lambda u: u["user_id"])
    return {"ok": True, "users": users}


@app.get("/api/workflow/sessions", tags=["Workflow"])
async def list_workflow_sessions(request: Request, all: bool = Query(False)):
    """저장된 Workflow 세션 목록을 반환합니다.
    - ?all=true: 같은 팀 내 전체 세션 (PM 뷰)
    - 기본: X-User-Id 헤더 또는 ?user= 파라미터와 일치하는 세션만 반환
    - 타 팀 데이터는 항상 숨김
    """
    _load_sessions_manifest()
    req_team_id = _get_team_id(request)
    auth_user = _get_auth_user(request)
    is_admin = auth_user.get("is_admin", False) if auth_user else False
    is_common = (req_team_id == "공통")  # 공통 멤버는 전체 팀 조회 가능

    # 팀 필터: 공통 멤버/Admin이면 전체, 일반 팀원이면 본인 팀만
    all_sessions = [
        v for k, v in _sessions_manifest.items()
        if k != "_current" and isinstance(v, dict)
        and (is_admin or is_common or v.get("team_id", "default") == req_team_id)
    ]
    all_sessions.sort(key=lambda s: s.get("updated_at", s.get("created_at", "")), reverse=True)

    if not all:
        user_id = _get_user_id(request) or request.query_params.get("user", "")
        if user_id and user_id != "unknown" and not is_admin and not is_common:
            all_sessions = [s for s in all_sessions if s.get("user_id", "unknown") == user_id]

    return {
        "ok": True,
        "current": _sessions_manifest.get("_current", _current_session_id),
        "sessions": all_sessions,
    }


@app.get("/api/workflow/sessions/{session_id}/files", tags=["Workflow"])
async def list_session_files(session_id: str):
    """세션에 저장된 모든 Excel·PPT 파일 목록을 반환합니다."""

    _load_sessions_manifest()
    if session_id not in _sessions_manifest:
        raise HTTPException(404, f"세션 '{session_id}'을 찾을 수 없습니다.")

    d = _get_session_dir(session_id)

    def _finfo(f: Path) -> dict:
        st = f.stat()
        return {
            "filename": f.name,
            "size_kb": round(st.st_size / 1024, 1),
            "modified": _mtime_kst(f),
            "is_current": str(f) == _wf_excel_path,
        }

    excels = sorted(d.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    ppts = sorted(d.glob("*.pptx"), key=lambda f: f.stat().st_mtime, reverse=True)
    json_exists = (d / "workflow.json").exists()

    return {
        "ok": True,
        "session_id": session_id,
        "excels": [_finfo(f) for f in excels],
        "ppts": [_finfo(f) for f in ppts],
        "has_json": json_exists,
    }


@app.post("/api/workflow/select-file", tags=["Workflow"])
async def select_workflow_file(request: Request):
    """세션 내 특정 Excel 파일을 명시적으로 선택하여 로드합니다."""
    from excel_reader import load_tasks, list_sheets
    body = await request.json()
    session_id = body.get("session_id", "")
    filename = body.get("filename", "")

    if not session_id or not filename:
        raise HTTPException(400, "session_id와 filename이 필요합니다.")

    import re as _re
    safe_sid = _re.sub(r'[^\w가-힣\-]', '_', session_id)[:80] or "default"
    file_path = _SESSIONS_DIR / safe_sid / Path(filename).name
    if not file_path.exists():
        raise HTTPException(404, f"파일을 찾을 수 없습니다: {filename}")

    global _wf_excel_tasks, _wf_excel_path, _wf_classification
    global _wf_chat_history, _wf_step1_cache, _wf_step2_cache, _wf_benchmark_table, _wf_user_resources

    sheets = list_sheets(str(file_path))
    recommended = next((s["name"] for s in sheets if s.get("recommended")), None)
    tasks = load_tasks(str(file_path), sheet_name=recommended)

    _wf_excel_tasks = tasks
    _wf_excel_path = str(file_path)
    _wf_classification = {}
    _wf_chat_history = []
    _wf_step1_cache = {}
    _wf_step2_cache = {}
    _wf_benchmark_table = {}
    _wf_user_resources = []

    # 분류 결과 재추출
    def _resolve_label(*candidates: str) -> str:
        for c in candidates:
            v = (c or "").strip()
            if v and v != "-":
                return v
        return ""

    has_cls = False
    for t in tasks:
        label = _resolve_label(t.cls_final_label, t.cls_doosan_label, t.cls_1st_label)
        if label:
            has_cls = True
            _wf_classification[t.id] = {
                "label": label, "reason": t.cls_1st_reason or "",
                "criterion": t.cls_1st_knockout or "",
                "ai_prerequisites": t.cls_1st_ai_prereq or "",
                "feedback": next((v for x in [t.cls_final_feedback, t.cls_doosan_feedback] if (v := (x or "").strip()) and v != "-"), ""),
                "task_name": t.name, "hybrid_note": "", "input_types": "", "output_types": "",
            }

    data_sheets = [s for s in sheets if not s.get("is_guide") and s.get("task_count", 0) > 0]
    return {
        "ok": True,
        "filename": filename,
        "task_count": len(tasks),
        "has_classification": has_cls,
        "classified_count": len(_wf_classification),
        "sheets": [{"name": s["name"], "recommended": s.get("recommended", False),
                    "row_count": s.get("task_count", 0), "l5_count": s.get("task_count", 0)}
                   for s in data_sheets],
    }


@app.get("/api/upload/history", tags=["Upload"])
async def get_upload_history():
    """Task 분류용으로 이전에 업로드된 엑셀 파일 목록을 반환합니다."""

    files = []
    if _UPLOAD_DIR.exists():
        for f in sorted(_UPLOAD_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
            st = f.stat()
            files.append({
                "filename": f.name,
                "size_kb": round(st.st_size / 1024, 1),
                "modified": _mtime_kst(f),
                "is_current": _current_excel_path is not None and str(f) == str(_current_excel_path),
            })
    return {"ok": True, "files": files}


@app.patch("/api/workflow/sessions/{session_id}/rename", tags=["Workflow"])
async def rename_workflow_session(session_id: str, request: Request):
    """세션 표시 이름을 변경합니다 (디렉토리 이름은 유지)."""
    body = await request.json()
    new_name = (body.get("name") or "").strip()
    if not new_name:
        raise HTTPException(400, "name 필드가 필요합니다.")

    _load_sessions_manifest()
    if session_id not in _sessions_manifest:
        raise HTTPException(404, f"세션 '{session_id}'을 찾을 수 없습니다.")

    session = _sessions_manifest[session_id]
    req_team = _get_team_id(request)
    auth_user = _get_auth_user(request)
    if (not auth_user or auth_user.get("is_admin") is True or req_team == "공통"):
        pass  # Admin / 공통 멤버는 허용
    elif session.get("team_id", "default") != req_team:
        raise HTTPException(403, "다른 팀의 세션에 접근할 수 없습니다.")

    session["name"] = new_name
    session["updated_at"] = _now_kst()
    _save_sessions_manifest()
    return {"ok": True, "session_id": session_id, "name": new_name}


@app.post("/api/workflow/sessions/create", tags=["Workflow"])
async def create_empty_session(request: Request):
    """이름만으로 빈 세션을 즉시 생성 (프로젝트 목록에 바로 등록).
    body: {name: "프로젝트 이름"}
    """
    body = await request.json() if True else {}
    name = str(body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "프로젝트 이름이 필요합니다.")

    # 세션 ID: 타임스탬프 기반 (기존 컨벤션과 동일)
    import time as _t
    sid = f"session_{int(_t.time() * 1000)}"

    # 세션 디렉토리 생성
    _get_session_dir(sid)

    # manifest 등록
    _load_sessions_manifest()
    now = _now_kst()
    user_id = _get_user_id(request)
    team_id = _get_team_id(request)
    _sessions_manifest[sid] = {
        "id": sid,
        "name": name,
        "created_at": now,
        "updated_at": now,
        "user_id": user_id,
        "team_id": team_id,
    }
    _sessions_manifest["_current"] = sid
    _save_sessions_manifest()

    # 현재 세션으로 설정 + 메모리 캐시 초기화 (새 프로젝트라 기존 데이터 없음)
    global _current_session_id
    global _wf_excel_tasks, _wf_excel_path, _wf_classification
    global _wf_benchmark_table, _wf_step1_cache, _wf_step2_cache, _wf_gap_analysis
    global _wf_chat_history, _wf_user_resources, _wf_tobe_flow_cache
    global _workflow_cache, _new_workflow_cache
    _current_session_id = sid
    _wf_excel_tasks = []
    _wf_excel_path = ""
    _wf_classification = {}
    _wf_benchmark_table = {}
    _wf_step1_cache = {}
    _wf_step2_cache = {}
    _wf_gap_analysis = {}
    _wf_chat_history = []
    _wf_user_resources = []
    _wf_tobe_flow_cache = {}
    _workflow_cache = {}
    _new_workflow_cache.clear()

    return {"ok": True, "session_id": sid, "name": name}


@app.post("/api/workflow/sessions/current/save", tags=["Workflow"])
async def save_current_session():
    """현재 세션 상태를 명시적으로 저장합니다."""
    if not _current_session_id:
        raise HTTPException(400, "활성 세션이 없습니다.")
    _save_session_data(_current_session_id)
    return {
        "ok": True,
        "session_id": _current_session_id,
        "name": _sessions_manifest.get(_current_session_id, {}).get("name", _current_session_id),
        "saved_at": _now_kst(),
    }


@app.post("/api/workflow/sessions/{session_id}/load", tags=["Workflow"])
async def load_workflow_session(session_id: str, request: Request):
    """다른 세션을 로드합니다 (파일 + 상태 전환)."""
    import asyncio

    _load_sessions_manifest()
    if session_id not in _sessions_manifest:
        raise HTTPException(404, f"세션 '{session_id}'을 찾을 수 없습니다.")

    session = _sessions_manifest[session_id]
    req_team = _get_team_id(request)
    auth_user2 = _get_auth_user(request)
    if not (auth_user2 and (auth_user2.get("is_admin") or req_team == "공통")):
        if session.get("team_id", "default") != req_team:
            raise HTTPException(403, "다른 팀의 세션에 접근할 수 없습니다.")

    global _current_session_id, _wf_excel_tasks, _wf_chat_history
    global _wf_step1_cache, _wf_step2_cache, _wf_benchmark_table, _wf_gap_analysis, _wf_user_resources

    # 현재 세션 저장 후 전환
    if _current_session_id:
        _save_session_data(_current_session_id)

    # 메모리 초기화
    _wf_chat_history = []
    _wf_step1_cache = {}
    _wf_step2_cache = {}
    _wf_benchmark_table = {}
    _wf_gap_analysis = {}
    _wf_user_resources = []

    # 새 세션 로드
    ok = await asyncio.to_thread(_load_session_data, session_id)
    if not ok:
        raise HTTPException(500, f"세션 '{session_id}' 로드에 실패했습니다.")

    _current_session_id = session_id
    _sessions_manifest["_current"] = session_id
    _save_sessions_manifest()

    # 응답: summary + 분류 현황
    from workflow_parser import get_workflow_summary
    summary = get_workflow_summary(_workflow_cache["parsed"]) if "parsed" in _workflow_cache else {}
    enriched = _enrich_summary_with_cls(summary)

    return {
        "ok": True,
        "session_id": session_id,
        "has_step1": bool(_wf_step1_cache),
        "has_step2": bool(_wf_step2_cache),
        "has_benchmark": bool(_wf_benchmark_table),
        "has_gap": bool(_wf_gap_analysis),
        "classified_count": len(_wf_classification),
        **enriched,
    }


@app.delete("/api/workflow/sessions/{session_id}", tags=["Workflow"])
async def delete_workflow_session(session_id: str, request: Request):
    """세션과 연관 파일을 삭제합니다."""
    import shutil

    _load_sessions_manifest()
    if session_id not in _sessions_manifest:
        raise HTTPException(404, f"세션 '{session_id}'을 찾을 수 없습니다.")

    session = _sessions_manifest[session_id]
    req_team = _get_team_id(request)
    auth_user3 = _get_auth_user(request)
    if not (auth_user3 and (auth_user3.get("is_admin") or req_team == "공통")):
        if session.get("team_id", "default") != req_team:
            raise HTTPException(403, "다른 팀의 세션에 접근할 수 없습니다.")

    # 디렉토리 삭제
    try:
        import re
        safe = re.sub(r'[^\w가-힣\-]', '_', session_id)[:80] or "default"
        sess_dir = _SESSIONS_DIR / safe
        if sess_dir.exists():
            shutil.rmtree(sess_dir)
    except Exception as e:
        print(f"[SESSION] 디렉토리 삭제 실패({session_id}): {e}", flush=True)

    # manifest에서 제거
    _sessions_manifest.pop(session_id, None)
    if _sessions_manifest.get("_current") == session_id:
        remaining = [k for k in _sessions_manifest if k != "_current" and isinstance(_sessions_manifest[k], dict)]
        _sessions_manifest["_current"] = remaining[0] if remaining else ""
    _save_sessions_manifest()

    return {"ok": True, "deleted": session_id}


# ─────────────────────────────────────────────────────────────────────────────
# Admin 전용 API
# ─────────────────────────────────────────────────────────────────────────────

def _require_admin(request: Request) -> dict:
    """Admin 권한 확인. Admin이 아니면 403."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    user = get_session_user(token)
    if not user:
        raise HTTPException(401, "인증이 필요합니다.")
    if user.get("email") != ADMIN_EMAIL:
        audit_log.log_event("admin_denied", email=user["email"], ip=_get_client_ip(request))
        raise HTTPException(403, "관리자 권한이 필요합니다.")
    return user


@app.get("/api/admin/dashboard", tags=["Admin"])
async def admin_dashboard(request: Request):
    """Admin 대시보드: 사용자/세션/활동/토큰 사용량 요약."""
    _require_admin(request)
    users = get_all_users_info()
    sessions = get_all_sessions()
    login_history = audit_log.get_login_history(limit=30)
    data_activity = audit_log.get_data_activity(limit=30)
    usage = get_usage()
    audit_log.log_event("admin_view", email=ADMIN_EMAIL, ip=_get_client_ip(request), detail="dashboard")
    return {
        "ok": True,
        "users": users,
        "active_sessions": sessions,
        "login_history": login_history,
        "data_activity": data_activity,
        "total_sessions": len(sessions),
        "total_users": len(users),
        "usage": usage,
        "team_projects": ALL_PROJECTS,
        "project_data": {
            proj: list_projects_for_user([proj])
            for proj in ALL_PROJECTS
        },
    }


@app.get("/api/admin/audit-log", tags=["Admin"])
async def admin_audit_log(
    request: Request,
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    email: str = Query(""),
    event: str = Query(""),
    ip: str = Query(""),
):
    """감사 로그 조회 (필터링 가능)."""
    _require_admin(request)
    logs, total = audit_log.get_logs(
        limit=limit, offset=offset,
        email_filter=email, event_filter=event, ip_filter=ip,
    )
    return {"ok": True, "logs": logs, "total": total, "limit": limit, "offset": offset}


@app.get("/api/admin/sessions", tags=["Admin"])
async def admin_sessions(request: Request):
    """모든 활성 세션 조회."""
    _require_admin(request)
    return {"ok": True, "sessions": get_all_sessions()}


@app.post("/api/admin/force-logout", tags=["Admin"])
async def admin_force_logout(request: Request):
    """특정 사용자의 모든 세션 강제 종료."""
    _require_admin(request)
    body = await request.json()
    target_email = body.get("email", "")
    if not target_email:
        raise HTTPException(400, "이메일이 필요합니다.")
    count = force_logout_user(target_email)
    audit_log.log_event("admin_force_logout", email=ADMIN_EMAIL, ip=_get_client_ip(request),
                        detail=f"{target_email}: {count}개 세션 종료")
    return {"ok": True, "email": target_email, "sessions_removed": count}


def _file_info(f: Path) -> dict:
    return {
        "filename": f.name,
        "size_kb": round(f.stat().st_size / 1024, 1),
        "modified": _mtime_kst(f),
        "path": str(f),
    }


@app.get("/api/admin/uploads", tags=["Admin"])
async def admin_list_uploads(request: Request):
    """업로드된 파일 목록 조회 (Task 분류용 엑셀)."""
    _require_admin(request)
    files = []
    if _UPLOAD_DIR.exists():
        for f in sorted(_UPLOAD_DIR.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
            if f.is_file():
                files.append(_file_info(f))
    return {"ok": True, "directory": str(_UPLOAD_DIR), "files": files}


@app.get("/api/admin/uploads-all", tags=["Admin"])
async def admin_list_uploads_all(request: Request):
    """카테고리별 업로드 파일 목록 조회."""
    _require_admin(request)

    def _dir_files(directory: Path, exts=None) -> list:
        if not directory.exists():
            return []
        result = []
        for f in sorted(directory.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
            if f.is_file() and (exts is None or f.suffix.lower() in exts):
                result.append(_file_info(f))
        return result

    # Task 분류 엑셀
    task_excel = _dir_files(_UPLOAD_DIR, {".xlsx", ".xls"})

    # Workflow 파일들 — 세션 디렉토리 직접 스캔 + manifest 보완
    import re as _re
    _load_sessions_manifest()
    wf_excel, wf_json, wf_ppt = [], [], []
    seen_sess: set[str] = set()  # 중복 방지

    # 1) _SESSIONS_DIR 직접 스캔 (manifest 없어도 파일 표시)
    if _SESSIONS_DIR.exists():
        for sess_dir in sorted(_SESSIONS_DIR.iterdir(), key=lambda d: d.stat().st_mtime, reverse=True):
            if not sess_dir.is_dir():
                continue
            sid = _sessions_manifest.get(sess_dir.name, {}).get("id", "") or sess_dir.name
            seen_sess.add(sess_dir.name)

            # 세션 내 모든 Excel 파일 표시 (최신순)
            for xf in sorted(sess_dir.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True):
                entry = _file_info(xf)
                entry["session_id"] = sid
                entry["display_name"] = f"[{sid}] {xf.name}"
                wf_excel.append(entry)

            jf = sess_dir / "workflow.json"
            if jf.exists():
                entry = _file_info(jf)
                entry["session_id"] = sid
                entry["display_name"] = f"[{sid}] workflow.json"
                wf_json.append(entry)

            # 세션 내 모든 PPT 파일 표시 (최신순)
            for pf in sorted(sess_dir.glob("*.pptx"), key=lambda f: f.stat().st_mtime, reverse=True):
                entry = _file_info(pf)
                entry["session_id"] = sid
                entry["display_name"] = f"[{sid}] {pf.name}"
                wf_ppt.append(entry)

    # New Workflow 결과 JSON들 (_NW_DIR 직접 스캔)
    nw_files = []
    if _NW_DIR.exists():
        for f in sorted(_NW_DIR.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
            if f.is_file() and f.suffix == ".json":
                nw_files.append(_file_info(f))

    return {
        "ok": True,
        "categories": {
            "task_excel": task_excel,
            "wf_excel": wf_excel,
            "wf_json": wf_json,
            "wf_ppt": wf_ppt,
            "new_workflow": nw_files,
        },
    }


@app.delete("/api/admin/workflow-session/{session_id}", tags=["Admin"])
async def admin_delete_workflow_session(session_id: str, request: Request):
    """Admin: Workflow 세션 삭제 (파일 + 상태 전체)."""
    import shutil, re as _re
    _require_admin(request)
    _load_sessions_manifest()

    safe_sid = _re.sub(r'[^\w가-힣\-]', '_', session_id)[:80] or "default"
    sess_dir = _SESSIONS_DIR / safe_sid
    if sess_dir.exists():
        try:
            shutil.rmtree(sess_dir)
        except Exception as e:
            raise HTTPException(500, f"세션 디렉토리 삭제 실패: {e}")

    _sessions_manifest.pop(session_id, None)
    if _sessions_manifest.get("_current") == session_id:
        remaining = [k for k in _sessions_manifest if k != "_current" and isinstance(_sessions_manifest[k], dict)]
        _sessions_manifest["_current"] = remaining[0] if remaining else ""
    _save_sessions_manifest()
    return {"ok": True, "deleted": session_id}


@app.delete("/api/admin/upload/{filename}", tags=["Admin"])
async def admin_delete_upload(filename: str, request: Request):
    """Admin: Task 분류용 엑셀 파일 삭제 + 대응하는 프로젝트 데이터 디렉토리도 삭제."""
    _require_admin(request)
    safe_name = Path(filename).name
    file_path = _UPLOAD_DIR / safe_name
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(404, f"파일을 찾을 수 없습니다: {safe_name}")
    file_path.unlink()
    # 대응하는 이전 프로젝트 데이터 폴더도 삭제 (이전 프로젝트 목록에서 제거)
    delete_project(safe_name)
    return {"ok": True, "deleted": safe_name}


@app.delete("/api/admin/workflow-file/{filename}", tags=["Admin"])
async def admin_delete_workflow_file(filename: str, request: Request):
    """Admin: Workflow 루트 디렉토리의 단일 파일 삭제 (세션 미존재 레거시 파일용)."""
    _require_admin(request)
    safe_name = Path(filename).name
    file_path = _WF_DIR / safe_name
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(404, f"파일을 찾을 수 없습니다: {safe_name}")
    file_path.unlink()
    return {"ok": True, "deleted": safe_name}


@app.get("/api/admin/projects", tags=["Admin"])
async def admin_list_projects(request: Request):
    """Admin: 이전 프로젝트 데이터 폴더 목록 조회."""
    _require_admin(request)
    return {"ok": True, "projects": list_projects()}


@app.delete("/api/admin/projects/{dirname}", tags=["Admin"])
async def admin_delete_project(dirname: str, request: Request):
    """Admin: 이전 프로젝트 데이터 폴더 삭제."""
    _require_admin(request)
    if not dirname:
        raise HTTPException(400, "디렉토리명이 필요합니다.")
    ok = delete_project(dirname)
    if not ok:
        raise HTTPException(404, "프로젝트를 찾을 수 없습니다.")
    return {"ok": True, "deleted": dirname}


@app.delete("/api/admin/workflow-reset", tags=["Admin"])
async def admin_reset_all_workflow(request: Request):
    """Admin: 전체 초기화 — Workflow 세션·파일 + Task 분류 엑셀 + 메모리 상태 삭제."""
    import shutil
    _require_admin(request)

    deleted = []
    errors = []

    # 1) 세션 디렉토리 전체 삭제
    if _SESSIONS_DIR.exists():
        try:
            shutil.rmtree(_SESSIONS_DIR)
            _SESSIONS_DIR.mkdir(exist_ok=True)
            deleted.append("sessions/")
        except Exception as e:
            errors.append(f"sessions/: {e}")

    # 2) _WF_DIR 루트의 파일들 삭제 (xlsx, json, pptx 등)
    for f in _WF_DIR.iterdir():
        if f.is_file():
            try:
                f.unlink()
                deleted.append(f.name)
            except Exception as e:
                errors.append(f"{f.name}: {e}")

    # 3) Task 분류용 엑셀 파일 삭제 (_UPLOAD_DIR)
    if _UPLOAD_DIR.exists():
        for f in _UPLOAD_DIR.iterdir():
            if f.is_file():
                try:
                    f.unlink()
                    deleted.append(f"uploads/{f.name}")
                except Exception as e:
                    errors.append(f"uploads/{f.name}: {e}")

    # 4) New Workflow 상태 파일 삭제 (/app/persist/new_workflow/)
    for f in _NW_DIR.iterdir() if _NW_DIR.exists() else []:
        if f.is_file():
            try:
                f.unlink()
                deleted.append(f"new_workflow/{f.name}")
            except Exception as e:
                errors.append(f"new_workflow/{f.name}: {e}")

    # 5) data_store 프로젝트 디렉토리 전체 삭제 (/app/persist/data/)
    from data_store import _BASE_DIR as _DS_BASE, _CURRENT_FILE as _DS_CURRENT
    if _DS_BASE.exists():
        try:
            shutil.rmtree(_DS_BASE)
            _DS_BASE.mkdir(exist_ok=True)
            deleted.append("persist/data/")
        except Exception as e:
            errors.append(f"persist/data/: {e}")
    if _DS_CURRENT.exists():
        try:
            _DS_CURRENT.unlink()
            deleted.append("current_project.json")
        except Exception as e:
            errors.append(f"current_project.json: {e}")

    # 6) 메모리 상태 전체 초기화 (Workflow + Task + New Workflow)
    global _workflow_cache, _wf_excel_tasks, _wf_excel_path
    global _wf_classification, _wf_chat_history
    global _wf_step1_cache, _wf_step2_cache, _wf_benchmark_table, _wf_gap_analysis, _wf_user_resources
    global _current_session_id, _sessions_manifest
    global _tasks_cache, _current_excel_path
    global _new_workflow_cache, _project_definition_cache, _project_design_cache
    global _nw_tasks_cache, _nw_projects_cache
    _workflow_cache = {}
    _wf_excel_tasks = []
    _wf_excel_path = ""
    _wf_classification = {}
    _wf_chat_history = []
    _wf_step1_cache = {}
    _wf_step2_cache = {}
    _wf_benchmark_table = {}
    _wf_gap_analysis = {}
    _wf_user_resources = []
    _current_session_id = ""
    _sessions_manifest = {}
    _tasks_cache = []
    _current_excel_path = None
    _new_workflow_cache.clear()
    _project_definition_cache.clear()
    _project_design_cache.clear()
    _nw_tasks_cache.clear()
    _nw_projects_cache.clear()

    return {"ok": True, "deleted": deleted, "errors": errors}


@app.get("/api/admin/download/{filename}", tags=["Admin"])
async def admin_download_file(
    filename: str,
    request: Request,
    session_id: str = Query(default=""),
):
    """업로드된 파일 다운로드 (Task 분류 엑셀 또는 Workflow 파일).
    session_id 쿼리 파라미터가 있으면 해당 세션 디렉토리에서 먼저 찾는다.
    없으면 uploads → workflow 세션 전체 검색 → workflow 루트 순서로 탐색.
    """
    _require_admin(request)
    # 경로 순회 방지
    import re as _re
    safe_name = Path(filename).name

    def _make_response(fp: Path) -> StreamingResponse:
        from urllib.parse import quote
        encoded = quote(safe_name)
        return StreamingResponse(
            open(fp, "rb"),
            media_type="application/octet-stream",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
        )

    # 1) Task 분류 엑셀
    fp = _UPLOAD_DIR / safe_name
    if fp.is_file():
        return _make_response(fp)

    # 2) session_id 지정 시 해당 세션 디렉토리 직접 확인
    if session_id:
        safe_sid = _re.sub(r'[^\w가-힣\-]', '_', session_id)[:80]
        fp = _SESSIONS_DIR / safe_sid / safe_name
        if fp.is_file():
            return _make_response(fp)

    # 3) 세션 디렉토리 전체 검색 (최신 수정 파일 우선)
    if _SESSIONS_DIR.exists():
        candidates = sorted(
            _SESSIONS_DIR.rglob(safe_name),
            key=lambda f: f.stat().st_mtime,
            reverse=True,
        )
        if candidates:
            return _make_response(candidates[0])

    # 4) New Workflow 디렉토리
    fp = _NW_DIR / safe_name
    if fp.is_file():
        return _make_response(fp)

    # 5) WF 루트 (레거시)
    fp = _WF_DIR / safe_name
    if fp.is_file():
        return _make_response(fp)

    raise HTTPException(404, f"파일을 찾을 수 없습니다: {safe_name}")


# ─────────────────────────────────────────────────────────────────────────────
# 서버 시작 시 persist 파일에서 캐시 자동 복구
# ─────────────────────────────────────────────────────────────────────────────

@app.on_event("startup")
async def _restore_from_persist():
    """서버 재시작 후 persist 파일에서 캐시 자동 복구 (비동기 — 이벤트 루프 블로킹 없음)."""
    import asyncio
    asyncio.create_task(_restore_task())


async def _restore_task():
    """실제 복구 작업 — 블로킹 I/O를 스레드풀에서 실행."""
    import asyncio
    global _wf_excel_tasks, _wf_classification, _tasks_cache, _current_excel_path
    global _current_session_id
    _load_manual_matches()

    from workflow_parser import parse_workflow_json, get_workflow_summary
    from excel_reader import load_tasks

    # 0) 세션 manifest 로드 후 마지막 사용 세션 복구 (있으면 세션 기반으로 복구)
    _load_sessions_manifest()
    last_sid = _sessions_manifest.get("_current", "")
    if last_sid and last_sid in _sessions_manifest:
        try:
            ok = await asyncio.to_thread(_load_session_data, last_sid)
            if ok:
                _current_session_id = last_sid
                print(f"[STARTUP] 세션 '{last_sid}' 복구 완료", flush=True)

                # Task 분류 엑셀 복구 (별도 경로)
                if not _tasks_cache:
                    task_excels = sorted(_UPLOAD_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime)
                    if task_excels:
                        try:
                            _tasks_cache = await asyncio.to_thread(load_tasks, str(task_excels[-1]))
                            _current_excel_path = task_excels[-1]
                        except Exception:
                            pass
                return  # 세션 복구 성공 → 이하 레거시 복구 불필요
        except Exception as e:
            print(f"[STARTUP] 세션 복구 실패({last_sid}): {e}", flush=True)

    # Task 분류 엑셀 복구 (세션과 무관하게 별도 경로)
    if not _tasks_cache:
        task_excels = sorted(_UPLOAD_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime)
        if task_excels:
            try:
                _tasks_cache = await asyncio.to_thread(load_tasks, str(task_excels[-1]))
                _current_excel_path = task_excels[-1]
                print(f"[STARTUP] Task Excel 복구: {task_excels[-1].name}", flush=True)
            except Exception as e:
                print(f"[STARTUP] Task Excel 복구 실패: {e}", flush=True)
