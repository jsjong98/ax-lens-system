"""
project_excel_reader.py — New Workflow용 과제 엑셀 파서

2행 병합 헤더 구조:
  Row 1: 과제번호 | 이름 | 과제개요 | 현황 및 Pain Point vs. 개선 방향 | 과제 수준 | 고려사항 | 기대효과 | 활용 Data/System
  Row 2:          |      |         | As-Is | Pain-Point | Needs | To-Be |          |          | 정량적 | 정성적 | Input내부 | Input외부 | Output

데이터는 Row 3부터 시작.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl


def _cell_value(ws: Any, row: int, col: int) -> str:
    """셀 값을 문자열로 반환. 병합 셀도 처리."""
    cell = ws.cell(row=row, column=col)
    val = cell.value
    if val is None:
        # 병합 셀이면 병합 범위의 첫 번째 셀 값 사용
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                break
    return str(val).strip() if val else ""


def _detect_columns(ws: Any) -> dict[str, int] | None:
    """
    1~2행에서 컬럼 매핑을 자동 감지합니다.
    키워드 기반으로 각 컬럼의 위치를 찾습니다.
    """
    mapping: dict[str, int] = {}

    # Row 1, 2의 모든 셀 값을 수집
    headers_r1: dict[int, str] = {}
    headers_r2: dict[int, str] = {}

    for col in range(1, ws.max_column + 1):
        v1 = _cell_value(ws, 1, col).lower().replace(" ", "")
        v2 = _cell_value(ws, 2, col).lower().replace(" ", "")
        headers_r1[col] = v1
        headers_r2[col] = v2

    # 디버그: 헤더 내용 출력
    print(f"[project_excel] Row 1 headers: { {c: v for c, v in headers_r1.items() if v} }")
    print(f"[project_excel] Row 2 headers: { {c: v for c, v in headers_r2.items() if v} }")

    # 키워드 매핑 (row1 또는 row2에서 찾기)
    keyword_map = {
        "project_no":       ["과제번호", "no", "번호"],
        "name":             ["이름", "과제명"],
        "overview":         ["과제개요", "주요과제내용", "주요과제", "개요"],
        "as_is":            ["업무현황", "as-is", "asis", "현황"],
        "pain_point":       ["pain-point", "painpoint", "pain_point", "pain"],
        "needs":            ["needs", "니즈"],
        "to_be":            ["개선모습", "to-be", "tobe", "개선방향", "개선"],
        "level":            ["과제수준", "수준", "level"],
        "considerations":   ["고려사항", "추진시고려", "고려"],
        "effect_quant":     ["정량적효과", "정량적"],
        "effect_qual":      ["정성적효과", "정성적"],
        "input_internal":   ["input(내부)", "input내부", "내부"],
        "input_external":   ["input(외부)", "input외부", "외부"],
        "output":           ["output", "산출물"],
    }

    # Row 2를 우선 검색 (세부 헤더), 못 찾으면 Row 1에서 검색
    for key, keywords in keyword_map.items():
        # 먼저 Row 2에서 찾기
        for col in sorted(headers_r2.keys()):
            val = headers_r2[col]
            for kw in keywords:
                if kw in val:
                    if key not in mapping:
                        mapping[key] = col
                    break
            if key in mapping:
                break
        # Row 2에서 못 찾으면 Row 1에서 (단, row1==row2인 병합셀만)
        if key not in mapping:
            for col in sorted(headers_r1.keys()):
                if headers_r1[col] == headers_r2[col]:  # 병합 셀 (row1==row2)
                    val = headers_r1[col]
                    for kw in keywords:
                        if kw in val:
                            mapping[key] = col
                            break
                if key in mapping:
                    break

    print(f"[project_excel] Column mapping result: {mapping}")

    # 최소한 이름 또는 overview가 있어야 유효
    if "name" not in mapping and "overview" not in mapping:
        return None

    return mapping


def parse_project_excel(filepath: str | Path, sheet_name: str | None = None) -> list[dict]:
    """
    과제 엑셀 파일을 파싱하여 프로젝트 목록을 반환합니다.

    Returns:
        [
            {
                "project_no": "35",
                "name": "의료비 판독 Agent",
                "overview": "...",
                "as_is": "...",
                "pain_point": "...",
                "needs": "...",
                "to_be": "...",
                "level": "Level 3",
                "considerations": "...",
                "effect_quant": "...",
                "effect_qual": "...",
                "input_internal": "...",
                "input_external": "...",
                "output": "...",
            },
            ...
        ]
    """
    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # 컬럼 매핑 감지
    mapping = _detect_columns(ws)
    if not mapping:
        wb.close()
        raise ValueError("과제 엑셀 형식을 인식할 수 없습니다. 헤더를 확인해 주세요.")

    # 데이터 행 파싱 (3행부터)
    projects = []
    for row in range(3, ws.max_row + 1):
        # 빈 행 건너뛰기
        name_col = mapping.get("name", mapping.get("overview", 1))
        name_val = _cell_value(ws, row, name_col)
        if not name_val:
            continue

        project: dict[str, str] = {}
        for key, col in mapping.items():
            project[key] = _cell_value(ws, row, col)

        # 최소 데이터 검증
        if project.get("name") or project.get("overview"):
            projects.append(project)

    wb.close()
    return projects


def projects_to_freeform_params(projects: list[dict]) -> dict:
    """
    파싱된 프로젝트 목록을 LLM 자유형식 입력 파라미터로 변환합니다.
    여러 프로젝트를 하나의 프롬프트로 통합합니다.
    """
    if not projects:
        return {"process_name": "", "inputs": "", "outputs": "", "systems": "", "pain_points": "", "additional_info": ""}

    # 프로세스명: 첫 프로젝트 이름 또는 전체 요약
    if len(projects) == 1:
        process_name = projects[0].get("name", "")
    else:
        names = [p.get("name", "") for p in projects if p.get("name")]
        process_name = f"{names[0]} 외 {len(names)-1}개 과제" if len(names) > 1 else names[0] if names else ""

    # 각 필드 통합
    all_pain_points = []
    all_inputs = []
    all_outputs = []
    all_additional = []

    for i, p in enumerate(projects, 1):
        prefix = f"[과제 {p.get('project_no', i)}] {p.get('name', '')}"

        if p.get("pain_point"):
            all_pain_points.append(f"{prefix}: {p['pain_point']}")

        if p.get("input_internal") or p.get("input_external"):
            inputs = []
            if p.get("input_internal"):
                inputs.append(f"내부: {p['input_internal']}")
            if p.get("input_external"):
                inputs.append(f"외부: {p['input_external']}")
            all_inputs.append(f"{prefix}: {', '.join(inputs)}")

        if p.get("output"):
            all_outputs.append(f"{prefix}: {p['output']}")

        # 추가 정보 통합 (As-Is, Needs, To-Be, 고려사항, 기대효과)
        details = []
        if p.get("overview"):
            details.append(f"과제개요: {p['overview']}")
        if p.get("as_is"):
            details.append(f"현황(As-Is): {p['as_is']}")
        if p.get("needs"):
            details.append(f"Needs: {p['needs']}")
        if p.get("to_be"):
            details.append(f"개선방향(To-Be): {p['to_be']}")
        if p.get("considerations"):
            details.append(f"고려사항: {p['considerations']}")
        if p.get("effect_quant"):
            details.append(f"정량적 효과: {p['effect_quant']}")
        if p.get("effect_qual"):
            details.append(f"정성적 효과: {p['effect_qual']}")
        if p.get("level"):
            details.append(f"과제 수준: {p['level']}")

        if details:
            all_additional.append(f"{prefix}\n  " + "\n  ".join(details))

    return {
        "process_name": process_name,
        "inputs": "\n".join(all_inputs),
        "outputs": "\n".join(all_outputs),
        "systems": "",
        "pain_points": "\n".join(all_pain_points),
        "additional_info": "\n\n".join(all_additional),
    }
